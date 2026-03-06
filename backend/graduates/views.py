"""
graduates/views.py
──────────────────
Vistas públicas y administrativas para el módulo de egresados.

Endpoints:
  PUBLIC:
    GET  /api/public/graduates/search/            → búsqueda pública
    GET  /api/public/graduates/<pk>/constancia/    → PDF constancia
  ADMIN CRUD:
    GET    /api/graduates/                         → listar (paginado, filtros)
    POST   /api/graduates/                         → crear
    GET    /api/graduates/<pk>/                     → detalle
    PUT    /api/graduates/<pk>/                     → actualizar
    PATCH  /api/graduates/<pk>/                     → actualizar parcial
    DELETE /api/graduates/<pk>/                     → soft delete
    GET    /api/graduates/stats/                    → estadísticas
    GET    /api/graduates/export/xlsx/              → exportar Excel
  CATÁLOGO:
    GET/POST   /api/graduates/grado-titulo-types/          → listar/crear
    GET/PUT/PATCH/DELETE /api/graduates/grado-titulo-types/<pk>/  → detalle
"""
import datetime
import os
from io import BytesIO

from django.conf import settings
from django.db.models import Q, Count
from django.http import HttpResponse
from rest_framework import generics, viewsets, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Graduate, GradoTituloType
from .serializers import (
    GraduatePublicSerializer,
    GraduateAdminSerializer,
    GradoTituloTypeSerializer,
)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGINATION
# ═══════════════════════════════════════════════════════════════════════════════

class GraduatePagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = "page_size"
    max_page_size = 100


# ═══════════════════════════════════════════════════════════════════════════════
# 1. BÚSQUEDA PÚBLICA — Solo egresados que YA SUSTENTARON (tienen título)
# ═══════════════════════════════════════════════════════════════════════════════


class GraduateSearchView(APIView):
    """
    GET /api/graduates/search/?dni=12345678
    GET /api/graduates/search/?nombre=ARAUJO MENDOZA

    Búsqueda pública de egresados titulados.
    Solo retorna egresados con fecha_sustentacion (ya tienen título).
    """
    permission_classes = [AllowAny]

    def get(self, request):
        dni = request.query_params.get("dni", "").strip()
        nombre = request.query_params.get("nombre", "").strip()

        if not dni and not nombre:
            return Response(
                {"detail": "Debe proporcionar 'dni' o 'nombre' como parámetro de búsqueda."},
                status=400,
            )

        # ★ FILTRO CLAVE: solo egresados que YA SUSTENTARON ★
        qs = Graduate.objects.filter(
            is_active=True,
            fecha_sustentacion__isnull=False,
        )

        if dni:
            if not dni.isdigit() or len(dni) != 8:
                return Response(
                    {"detail": "El DNI debe tener exactamente 8 dígitos numéricos."},
                    status=400,
                )
            qs = qs.filter(dni=dni)
        else:
            terms = nombre.upper().split()
            for term in terms:
                if len(term) >= 2:
                    qs = qs.filter(apellidos_nombres__icontains=term)

        qs = qs.select_related("grado_titulo_type").order_by("apellidos_nombres")[:20]
        serializer = GraduatePublicSerializer(qs, many=True)
        return Response({"results": serializer.data})


# ═══════════════════════════════════════════════════════════════════════════════
# 2. ADMIN: CRUD ViewSet de egresados
# ═══════════════════════════════════════════════════════════════════════════════


class GraduateAdminViewSet(viewsets.ModelViewSet):
    """
    CRUD completo de egresados para el panel administrativo.
    Soporta búsqueda, filtros por especialidad/año/sustentación/diploma.
    Incluye endpoints custom: stats y export/xlsx.
    """
    serializer_class = GraduateAdminSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = GraduatePagination

    def get_queryset(self):
        qs = Graduate.objects.select_related("grado_titulo_type").filter(is_active=True)

        # Filtro por búsqueda general (DNI o nombre)
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(dni__icontains=search) |
                Q(apellidos_nombres__icontains=search)
            )

        # Filtro por especialidad
        especialidad = self.request.query_params.get("especialidad", "").strip()
        if especialidad:
            qs = qs.filter(especialidad__iexact=especialidad)

        # Filtro por año de egreso
        anio_egreso = self.request.query_params.get("anio_egreso", "").strip()
        if anio_egreso:
            qs = qs.filter(anio_egreso=anio_egreso)

        # Filtro por presencia de sustentación
        has_sustentacion = self.request.query_params.get("has_sustentacion", "").strip()
        if has_sustentacion == "true":
            qs = qs.filter(fecha_sustentacion__isnull=False)
        elif has_sustentacion == "false":
            qs = qs.filter(fecha_sustentacion__isnull=True)

        # Filtro por presencia de diploma
        has_diploma = self.request.query_params.get("has_diploma", "").strip()
        if has_diploma == "true":
            qs = qs.exclude(codigo_diploma="")
        elif has_diploma == "false":
            qs = qs.filter(codigo_diploma="")

        return qs.order_by("apellidos_nombres")

    def perform_destroy(self, instance):
        """Soft delete: desactivar en vez de eliminar físicamente."""
        instance.is_active = False
        instance.save(update_fields=["is_active"])

    # ── Stats ──────────────────────────────────────────────────────────────
    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Estadísticas generales de egresados activos."""
        qs = Graduate.objects.filter(is_active=True)
        total = qs.count()
        with_sustentacion = qs.filter(fecha_sustentacion__isnull=False).count()
        with_diploma = qs.exclude(codigo_diploma="").count()
        with_resolucion = qs.exclude(resolucion_acta="").count()
        with_dni = qs.exclude(dni="").count()
        with_registro_ped = qs.exclude(registro_pedagogico="").count()

        # Por especialidad
        by_especialidad = list(
            qs.values("especialidad")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Por año de egreso (últimos 10)
        by_anio = list(
            qs.values("anio_egreso")
            .annotate(count=Count("id"))
            .order_by("-anio_egreso")[:10]
        )

        return Response({
            "total": total,
            "total_graduates": total,
            "with_sustentacion": with_sustentacion,
            "with_diploma": with_diploma,
            "with_resolucion": with_resolucion,
            "with_dni": with_dni,
            "with_registro_pedagogico": with_registro_ped,
            "by_especialidad": by_especialidad,
            "by_anio_egreso": by_anio,
        })

    # ── Export Excel ───────────────────────────────────────────────────────
    @action(detail=False, methods=["get"], url_path="export/xlsx")
    def export_xlsx(self, request):
        """Exportar todos los egresados (filtrados) a Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {"detail": "openpyxl no está instalado"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        qs = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "Egresados"

        # ── Headers ──
        headers = [
            "N°", "DNI", "Apellidos y Nombres", "Grado/Título",
            "Especialidad", "Nivel", "Año Ingreso", "Año Egreso",
            "Fecha Sustentación", "Resolución/Acta", "Código Diploma",
            "Registro Pedagógico", "Director General", "Secretario Académico",
        ]

        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ── Data ──
        for row_idx, grad in enumerate(qs.iterator(), start=2):
            vals = [
                row_idx - 1,
                grad.dni,
                grad.apellidos_nombres,
                grad.grado_titulo_display,
                grad.especialidad,
                grad.nivel,
                grad.anio_ingreso,
                grad.anio_egreso,
                grad.fecha_sustentacion.strftime("%d/%m/%Y") if grad.fecha_sustentacion else "",
                grad.resolucion_acta,
                grad.codigo_diploma,
                grad.registro_pedagogico,
                grad.director_general,
                grad.secretario_academico,
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border

        # ── Column widths ──
        col_letters = [
            "A", "B", "C", "D", "E", "F", "G", "H",
            "I", "J", "K", "L", "M", "N",
        ]
        widths = [5, 10, 35, 35, 25, 15, 12, 12, 14, 25, 12, 25, 30, 30]
        for letter, w in zip(col_letters, widths):
            ws.column_dimensions[letter].width = w

        ws.freeze_panes = "A2"

        # ── Write ──
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="egresados.xlsx"'
        return response


# ═══════════════════════════════════════════════════════════════════════════════
# 3. CRUD DE GRADO/TÍTULO TYPES (Admin)
# ═══════════════════════════════════════════════════════════════════════════════


class GradoTituloTypeListCreateView(generics.ListCreateAPIView):
    """GET: listar todos los tipos | POST: crear nuevo tipo."""
    permission_classes = [IsAuthenticated]
    serializer_class = GradoTituloTypeSerializer
    queryset = GradoTituloType.objects.all()

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        serializer = self.get_serializer(qs, many=True)
        return Response({"items": serializer.data})


class GradoTituloTypeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE un tipo de grado/título."""
    permission_classes = [IsAuthenticated]
    serializer_class = GradoTituloTypeSerializer
    queryset = GradoTituloType.objects.all()


# ═══════════════════════════════════════════════════════════════════════════════
# 4. GENERACIÓN DE CONSTANCIA PDF
# ═══════════════════════════════════════════════════════════════════════════════

MESES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

DEFAULT_DIRECTOR = "MG. MARIA ELVIRA GARCIA PORRAS"
DEFAULT_SECRETARIO = "MG. CASTRO MENDOZA, NEISY MARLENI"


def _resolve_media_path(url_or_path):
    if not url_or_path:
        return None
    url_or_path = str(url_or_path).strip()
    if not url_or_path:
        return None
    if os.path.isabs(url_or_path) and os.path.isfile(url_or_path):
        return url_or_path

    path = url_or_path
    for prefix in ("http://", "https://"):
        if path.lower().startswith(prefix):
            path = path[len(prefix):]
            idx = path.find("/")
            path = path[idx:] if idx >= 0 else path
            break

    media_url = getattr(settings, "MEDIA_URL", "/media/")
    if path.startswith(media_url):
        path = path[len(media_url):]
    elif path.startswith("/media/"):
        path = path[7:]
    path = path.lstrip("/")
    if not path:
        return None

    media_root = getattr(settings, "MEDIA_ROOT", "")
    if media_root:
        c = os.path.join(media_root, path)
        if os.path.isfile(c):
            return c
    c = os.path.join(settings.BASE_DIR, "media", path)
    if os.path.isfile(c):
        return c
    return None


def _load_institution_data():
    try:
        from catalogs.models import InstitutionSetting
        obj = InstitutionSetting.objects.filter(pk=1).first()
        if obj and isinstance(getattr(obj, "data", None), dict):
            return obj.data
    except (ImportError, Exception):
        pass
    try:
        from catalogs.models import InstitutionSettings
        obj = InstitutionSettings.objects.first()
        if obj:
            if hasattr(obj, "data") and isinstance(obj.data, dict):
                return obj.data
            return {
                "logo_url": getattr(obj, "logo_url", ""),
                "responsible_signature_url": getattr(obj, "responsible_signature_url", ""),
                "name": getattr(obj, "name", ""),
            }
    except (ImportError, Exception):
        pass
    return {}


class GraduateConstanciaView(APIView):
    """
    GET /api/graduates/<pk>/constancia/
    Genera y descarga la Constancia de Inscripción en PDF.
    """
    permission_classes = [AllowAny]

    def get(self, request, pk):
        try:
            grad = Graduate.objects.select_related("grado_titulo_type").get(
                pk=pk, is_active=True
            )
        except Graduate.DoesNotExist:
            return Response({"detail": "Egresado no encontrado."}, status=404)

        pdf_bytes = self._generate_pdf(grad)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        filename = f"Constancia_Inscripcion_{grad.dni or grad.id}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _generate_pdf(self, grad):
        from reportlab.lib.colors import HexColor
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=2.2 * cm, rightMargin=2.2 * cm,
            topMargin=1.0 * cm, bottomMargin=1.0 * cm,
        )

        styles = getSampleStyleSheet()
        inst = _load_institution_data()

        DK = HexColor("#0c2340")
        GY = HexColor("#333333")
        LT = HexColor("#555555")
        MU = HexColor("#999999")
        LN = HexColor("#cccccc")
        W = 16.6 * cm

        s_title = ParagraphStyle("T", parent=styles["Title"], fontSize=13, fontName="Helvetica-Bold", textColor=DK, alignment=TA_CENTER, spaceBefore=0, spaceAfter=4, leading=16)
        s_intro = ParagraphStyle("I", parent=styles["Normal"], fontSize=8, alignment=TA_JUSTIFY, textColor=GY, leading=11, spaceBefore=0, spaceAfter=2)
        s_sect = ParagraphStyle("S", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold", textColor=DK, spaceBefore=6, spaceAfter=2)
        s_lbl = ParagraphStyle("L", parent=styles["Normal"], fontSize=7.5, textColor=LT, leading=9)
        s_val = ParagraphStyle("V", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold", textColor=HexColor("#111111"), leading=9)
        s_hdr = ParagraphStyle("H", parent=styles["Normal"], fontSize=6.5, alignment=TA_CENTER, textColor=HexColor("#666666"), leading=8)
        s_date = ParagraphStyle("D", parent=styles["Normal"], fontSize=8, alignment=TA_CENTER, textColor=DK, fontName="Helvetica-Bold", leading=11)
        s_firma = ParagraphStyle("F", parent=styles["Normal"], fontSize=7, alignment=TA_CENTER, textColor=GY, leading=9)
        s_foot = ParagraphStyle("Fo", parent=styles["Normal"], fontSize=6, textColor=MU, alignment=TA_CENTER, leading=7.5)

        els = []

        def hline(color=DK, width=1.5):
            t = Table([[""]], colWidths=[W], rowHeights=[1], style=TableStyle([("LINEABOVE", (0, 0), (-1, 0), width, color)]))
            t.hAlign = "CENTER"
            return t

        def ftable(pairs):
            rows = []
            for lbl, val in pairs:
                if val:
                    rows.append([Paragraph(lbl, s_lbl), Paragraph(str(val), s_val)])
            if not rows:
                return
            t = Table(rows, colWidths=[4.8 * cm, W - 4.8 * cm], style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
            ]))
            els.append(t)

        # ── Logo ──
        logo_path = _resolve_media_path(inst.get("logo_url"))
        if not logo_path:
            for lp in [
                os.path.join(settings.BASE_DIR, "static", "img", "logo.png"),
                os.path.join(settings.BASE_DIR, "staticfiles", "img", "logo.png"),
            ]:
                if lp and os.path.exists(lp):
                    logo_path = lp
                    break

        if logo_path:
            img = Image(logo_path, width=2 * cm, height=2 * cm)
            img.hAlign = "CENTER"
            els.append(img)
            els.append(Spacer(1, 2))

        # ── Encabezado ──
        els.append(Paragraph(
            "INSTITUTO DE EDUCACIÓN SUPERIOR PEDAGÓGICO PÚBLICO<br/>"
            '<b>"GUSTAVO ALLENDE LLAVERÍA"</b><br/>'
            "Tarma — Junín — Perú", s_hdr))
        els.append(Spacer(1, 6))
        els.append(hline(DK, 1.5))
        els.append(Spacer(1, 6))

        # ── Título ──
        els.append(Paragraph(
            "CONSTANCIA DE INSCRIPCIÓN EN EL<br/>"
            "REGISTRO DE GRADOS Y TÍTULOS", s_title))

        # ── Introducción ──
        els.append(Paragraph(
            "La Dirección General del IESPP \"Gustavo Allende Llavería\", a través "
            "de la Secretaría Académica, deja constancia que la información contenida "
            "en este documento se encuentra inscrita en el Registro de Grados y Títulos "
            "administrado por la institución.", s_intro))

        grado_titulo = grad.grado_titulo_display
        diploma_label = grad.diploma_label

        # ── Ciudadano ──
        els.append(Paragraph("INFORMACIÓN DEL CIUDADANO", s_sect))
        ftable([
            ("Apellidos y Nombres", grad.apellidos_nombres),
            ("Tipo de Documento de Identidad", "DNI"),
            ("Número de Documento de Identidad", grad.dni or "Sin información"),
        ])

        # ── Institución ──
        els.append(Paragraph("INFORMACIÓN DE LA INSTITUCIÓN", s_sect))
        director = grad.director_general or inst.get("director_general", "") or DEFAULT_DIRECTOR
        secretario = grad.secretario_academico or inst.get("secretario_academico", "") or DEFAULT_SECRETARIO
        ftable([
            ("Nombre", 'IESPP "GUSTAVO ALLENDE LLAVERÍA"'),
            ("Director General", director),
            ("Secretario Académico", secretario),
        ])

        # ── Diploma ──
        els.append(Paragraph("INFORMACIÓN DEL DIPLOMA", s_sect))
        dip = [
            ("Grado / Título", diploma_label),
            ("Denominación", grado_titulo),
            ("Especialidad", grad.especialidad),
            ("Año de Ingreso", grad.anio_ingreso),
            ("Año de Egreso", grad.anio_egreso),
        ]
        if grad.fecha_sustentacion:
            dip.append(("Fecha de Sustentación", grad.fecha_sustentacion.strftime("%d/%m/%Y")))
        if grad.resolucion_acta:
            dip.append(("Resolución / Acta", grad.resolucion_acta))
        if grad.codigo_diploma:
            dip.append(("Diploma", grad.codigo_diploma))
        if grad.registro_pedagogico:
            dip.append(("Registro Pedagógico", grad.registro_pedagogico))
        ftable(dip)

        # ── Fecha emisión ──
        els.append(Spacer(1, 10))
        now = datetime.datetime.now()
        els.append(Paragraph(
            f"Fecha de emisión de la constancia:<br/>"
            f"{now.day} de {MESES[now.month]} de {now.year}", s_date))

        # ── Firma ──
        els.append(Spacer(1, 8))
        firma_path = _resolve_media_path(inst.get("responsible_signature_url"))
        firmante = grad.director_general or inst.get("director_general", "") or DEFAULT_DIRECTOR

        if firma_path:
            fi = Image(firma_path, width=3.5 * cm, height=1.5 * cm)
            fi.hAlign = "CENTER"
            els.append(fi)
            els.append(Spacer(1, 1))
            fl = Table([[""]], colWidths=[6.5 * cm], rowHeights=[1],
                       style=TableStyle([("LINEABOVE", (0, 0), (-1, 0), 0.6, GY)]))
            fl.hAlign = "CENTER"
            els.append(fl)
            els.append(Spacer(1, 1))
            els.append(Paragraph(
                f"<b>{firmante}</b><br/>DIRECTORA GENERAL<br/>"
                'IESPP "Gustavo Allende Llavería"', s_firma))
        else:
            els.append(Spacer(1, 12))
            els.append(Paragraph(
                "_" * 40 + "<br/><br/>"
                f"<b>{firmante}</b><br/>DIRECTORA GENERAL<br/>"
                'IESPP "Gustavo Allende Llavería"', s_firma))

        # ── Nota al pie ──
        els.append(Spacer(1, 10))
        els.append(hline(LN, 0.4))
        els.append(Spacer(1, 2))
        els.append(Paragraph(
            "Esta constancia puede ser verificada en el sistema académico del IESPP "
            '"Gustavo Allende Llavería".<br/>'
            "(*) El presente documento deja constancia únicamente del registro del "
            "Grado o Título que se señala.", s_foot))

        doc.build(els)
        buf.seek(0)
        return buf.getvalue()