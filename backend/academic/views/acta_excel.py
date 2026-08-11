"""
acta_excel.py — Registro por Excel estilo SIAGIE
─────────────────────────────────────────────────
Flujo (como SIAGIE → Evaluación → Registro Excel / Asistencia Mensual):

  NOTAS
    GET  /academic/sections/<id>/grades/plantilla
         → Excel de UNA sección, prellenado con los alumnos matriculados
    GET  /academic/acta-excel/notas/plantilla?plan_id&semester&period
         → Excel del CICLO completo: una hoja por curso (sección), cada una
           con sus alumnos matriculados
    POST /academic/acta-excel/notas/importar
         → carga cualquiera de los dos formatos (detecta SECCION_ID por hoja)
    POST /academic/sections/<id>/grades/importar   (compatibilidad)

  ASISTENCIA
    GET  /academic/sections/<id>/attendance/plantilla?month=YYYY-MM
    POST /academic/sections/<id>/attendance/importar

Los docentes registran directo en su panel; estos endpoints son para el
flujo administrativo (un docente solo puede usarlos sobre SUS secciones).
"""
import io
import re
from calendar import monthrange
from datetime import date as date_cls

from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from rest_framework_simplejwt.authentication import JWTAuthentication

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from academic.models import (
    Section, SectionGrades, AttendanceSession, AttendanceRow,
    EnrollmentItem,
)
from students.name_utils import (apellidos_de, clave_orden, nombre_oficial,
                                 nombres_de)
from .utils import ok

# Marcas de asistencia estilo SIAGIE ↔ estados internos
MARK_TO_STATUS = {
    "P": "PRESENT", "T": "LATE", "F": "ABSENT", "J": "EXCUSED", "0": "HOLIDAY",
    "PRESENT": "PRESENT", "LATE": "LATE", "ABSENT": "ABSENT", "EXCUSED": "EXCUSED",
    "PRESENTE": "PRESENT", "TARDANZA": "LATE", "FALTA": "ABSENT", "JUSTIFICADO": "EXCUSED",
    "FERIADO": "HOLIDAY", "HOLIDAY": "HOLIDAY",
}
STATUS_TO_MARK = {"PRESENT": "P", "LATE": "T", "ABSENT": "F", "EXCUSED": "J", "HOLIDAY": "0"}

GRADE_COLS = ["C1_LEVEL", "C1_REC", "C2_LEVEL", "C2_REC", "C3_LEVEL", "C3_REC"]

# ── Formato oficial "ACTA DE CALIFICACIÓN" (modelo institucional) ──
ACTA_LEVELS_ORD = ["PI", "I", "P", "L", "D"]
# columnas (1-index): bloques de niveles por competencia y recomendaciones
ACTA_BLOCKS = [
    {"levels_start": 4,  "rec": 9},    # Comp 1: D-H, rec I
    {"levels_start": 10, "rec": 15},   # Comp 2: J-N, rec O
    {"levels_start": 16, "rec": 21},   # Comp 3: P-T, rec U
]
ACTA_HEAD_ROW = 12      # fila donde inicia el encabezado de la tabla
ACTA_FIRST_STUDENT = 16 # primera fila de alumnos

_ROMANS = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X",
           "XI", "XII"]


def _roman(n):
    try:
        return _ROMANS[int(n)]
    except (TypeError, ValueError, IndexError):
        return str(n or "")

# Estilos compartidos
_HEAD_FILL = PatternFill("solid", start_color="1F4E79")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FINDE_FILL = PatternFill("solid", start_color="E7EBF0")


def _items_de_seccion(sec):
    """
    Ítems de matrícula (CONFIRMED, mismo período) que pertenecen a la sección.

    Devuelve (items_de_la_seccion, items_sin_seccion_ambiguos).

    Reglas — el respaldo antes era TODO O NADA (`if not base.exists()`), así que
    bastaba un alumno con sección asignada para que TODOS los que la tenían en
    NULL desaparecieran del acta sin ningún aviso:

      · ítem con section = esta sección          → siempre entra
      · ítem con section = NULL                  → entra si no hay ambigüedad,
        es decir si el curso tiene una sola sección en el período, o si nadie
        tiene sección asignada (comportamiento histórico, para no vaciar actas)
      · ítem con section = otra sección          → nunca entra (es de la otra)

    Cuando hay 2+ secciones y algunos ya están asignados, los de section NULL
    quedan fuera porque no se puede saber a cuál pertenecen: se devuelven como
    ambiguos para poder avisarlo (ver `academic auditar_datos`).
    """
    items = list(
        EnrollmentItem.objects
        .select_related("enrollment__student")
        .filter(
            plan_course_id=sec.plan_course_id,
            enrollment__status="CONFIRMED",
            enrollment__period=sec.period,
        )
    )
    de_esta = [i for i in items if i.section_id == sec.id]
    sin_sec = [i for i in items if i.section_id is None]
    de_otras = [i for i in items if i.section_id not in (None, sec.id)]

    n_secciones = (Section.objects
                   .filter(plan_course_id=sec.plan_course_id, period=sec.period)
                   .count())

    if n_secciones <= 1 or not (de_esta or de_otras):
        return de_esta + sin_sec, []
    return de_esta, sin_sec


def roster_sin_seccion(sec):
    """Matriculados del curso que el acta NO puede ubicar (section NULL con
    varias secciones en juego). Para avisar en pantalla / auditoría."""
    _base, ambiguos = _items_de_seccion(sec)
    return [i.enrollment.student for i in ambiguos]


def _roster(sec):
    """
    Alumnos matriculados (CONFIRMED) de la sección.
    Retorna lista de dicts {id, pk, dni, nombre, nombres, apellidos, estado}.
    `id` = clave usada en el acta (user_id si existe, si no el pk del Student).
    """
    base, _ambiguos = _items_de_seccion(sec)

    out, seen = [], set()
    for item in base:
        st = item.enrollment.student
        if st.id in seen:
            continue
        seen.add(st.id)
        out.append({
            "id": st.user_id if getattr(st, "user_id", None) else st.id,
            "pk": st.id,
            "dni": getattr(st, "num_documento", "") or "",
            # Formato oficial único: "APELLIDOS, NOMBRES" (students/name_utils.py)
            "nombre": nombre_oficial(st),
            "nombres": nombres_de(st),
            "apellidos": apellidos_de(st),
            "estado": (getattr(st, "estado_academico", "") or "").upper(),
            "estado_rd": getattr(st, "estado_rd", "") or "",
        })
    # Alfabético español: sin `clave_orden` la Ñ (U+00D1) cae después de la Z
    # y los "ÑAUPARI" quedaban al final de la nómina en vez de entre N y O.
    out.sort(key=lambda x: clave_orden(x["nombre"]))
    return out


def _institution_logo_paths():
    """(logo_institución, logo_sistema) como paths absolutos, o None.
    Prioriza los logos subidos en Configuración de Institución; si no,
    usa los empaquetados en templates/kardex/images."""
    import os
    from django.conf import settings as dj_settings
    inst_logo = sist_logo = None
    try:
        from catalogs.models import InstitutionSetting
        from .utils import _media_url_to_abs_path
        obj = InstitutionSetting.objects.filter(pk=1).first()
        data = (obj.data or {}) if obj else {}
        p = _media_url_to_abs_path((data.get("logo_url") or "").strip())
        if p and os.path.isfile(p):
            inst_logo = p
        p2 = _media_url_to_abs_path((data.get("second_logo_url") or "").strip())
        if p2 and os.path.isfile(p2):
            sist_logo = p2
    except Exception:
        pass
    base = os.path.join(str(dj_settings.BASE_DIR), "academic", "templates", "kardex", "images")
    if not inst_logo:
        p = os.path.join(base, "logo.png")
        inst_logo = p if os.path.isfile(p) else None
    if not sist_logo:
        p = os.path.join(base, "logo_SIST.png")
        sist_logo = p if os.path.isfile(p) else None
    return inst_logo, sist_logo


def _add_logo(ws, path, anchor, height=55):
    """Incrusta un logo (manteniendo proporción) anclado a una celda."""
    if not path:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(path)
        ratio = (img.width / img.height) if img.height else 1
        img.height = height
        img.width = int(height * ratio)
        ws.add_image(img, anchor)
    except Exception:
        pass


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _turno_de(secciones):
    """Turno según el HORARIO registrado de la(s) sección(es), por mayoría de
    bloques: antes de 13:00 = MAÑANA, 13:00-18:59 = TARDE, desde 19:00 = NOCHE.

    Definición de Secretaría: el turno de las actas "se jala de lo que se
    estableció en el horario" — antes salía "MAÑANA" fijo en todas (y el
    acta institucional real de 2025-II era de TARDE). Sin horario cargado se
    mantiene MAÑANA como antes.

    Acepta una Section o un iterable de Sections.
    """
    from academic.models import SectionScheduleSlot
    if isinstance(secciones, Section):
        secciones = [secciones]
    conteo = {"MAÑANA": 0, "TARDE": 0, "NOCHE": 0}
    for ini in SectionScheduleSlot.objects.filter(
            section__in=list(secciones)).values_list("start", flat=True):
        h = getattr(ini, "hour", None)
        if h is None:
            continue
        conteo["MAÑANA" if h < 13 else ("TARDE" if h < 19 else "NOCHE")] += 1
    if not any(conteo.values()):
        return "MAÑANA"
    return max(conteo, key=conteo.get)


def _section_header_info(sec):
    pc = sec.plan_course
    curso = (pc.display_name or (pc.course.name if pc.course_id else "")) if pc else ""
    codigo = (pc.display_code or (pc.course.code if pc.course_id else "")) if pc else ""
    docente = ""
    if sec.teacher and sec.teacher.user:
        u = sec.teacher.user
        docente = (getattr(u, "full_name", "") or u.username or "").strip()
    return curso, codigo, docente


def _sheet_title(codigo, label, idx):
    base = re.sub(r"[\[\]:*?/\\]", "-", f"{codigo or 'CURSO'} {label or 'A'}").strip()
    # idx siempre al final: garantiza títulos únicos (límite Excel: 31 chars)
    return f"{(base or 'HOJA')[:26]} {idx}"


# ══════════════════════════════════════════════════════════════
# NOTAS — construcción de hoja (compartida)
# ══════════════════════════════════════════════════════════════

def _write_grades_sheet(ws, sec):
    """
    Escribe la hoja en el formato oficial "ACTA DE CALIFICACIÓN":
      - Cabecera: programa, docente, curso, período, ciclo-sección, turno, modalidad
      - 3 competencias con niveles PI/I/P/L/D (se marca con X) + recomendación
      - C1-C3, escala, PROMEDIO FINAL y CALIFICACIÓN CUALITATIVA con fórmulas vivas
      - Resumen (matriculados/aprobados/desaprobados/promedio del aula)
    Retorna n° de alumnos.
    """
    curso, codigo, docente = _section_header_info(sec)
    bundle = SectionGrades.objects.filter(section=sec).first()
    existing = (bundle.grades or {}) if bundle else {}

    pc = sec.plan_course
    career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "").upper()
    ciclo = _roman(pc.semester if pc else "")

    ini = fin = ""
    try:
        from academic.models import AcademicPeriod
        per = AcademicPeriod.get_or_none(sec.period)
        if per:
            ini, fin = per.start.isoformat(), per.end.isoformat()
    except Exception:
        pass

    thin = _BORDER
    lab_font = Font(bold=True, size=9)
    val_font = Font(size=9)
    white = Font(color="FFFFFF", size=8)

    # ── Marcadores para el importador (texto blanco, no estorban) ──
    ws["A2"] = f"SECCION_ID: {sec.id}"
    ws["A2"].font = white
    ws["A3"] = "FORMATO: ACTA_OFICIAL_V1"
    ws["A3"].font = white

    # ── Membrete: logos + título (como el modelo oficial) ──
    inst_logo, sist_logo = _institution_logo_paths()
    _add_logo(ws, sist_logo, "C1", height=58)
    _add_logo(ws, inst_logo, "T1", height=58)
    ws.merge_cells(start_row=2, start_column=7, end_row=4, end_column=18)
    tcell = ws.cell(row=2, column=7, value="ACTA DE CALIFICACIÓN DEL CURSO O MÓDULO")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=21, end_row=1, end_column=27)
    icell = ws.cell(row=1, column=21,
                    value='INSTITUTO DE EDUCACIÓN SUPERIOR PEDAGÓGICO PÚBLICO')
    icell.font = Font(bold=True, size=7)
    icell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=22, end_row=2, end_column=27)
    i2 = ws.cell(row=2, column=22, value='"Gustavo Allende Llavería" — TARMA')
    i2.font = Font(italic=True, size=8)
    i2.alignment = Alignment(horizontal="center")

    # ── Tabla de conversión (AC1:AE5, como el modelo) ──
    conv = [(1, 1.9, "Previo al Inicio"), (2, 2.9, "Inicio"), (3, 3.9, "En proceso"),
            (4, 4.9, "Logrado"), (5, 5, "Destacado")]
    for i, (a, b, txt) in enumerate(conv, 1):
        ws.cell(row=i, column=29, value=a).font = Font(size=8)
        ws.cell(row=i, column=30, value=b).font = Font(size=8)
        ws.cell(row=i, column=31, value=txt).font = Font(size=8)

    # ── Cabecera institucional ──
    def _kv(row, lcol_a, lcol_b, label, vcol_a, vcol_b, value):
        ws.merge_cells(start_row=row, start_column=lcol_a, end_row=row, end_column=lcol_b)
        ws.merge_cells(start_row=row, start_column=vcol_a, end_row=row, end_column=vcol_b)
        lc = ws.cell(row=row, column=lcol_a, value=label)
        lc.font = lab_font
        vc = ws.cell(row=row, column=vcol_a, value=value)
        vc.font = val_font
        vc.border = Border(bottom=_THIN)

    _kv(6, 1, 2, "Programa de Estudios", 3, 10, career)
    _kv(6, 21, 24, "Periodo Académico", 25, 27, sec.period)
    _kv(8, 1, 2, "Docente Formador(a)", 3, 13, docente or "")
    ws.cell(row=8, column=15, value="INICIO").font = lab_font
    ws.cell(row=8, column=16, value=ini).font = val_font
    _kv(8, 21, 24, "Ciclo - Sección", 25, 27, f'{ciclo} - "{sec.label or "A"}"')
    _kv(9, 1, 2, "Curso / Módulo", 3, 13, f"{curso}" + (f" ({codigo})" if codigo else ""))
    ws.cell(row=9, column=15, value="FIN").font = lab_font
    ws.cell(row=9, column=16, value=fin).font = val_font
    _kv(9, 21, 24, "Turno", 25, 27, _turno_de(sec))
    _kv(10, 21, 24, "Modalidad de estudios", 25, 27, "PRESENCIAL")

    # ── Encabezado de tabla (filas 12-15) ──
    H = ACTA_HEAD_ROW
    head_fill = PatternFill("solid", start_color="D9E2F3")
    head_font = Font(bold=True, size=8)

    def _mh(r1, c1, r2, c2, text, fill=True):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=text)
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill:
            cell.fill = head_fill
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = thin

    _mh(H, 1, H + 3, 1, "N°")
    _mh(H, 2, H + 3, 2, "N° de\nMatrícula")
    _mh(H, 3, H + 3, 3, "Apellidos y Nombres del Estudiante")
    for i, blk in enumerate(ACTA_BLOCKS, 1):
        s = blk["levels_start"]
        _mh(H, s, H, s + 4, f"COMPETENCIA {i}:")
        _mh(H + 1, s, H + 1, s + 4, "CRITERIO DE EVALUACIÓN:")
        _mh(H + 2, s, H + 2, s + 4, "Nivel de Desempeño")
        for j, lv in enumerate(ACTA_LEVELS_ORD):
            _mh(H + 3, s + j, H + 3, s + j, lv)
        _mh(H, blk["rec"], H + 3, blk["rec"], "Recomendación\nComentario*")
    _mh(H, 22, H + 2, 24, "CONCLUSIÓN\nDESCRIPTIVA")
    for j, t in enumerate(["C1", "C2", "C3"]):
        _mh(H + 3, 22 + j, H + 3, 22 + j, t)
    _mh(H, 25, H + 3, 25, "CALIFI-CACIÓN\nPARA EL SISTEMA\nDE EDUCACIÓN\nSUPERIOR")
    _mh(H, 26, H + 3, 26, "PROMEDIO\nFINAL")
    _mh(H, 27, H + 3, 27, "CALIFICACIÓN\nCUALITATIVA")

    # ── Alumnos ──
    alumnos = _roster(sec)
    r0 = ACTA_FIRST_STUDENT
    for idx, st in enumerate(alumnos):
        r = r0 + idx
        prev = existing.get(str(st["id"]), {})
        ws.cell(row=r, column=1, value=idx + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=st["dni"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=st["nombre"]).font = Font(size=9)

        es_licencia = st.get("estado") == "LICENCIA"
        estado_txt = ""
        if st.get("estado"):
            estado_txt = st["estado"] + (f" (RD {st['estado_rd']})" if st.get("estado_rd") else "")

        for i, blk in enumerate(ACTA_BLOCKS, 1):
            s = blk["levels_start"]
            if es_licencia:
                # Alumno con licencia: se anota LICENCIA en el comentario y
                # no se registran niveles (fila bloqueada, como el modelo)
                ws.cell(row=r, column=blk["rec"], value=estado_txt or "LICENCIA")\
                  .font = Font(size=8, bold=True, color="B91C1C")
                continue
            lv = (prev.get(f"C{i}_LEVEL") or "").upper()
            if lv in ACTA_LEVELS_ORD:
                # Se prellena con la puntuación del nivel (como el modelo oficial)
                idx_lv = ACTA_LEVELS_ORD.index(lv)
                ws.cell(row=r, column=s + idx_lv, value=idx_lv + 1)\
                  .alignment = Alignment(horizontal="center")
            rec = prev.get(f"C{i}_REC") or ""
            if not rec and i == 1 and estado_txt:
                rec = estado_txt   # otros estados: comentario informativo
            if rec:
                ws.cell(row=r, column=blk["rec"], value=rec).font = Font(size=8)

        # Fórmulas: C1-C3 = puntuación del casillero marcado (acepta X o el número),
        # escala, promedio final (tabla oficial) y calificación cualitativa
        for j, blk in enumerate(ACTA_BLOCKS):
            s = blk["levels_start"]
            a, b = get_column_letter(s), get_column_letter(s + 4)
            ws.cell(row=r, column=22 + j, value=(
                f'=IF(COUNTA({a}{r}:{b}{r})=0,"",'
                f'SUMPRODUCT(({a}{r}:{b}{r}<>"")*{{1,2,3,4,5}}))'
            ))
        ws.cell(row=r, column=25, value=f'=IF(COUNT(V{r}:X{r})<3,"",ROUND(AVERAGE(V{r}:X{r}),1))')
        # Tabla oficial RVM 123-2022 pág. 28 (escala → vigesimal)
        ws.cell(row=r, column=26, value=(
            f'=IF(Y{r}="","",LOOKUP(Y{r},'
            "{1;1.2;1.4;1.6;1.8;2;2.2;2.4;2.6;2.8;3;3.3;3.6;3.8;4;4.2;4.4;4.6;4.8;5},"
            "{1;2;3;4;5;6;7;8;9;10;11;12;13;14;15;16;17;18;19;20}))"
        ))
        ws.cell(row=r, column=27, value=(
            f'=IF(Y{r}="","",IF(Y{r}>=5,"Destacado",IF(Y{r}>=4,"Logrado",'
            f'IF(Y{r}>=3,"En proceso",IF(Y{r}>=2,"Inicio","Previo al Inicio")))))'
        ))
        for c in range(1, 28):
            ws.cell(row=r, column=c).border = thin
        for c in (22, 23, 24, 25, 26, 27):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    last = r0 + len(alumnos) - 1
    if alumnos:
        # Cada columna de nivel acepta su puntuación (1-5) o una X
        for j in range(5):
            dv = DataValidation(
                type="list", formula1=f'"X,x,{j + 1}"', allow_blank=True,
                showErrorMessage=True, errorTitle="Marca inválida",
                error=(f"En esta columna marca con X o con su puntuación ({j + 1}). "
                       "Solo un casillero por competencia."),
            )
            ws.add_data_validation(dv)
            for blk in ACTA_BLOCKS:
                col = get_column_letter(blk["levels_start"] + j)
                dv.add(f"{col}{r0}:{col}{last}")

        # ── Resumen ──
        fr = last + 3
        resumen = [
            ("Matriculados", len(alumnos)),
            ("Aprobados", f'=COUNTIF(Z{r0}:Z{last},">=11")'),
            ("Desaprobados", f'=COUNTIF(Z{r0}:Z{last},"<11")'),
            ("Con Licencia", ""),
            ("Promedio del aula o sección", f'=IFERROR(ROUND(AVERAGE(Z{r0}:Z{last}),1),"")'),
        ]
        ws.merge_cells(start_row=fr, start_column=2, end_row=fr, end_column=3)
        ws.cell(row=fr, column=2, value="Resumen").font = Font(bold=True, size=9)
        ws.cell(row=fr, column=4, value="N°").font = Font(bold=True, size=9)
        for i, (lbl, val) in enumerate(resumen, 1):
            ws.merge_cells(start_row=fr + i, start_column=2, end_row=fr + i, end_column=3)
            ws.cell(row=fr + i, column=2, value=lbl).font = Font(size=9)
            c = ws.cell(row=fr + i, column=4, value=val)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center")
            ws.cell(row=fr + i, column=2).border = thin
            ws.cell(row=fr + i, column=3).border = thin
            ws.cell(row=fr + i, column=4).border = thin

        # Firmas — Anexo 3 RVM 123-2022: Director + Secretario + Docente
        # (antes solo firmaba el docente)
        fs = fr + len(resumen) + 3
        FIRMAS = [
            (2,  7,  "DIRECTOR(A) GENERAL",       "Firma, Post Firma y Sello"),
            (10, 15, "SECRETARIO(A) ACADÉMICO(A)", "Firma, Post Firma y Sello"),
            (18, 23, "DOCENTE FORMADOR(A)",        "Firma"),
        ]
        for c1, c2, cargo, sub in FIRMAS:
            ws.merge_cells(start_row=fs, start_column=c1, end_row=fs, end_column=c2)
            ws.cell(row=fs, column=c1, value="_______________________________")\
              .alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=fs + 1, start_column=c1, end_row=fs + 1, end_column=c2)
            fc = ws.cell(row=fs + 1, column=c1, value=cargo)
            fc.font = Font(bold=True, size=9)
            fc.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=fs + 2, start_column=c1, end_row=fs + 2, end_column=c2)
            sc = ws.cell(row=fs + 2, column=c1, value=sub)
            sc.font = Font(size=8)
            sc.alignment = Alignment(horizontal="center")

    # ── Anchos de columna ──
    widths = {1: 4, 2: 11, 3: 36, 9: 22, 15: 22, 21: 22,
              22: 4.5, 23: 4.5, 24: 4.5, 25: 9, 26: 9, 27: 13,
              29: 4, 30: 4, 31: 14}
    for blk in ACTA_BLOCKS:
        for j in range(5):
            widths[blk["levels_start"] + j] = 3.6
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"D{ACTA_FIRST_STUDENT}"

    return len(alumnos)


def _import_grades_ws(ws, request):
    """
    Importa una hoja (detecta SECCION_ID en las primeras filas).
    Retorna dict {seccion, importados, errores} — errores incluye los de
    permiso/estado, sin lanzar excepción.
    """
    from .teachers import (
        _grades_section_access_denied, _is_grades_admin,
        _check_grades_window, _normalize_acta_student_payload,
    )

    res = {"hoja": ws.title, "seccion": "", "importados": 0, "errores": []}

    sec_id = None
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=3):
        for cell in row:
            m = re.match(r"^SECCION_ID:\s*(\d+)$", str(cell.value or "").strip())
            if m:
                sec_id = int(m.group(1))
                break
        if sec_id:
            break
    if not sec_id:
        res["errores"].append("No se encontró la celda 'SECCION_ID' (usa la plantilla generada)")
        return res

    sec = Section.objects.filter(id=sec_id).select_related(
        "plan_course__course", "teacher__user"
    ).first()
    if not sec:
        res["errores"].append(f"Sección {sec_id} no existe")
        return res
    curso, codigo, _ = _section_header_info(sec)
    res["seccion"] = f"{curso} — {sec.label or 'A'}"

    if _grades_section_access_denied(request, sec):
        res["errores"].append("Sin permiso sobre esta sección")
        return res
    ok_win, _werr = _check_grades_window(sec, request.user)
    if not ok_win:
        res["errores"].append("Ventana de notas cerrada para este período")
        return res

    bundle, _ = SectionGrades.objects.get_or_create(section=sec)
    if bundle.submitted and not _is_grades_admin(request.user):
        res["errores"].append("El acta ya está cerrada (solo administradores)")
        return res

    # ¿Formato oficial (ACTA_OFICIAL_V1) o legado (columnas C1_LEVEL...)?
    es_oficial = False
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=3):
        for cell in row:
            if "ACTA_OFICIAL" in str(cell.value or "").upper():
                es_oficial = True
                break

    roster = _roster(sec)
    by_dni = {st["dni"]: st for st in roster if st["dni"]}
    grades = dict(bundle.grades or {})

    def _guardar_fila(nrow, dni, payload, vacio):
        st = by_dni.get(dni)
        if not st:
            res["errores"].append(f"Fila {nrow}: DNI {dni} no está matriculado")
            return
        if (st.get("estado") or "") == "LICENCIA":
            # Solo es error si intentaron MARCAR niveles (los comentarios
            # "LICENCIA" prellenados por la plantilla no cuentan)
            if any(k.endswith("_LEVEL") and payload.get(k) for k in payload):
                res["errores"].append(
                    f"{st['nombre']} (DNI {dni}): alumno con LICENCIA — bloqueado, no se puede calificar")
            return
        if vacio:
            return
        merged = dict(grades.get(str(st["id"]), {}))
        merged.update(payload)
        normalized, errs = _normalize_acta_student_payload(merged)
        if errs:
            res["errores"].append(f"{st['nombre']} (DNI {dni}): {'; '.join(errs)}")
            return
        grades[str(st["id"])] = normalized
        res["importados"] += 1

    def _dni_de(cell_value):
        if isinstance(cell_value, float) and cell_value.is_integer():
            return str(int(cell_value))
        return str(cell_value or "").strip()

    if es_oficial:
        # ── Formato oficial: X en columnas de nivel, recomendaciones en I/O/U ──
        for row in ws.iter_rows(min_row=ACTA_FIRST_STUDENT, max_col=27):
            nrow = row[0].row
            dni = _dni_de(row[1].value)
            # Solo filas de alumnos: el documento debe tener dígitos y sin espacios
            # (así se saltan las filas del Resumen / firma al final del acta)
            if not dni or " " in dni or not re.search(r"\d", dni):
                continue

            payload, vacio, err_marca = {}, True, False
            for i, blk in enumerate(ACTA_BLOCKS, 1):
                s = blk["levels_start"] - 1   # index 0-based en row
                marcas = []
                for j in range(5):
                    raw = row[s + j].value
                    if isinstance(raw, float) and raw.is_integer():
                        raw = int(raw)
                    v = str(raw or "").strip().upper()
                    # Vale: X, la letra del nivel, o su puntuación (1-5)
                    if v and (v == "X" or v == ACTA_LEVELS_ORD[j] or v == str(j + 1)):
                        marcas.append(j)
                    elif v:
                        res["errores"].append(
                            f"Fila {nrow} (DNI {dni}): valor '{v}' no válido en la columna "
                            f"{ACTA_LEVELS_ORD[j]} de Competencia {i} (usa X o {j + 1})")
                        err_marca = True
                if err_marca:
                    break
                if len(marcas) > 1:
                    res["errores"].append(
                        f"Fila {nrow} (DNI {dni}): más de un nivel marcado en Competencia {i}")
                    err_marca = True
                    break
                if marcas:
                    payload[f"C{i}_LEVEL"] = ACTA_LEVELS_ORD[marcas[0]]
                    vacio = False
                rec = str(row[blk["rec"] - 1].value or "").strip()
                if rec:
                    payload[f"C{i}_REC"] = rec
                    vacio = False
            if err_marca:
                continue
            _guardar_fila(nrow, dni, payload, vacio)
    else:
        # ── Formato legado: encabezados C1_LEVEL / C1_REC ... ──
        head_row, col_of = None, {}
        for row in ws.iter_rows(min_row=1, max_row=15):
            vals = [str(c.value or "").strip().upper() for c in row]
            if "DNI" in vals:
                head_row = row[0].row
                col_of = {v: i + 1 for i, v in enumerate(vals) if v}
                break
        if not head_row or "DNI" not in col_of:
            res["errores"].append("No se encontró la fila de encabezados (columna DNI)")
            return res

        for row in ws.iter_rows(min_row=head_row + 1):
            dni = _dni_de(row[col_of["DNI"] - 1].value)
            if not dni:
                continue
            payload, vacio = {}, True
            for key in GRADE_COLS:
                if key not in col_of:
                    continue
                v = row[col_of[key] - 1].value
                v = "" if v is None else str(v).strip()
                if key.endswith("_LEVEL"):
                    v = v.upper()
                if v:
                    vacio = False
                payload[key] = v
            _guardar_fila(row[0].row, dni, payload, vacio)

    if res["importados"]:
        bundle.grades = grades
        bundle.save()
    return res


# ══════════════════════════════════════════════════════════════
# NOTAS — vistas
# ══════════════════════════════════════════════════════════════

class SectionGradesTemplateView(APIView):
    """Plantilla de UNA sección."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id: int):
        from .teachers import _grades_section_access_denied

        sec = get_object_or_404(
            Section.objects.select_related("plan_course__course", "teacher__user"),
            id=section_id,
        )
        if err := _grades_section_access_denied(request, sec):
            return err

        wb = Workbook()
        ws = wb.active
        ws.title = "REGISTRO DE NOTAS"
        n = _write_grades_sheet(ws, sec)
        if n == 0:
            return Response(
                {"detail": (
                    f"La sección no tiene alumnos matriculados en {sec.period}. "
                    f"Matricula a los alumnos primero (Módulo Académico → Matrícula) "
                    f"o elige una sección de un período con matrículas."
                )},
                status=400,
            )

        _, codigo, _ = _section_header_info(sec)
        fname = f"NOTAS_{codigo or 'CURSO'}_{sec.label or 'A'}_{sec.period}.xlsx"
        return _xlsx_response(wb, fname)


class ActaNotasCicloTemplateView(APIView):
    """
    Plantilla del CICLO completo: una hoja por curso (sección) del
    plan + ciclo + período, cada una con sus alumnos matriculados.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .teachers import _is_grades_admin

        try:
            plan_id = int(request.query_params.get("plan_id"))
            semester = int(request.query_params.get("semester"))
        except (TypeError, ValueError):
            return Response({"detail": "plan_id y semester son requeridos"}, status=400)
        period = (request.query_params.get("period") or "").strip()
        if not period:
            return Response({"detail": "period es requerido (ej: 2026-I)"}, status=400)

        qs = (
            Section.objects
            .select_related("plan_course__course", "teacher__user")
            .filter(
                period=period,
                plan_course__plan_id=plan_id,
                plan_course__semester=semester,
            )
            .order_by("plan_course__display_code", "label", "id")
        )
        # Un docente solo recibe SUS secciones del ciclo
        if not _is_grades_admin(request.user):
            qs = qs.filter(teacher__user_id=request.user.id)

        secciones = list(qs)
        if not secciones:
            return Response(
                {"detail": f"No hay secciones del ciclo {semester} en {period}."},
                status=400,
            )

        wb = Workbook()
        wb.remove(wb.active)
        sin_alumnos = []
        for i, sec in enumerate(secciones, 1):
            _, codigo, _ = _section_header_info(sec)
            ws = wb.create_sheet(title=_sheet_title(codigo, sec.label, i))
            n = _write_grades_sheet(ws, sec)
            if n == 0:
                curso, _, _ = _section_header_info(sec)
                sin_alumnos.append(f"{curso} — {sec.label or 'A'}")
                wb.remove(ws)

        if not wb.sheetnames:
            return Response(
                {"detail": (
                    f"Ningún curso del ciclo {semester} tiene alumnos matriculados "
                    f"en {period}. Matricula a los alumnos primero."
                )},
                status=400,
            )

        fname = f"NOTAS_CICLO{semester}_{period}.xlsx"
        return _xlsx_response(wb, fname)


class ActaNotasImportView(APIView):
    """
    Importa un Excel de notas: recorre TODAS las hojas y detecta la sección
    de cada una por su celda SECCION_ID (sirve para plantilla de una sección
    o del ciclo completo).
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Adjunta el archivo Excel en el campo 'file'"}, status=400)

        try:
            wb = load_workbook(f, data_only=True)
        except Exception as exc:
            return Response({"detail": f"No se pudo leer el Excel: {exc}"}, status=400)

        resultados = [_import_grades_ws(ws, request) for ws in wb.worksheets]
        total = sum(r["importados"] for r in resultados)
        errores = [
            f"[{r['seccion'] or r['hoja']}] {e}"
            for r in resultados for e in r["errores"]
        ]
        hojas_ok = sum(1 for r in resultados if r["importados"])

        return ok(
            success=True,
            importados=total,
            hojas=len(resultados),
            errores=errores,
            resultados=resultados,
            message=(
                f"{total} nota(s) importada(s) en {hojas_ok} curso(s)"
                + (f" · {len(errores)} error(es)" if errores else "")
                + " · las actas siguen en borrador"
            ),
        )


class SectionGradesImportView(APIView):
    """Compatibilidad: importa sobre una sección (usa el mismo parser)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, section_id: int):
        return ActaNotasImportView().post(request)


# ══════════════════════════════════════════════════════════════
# ACTA DE EVALUACIÓN DE ÁREA (formato oficial por curso/docente)
# ══════════════════════════════════════════════════════════════

# Cualitativa oficial a partir del vigesimal (leyenda RVM 123-2022 pág. 28)
def _cualitativa_de_vigesimal(v):
    if v is None:
        return ""
    v = float(v)
    if v >= 20:
        return "Destacado"
    if v >= 15:
        return "Logrado"
    if v >= 11:
        return "En proceso"
    if v >= 6:
        return "Inicio"
    return "Previo al inicio"


# Bandas de la leyenda (para imprimirla en el acta)
_LEYENDA_BANDAS = [
    ("Previo al inicio", ["1 a 1.1", "1.2 a 1.3", "1.4 a 1.5", "1.6 a 1.7", "1.8 a 1.9"], [1, 2, 3, 4, 5]),
    ("Inicio",           ["2.0 a 2.1", "2.2 a 2.3", "2.4 a 2.5", "2.6 a 2.7", "2.8 a 2.9"], [6, 7, 8, 9, 10]),
    ("En proceso",       ["3.0 a 3.2", "3.3 a 3.5", "3.6 a 3.7", "3.8 a 3.9"], [11, 12, 13, 14]),
    ("Logrado",          ["4.0 a 4.1", "4.2 a 4.3", "4.4 a 4.5", "4.6 a 4.7", "4.8 a 4.9"], [15, 16, 17, 18, 19]),
    ("Destacado",        ["5.0"], [20]),
]


def _acta_area_inst():
    """Datos institucionales para el acta de área (InstitutionSetting.data
    con defaults del IESP)."""
    try:
        from catalogs.models import InstitutionSetting
        data = (InstitutionSetting.objects.filter(pk=1).first() or None)
        data = (data.data or {}) if data else {}
    except Exception:
        data = {}
    g = data.get
    return {
        "nombre": g("institution_name") or '"GUSTAVO ALLENDE  LLAVERIA"',
        "drej": g("drej") or "Junín",
        "uge": g("ugel") or "DREJ",
        "codigo_modular": g("codigo_modular") or "0609370",
        "denominacion": g("denominacion") or "IESP",
        "gestion": g("gestion") or "Público",
        "ds_creacion": g("ds_creacion") or "D.S. 059-1984-ED",
        "direccion": g("direccion") or ("Av. Hiroshi Takahashi Takahashi N° 162 "
                                        "Km. 4  Carretera Central – Pomachaca "
                                        "Tarma – Junín – Perú"),
        "provincia": g("provincia") or "Tarma",
        "distrito": g("distrito") or "Tarma",
        # R.D. de revalidación oficial (confirmada por Secretaría, 2026)
        "resolucion": (g("resolucion_autorizacion")
                       or "R.D. N° 306-2016-MINEDU/VMGP/DIGEDD/DIFOID"),
        "director": (g("director_name") or "GARCIA PORRAS, MARIA ELVIRA").upper(),
        "rd_encargatura": g("rd_encargatura") or "R.D.R. N°  017-2026-DREJ",
    }


def build_acta_area_workbook(sec):
    """Construye el Acta de Evaluación de Área de una sección.
    Retorna (Workbook, filename) o (None, detail_error) si no hay alumnos."""
    curso, codigo, docente = _section_header_info(sec)
    pc = sec.plan_course
    career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "").upper()
    creditos = getattr(pc, "credits", 0) or 0
    ciclo = _roman(pc.semester if pc else "")
    inst = _acta_area_inst()

    bundle = SectionGrades.objects.filter(section=sec).first()
    grades = (bundle.grades or {}) if bundle else {}
    alumnos = _roster(sec)
    if not alumnos:
        return None, f"La sección no tiene alumnos matriculados en {sec.period}."

    def _entry(st):
        for key in (st["id"], st.get("pk")):
            if key is not None and isinstance(grades.get(str(key)), dict):
                return grades[str(key)]
        return {}

    def _final(entry):
        if (entry.get("status") or "").upper() == "DPI":
            return 0
        for k in ("final_grade", "PROMEDIO_FINAL", "FINAL"):
            try:
                f = float(entry.get(k))
                if 0 <= f <= 20:
                    return int(round(f))
            except (TypeError, ValueError):
                continue
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"{ciclo or 'ACTA'}"[:31]

    bold = Font(bold=True, size=9)
    val9 = Font(size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _box(r1, c1, r2, c2, text, is_bold=True, fill=None, size=9):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=text)
        cell.font = Font(bold=is_bold, size=size)
        cell.alignment = center
        if fill:
            cell.fill = PatternFill("solid", start_color=fill)
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = _BORDER

    # ── Logos (instituto a la izquierda, como el modelo; sistema a la derecha) ──
    inst_logo, sist_logo = _institution_logo_paths()
    ws.row_dimensions[1].height = 48
    _add_logo(ws, inst_logo, "A1", height=58)
    _add_logo(ws, sist_logo, "H1", height=58)

    # ── Título ──
    _box(2, 1, 2, 8, f"ACTA DE EVALUACIÓN DE ÁREA {curso}".upper(), size=12)

    # ── Bloque institucional ──
    _box(4, 1, 5, 2, "Nombre de la Institución")
    _box(4, 3, 5, 4, inst["nombre"])
    _box(4, 5, 4, 5, "DREJ"); _box(4, 6, 4, 8, inst["drej"], is_bold=False)
    _box(5, 5, 5, 5, "UGE");  _box(5, 6, 5, 8, inst["uge"], is_bold=False)
    _box(6, 1, 6, 1, "Código Modular")
    _box(6, 2, 6, 2, "Denominación")
    _box(6, 3, 6, 3, "Gestión")
    _box(6, 4, 6, 4, "D.S./R.M.\nde Creación y R.D\nde Revalidación")
    _box(6, 5, 6, 5, "Dirección")
    _box(6, 6, 7, 8, inst["direccion"], is_bold=False)
    _box(7, 1, 7, 1, inst["codigo_modular"], is_bold=False)
    _box(7, 2, 7, 2, inst["denominacion"], is_bold=False)
    _box(7, 3, 7, 3, inst["gestion"], is_bold=False)
    _box(7, 4, 7, 4, inst["ds_creacion"], is_bold=False)
    _box(7, 5, 7, 5, "Provincia")
    _box(8, 5, 8, 5, inst["provincia"], is_bold=False)
    _box(8, 6, 8, 6, "Distrito")
    _box(8, 7, 8, 8, inst["distrito"], is_bold=False)
    _box(8, 1, 8, 4, "")

    _box(9, 1, 9, 2, "Programa de estudios / Turno")
    _box(9, 3, 9, 5, f"{career} / TURNO: {_turno_de(sec)}", is_bold=False)
    _box(9, 6, 9, 7, "Periodo Académico")
    _box(9, 8, 9, 8, sec.period, is_bold=False)
    _box(10, 1, 10, 2, "Resolución de Autorización")
    _box(10, 3, 10, 5, inst["resolucion"], is_bold=False)
    _box(10, 6, 10, 7, "Ciclo - Sección")
    _box(10, 8, 10, 8, f'{ciclo} - "{sec.label or "A"}"', is_bold=False)
    _box(11, 1, 11, 2, "Director General (e)")
    _box(11, 3, 11, 5, inst["director"], is_bold=False)
    _box(11, 6, 11, 7, "R.D. de Encargatura")
    _box(11, 8, 11, 8, inst["rd_encargatura"], is_bold=False)
    _box(12, 1, 12, 2, "Docente")
    _box(12, 3, 12, 8, docente.upper(), is_bold=False)

    # ── Tabla de alumnos ──
    H = 14
    heads = ["N°\nOrden", "N° Matrícula\n(DNI)", "APELLIDOS Y NOMBRES\n(Por orden Alfabético)",
             "Calificativo", "Crédito", "Puntaje", "Calificación\nCualitativa"]
    cols = [1, 2, 3, 4, 5, 6, 7]
    for c, h in zip(cols, heads):
        if c == 3:
            _box(H, 3, H, 4, h, fill="D9E2F3")
        else:
            cc = c if c < 3 else c + 1
            _box(H, cc, H, cc, h, fill="D9E2F3")

    r = H + 1
    for n, st in enumerate(alumnos, 1):
        entry = _entry(st)
        final = _final(entry)
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=st["dni"]).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(row=r, column=3, value=st["nombre"]).font = val9
        if final is not None:
            ws.cell(row=r, column=5, value=final)
            ws.cell(row=r, column=6, value=creditos)
            ws.cell(row=r, column=7, value=round(final * creditos))
            ws.cell(row=r, column=8, value=_cualitativa_de_vigesimal(final)).font = val9
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if c in (1, 2, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center")
        r += 1

    # ── Resumen (Anexo 3 RVM 123-2022) ──
    finales = [(_final(_entry(st))) for st in alumnos]
    con_nota = [f for f in finales if f is not None]
    promedio_aula = round(sum(con_nota) / len(con_nota), 2) if con_nota else ""
    resumen = [
        ("Matriculados", len(alumnos)),
        ("Aprobados", sum(1 for f in con_nota if f >= 11)),
        ("Desaprobados", sum(1 for f in con_nota if f < 11)),
        ("Con Licencia", sum(1 for st in alumnos
                             if st.get("estado") == "LICENCIA")),
        # Última fila según el Anexo 3 (antes decía "Límite de Inasistencia")
        ("Promedio del aula o sección en la calificación\n"
         "para el sistema de educación superior", promedio_aula),
    ]
    rr = r + 2
    _box(rr, 1, rr, 2, "Resumen", fill="E2EFDA")
    _box(rr, 3, rr, 3, "N°", fill="E2EFDA")
    for i, (lbl, val) in enumerate(resumen, 1):
        _box(rr + i, 1, rr + i, 2, lbl, is_bold=False)
        _box(rr + i, 3, rr + i, 3, val, is_bold=False)

    # ── Firmas — Anexo 3: Director + Secretario + Docente ──
    fr = rr + len(resumen) + 3
    FIRMAS = [
        (1, 3,  "DIRECTOR(A) GENERAL",        "Firma, Post Firma y Sello"),
        (5, 8,  "SECRETARIO(A) ACADÉMICO(A)", "Firma, Post Firma y Sello"),
        (10, 12, "DOCENTE FORMADOR(A)",       f"Firma · {docente.upper()}"),
    ]
    for c1, c2, cargo, sub in FIRMAS:
        _box(fr, c1, fr, c2, "_______________________________", is_bold=False)
        _box(fr + 1, c1, fr + 1, c2, cargo, size=8)
        _box(fr + 2, c1, fr + 2, c2, sub, is_bold=False, size=7)

    widths = {1: 6, 2: 13, 3: 30, 4: 16, 5: 11, 6: 8, 7: 9, 8: 13, 9: 3,
              10: 15, 11: 26, 12: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    fname = f"ACTA_AREA_{codigo or 'CURSO'}_{sec.label or 'A'}_{sec.period}.xlsx"
    return wb, fname


class SectionActaAreaView(APIView):
    """
    GET /api/academic/sections/<id>/acta-area.xlsx
    Acta de Evaluación de Área del curso (formato oficial): por alumno,
    Calificativo (promedio final del acta), Crédito (del plan), Puntaje
    (calificativo × crédito) y Calificación Cualitativa; con la leyenda
    oficial al costado y el resumen final. Docente (su sección) o admin.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id: int):
        from .teachers import _grades_section_access_denied

        sec = get_object_or_404(
            Section.objects.select_related(
                "plan_course__course", "plan_course__plan__career", "teacher__user"),
            id=section_id,
        )
        if err := _grades_section_access_denied(request, sec):
            return err

        wb, fname_or_err = build_acta_area_workbook(sec)
        if wb is None:
            return Response({"detail": fname_or_err}, status=400)
        return _xlsx_response(wb, fname_or_err)


# ══════════════════════════════════════════════════════════════
# ASISTENCIA — plantilla mensual + importación (estilo SIAGIE)
# ══════════════════════════════════════════════════════════════

def _parse_month(raw):
    m = re.match(r"^(\d{4})-(\d{1,2})$", str(raw or "").strip())
    if not m:
        return None, None
    y, mo = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12):
        return None, None
    return y, mo


class SectionAttendanceTemplateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id: int):
        from .attendance import _section_access_denied

        if err := _section_access_denied(request, section_id):
            return err
        sec = get_object_or_404(
            Section.objects.select_related("plan_course__course", "teacher__user"),
            id=section_id,
        )

        y, mo = _parse_month(request.query_params.get("month"))
        if not y:
            hoy = date_cls.today()
            y, mo = hoy.year, hoy.month
        dias = monthrange(y, mo)[1]

        alumnos = _roster(sec)
        if not alumnos:
            return Response(
                {"detail": (
                    f"La sección no tiene alumnos matriculados en {sec.period}. "
                    f"Matricula a los alumnos primero."
                )},
                status=400,
            )

        curso, codigo, docente = _section_header_info(sec)

        existentes = {}
        sesiones = AttendanceSession.objects.filter(
            section=sec, date__year=y, date__month=mo
        ).prefetch_related("rows")
        for s in sesiones:
            for rrow in s.rows.all():
                existentes[(str(rrow.student_id), s.date.day)] = STATUS_TO_MARK.get(rrow.status, "")

        wb = Workbook()
        ws = wb.active
        ws.title = "ASISTENCIA MENSUAL"

        ws["A1"] = "REGISTRO DE ASISTENCIA MENSUAL — SISTEMA ACADÉMICO"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"SECCION_ID: {sec.id}"
        ws["A3"] = f"MES: {y}-{mo:02d}"
        ws["A4"] = f"Curso: {curso} ({codigo})   ·   Sección: {sec.label or 'A'}   ·   Período: {sec.period}   ·   Docente: {docente or '—'}"
        ws["A5"] = "Marcas: P = Presente · T = Tardanza · F = Falta · J = Justificado · 0 = Feriado · (vacío = sin registrar)"

        HEAD_ROW = 7
        # Martes y miércoles con dos letras (antes el miércoles era "X")
        DIAS_SEMANA = ["L", "Ma", "Mi", "J", "V", "S", "D"]
        headers = ["N°", "DNI", "APELLIDOS Y NOMBRES"] + [str(d) for d in range(1, dias + 1)]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=HEAD_ROW, column=c, value=h)
            cell.fill = _HEAD_FILL
            cell.font = _HEAD_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER
        for d in range(1, dias + 1):
            wd = date_cls(y, mo, d).weekday()
            cell = ws.cell(row=HEAD_ROW + 1, column=3 + d, value=DIAS_SEMANA[wd])
            cell.alignment = _CENTER
            cell.border = _BORDER
            cell.font = Font(bold=True, size=8, color="666666")
            if wd >= 5:
                cell.fill = _FINDE_FILL

        ws.column_dimensions["A"].width = 4.5
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 40
        for d in range(1, dias + 1):
            ws.column_dimensions[get_column_letter(3 + d)].width = 3.4

        r = HEAD_ROW + 2
        for idx, st in enumerate(alumnos, 1):
            ws.cell(row=r, column=1, value=idx).border = _BORDER
            ws.cell(row=r, column=2, value=st["dni"]).border = _BORDER
            ws.cell(row=r, column=3, value=st["nombre"]).border = _BORDER
            for d in range(1, dias + 1):
                cell = ws.cell(row=r, column=3 + d,
                               value=existentes.get((str(st["id"]), d), ""))
                cell.alignment = Alignment(horizontal="center")
                cell.border = _BORDER
                if date_cls(y, mo, d).weekday() >= 5:
                    cell.fill = _FINDE_FILL
            r += 1

        dv = DataValidation(
            type="list", formula1='"P,T,F,J,0"', allow_blank=True,
            showErrorMessage=True, errorTitle="Marca inválida",
            error="Use: P, T, F, J o 0 (feriado)",
        )
        ws.add_data_validation(dv)
        first = HEAD_ROW + 2
        last = first + len(alumnos) - 1
        dv.add(f"{get_column_letter(4)}{first}:{get_column_letter(3 + dias)}{last}")

        fname = f"ASISTENCIA_{codigo or 'CURSO'}_{sec.label or 'A'}_{y}-{mo:02d}.xlsx"
        return _xlsx_response(wb, fname)


class SectionAttendanceImportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, section_id: int):
        from .attendance import _section_access_denied

        if err := _section_access_denied(request, section_id):
            return err
        sec = get_object_or_404(Section, id=section_id)

        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Adjunta el archivo Excel en el campo 'file'"}, status=400)

        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as exc:
            return Response({"detail": f"No se pudo leer el Excel: {exc}"}, status=400)

        # Mes: celda "MES: YYYY-MM" en las primeras filas
        y = mo = None
        for row in ws.iter_rows(min_row=1, max_row=6, max_col=3):
            for cell in row:
                m = re.match(r"^MES:\s*(\d{4})-(\d{1,2})$", str(cell.value or "").strip())
                if m:
                    y, mo = int(m.group(1)), int(m.group(2))
                    break
            if y:
                break
        if not y:
            return Response({"detail": "No se encontró la celda 'MES: YYYY-MM' (usa la plantilla generada)"}, status=400)
        dias = monthrange(y, mo)[1]

        head_row = None
        for row in ws.iter_rows(min_row=1, max_row=15):
            vals = [str(c.value or "").strip().upper() for c in row]
            if "DNI" in vals:
                head_row = row[0].row
                col_of = {v: i + 1 for i, v in enumerate(vals) if v}
                break
        if not head_row:
            return Response({"detail": "No se encontró la fila de encabezados (columna DNI)"}, status=400)

        day_cols = {d: col_of[str(d)] for d in range(1, dias + 1) if str(d) in col_of}

        roster = _roster(sec)
        by_dni = {st["dni"]: st for st in roster if st["dni"]}

        # Días de dictado de la sección (misma regla que el registro mensual:
        # horario configurado, o L-V si no hay). Fuera de ellos no se importa.
        from .attendance import _schedule_weekdays
        horario_wd = set(_schedule_weekdays(sec.id))

        # Vigencia del semestre (Configuración → Periodos Académicos): el
        # import tampoco graba días fuera del rango del período.
        from .attendance import _vigencia_de_periodo, _fuera_de_vigencia
        vig_ini, vig_fin = _vigencia_de_periodo(sec.period)

        def _es_dia_dictado(d):
            fecha = date_cls(y, mo, d)
            if _fuera_de_vigencia(fecha, vig_ini, vig_fin):
                return False
            wd = fecha.weekday()                   # 0=Lunes … 6=Domingo
            return (wd + 1) in horario_wd if horario_wd else wd < 5

        por_dia, errores = {}, []
        dias_no_dictado, lic_omitidos = set(), set()
        for row in ws.iter_rows(min_row=head_row + 1):
            dni = str(row[col_of["DNI"] - 1].value or "").strip()
            if not dni:
                continue
            st = by_dni.get(dni)
            if not st:
                errores.append(f"Fila {row[0].row}: DNI {dni} no está matriculado en esta sección")
                continue
            for d, col in day_cols.items():
                raw = str(row[col - 1].value or "").strip().upper()
                if not raw:
                    continue
                if not _es_dia_dictado(d):
                    dias_no_dictado.add(d)
                    continue
                if (st.get("estado") or "").upper() == "LICENCIA":
                    lic_omitidos.add(st["nombre"])
                    continue
                status = MARK_TO_STATUS.get(raw)
                if not status:
                    errores.append(f"DNI {dni} día {d}: marca inválida '{raw}'")
                    continue
                por_dia.setdefault(d, {})[str(st["id"])] = status
        if dias_no_dictado:
            errores.append(
                "Días sin dictado omitidos (la sección no dicta ese día): "
                + ", ".join(map(str, sorted(dias_no_dictado))))
        if lic_omitidos:
            errores.append(
                "Alumnos con LICENCIA omitidos (no se les registra asistencia): "
                + ", ".join(sorted(lic_omitidos)))

        sesiones_ok, cerradas = 0, []
        with transaction.atomic():
            for d, marcas in sorted(por_dia.items()):
                fecha = date_cls(y, mo, d)
                sess, _ = AttendanceSession.objects.get_or_create(section=sec, date=fecha)
                if sess.closed:
                    cerradas.append(str(d))
                    continue
                ids = [int(k) for k in marcas]
                AttendanceRow.objects.filter(session=sess, student_id__in=ids).delete()
                AttendanceRow.objects.bulk_create([
                    AttendanceRow(session=sess, student_id=int(k), status=v)
                    for k, v in marcas.items()
                ])
                sesiones_ok += 1

        msg = f"{sesiones_ok} día(s) de asistencia importado(s) para {y}-{mo:02d}"
        if cerradas:
            msg += f" · días omitidos por estar cerrados: {', '.join(cerradas)}"
        return ok(success=True, dias_importados=sesiones_ok,
                  dias_cerrados=cerradas, errores=errores, message=msg)
