"""
Centro de Evaluación (tipo SIAGIE)
──────────────────────────────────
Administra el proceso completo de evaluación de un período:

  1. Apertura / cierre del registro de calificaciones (AcademicPeriod.grades_state)
  2. Monitoreo por sección: notas cargadas, actas cerradas, procesadas
  3. "Procesar calificaciones": consolida SectionGrades (acta del docente)
     → AcademicGradeRecord (kárdex oficial del alumno), cerrando las actas
  4. Boletas de información en ZIP masivo por período/carrera

Las actas del docente (SectionGrades.grades) usan como clave
`str(student.user_id or student.id)`; el kárdex usa Student.pk. Este módulo
resuelve ambas claves (ver _entry_for).
"""
import re
import zipfile
from io import BytesIO

from django.db import transaction
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from academic.models import (
    AcademicPeriod, Section, SectionGrades, AcademicGradeRecord,
    EnrollmentItem, Enrollment,
)
from students.models import Student
from students.name_utils import clave_orden

from .teachers import _is_grades_admin
from ..pdf_render import html_to_pdf_bytes
from .kardex_helpers import _build_reporte_periodo_ctx

import logging
logger = logging.getLogger(__name__)

_PERIOD_RE = re.compile(r"^(\d{4})-(I|II)$")


# ══════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════

def _require_grades_admin(request):
    if not _is_grades_admin(request.user):
        return Response({"detail": "No autorizado."}, status=403)
    return None


def _period_defaults(code: str):
    """Fechas por defecto para crear un AcademicPeriod (Sem I: mar–jul, Sem II: ago–dic)."""
    from datetime import date
    m = _PERIOD_RE.match(code)
    if not m:
        return None
    y, t = int(m.group(1)), m.group(2)
    if t == "I":
        return {"start": date(y, 3, 1), "end": date(y, 7, 31)}
    return {"start": date(y, 8, 1), "end": date(y, 12, 15)}


def _roster_students(sec, incluir_licencia=False):
    """Alumnos matriculados (CONFIRMED) de la sección — EXACTAMENTE el mismo
    roster que el acta (`acta_excel._items_de_seccion`).

    Antes usaba un respaldo todo-o-nada propio: bastaba que UN alumno tuviera
    sección asignada para que todos los ítems con section NULL (típico de los
    ingresantes matriculados antes de crear las secciones) desaparecieran del
    procesamiento en silencio — figuraban en el acta con notas, pero nunca se
    les escribía el kárdex y su boleta respondía 404."""
    from .acta_excel import _items_de_seccion
    base, _ambiguos = _items_de_seccion(sec)
    out, seen = [], set()
    for item in base:
        st = item.enrollment.student
        if st.id in seen:
            continue
        seen.add(st.id)
        # Marcas: subsanación (por tipo de MATRÍCULA del período o por
        # estado) y licencia. Pueden convivir (pidió licencia de su
        # subsanación).
        st._subsa = (
            (getattr(item.enrollment, "tipo_matricula", "") or "").upper()
            == "SUBSANACION"
            or (getattr(st, "estado_academico", "") or "").upper()
            == "SUBSANACION")
        st._licencia = ((getattr(st, "estado_academico", "") or "").upper()
                        == "LICENCIA")
        # Con LICENCIA no se puede calificar (el acta bloquea su nota), así
        # que tampoco cuenta como "esperado" — si sumara al denominador, la
        # sección quedaría 24/25 en naranja para siempre. Con el flag sí se
        # devuelve (marcada) para las actas donde DEBE figurar.
        if st._licencia and not incluir_licencia:
            continue
        out.append(st)
    return out


def _entry_for(grades: dict, st) -> dict | None:
    """Entrada del acta para un alumno, probando la clave por user_id y por pk."""
    if not isinstance(grades, dict):
        return None
    for key in (getattr(st, "user_id", None), st.id):
        if key is None:
            continue
        e = grades.get(str(key))
        if isinstance(e, dict):
            return e
    return None


def _final_of(entry: dict):
    """Nota final 0-20 de una entrada del acta. DPI → 0. None si no hay nota."""
    if not isinstance(entry, dict):
        return None
    if (entry.get("status") or "").upper() == "DPI":
        return 0
    for key in ("final_grade", "PROMEDIO_FINAL", "FINAL"):
        v = entry.get(key)
        try:
            f = float(v)
            if 0 <= f <= 20:
                return f
        except (TypeError, ValueError):
            continue
    return None


def _section_eval_row(sec, bundle_map=None):
    """Fila de estado de evaluación de una sección (para el monitor y el proceso)."""
    if bundle_map is None:
        bundle = SectionGrades.objects.filter(section=sec).first()
    else:
        bundle = bundle_map.get(sec.id)
    grades = (bundle.grades or {}) if bundle else {}
    _full = _roster_students(sec, incluir_licencia=True)
    students = [st for st in _full if not getattr(st, "_licencia", False)]

    n_loaded = 0
    finals = {}   # student.pk -> final
    for st in students:
        f = _final_of(_entry_for(grades, st))
        if f is not None:
            n_loaded += 1
            finals[st.id] = f

    course = getattr(sec.plan_course, "course", None) if sec.plan_course else None
    n_processed = 0
    if course and finals:
        n_processed = AcademicGradeRecord.objects.filter(
            student_id__in=list(finals.keys()),
            course=course, term=sec.period,
        ).count()

    teacher_name = ""
    if sec.teacher and sec.teacher.user:
        teacher_name = (getattr(sec.teacher.user, "full_name", "")
                        or sec.teacher.user.username or "")

    return {
        "section_id": sec.id,
        "label": sec.label,
        "period": sec.period,
        "course_name": sec.plan_course.effective_name if sec.plan_course else "",
        "career_name": (sec.plan_course.plan.career.name
                        if sec.plan_course and sec.plan_course.plan
                        and sec.plan_course.plan.career else ""),
        "semester": sec.plan_course.semester if sec.plan_course else None,
        "teacher_name": teacher_name,
        "n_students": len(students),
        "n_loaded": n_loaded,
        "submitted": bool(bundle and bundle.submitted),
        "n_processed": n_processed,
        "processed": bool(finals) and n_processed >= len(finals),
        "_students": students,
        "_students_lic": [st for st in _full
                          if getattr(st, "_licencia", False)],
        "_finals": finals,
        "_grades": grades,
    }


def _sections_for(period, career_id=None, semester=None, section_id=None, anio=None):
    qs = Section.objects.filter(period=period).select_related(
        "plan_course", "plan_course__course", "plan_course__plan",
        "plan_course__plan__career", "teacher", "teacher__user",
    )
    if career_id:
        try:
            qs = qs.filter(plan_course__plan__career_id=int(career_id))
        except (TypeError, ValueError):
            pass
    if semester:
        try:
            qs = qs.filter(plan_course__semester=int(semester))
        except (TypeError, ValueError):
            pass
    if section_id:
        try:
            qs = qs.filter(id=int(section_id))
        except (TypeError, ValueError):
            pass
    if anio:
        # Año académico: 1° = ciclos 1-2, 2° = 3-4, … 5° = 9-10
        try:
            n = int(anio)
            qs = qs.filter(plan_course__semester__in=[2 * n - 1, 2 * n])
        except (TypeError, ValueError):
            pass
    return qs.order_by("plan_course__plan__career__name",
                       "plan_course__semester", "plan_course__id")


def _bundle_map(sections):
    return {b.section_id: b for b in SectionGrades.objects.filter(section__in=sections)}


def _public_row(row):
    return {k: v for k, v in row.items() if not k.startswith("_")}


# ══════════════════════════════════════════════════════════════
# 1. Estado del período: apertura / cierre
# ══════════════════════════════════════════════════════════════

class EvaluationStateView(APIView):
    """
    GET  /api/academic/admin/evaluation/state?period=2026-I
    POST /api/academic/admin/evaluation/state   body: {period, action: "open"|"close"}
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        per = AcademicPeriod.objects.filter(code=period).first()
        sections = list(_sections_for(period))
        bmap = _bundle_map(sections)

        rows = [_section_eval_row(s, bmap) for s in sections]
        stats = {
            "sections": len(rows),
            "students": sum(r["n_students"] for r in rows),
            "with_grades": sum(1 for r in rows if r["n_loaded"] > 0),
            "grades_loaded": sum(r["n_loaded"] for r in rows),
            "actas_closed": sum(1 for r in rows if r["submitted"]),
            "processed": sum(1 for r in rows if r["processed"]),
        }
        return Response({
            "period": period,
            "exists": bool(per),
            "grades_state": per.grades_state if per else AcademicPeriod.GRADES_OPEN,
            "window": per.grades_window_info() if per else {
                "grades_start": None, "grades_end": None,
                "is_open": True, "has_window": False,
                "grades_state": AcademicPeriod.GRADES_OPEN, "grades_closed_at": None,
            },
            "stats": stats,
        })

    def post(self, request):
        if err := _require_grades_admin(request):
            return err
        body = request.data or {}
        period = (body.get("period") or "").strip().upper()
        action = (body.get("action") or "").strip().lower()
        if not period or action not in ("open", "close"):
            return Response({"detail": "period y action ('open'|'close') son requeridos"}, status=400)

        per = AcademicPeriod.objects.filter(code=period).first()
        if not per:
            defaults = _period_defaults(period)
            if not defaults:
                return Response({"detail": f"Formato de período inválido: {period!r}"}, status=400)
            per = AcademicPeriod.objects.create(code=period, **defaults)

        if action == "open":
            per.grades_state = AcademicPeriod.GRADES_OPEN
            per.grades_closed_at = None
            msg = f"Registro de calificaciones de {period} habilitado."
        else:
            per.grades_state = AcademicPeriod.GRADES_CLOSED
            per.grades_closed_at = timezone.now()
            msg = f"Período {period} cerrado: los docentes ya no pueden cargar notas."
        per.save(update_fields=["grades_state", "grades_closed_at"])

        return Response({"success": True, "message": msg,
                         "grades_state": per.grades_state,
                         "window": per.grades_window_info()})


# ══════════════════════════════════════════════════════════════
# 2. Monitoreo de secciones para procesar
# ══════════════════════════════════════════════════════════════

class EvaluationSectionsView(APIView):
    """GET /api/academic/admin/evaluation/sections?period=2026-I[&career_id=5]"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        sections = list(_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            anio=request.query_params.get("anio")))
        bmap = _bundle_map(sections)
        rows = [_public_row(_section_eval_row(s, bmap)) for s in sections]
        return Response({"period": period, "sections": rows})


# ══════════════════════════════════════════════════════════════
# 3. Procesar calificaciones: acta → kárdex
# ══════════════════════════════════════════════════════════════

class EvaluationProcessView(APIView):
    """
    POST /api/academic/admin/evaluation/process
      body: {
        period: "2026-I",
        section_ids?: [1,2,3],   // omitir = todas las del período (con filtro career_id)
        career_id?: 5,
        close_actas?: true       // default true: cierra el acta al procesar
      }
    Por cada sección: copia la nota final de cada alumno del acta del docente
    (SectionGrades) al kárdex oficial (AcademicGradeRecord, update_or_create por
    student+course+term) y cierra el acta.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if err := _require_grades_admin(request):
            return err
        body = request.data or {}
        period = (body.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        close_actas = body.get("close_actas", True)
        section_ids = body.get("section_ids") or None

        sections = _sections_for(period, body.get("career_id"), body.get("semester"))
        if section_ids:
            try:
                sections = sections.filter(id__in=[int(x) for x in section_ids])
            except (TypeError, ValueError):
                return Response({"detail": "section_ids inválido"}, status=400)
        sections = list(sections)
        if not sections:
            return Response({"detail": "No hay secciones para procesar"}, status=404)

        bmap = _bundle_map(sections)
        results, total_processed, total_skipped = [], 0, 0

        with transaction.atomic():
            for sec in sections:
                row = _section_eval_row(sec, bmap)
                course = getattr(sec.plan_course, "course", None) if sec.plan_course else None
                if not course:
                    results.append({**_public_row(row), "result": "SIN_CURSO", "n_saved": 0})
                    continue
                if not row["_finals"]:
                    total_skipped += 1
                    results.append({**_public_row(row), "result": "SIN_NOTAS", "n_saved": 0})
                    continue

                n_saved = 0
                st_by_id = {s.id: s for s in row["_students"]}
                for st_id, final in row["_finals"].items():
                    st = st_by_id.get(st_id)
                    if not st:
                        continue
                    entry = _entry_for(row["_grades"], st) or {}
                    AcademicGradeRecord.objects.update_or_create(
                        student=st, course=course, term=sec.period,
                        defaults={
                            "final_grade": final,
                            "components": entry,
                            "plan_course": sec.plan_course,
                        },
                    )
                    n_saved += 1

                # Entradas del acta que no corresponden a ningún alumno del
                # roster (p. ej. matrícula reasignada de sección después de
                # registrar la nota): antes se saltaban EN SILENCIO y el curso
                # "desaparecía" de la boleta del alumno. Ahora se reportan.
                claves_roster = set()
                for s in list(row["_students"]) + list(row["_students_lic"]):
                    claves_roster.add(str(s.id))
                    if getattr(s, "user_id", None):
                        claves_roster.add(str(s.user_id))
                huerfanas = [k for k, v in (row["_grades"] or {}).items()
                             if isinstance(v, dict) and k not in claves_roster]

                if close_actas:
                    bundle = bmap.get(sec.id)
                    if bundle and not bundle.submitted:
                        bundle.submitted = True
                        bundle.submitted_at = timezone.now()
                        bundle.save(update_fields=["submitted", "submitted_at"])

                total_processed += 1
                results.append({**_public_row(row), "result": "PROCESADO",
                                "n_saved": n_saved,
                                "notas_sin_alumno": len(huerfanas),
                                "claves_sin_alumno": huerfanas,
                                "submitted": True if close_actas else row["submitted"]})

        total_huerfanas = sum(r.get("notas_sin_alumno", 0) for r in results)
        return Response({
            "success": True,
            "period": period,
            "message": (f"{total_processed} sección(es) procesada(s) al kárdex"
                        + (f" · {total_skipped} sin notas (omitidas)" if total_skipped else "")
                        + (f" · ADVERTENCIA: {total_huerfanas} nota(s) del acta sin "
                           "alumno en el roster (revisar la sección del alumno en "
                           "Matrícula)" if total_huerfanas else "")),
            "processed": total_processed,
            "skipped": total_skipped,
            "notas_sin_alumno": total_huerfanas,
            "results": results,
        })


# ══════════════════════════════════════════════════════════════
# 4a. Acta consolidada de evaluación — Excel por período
# ══════════════════════════════════════════════════════════════

def _abrev_calif(nota):
    """Abreviatura cualitativa del modelo oficial: D, L, EP, I, PI."""
    if nota is None:
        return ""
    n = float(nota)
    if n >= 20:
        return "D"
    if n >= 15:
        return "L"
    if n >= 11:
        return "EP"
    if n >= 6:
        return "I"
    return "PI"


def _inst_data():
    """Datos institucionales desde catalogs.InstitutionSetting.data (con defaults)."""
    try:
        from catalogs.models import InstitutionSetting
        data = (InstitutionSetting.objects.filter(pk=1).first() or InstitutionSetting()).data or {}
    except Exception:
        data = {}
    # Defaults = datos oficiales confirmados por Secretaría (2026). Antes eran
    # cadenas vacías: si faltaba la configuración, el acta salía con la
    # cabecera en blanco.
    return {
        "nombre": ((data.get("institution_name") or data.get("name")
                    or 'INSTITUTO DE EDUCACIÓN SUPERIOR PEDAGÓGICO PÚBLICO '
                       '"GUSTAVO ALLENDE LLAVERIA"').upper().rstrip()
                   + (" - TARMA" if "TARMA" not in
                      (data.get("institution_name") or data.get("name") or "").upper()
                      else "")),
        "licenciamiento": (data.get("licenciamiento") or data.get("resolucion")
                           or "REVALIDADO: R.D. N° 306-2016-MINEDU/VMGP/DIGEDD/DIFOID"),
        "director": (data.get("director_name") or "GARCIA PORRAS, MARIA ELVIRA").upper(),
        "codigo_modular": data.get("codigo_modular") or "0609370",
        "direccion": (data.get("direccion") or data.get("address")
                      or "Calle Tacajashi Tacajashi Km. 4 - Pomachaca"),
        "rd_encargatura": data.get("rd_encargatura") or "R.D. N° 017-2026-DREJ",
    }


class EvaluationActaConsolidadaView(APIView):
    """
    GET /api/academic/admin/evaluation/actas.xlsx?period=2026-I[&career_id=5]
    Acta consolidada oficial: una hoja por programa+ciclo+sección, cursos en
    columnas (C = calificación cualitativa, CS = nota, PTJ = nota × créditos),
    promedio ponderado del semestre y cuadro de firmas de docentes.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from .acta_excel import _roman

        sections = list(_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            anio=request.query_params.get("anio")))
        if not sections:
            return Response({"detail": f"No hay secciones en {period}"}, status=404)
        bmap = _bundle_map(sections)
        inst = _inst_data()

        # Agrupar por (carrera, ciclo, sección)
        groups = {}
        for sec in sections:
            pc = sec.plan_course
            career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "?")
            sem = pc.semester if pc else 0
            key = (career, sem or 0, sec.label or "A")
            groups.setdefault(key, []).append(sec)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Sombreado "plomo claro" (pedido de Secretaría; antes era celeste)
        head_fill = PatternFill("solid", start_color="E7E6E6")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        subsa = str(request.query_params.get("subsanacion", "")).lower() in ("1", "true", "si")

        wb = Workbook()
        wb.remove(wb.active)
        used_titles = set()
        any_sheet = False

        for (career, sem, label), secs in sorted(groups.items()):
            rows = [_section_eval_row(s, bmap) for s in secs]
            # Roster consolidado del ciclo (unión de todas las secciones).
            # Subsanación lleva acta consolidada APARTE (?subsanacion=1).
            by_id = {}
            for rw in rows:
                for st in rw["_students"] + rw.get("_students_lic", []):
                    if bool(getattr(st, "_subsa", False)) != bool(subsa):
                        continue
                    # En el acta regular los de licencia no se consolidan;
                    # en la de subsanación SÍ figuran (con marca LICENCIA).
                    if getattr(st, "_licencia", False) and not subsa:
                        continue
                    by_id[st.id] = st
            if not by_id:
                continue
            any_sheet = True
            students = sorted(by_id.values(), key=clave_orden)
            ncur = len(secs)
            # Columnas: A N°, B matrícula, C nombres; por curso 3 (C, CS, PTJ);
            # luego Puntaje, Créditos, Promedio, Calificación, Observaciones
            first_c = 4
            tot_c = first_c + ncur * 3
            last_c = tot_c + 4

            title = f"{career[:22]} {_roman(sem)}" + ("" if (label or "A") == "A" else f" {label}")
            title = title.strip()[:31] or "HOJA"
            while title in used_titles:
                title = (title[:29] + "_")[:31]
            used_titles.add(title)
            ws = wb.create_sheet(title=title)

            # ── Logos + Título ──
            # MINEDU + instituto a la izquierda; el del sistema al otro extremo
            import os as _os
            from .acta_excel import _institution_logo_paths, _add_logo
            inst_logo, sist_logo = _institution_logo_paths()
            minedu_logo = _os.path.join(
                _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                "..", "common", "assets", "logo_minedu_peru.png")
            minedu_logo = _os.path.normpath(minedu_logo)
            ws.row_dimensions[1].height = 46
            if _os.path.exists(minedu_logo):
                _add_logo(ws, minedu_logo, "A1", height=30)
                _add_logo(ws, inst_logo, "C1", height=56)
            else:
                _add_logo(ws, inst_logo, "A1", height=56)
            if last_c >= 8:
                from openpyxl.utils import get_column_letter as _gcl
                _add_logo(ws, sist_logo, f"{_gcl(last_c - 1)}1", height=56)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_c)
            t = ws.cell(row=2, column=1,
                        value=f"ACTA CONSOLIDADA DE EVALUACIÓN DEL PERÍODO ACADÉMICO {period}"
                              + (" — SUBSANACIÓN" if subsa else ""))
            t.font = Font(bold=True, size=13)
            t.alignment = Alignment(horizontal="center")

            # ── Cabecera institucional (par de columnas izquierda/derecha) ──
            mid_l = min(first_c + 5, tot_c - 1)
            rlab_a = mid_l + 2
            rlab_b = min(rlab_a + 2, last_c - 2)
            rval_a = rlab_b + 1

            # Todo el texto de la cabecera va CENTRADO (pedido de Secretaría)
            _hdr_center = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)

            def _hdr(row, llabel, lvalue, rlabel, rvalue):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                c = ws.cell(row=row, column=1, value=llabel)
                c.font = Font(bold=True, size=9)
                c.alignment = _hdr_center
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=mid_l)
                c = ws.cell(row=row, column=4, value=lvalue)
                c.font = Font(size=9)
                c.alignment = _hdr_center
                ws.merge_cells(start_row=row, start_column=rlab_a, end_row=row, end_column=rlab_b)
                c = ws.cell(row=row, column=rlab_a, value=rlabel)
                c.font = Font(bold=True, size=9)
                c.alignment = _hdr_center
                ws.merge_cells(start_row=row, start_column=rval_a, end_row=row, end_column=last_c)
                c = ws.cell(row=row, column=rval_a, value=rvalue)
                c.font = Font(size=9)
                c.alignment = _hdr_center

            ciclo_sec = f'{_roman(sem)} - "{label or "A"}"'
            # El turno se jala del horario de las secciones. La fecha de
            # emisión va SEPARADA del período, en su propia línea.
            from django.utils import timezone as _tz
            from .acta_excel import _turno_de
            hoy = _tz.localtime(_tz.now()).strftime("%d/%m/%Y")
            _hdr(4, "Nombre de la Institución", inst["nombre"], "Código Modular", inst["codigo_modular"])
            _hdr(5, "R.M. de Licenciamiento o R.D.", inst["licenciamiento"], "Dirección", inst["direccion"])
            _hdr(6, "Directora General", inst["director"], "R.D. de Encargatura", inst["rd_encargatura"])
            _hdr(7, "Programa de Estudios", career.upper(), "Periodo Académico",
                 f"{period}   ·   FECHA: {hoy}")
            _hdr(8, "Ciclo - Sección", ciclo_sec, "Número de Estudiantes", len(students))
            _hdr(9, "Modalidad de Estudios", "PRESENCIAL", "Turno", _turno_de(secs))

            # ── Encabezado de la tabla (filas 11-16) ──
            H = 11

            def _mh(r1, c1, r2, c2, text, fill=True, size=8):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                cell = ws.cell(row=r1, column=c1, value=text)
                cell.font = Font(bold=True, size=size)
                cell.alignment = center
                if fill:
                    cell.fill = head_fill
                for rr in range(r1, r2 + 1):
                    for cc in range(c1, c2 + 1):
                        ws.cell(row=rr, column=cc).border = border

            _mh(H, 1, H + 5, 1, "N°")
            _mh(H, 2, H + 5, 2, "N° de\nMatrícula")
            _mh(H, 3, H + 5, 3, "Apellidos y Nombres del Estudiante\n(según nóminas)")
            _mh(H, first_c, H, tot_c - 1, "ASIGNATURAS / ÁREAS")
            # Nombres de curso "echados" (verticales), completos y con la
            # fila alta para que entren — como el acta institucional.
            ws.row_dimensions[H + 2].height = 110
            vertical = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True, text_rotation=90)
            for i, rw in enumerate(rows):
                c0 = first_c + i * 3
                _mh(H + 1, c0, H + 1, c0 + 2, str(i + 1))
                _mh(H + 2, c0, H + 2, c0 + 2, (rw["course_name"] or "").upper())
                ws.cell(row=H + 2, column=c0).alignment = vertical
                _mh(H + 3, c0, H + 3, c0 + 2, "Créditos")
                pc = secs[i].plan_course
                _mh(H + 4, c0, H + 4, c0 + 2, getattr(pc, "credits", "") or "")
                for j, sub in enumerate(["C", "CS", "PTJ"]):
                    _mh(H + 5, c0 + j, H + 5, c0 + j, sub)
            _mh(H, tot_c, H + 5, tot_c, "Puntaje del\nSemestre")
            _mh(H, tot_c + 1, H + 5, tot_c + 1, "Crédito del\nSemestre")
            _mh(H, tot_c + 2, H + 5, tot_c + 2, "Promedio\nPonderado")
            _mh(H, tot_c + 3, H + 5, tot_c + 3, "Calificación\nCualitativa")
            _mh(H, tot_c + 4, H + 5, tot_c + 4, "Observaciones")

            # ── Filas de alumnos ──
            r0 = H + 6
            for n, st in enumerate(students, 1):
                r = r0 + n - 1
                nombre = (f"{st.apellido_paterno or ''} {st.apellido_materno or ''}, "
                          f"{st.nombres or ''}").strip(", ").strip().upper()
                ws.cell(row=r, column=1, value=n)
                ws.cell(row=r, column=2, value=st.num_documento or "")
                # Apellidos y nombres en UNA sola fila (sin envolver texto)
                cn = ws.cell(row=r, column=3, value=nombre)
                cn.font = Font(size=8)
                cn.alignment = Alignment(vertical="center", wrap_text=False)

                puntaje = creditos = 0
                tiene_nota = False
                for i, rw in enumerate(rows):
                    c0 = first_c + i * 3
                    final = _final_of(_entry_for(rw["_grades"], st))
                    creds = getattr(secs[i].plan_course, "credits", 0) or 0
                    if final is not None:
                        tiene_nota = True
                        ptj = round(float(final) * float(creds))
                        ws.cell(row=r, column=c0, value=_abrev_calif(final))
                        ws.cell(row=r, column=c0 + 1, value=round(float(final)))
                        ws.cell(row=r, column=c0 + 2, value=ptj)
                        puntaje += ptj
                        creditos += creds
                if tiene_nota and creditos:
                    prom = puntaje / creditos
                    ws.cell(row=r, column=tot_c, value=puntaje)
                    ws.cell(row=r, column=tot_c + 1, value=creditos)
                    # Secretaría pidió el promedio con DOS decimales (el
                    # Anexo 4 dice "hasta con tres", así que dos cumple).
                    ws.cell(row=r, column=tot_c + 2, value=round(prom, 2))
                    ws.cell(row=r, column=tot_c + 3, value=_abrev_calif(prom))
                if getattr(st, "_licencia", False):
                    obs = "LICENCIA" + (f" ({st.estado_rd})" if st.estado_rd else "")
                    ws.cell(row=r, column=last_c, value=obs).font = Font(size=8)
                for c in range(1, last_c + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    if c != 3:
                        cell.alignment = Alignment(horizontal="center")

            # ── Cuadro de firmas de docentes ──
            # +5: "un poco de distancia entre el acta y la firma de los
            # docentes" (pedido de Secretaría)
            fr = r0 + len(students) + 5
            _mh(fr, 1, fr, 2, "Número de\nCurso o Área", size=8)
            _mh(fr, 3, fr, 5, "APELLIDOS Y NOMBRES DEL DOCENTE", size=8)
            _mh(fr, 6, fr, 8, "FIRMA", size=8)
            for i, rw in enumerate(rows, 1):
                rr = fr + i
                # Filas más separadas para que la firma quepa cómoda
                ws.row_dimensions[rr].height = 26
                ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
                ws.cell(row=rr, column=1, value=i).alignment = Alignment(horizontal="center")
                ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=5)
                ws.cell(row=rr, column=3, value=(rw["teacher_name"] or "").upper()).font = Font(size=9)
                ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=8)
                for c in range(1, 9):
                    ws.cell(row=rr, column=c).border = border

            # ── Firmas de Dirección y Secretaría Académica ──
            # +6 filas de aire: van a poner SELLO REDONDO entre el cuadro de
            # docentes y estas firmas. Línea corta ("casi el tamaño de las
            # letras nomas").
            fir = fr + len(rows) + 6
            for col, cargo in ((1, "DIRECTORA GENERAL"),
                               (6, "SECRETARIO(A) ACADÉMICO(A)"),
                               (11, "V°B° DRE / UGEL")):
                ws.merge_cells(start_row=fir, start_column=col,
                               end_row=fir, end_column=col + 3)
                linea = ws.cell(row=fir, column=col, value="_" * 20)
                linea.font = Font(size=8)
                linea.alignment = Alignment(horizontal="center")

                ws.merge_cells(start_row=fir + 1, start_column=col,
                               end_row=fir + 1, end_column=col + 3)
                c1 = ws.cell(row=fir + 1, column=col, value=cargo)
                c1.font = Font(bold=True, size=9)
                c1.alignment = Alignment(horizontal="center")

                ws.merge_cells(start_row=fir + 2, start_column=col,
                               end_row=fir + 2, end_column=col + 3)
                c2 = ws.cell(row=fir + 2, column=col, value="Firma, Post Firma y Sello")
                c2.font = Font(size=8)
                c2.alignment = Alignment(horizontal="center")

            # ── Anchos ──
            # C/CS/PTJ angostas (los nombres de curso van en vertical);
            # la columna de apellidos más ancha para que quepan en una fila.
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 11
            ws.column_dimensions["C"].width = 42
            for i in range(ncur):
                c0 = first_c + i * 3
                for j, w in enumerate((3.6, 4.2, 4.8)):
                    ws.column_dimensions[get_column_letter(c0 + j)].width = w
            for j, w in enumerate((9, 9, 10, 12, 14)):
                ws.column_dimensions[get_column_letter(tot_c + j)].width = w
            ws.freeze_panes = f"D{r0}"

        if not any_sheet:
            det = (f"Ningún alumno de SUBSANACIÓN en {period} con el filtro."
                   if subsa else
                   f"Ninguna sección de {period} tiene alumnos matriculados")
            return Response({"detail": det}, status=404)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f'attachment; filename="acta-consolidada{"-subsanacion" if subsa else ""}-{period}.xlsx"'},
        )


# ══════════════════════════════════════════════════════════════
# 4b. Boletas de información — ZIP masivo por período
# ══════════════════════════════════════════════════════════════

class EvaluationBoletasZipView(APIView):
    """
    GET /api/academic/admin/evaluation/boletas.zip?period=2026-I[&career_id][&semester][&anio=1]
    Una boleta de calificaciones (reporte_calificaciones) en PDF por cada alumno
    con notas procesadas en el período, empaquetadas en un ZIP.
    Con anio=1 la boleta de cada alumno incluye AMBOS semestres del año
    (YYYY-I + YYYY-II concatenados en un solo PDF).
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        anio_mode = str(request.query_params.get("anio", "")).lower() in ("1", "true", "yes")
        year = period.split("-")[0]
        terms = [f"{year}-I", f"{year}-II"] if anio_mode else [period]

        # El filtro de ciclo va sobre los CURSOS del kárdex del período, no
        # sobre Student.ciclo (que es el ciclo ACTUAL): al promover ciclos,
        # "2026-I ciclo 2" devolvía a los promovidos equivocados o a nadie
        # ("No hay alumnos con notas procesadas" con 1867 registros cargados).
        # Ambas condiciones en el MISMO filter() → aplican al mismo registro.
        from django.db.models import Q as _Q
        cond = _Q(grade_records__term__in=terms)
        semester = request.query_params.get("semester")
        if semester:
            try:
                cond &= _Q(grade_records__plan_course__semester=int(semester))
            except (TypeError, ValueError):
                return Response({"detail": "semester inválido"}, status=400)
        else:
            anio_acad = request.query_params.get("anio_academico")
            if anio_acad:
                try:
                    n = int(anio_acad)
                    cond &= _Q(grade_records__plan_course__semester__in=[2 * n - 1, 2 * n])
                except (TypeError, ValueError):
                    pass
        qs = (Student.objects.filter(cond)
              .select_related("plan", "plan__career")
              .distinct())
        career_id = request.query_params.get("career_id")
        if career_id:
            try:
                qs = qs.filter(plan__career_id=int(career_id))
            except (TypeError, ValueError):
                return Response({"detail": "career_id inválido"}, status=400)

        students = list(qs.order_by("apellido_paterno", "apellido_materno", "nombres"))
        if not students:
            return Response(
                {"detail": f"No hay alumnos con notas procesadas en "
                           f"{year if anio_mode else period}. "
                           "Primero usa 'Procesar calificaciones'."},
                status=404)

        etiqueta = year if anio_mode else period
        zip_buf = BytesIO()
        generated, errors_list = 0, []
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for st in students:
                try:
                    pdfs = []
                    for term in terms:
                        ctx, err = _build_reporte_periodo_ctx(request, st, term)
                        if err:
                            continue
                        html = render_to_string("kardex/reporte_calificaciones.html", ctx)
                        pdfs.append(html_to_pdf_bytes(html))
                    if not pdfs:
                        errors_list.append(f"{st.num_documento or st.id}: sin notas en {etiqueta}")
                        continue
                    if len(pdfs) == 1:
                        pdf_bytes = pdfs[0]
                    else:
                        # Año completo: concatenar I + II en un solo PDF
                        from pypdf import PdfReader, PdfWriter
                        writer = PdfWriter()
                        for p in pdfs:
                            for page in PdfReader(BytesIO(p)).pages:
                                writer.add_page(page)
                        out = BytesIO()
                        writer.write(out)
                        pdf_bytes = out.getvalue()
                    doc = st.num_documento or str(st.id)
                    safe = (f"{st.apellido_paterno or ''}_{st.apellido_materno or ''}_{st.nombres or ''}"
                            .replace(" ", "_").replace("/", "-").strip("_"))
                    zf.writestr(f"boleta-{etiqueta}-{doc}-{safe}.pdf", pdf_bytes)
                    generated += 1
                except Exception as exc:
                    logger.exception("Error boleta bulk student=%s period=%s", st.id, etiqueta)
                    errors_list.append(f"{st.num_documento or st.id}: {exc}")

            if errors_list:
                zf.writestr("errores.txt", "\n".join(errors_list))

        if not generated:
            return Response({"detail": "No se pudo generar ninguna boleta",
                             "errors": errors_list}, status=500)

        zip_buf.seek(0)
        return HttpResponse(
            zip_buf.getvalue(),
            content_type="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="boletas-{etiqueta}.zip"'},
        )


# ══════════════════════════════════════════════════════════════
# 4c. Actas de Evaluación de Área — ZIP masivo con filtros
# ══════════════════════════════════════════════════════════════

class EvaluationSubsanacionListView(APIView):
    """
    GET /academic/admin/evaluation/subsanacion?period=2026-II[&career_id]
    Estudiantes con estado SUBSANACIÓN (ciclos II al VIII) matriculados en el
    período, con sus cursos — para el panel de Subsanación de Evaluación.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from django.db.models import Q
        # Subsanación por estado del alumno O por tipo de matrícula del
        # período (caso real: alumna con matrícula de subsanación que pidió
        # LICENCIA — sigue siendo de subsanación, con su marca de licencia).
        qs = (Student.objects
              .select_related("plan", "plan__career")
              .filter(Q(estado_academico=Student.ESTADO_SUBSANACION)
                      | Q(enrollments__tipo_matricula="SUBSANACION"),
                      enrollments__period=period,
                      enrollments__status=Enrollment.STATUS_CONFIRMED)
              .distinct())
        career_id = request.query_params.get("career_id")
        if career_id:
            try:
                qs = qs.filter(plan__career_id=int(career_id))
            except (TypeError, ValueError):
                pass

        items_map = {}
        for it in (EnrollmentItem.objects
                   .select_related("plan_course__course", "section",
                                   "enrollment")
                   .filter(enrollment__student__in=qs,
                           enrollment__period=period,
                           enrollment__status=Enrollment.STATUS_CONFIRMED)):
            pc = it.plan_course
            items_map.setdefault(it.enrollment.student_id, []).append({
                "curso": (getattr(pc, "display_name", "")
                          or (pc.course.name if pc and pc.course else "")),
                "semestre": getattr(pc, "semester", None),
                "creditos": getattr(pc, "credits", 0) or 0,
                "section_id": it.section_id,
            })

        out = []
        for st in sorted(qs, key=clave_orden):
            cursos = items_map.get(st.id, [])
            sems = [c["semestre"] for c in cursos if c["semestre"]]
            ciclo = max(sems) if sems else (st.ciclo or None)
            # Subsanación solo existe del II al VIII (regla de Secretaría)
            if ciclo and not (2 <= int(ciclo) <= 8):
                continue
            out.append({
                "id": st.id,
                "dni": st.num_documento or "",
                "nombre": (f"{st.apellido_paterno or ''} "
                           f"{st.apellido_materno or ''}, "
                           f"{st.nombres or ''}").strip(", ").strip().upper(),
                "carrera": (st.plan.career.name
                            if st.plan and st.plan.career else ""),
                "ciclo": ciclo,
                "estado_rd": st.estado_rd or "",
                "creditos": sum(c["creditos"] for c in cursos),
                "cursos": cursos,
            })
        return Response({"period": period, "total": len(out), "students": out})


class EvaluationSilabosSesionesView(APIView):
    """
    GET /academic/admin/evaluation/silabos-sesiones
        ?period=2026-II[&career_id][&semester][&anio][&section_id]
    Monitor admin de lo que suben los docentes: por cada sección del filtro,
    el sílabo (con enlace) y las sesiones de aprendizaje (fecha, semana,
    tema y archivo). Mismos filtros que los reportes de evaluación.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from .utils import romanos_mayusculas
        secs = (_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            section_id=request.query_params.get("section_id"),
            anio=request.query_params.get("anio"))
            .select_related("syllabus")
            .prefetch_related("sesiones_aprendizaje"))

        def _url(f):
            try:
                return request.build_absolute_uri(f.url) if f else ""
            except Exception:
                return ""

        rows, con_silabo, total_sesiones = [], 0, 0
        for sec in secs:
            pc = sec.plan_course
            try:
                silabo_url = _url(sec.syllabus.file)
            except Exception:
                silabo_url = ""
            sesiones = [{
                "fecha": str(s.fecha), "semana": s.semana, "tema": s.tema,
                "archivo_url": _url(s.archivo),
            } for s in sec.sesiones_aprendizaje.all()]
            docente = ""
            if sec.teacher:
                u = getattr(sec.teacher, "user", None)
                docente = ((getattr(u, "full_name", "") or "").strip()
                           or str(sec.teacher))
            if silabo_url:
                con_silabo += 1
            total_sesiones += len(sesiones)
            rows.append({
                "section_id": sec.id,
                "curso": romanos_mayusculas(pc.effective_name if pc else ""),
                "codigo": (getattr(pc, "display_code", "")
                           or (pc.course.code if pc and pc.course else "")) if pc else "",
                "ciclo": pc.semester if pc else None,
                "carrera": (pc.plan.career.name
                            if pc and pc.plan and pc.plan.career else ""),
                "seccion": sec.label or "A",
                "docente": docente,
                "silabo_url": silabo_url,
                "sesiones": sesiones,
                "n_sesiones": len(sesiones),
            })

        return Response({
            "period": period,
            "total_secciones": len(rows),
            "con_silabo": con_silabo,
            "sin_silabo": len(rows) - con_silabo,
            "total_sesiones": total_sesiones,
            "rows": rows,
        })


class EvaluationActasCalificacionSubsaZipView(APIView):
    """
    GET /academic/admin/evaluation/actas-calificacion-subsanacion.zip
        ?period=2026-II[&career_id][&semester][&anio]
    Un Excel de ACTA DE CALIFICACIÓN (Anexo 3, competencias) en versión
    SUBSANACIÓN por cada curso del filtro que tenga alumnos de subsanación,
    empaquetados en ZIP — para no obligar a emitirlas curso por curso.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from openpyxl import Workbook
        from .acta_excel import _write_grades_sheet, _section_header_info

        sections = list(_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            anio=request.query_params.get("anio")))
        if not sections:
            return Response(
                {"detail": f"No hay secciones para el filtro en {period}"},
                status=404)

        zip_buf = BytesIO()
        generados = 0
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for sec in sections:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "ACTA SUBSANACIÓN"
                    n = _write_grades_sheet(ws, sec, subsanacion=True)
                    if n == 0:
                        continue   # este curso no tiene alumnos de subsanación
                    _, codigo, _ = _section_header_info(sec)
                    buf = BytesIO()
                    wb.save(buf)
                    zf.writestr(
                        f"ACTA_SUBSANACION_{codigo or sec.id}_"
                        f"{sec.label or 'A'}_{period}.xlsx",
                        buf.getvalue())
                    generados += 1
                except Exception:
                    logger.exception("Acta subsanación Anexo 3 sec=%s", sec.id)
                    continue
        if not generados:
            return Response(
                {"detail": "Ningún curso del filtro tiene alumnos de "
                           "subsanación (el estado/tipo se asigna en "
                           "Matrícula → Padrón de Alumnos)."}, status=404)

        zip_buf.seek(0)
        return HttpResponse(
            zip_buf.getvalue(), content_type="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="actas-calificacion-subsanacion-{period}.zip"'})


class EvaluationActasAreaZipView(APIView):
    """
    GET /api/academic/admin/evaluation/actas-area.zip?period=2026-I[&career_id][&semester]
    Un Excel de Acta de Evaluación de Área por cada curso (sección) del filtro,
    empaquetados en ZIP.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from .acta_excel import build_acta_area_workbook

        subsa = str(request.query_params.get("subsanacion", "")).lower() in ("1", "true", "si")
        sections = list(_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            anio=request.query_params.get("anio")))
        if not sections:
            return Response({"detail": f"No hay secciones para el filtro en {period}"}, status=404)

        zip_buf = BytesIO()
        generated, errors_list = 0, []
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for sec in sections:
                try:
                    wb, fname_or_err = build_acta_area_workbook(sec, subsanacion=subsa)
                    if wb is None:
                        errors_list.append(f"Sección {sec.id}: {fname_or_err}")
                        continue
                    buf = BytesIO()
                    wb.save(buf)
                    zf.writestr(fname_or_err, buf.getvalue())
                    generated += 1
                except Exception as exc:
                    logger.exception("Error acta-area bulk section=%s", sec.id)
                    errors_list.append(f"Sección {sec.id}: {exc}")
            if errors_list:
                zf.writestr("errores.txt", "\n".join(errors_list))

        if not generated:
            return Response({"detail": "Ninguna sección del filtro tiene alumnos",
                             "errors": errors_list}, status=404)

        zip_buf.seek(0)
        return HttpResponse(
            zip_buf.getvalue(),
            content_type="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="actas-area-{period}.zip"'},
        )


# ══════════════════════════════════════════════════════════════
# 4d. Reporte de Rendimiento — Excel real con filtros
# ══════════════════════════════════════════════════════════════

class EvaluationReporteRendimientoView(APIView):
    """
    GET /api/academic/admin/evaluation/rendimiento.xlsx
        ?period=2026-I[&career_id][&semester][&section_id]
    Reporte de rendimiento académico: hoja "Resumen por curso" (matriculados,
    con notas, aprobados, desaprobados, DPI, promedio) y hoja "Detalle por
    alumno" (curso, alumno, nota, calificación cualitativa).
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from .acta_excel import _cualitativa_de_vigesimal

        sections = list(_sections_for(
            period, request.query_params.get("career_id"),
            request.query_params.get("semester"),
            request.query_params.get("section_id"),
            anio=request.query_params.get("anio")))
        if not sections:
            return Response({"detail": f"No hay secciones para el filtro en {period}"}, status=404)
        bmap = _bundle_map(sections)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen por curso"
        head_fill = PatternFill("solid", start_color="1F4E79")
        head_font = Font(bold=True, color="FFFFFF")

        from .acta_excel import _institution_logo_paths, _add_logo
        inst_logo, _sist = _institution_logo_paths()
        _add_logo(ws, inst_logo, "L1", height=50)
        ws["A1"] = f"REPORTE DE RENDIMIENTO ACADÉMICO — {period}"
        ws["A1"].font = Font(bold=True, size=13)

        headers = ["CARRERA", "CICLO", "CURSO", "SEC.", "DOCENTE", "MATRICULADOS",
                   "CON NOTAS", "APROBADOS", "DESAPROBADOS", "DPI", "PROMEDIO", "ACTA"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")

        ws2 = wb.create_sheet(title="Detalle por alumno")
        headers2 = ["CARRERA", "CICLO", "CURSO", "SEC.", "DNI", "ALUMNO",
                    "NOTA (0-20)", "CALIFICACIÓN", "ESTADO"]
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.fill = head_fill
            cell.font = head_font

        r1, r2 = 4, 2
        for sec in sections:
            row = _section_eval_row(sec, bmap)
            finals = list(row["_finals"].values())
            aprob = sum(1 for f in finals if f >= 11)
            desap = sum(1 for f in finals if f < 11)
            n_dpi = 0
            for st in row["_students"]:
                e = _entry_for(row["_grades"], st) or {}
                if (e.get("status") or "").upper() == "DPI":
                    n_dpi += 1
            prom = round(sum(finals) / len(finals), 1) if finals else ""
            vals = [row["career_name"], row["semester"], row["course_name"], row["label"],
                    row["teacher_name"], row["n_students"], row["n_loaded"],
                    aprob, desap, n_dpi, prom,
                    "CERRADA" if row["submitted"] else "ABIERTA"]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r1, column=c, value=v)
            r1 += 1

            for st in sorted(row["_students"],
                             key=lambda s: f"{s.apellido_paterno or ''} {s.apellido_materno or ''} {s.nombres or ''}"):
                e = _entry_for(row["_grades"], st) or {}
                final = row["_finals"].get(st.id)
                es_dpi = (e.get("status") or "").upper() == "DPI"
                estado = ("DPI" if es_dpi else
                          "" if final is None else
                          "APROBADO" if final >= 11 else "DESAPROBADO")
                nombre = (f"{st.apellido_paterno or ''} {st.apellido_materno or ''}, "
                          f"{st.nombres or ''}").strip(", ").strip().upper()
                vals2 = [row["career_name"], row["semester"], row["course_name"], row["label"],
                         st.num_documento or "", nombre,
                         final if final is not None else "",
                         _cualitativa_de_vigesimal(final) if final is not None else "",
                         estado or "SIN NOTA"]
                for c, v in enumerate(vals2, 1):
                    ws2.cell(row=r2, column=c, value=v)
                r2 += 1

        for wcol, w in zip("ABCDEFGHIJKL", (24, 6, 38, 5, 28, 13, 10, 11, 14, 6, 10, 10)):
            ws.column_dimensions[wcol].width = w
        for wcol, w in zip("ABCDEFGHI", (24, 6, 38, 5, 11, 38, 11, 16, 12)):
            ws2.column_dimensions[wcol].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f'attachment; filename="rendimiento-{period}.xlsx"'},
        )
