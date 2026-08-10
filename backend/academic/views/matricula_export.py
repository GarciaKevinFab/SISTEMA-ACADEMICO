"""
Listas de matrícula por TIPO — panel de descargas del módulo Matrícula.

  GET /academic/reports/matricula-tipo.xlsx?period=2026-I&tipo=regular
  GET /academic/reports/matricula-tipo.pdf?period=2026-I&tipo=subsanacion

`tipo` agrupa según la nómina oficial (pedido de Secretaría):
    regular                  → REGULAR + INGRESANTE
    subsanacion              → SUBSANACION
    reincorporacion-traslado → REINCORPORACION + TRASLADO

Filtros opcionales: career_id, semester (ciclo del alumno).

El tipo sale de `Enrollment.tipo_matricula`; para matrículas anteriores a la
migración (campo vacío) se deriva con las mismas reglas del modelo, así las
listas no salen vacías el primer día.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from rest_framework_simplejwt.authentication import JWTAuthentication

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from academic.models import Enrollment, AcademicGradeRecord
from students.name_utils import clave_orden, nombre_oficial
from .utils import _can_admin_enroll

GRUPOS = {
    "regular": {
        "tipos": {Enrollment.TIPO_REGULAR, Enrollment.TIPO_INGRESANTE},
        "titulo": "MATRÍCULA REGULAR E INGRESANTES",
    },
    "subsanacion": {
        "tipos": {Enrollment.TIPO_SUBSANACION},
        "titulo": "MATRÍCULA POR SUBSANACIÓN",
    },
    "reincorporacion-traslado": {
        "tipos": {Enrollment.TIPO_REINCORPORACION, Enrollment.TIPO_TRASLADO},
        "titulo": "REINCORPORACIÓN Y TRASLADO",
    },
}

_ESPECIALES = {Enrollment.TIPO_SUBSANACION, Enrollment.TIPO_TRASLADO,
               Enrollment.TIPO_REINCORPORACION}

_ROMANS = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]

TIPO_LABEL = dict(Enrollment.TIPO_CHOICES)


def _filas_por_tipo(period, grupo, career_id=None, semester=None):
    """[(enrollment, student, tipo_efectivo)] del grupo pedido, ordenado
    alfabéticamente (español) por apellidos."""
    qs = (Enrollment.objects
          .filter(period=period, status=Enrollment.STATUS_CONFIRMED)
          .select_related("student", "student__plan", "student__plan__career"))
    if career_id:
        try:
            qs = qs.filter(student__plan__career_id=int(career_id))
        except (TypeError, ValueError):
            pass
    if semester:
        try:
            qs = qs.filter(student__ciclo=int(semester))
        except (TypeError, ValueError):
            pass
    enrs = list(qs)

    # Derivación en batch para los que no tienen tipo guardado
    sin_tipo = [e for e in enrs if not e.tipo_matricula]
    con_historia = set()
    if sin_tipo:
        ids = [e.student_id for e in sin_tipo]
        con_historia = set(
            Enrollment.objects
            .filter(student_id__in=ids, status=Enrollment.STATUS_CONFIRMED)
            .exclude(period=period)
            .values_list("student_id", flat=True)
        ) | set(
            AcademicGradeRecord.objects.filter(student_id__in=ids)
            .exclude(term=period).values_list("student_id", flat=True))

    tipos_grupo = GRUPOS[grupo]["tipos"]
    out = []
    for e in enrs:
        tipo = e.tipo_matricula
        if not tipo:
            est = (getattr(e.student, "estado_academico", "") or "").upper()
            if est in _ESPECIALES:
                tipo = est
            else:
                tipo = (Enrollment.TIPO_REGULAR if e.student_id in con_historia
                        else Enrollment.TIPO_INGRESANTE)
        if tipo in tipos_grupo:
            out.append((e, e.student, tipo))
    out.sort(key=lambda t: clave_orden(nombre_oficial(t[1])))
    return out


def _fila_datos(n, enr, st, tipo):
    carrera = (st.plan.career.name if st.plan and st.plan.career
               else (st.programa_carrera or ""))
    try:
        ciclo = _ROMANS[int(st.ciclo)]
    except (TypeError, ValueError, IndexError):
        ciclo = str(st.ciclo or "")
    return [
        n,
        st.num_documento or "",
        nombre_oficial(st),
        carrera.upper(),
        ciclo,
        st.seccion or "A",
        enr.total_credits or "",
        TIPO_LABEL.get(tipo, tipo),
        (getattr(st, "estado_rd", "") or ""),
    ]


HEADERS = ["N°", "N° de Matrícula\n(DNI)", "Apellidos y Nombres", "Programa de Estudios",
           "Ciclo", "Sección", "Créditos", "Tipo", "N° de R.D."]


class MatriculaPorTipoView(APIView):
    """Lista de matriculados de un grupo de tipo, en Excel o PDF (?fmt=)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    fmt = "xlsx"   # sobreescrito por urls.py

    def get(self, request):
        if not _can_admin_enroll(request.user):
            return Response({"detail": "No autorizado."}, status=403)

        period = (request.query_params.get("period") or "").strip().upper()
        grupo = (request.query_params.get("tipo") or "").strip().lower()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        if grupo not in GRUPOS:
            return Response(
                {"detail": f"tipo debe ser uno de: {', '.join(GRUPOS)}"},
                status=400)

        filas = _filas_por_tipo(
            period, grupo,
            career_id=request.query_params.get("career_id"),
            semester=request.query_params.get("semester"))
        if not filas:
            return Response(
                {"detail": f"No hay matriculados de ese tipo en {period} "
                           "con los filtros elegidos."}, status=404)

        titulo = GRUPOS[grupo]["titulo"]
        datos = [_fila_datos(n, e, st, t)
                 for n, (e, st, t) in enumerate(filas, 1)]

        if self.fmt == "pdf":
            return self._pdf(titulo, period, datos)
        return self._xlsx(titulo, period, datos)

    # ── Excel ──
    def _xlsx(self, titulo, period, datos):
        from .acta_excel import (_institution_logo_paths, _add_logo,
                                 _acta_area_inst)
        inst = _acta_area_inst()

        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ncols = len(HEADERS)

        logo, logo2 = _institution_logo_paths()
        ws.row_dimensions[1].height = 46
        _add_logo(ws, logo, "A1", height=56)
        if logo2:
            _add_logo(ws, logo2, f"{get_column_letter(ncols - 1)}1", height=56)

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        t = ws.cell(row=2, column=1, value=inst["nombre"])
        t.font = Font(bold=True, size=12)
        t.alignment = center
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        hoy = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
        s = ws.cell(row=3, column=1,
                    value=f"{titulo} — PERÍODO {period}   ·   Emitido: {hoy}")
        s.font = Font(bold=True, size=10)
        s.alignment = center

        H = 5
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=H, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", start_color="1F4E79")
            cell.alignment = center
            cell.border = border

        for r, fila in enumerate(datos, H + 1):
            for c, v in enumerate(fila, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(size=9)
                cell.border = border
                if c != 3:
                    cell.alignment = Alignment(horizontal="center")

        for c, w in enumerate((5, 13, 38, 26, 7, 8, 9, 15, 16), 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{H + 1}"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        slug = titulo.lower().replace(" ", "-")
        return HttpResponse(
            buf.getvalue(),
            content_type=("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     f'attachment; filename="{slug}-{period}.xlsx"'})

    # ── PDF ──
    def _pdf(self, titulo, period, datos):
        from html import escape as esc
        from .evaluation_pdf import _pdf_shell
        from ..pdf_render import html_to_pdf_bytes

        filas = "".join(
            "<tr>" + "".join(
                f"<td{' class=' + chr(39) + 'c' + chr(39) if i != 2 else ''}>"
                f"{esc(str(v))}</td>"
                for i, v in enumerate(fila)) + "</tr>"
            for fila in datos)
        heads = "".join(f"<th>{esc(h)}</th>" for h in
                        [h.replace("\n", " ") for h in HEADERS])
        hoy = timezone.localtime(timezone.now()).strftime(
            "Documento generado el %d/%m/%Y a las %H:%M")
        cuerpo = f"""
<table>
  <thead><tr>{heads}</tr></thead>
  <tbody>{filas}</tbody>
</table>
<p style="margin-top:6px; font-size:8px; color:#777; text-align:right">
  Total: {len(datos)} estudiante(s) · {hoy}</p>
"""
        html = _pdf_shell(f"{titulo} — {period}", cuerpo, landscape=True)
        slug = titulo.lower().replace(" ", "-")
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="{slug}-{period}.pdf"'})
