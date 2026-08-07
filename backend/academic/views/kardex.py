"""
Vistas para Kardex, Boletas y Expediente Académico
"""
import os
import re
import logging
from datetime import datetime
from io import BytesIO
from django.conf import settings
from django.http import FileResponse, HttpResponse
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework import permissions
from rest_framework_simplejwt.authentication import JWTAuthentication
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

from students.models import Student as StudentProfile
from academic.models import AcademicGradeRecord, PlanCourse
from academic.pdf_render import html_to_pdf_bytes
from .utils import ok, _norm_term, _norm_text, _term_sort_key
from .kardex_helpers import (
    _student_lookup,
    _resolve_plan_for_student,
    _credits_for_student_course,
    _build_pc_name_cache,
    _status_text_from_record,
    _safe_float,
    _get_institution_media_datauris,
    _build_reporte_periodo_ctx,
    _build_record_notas_ctx,
    _build_ficha_rendimiento_ctx,
    build_boleta_full,
    build_boleta_for_period,
    _pick_kardex_template,
    _draw_text,
    _make_overlay_pdf,
    _merge_overlay,
    KARDEX_POS
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════
# HELPER: DETECCIÓN DE STINT ACTIVO (REINGRESO)
# ══════════════════════════════════════════════════════════════

def _period_to_num(p: str):
    """
    Convierte un período a número comparable.
    Usa base *3 por año: I=+0, II=+1, VERANO/0/EXTRA=+2.

    2025-I  → 6075    2025-II → 6076    2025-VERANO → 6077
    2026-0  → 6077  (verano entre 2025-II y 2026-I, ubica como fin 2025)
    2026-I  → 6078    2026-II → 6079
    2025-EXTRAORDINARIO → 6077
    """
    s = (p or "").strip().upper()
    # Regular: YYYY-I, YYYY-II, YYYY-1, YYYY-2
    m = re.match(r'^(\d{4})-(I{1,2}|[12])$', s)
    if m:
        year = int(m.group(1))
        sem = 0 if m.group(2) in ('I', '1') else 1
        return year * 3 + sem
    # YYYY-0: verano entre (YYYY-1)-II y YYYY-I → ubica como fin del año anterior
    m0 = re.match(r'^(\d{4})-0$', s)
    if m0:
        return (int(m0.group(1)) - 1) * 3 + 2
    # VERANO/EXTRAORDINARIO: después de II del mismo año
    m2 = re.match(r'^(\d{4})-(VERANO|EXTRAORDINARIO)$', s)
    if m2:
        return int(m2.group(1)) * 3 + 2
    # Fallback: cualquier formato YYYY-XXX → fin del año
    m3 = re.match(r'^(\d{4})-', s)
    if m3:
        return int(m3.group(1)) * 3 + 2
    return None


def _detect_active_stint_periods(all_periods: set) -> set:
    """
    Detecta el "stint activo": la cadena continua de períodos MÁS RECIENTE.

    Un gap > GAP_THRESHOLD semestres consecutivos sin actividad académica
    indica que el alumno se retiró y luego reingresó → nuevo stint.

    Ejemplo:
        Períodos: {2019-I, 2019-II, 2025-II}
        Gap entre 2019-II y 2025-II = 12 semestres → nuevo stint
        Stint activo = {2025-II}

    GAP_THRESHOLD = 3 semestres (≈ 1.5 años sin cursos).
    """
    GAP_THRESHOLD = 3

    if not all_periods:
        return set(all_periods)

    sortable = []
    for p in all_periods:
        num = _period_to_num(p)
        if num is not None:
            sortable.append((num, p))
    sortable.sort()

    if not sortable:
        return set(all_periods)

    # Construir lista de stints
    stints = []
    current = [sortable[0][1]]

    for i in range(1, len(sortable)):
        prev_num = sortable[i - 1][0]
        curr_num = sortable[i][0]
        gap = curr_num - prev_num

        if gap > GAP_THRESHOLD:
            stints.append(current)
            current = [sortable[i][1]]
        else:
            current.append(sortable[i][1])

    stints.append(current)

    # El stint más reciente es el último de la lista
    return set(stints[-1])


# ══════════════════════════════════════════════════════════════
# CONTROL DE ACCESO AL KÁRDEX
# ══════════════════════════════════════════════════════════════

def _kardex_access_denied(request, student_id):
    """
    El kárdex contiene datos personales y académicos: un ESTUDIANTE solo
    puede consultar el SUYO. Administración (staff/roles académicos) y
    docentes con permiso de kárdex pueden consultar cualquiera.

    Retorna Response 403 si está prohibido, o None si está permitido.
    """
    from .utils import user_has_any_role

    u = request.user
    if getattr(u, "is_staff", False) or getattr(u, "is_superuser", False):
        return None
    if user_has_any_role(u, (
        "ADMIN_SYSTEM", "ADMIN_ACADEMIC", "ADMIN_ACADEMICO", "REGISTRAR",
        "SECRETARIA", "SECRETARIA_ACADEMICA", "DIRECTOR", "TEACHER",
        "DOCENTE", "PROFESOR",
    )):
        return None

    # Docentes: se reconocen también por su perfil (no dependen del rol ACL)
    try:
        from academic.models import Teacher as AcademicTeacher
        from catalogs.models import Teacher as CatalogTeacher
        if (AcademicTeacher.objects.filter(user=u).exists()
                or CatalogTeacher.objects.filter(user=u).exists()):
            return None
    except Exception:
        pass

    # Resto (alumnos): solo su propio kárdex
    mi = StudentProfile.objects.filter(user=u).first()
    if mi:
        pedido = str(student_id).strip()
        if pedido == str(mi.id) or pedido == (mi.num_documento or "").strip():
            return None
    return Response(
        {"detail": "Solo puedes consultar tu propio kárdex."},
        status=403,
    )


# ══════════════════════════════════════════════════════════════
# VISTA PRINCIPAL DE KARDEX (DATOS JSON)
# ══════════════════════════════════════════════════════════════

class KardexView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        # 1) Buscar estudiante
        st = None
        doc = str(student_id).strip()
        if doc.isdigit():
            st = StudentProfile.objects.filter(num_documento=doc).first()
            if not st:
                st = StudentProfile.objects.filter(id=int(doc)).first()
        else:
            st = StudentProfile.objects.filter(num_documento=doc).first()

        if not st:
            return ok(
                student_id=student_id,
                student_name="No encontrado",
                career_name="",
                credits_earned=0,
                gpa=None,
                plan_id=None,
                items=[],
                active_since=None,
                has_prior_enrollment=False,
            )

        # 2) Asegurar plan + refrescar DB
        if not getattr(st, "plan_id", None):
            try:
                _resolve_plan_for_student(st)
            except Exception:
                pass

        st.refresh_from_db()

        student_name = f"{st.apellido_paterno} {st.apellido_materno} {st.nombres}".strip()
        career_name = st.programa_carrera or ""

        # ✅ Pre-cachear PlanCourse por nombre normalizado
        _pc_name_map = _build_pc_name_cache(st.plan_id)

        best_by_key = {}

        recs = (
            AcademicGradeRecord.objects
            .select_related("course", "plan_course")
            .filter(student=st)
            .order_by("-id")
        )

        for rec in recs:
            crs = rec.course
            period = _norm_term(getattr(rec, "term", "") or "")
            course_code = (getattr(crs, "code", "") or "").strip() or f"CRS-{crs.id}"
            code_norm = course_code.strip().upper()

            semester = 0
            weekly_hours = 0
            pc = None

            # Buscar PlanCourse
            if hasattr(rec, "plan_course") and rec.plan_course:
                pc = rec.plan_course
            elif st.plan_id:
                pc = PlanCourse.objects.filter(plan_id=st.plan_id, course_id=crs.id).first()

                if not pc and code_norm:
                    from django.db.models import Q
                    pc = (
                        PlanCourse.objects
                        .select_related("course")
                        .filter(plan_id=st.plan_id)
                        .filter(Q(display_code__iexact=code_norm) | Q(course__code__iexact=code_norm))
                        .first()
                    )

                # Fallback por nombre normalizado (usando cache)
                if not pc:
                    crs_name_norm = _norm_text(getattr(crs, "name", "") or "")
                    if crs_name_norm and crs_name_norm in _pc_name_map:
                        pc = _pc_name_map[crs_name_norm]

            if pc:
                semester = int(getattr(pc, "semester", 0) or 0)
                weekly_hours = int(getattr(pc, "weekly_hours", 0) or 0)

            try:
                grade = None if rec.final_grade is None else float(rec.final_grade)
            except Exception:
                grade = None

            status_txt = _status_text_from_record(rec)

            credits = 0
            if pc:
                credits = int(getattr(pc, "credits", 0) or 0)
            if credits <= 0:
                credits = int(getattr(crs, "credits", 0) or 0)
            credits = max(0, credits)

            item = {
                "period": period,
                "semester": semester,
                "weekly_hours": weekly_hours,
                "course_code": course_code,
                "course_name": getattr(crs, "name", "") or "",
                "credits": credits,
                "grade": grade,
                "status": status_txt,
            }

            key = (period, course_code)

            def score(x):
                g = x.get("grade")
                return -1 if g is None else float(g)

            prev = best_by_key.get(key)
            if (prev is None) or (score(item) > score(prev)):
                best_by_key[key] = item

        # ── Agregar cursos de matrículas activas sin notas aún ──
        try:
            from academic.models import Enrollment, EnrollmentItem
            active_enrollments = Enrollment.objects.filter(
                student=st
            ).exclude(status="CANCELLED")

            for enr in active_enrollments:
                enr_period = _norm_term(enr.period or "")
                if not enr_period:
                    continue
                for ei in EnrollmentItem.objects.select_related(
                    "section__course", "plan_course", "plan_course__course"
                ).filter(enrollment=enr):
                    crs = None
                    pc = ei.plan_course
                    if pc and hasattr(pc, "course"):
                        crs = pc.course
                    elif hasattr(ei, "section") and ei.section and hasattr(ei.section, "course"):
                        crs = ei.section.course

                    if not crs:
                        continue

                    course_code = (getattr(crs, "code", "") or "").strip() or f"CRS-{crs.id}"
                    key = (enr_period, course_code)

                    # Solo agregar si no existe ya un registro de notas para este curso/periodo
                    if key not in best_by_key:
                        semester = int(getattr(pc, "semester", 0) or 0) if pc else 0
                        weekly_hours = int(getattr(pc, "weekly_hours", 0) or 0) if pc else 0
                        credits = int(getattr(ei, "credits", 0) or 0)
                        if credits <= 0 and pc:
                            credits = int(getattr(pc, "credits", 0) or 0)
                        if credits <= 0:
                            credits = int(getattr(crs, "credits", 0) or 0)

                        best_by_key[key] = {
                            "period": enr_period,
                            "semester": semester,
                            "weekly_hours": weekly_hours,
                            "course_code": course_code,
                            "course_name": getattr(crs, "name", "") or "",
                            "credits": max(0, credits),
                            "grade": None,
                            "status": "EN PROCESO",
                        }
        except Exception as exc:
            logger.warning(f"Error agregando cursos matriculados al kardex: {exc}")

        items = list(best_by_key.values())
        items.sort(key=lambda x: (
            str(x.get("period") or ""),
            int(x.get("semester") or 0),
            str(x.get("course_code") or "")
        ))

        # ══════════════════════════════════════════════════════
        # CÁLCULO DE TOTALES — solo desde el stint activo
        # ══════════════════════════════════════════════════════

        # Obtener todos los períodos con registros
        all_periods = {it.get("period") for it in items if it.get("period")}

        # Detectar stint activo (maneja reingreso tras gap)
        active_periods = _detect_active_stint_periods(all_periods)

        # ¿Hay períodos previos al stint activo? → hubo retiro y reingreso
        has_prior_enrollment = len(all_periods - active_periods) > 0

        # El período más antiguo del stint activo
        active_sorted = sorted(active_periods, key=_term_sort_key)
        active_since = active_sorted[0] if active_sorted else None

        credits_earned = 0
        sum_w = 0.0
        sum_c = 0

        for it in items:
            period = it.get("period", "")
            in_active = period in active_periods
            cr = int(it.get("credits") or 0)
            g = it.get("grade")

            # Créditos aprobados: solo del stint activo.
            # Aprobado = nota vigesimal >= 11 (incluye EN PROCESO, LOGRADO y DESTACADO
            # según escala MINEDU RVM N° 123-2022).
            if in_active and it.get("status") in ("LOGRADO", "EN PROCESO", "DESTACADO"):
                credits_earned += cr

            # Ponderado para GPA: solo del stint activo
            if g is not None and cr > 0 and in_active:
                sum_w += float(g) * cr
                sum_c += cr

        gpa = round(sum_w / sum_c, 2) if sum_c > 0 else None

        return ok(
            student_id=st.num_documento,
            student_name=student_name,
            career_name=career_name,
            credits_earned=credits_earned,
            gpa=gpa,
            plan_id=st.plan_id,
            items=items,
            # ── campos adicionales para el frontend ──
            active_since=active_since,
            has_prior_enrollment=has_prior_enrollment,
            active_periods=sorted(active_periods, key=_term_sort_key),
        )


# ══════════════════════════════════════════════════════════════
# EXPORTACIÓN A XLSX (KARDEX)
# ══════════════════════════════════════════════════════════════

class KardexExportXlsxView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        period_q = (request.query_params.get("period") or "").strip()
        
        kv = KardexView()
        kv.request = request
        resp = kv.get(request, student_id)
        data = getattr(resp, "data", None) or {}
        items = data.get("items") or []
        
        if period_q:
            items = [x for x in items if str(x.get("period") or "").strip() == period_q]
        
        template_path = os.path.join(settings.BASE_DIR, "academic", "templates", "kardex_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]
        
        start_row = 14
        for r in range(start_row, start_row + 200):
            for c in range(1, 10):
                ws.cell(r, c).value = None
        
        today = timezone.now().date().isoformat()
        
        def s_int(v):
            try:
                return int(v)
            except Exception:
                return 0
        
        def s_float(v):
            try:
                return float(v)
            except Exception:
                return None
        
        items_sorted = sorted(items, key=lambda x: (s_int(x.get("semester") or 0), str(x.get("course_code") or "")))
        
        for idx, it in enumerate(items_sorted, start=1):
            r = start_row + (idx - 1)
            semester = s_int(it.get("semester") or 0) or ""
            hours = s_int(it.get("weekly_hours") or 0) or ""
            credits = s_int(it.get("credits") or 0) or ""
            grade = s_float(it.get("grade"))
            status_txt = it.get("status") or ""
            
            ws.cell(r, 1).value = semester
            ws.cell(r, 2).value = idx
            ws.cell(r, 3).value = it.get("course_name")
            ws.cell(r, 4).value = hours
            ws.cell(r, 5).value = credits
            ws.cell(r, 6).value = "" if grade is None else grade
            ws.cell(r, 7).value = today
            ws.cell(r, 8).value = status_txt
            ws.cell(r, 9).value = "" if grade is None else grade
        
        filename = f"kardex-{student_id}{('-' + period_q) if period_q else ''}.xlsx"
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        
        return FileResponse(
            bio,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════
# BOLETAS PDF
# ══════════════════════════════════════════════════════════════

class KardexBoletaPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)
        
        student_name = f"{st.apellido_paterno} {st.apellido_materno} {st.nombres}".strip()
        student_code = st.num_documento
        career_name = (st.programa_carrera or "").upper()
        
        grouped = build_boleta_full(st)
        
        page1 = [x for x in grouped if 1 <= int(x["semester"]) <= 5]
        page2 = [x for x in grouped if 6 <= int(x["semester"]) <= 10]
        
        logo_data, second_logo_data, secretary_sig_data = _get_institution_media_datauris(request)
        
        ctx = {
            "student_name": student_name,
            "student_code": student_code,
            "career_name": career_name,
            "logo_url": logo_data,
            "signature_url": secretary_sig_data,
            "page1": page1,
            "page2": page2,
            "period_label": "",
        }
        
        html = render_to_string("kardex/boleta_comunicacion.html", ctx)
        pdf_bytes = html_to_pdf_bytes(html)
        filename = f"boleta-{student_id}.pdf"
        
        return HttpResponse(pdf_bytes, content_type="application/pdf",
                            headers={"Content-Disposition": f'attachment; filename="{filename}"'})


class KardexConstanciaPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        kv = KardexView()
        kv.request = request
        resp = kv.get(request, student_id)
        data = getattr(resp, "data", None) or {}
        
        student_name = data.get("student_name", "") or ""
        student_code = data.get("student_id", "") or ""
        career_name = data.get("career_name", "") or ""
        
        template_name = _pick_kardex_template(career_name)
        template_path = os.path.join(settings.BASE_DIR, "templates", "kardex", template_name)
        
        if not os.path.exists(template_path):
            template_path = os.path.join(settings.BASE_DIR, "templates", "kardex", "inicial.pdf")
        
        tpl_pages = len(PdfReader(template_path).pages)
        
        def draw_fn(c, page_i):
            if page_i == 0:
                _draw_text(c, *KARDEX_POS["name"], student_name, size=10, bold=True)
                _draw_text(c, *KARDEX_POS["code"], student_code, size=9, bold=False)
        
        overlay_reader = _make_overlay_pdf(tpl_pages, draw_fn)
        pdf_bytes = _merge_overlay(template_path, overlay_reader)
        filename = f"constancia-{student_id}.pdf"
        
        return HttpResponse(
            pdf_bytes,
            content_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    
    def post(self, request, student_id):
        return self.get(request, student_id)


class KardexBoletaPeriodoPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        period_q = (request.query_params.get("period") or "").strip()
        if not period_q:
            return Response({"detail": "Falta period (ej: 2018-II)"}, status=400)

        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        pq = _norm_term(period_q)

        try:
            ctx, err = _build_reporte_periodo_ctx(request, st, pq)
            if err:
                raise ValueError(err)
            
            html = render_to_string("kardex/reporte_calificaciones.html", ctx)
            pdf_bytes = html_to_pdf_bytes(html)
            filename = f"reporte-calificaciones-{student_id}-{pq}.pdf"
            
            return HttpResponse(
                pdf_bytes,
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except ValueError as ve:
            logger.info("Sin registros PDF periodo student=%s period=%s", student_id, pq)
            return Response({"detail": "No hay registros", "error": str(ve)}, status=404)
        except Exception as e:
            logger.exception("Error PDF periodo student=%s period=%s", student_id, pq)
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)


class KardexBoletaAnioPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        period_q = (request.query_params.get("period") or "").strip()
        if not period_q:
            return Response({"detail": "Falta period (ej: 2018-I)"}, status=400)

        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        p = _norm_term(period_q)
        m = re.match(r"^(\d{4})-(I|II|1|2)$", p)
        if not m:
            return Response({"detail": "period inválido (ej: 2018-I / 2018-II)"}, status=400)

        year = int(m.group(1))
        periods = [f"{year}-I", f"{year}-II"]

        writer = PdfWriter()
        any_pages = False

        try:
            for per in periods:
                pq = _norm_term(per)
                ctx, err = _build_reporte_periodo_ctx(request, st, pq)
                if err:
                    continue

                html = render_to_string("kardex/reporte_calificaciones.html", ctx)
                pdf_bytes = html_to_pdf_bytes(html)

                reader = PdfReader(BytesIO(pdf_bytes))
                for pg in reader.pages:
                    writer.add_page(pg)
                    any_pages = True

            if not any_pages:
                return Response({"detail": f"No hay reportes para el año {year}."}, status=404)

            out = BytesIO()
            writer.write(out)
            out.seek(0)

            filename = f"reporte-calificaciones-{student_id}-{year}.pdf"
            return HttpResponse(
                out.getvalue(),
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except Exception as e:
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)


class KardexRecordNotasPDFView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        try:
            ctx, err = _build_record_notas_ctx(request, st)
            if err:
                raise ValueError(err)

            html = render_to_string("kardex/record_notas.html", ctx)
            pdf_bytes = html_to_pdf_bytes(html)
            filename = f"record_notas-{student_id}.pdf"

            return HttpResponse(
                pdf_bytes,
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except ValueError as ve:
            return Response({"detail": "No hay registros", "error": str(ve)}, status=404)
        except Exception as e:
            logger.exception("Error PDF certificado student=%s", student_id)
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)


class KardexFichaRendimientoPDFView(APIView):
    """Ficha de Rendimiento Académico (formato oficial MINEDU)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        if err := _kardex_access_denied(request, student_id):
            return err
        st = _student_lookup(student_id)
        if not st:
            return Response({"detail": "Estudiante no encontrado"}, status=404)

        try:
            ctx, err = _build_ficha_rendimiento_ctx(request, st)
            if err:
                raise ValueError(err)

            html = render_to_string("kardex/ficha_rendimiento.html", ctx)
            pdf_bytes = html_to_pdf_bytes(html)
            doc = st.num_documento or student_id
            filename = f"ficha_rendimiento-{doc}.pdf"

            return HttpResponse(
                pdf_bytes,
                content_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except ValueError as ve:
            return Response({"detail": "No hay registros", "error": str(ve)}, status=404)
        except Exception as e:
            logger.exception("Error PDF ficha rendimiento student=%s", student_id)
            return Response({"detail": "Error interno generando PDF", "error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════
# FICHA DE RENDIMIENTO — BULK (ZIP)
# ══════════════════════════════════════════════════════════════

class FichaRendimientoBulkZipView(APIView):
    """
    GET /api/academic/reports/fichas-rendimiento.zip
    Query params (todos opcionales):
        ?career_id=ID       Filtra por carrera
        ?semester=N         Filtra por ciclo actual del alumno
        ?period=2026-I      Filtra por período actual del alumno
        ?only_with_grades=1 (default) Solo alumnos con notas registradas

    Empaqueta una ficha de rendimiento PDF por estudiante en un ZIP.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        import zipfile
        from io import BytesIO as ZipBuf

        career_id = request.query_params.get("career_id")
        semester = request.query_params.get("semester")
        period = (request.query_params.get("period") or "").strip()
        only_with_grades = str(
            request.query_params.get("only_with_grades", "1")
        ).lower() in ("1", "true", "yes")

        qs = StudentProfile.objects.select_related("plan", "plan__career")

        if career_id:
            try:
                qs = qs.filter(plan__career_id=int(career_id))
            except (TypeError, ValueError):
                return Response({"detail": "career_id inválido"}, status=400)

        if semester:
            try:
                qs = qs.filter(ciclo=int(semester))
            except (TypeError, ValueError):
                return Response({"detail": "semester inválido"}, status=400)

        if period:
            qs = qs.filter(periodo=period)

        if only_with_grades:
            qs = qs.filter(grade_records__isnull=False).distinct()

        qs = qs.order_by("apellido_paterno", "apellido_materno", "nombres")
        students = list(qs)

        if not students:
            return Response(
                {"detail": "No hay estudiantes con notas para los filtros indicados"},
                status=404,
            )

        zip_buf = ZipBuf()
        generated = 0
        errors_list = []

        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for st in students:
                try:
                    ctx, err = _build_ficha_rendimiento_ctx(request, st)
                    if err:
                        errors_list.append({
                            "dni": st.num_documento or str(st.id),
                            "name": f"{st.apellido_paterno} {st.nombres}".strip(),
                            "error": err,
                        })
                        continue

                    html = render_to_string("kardex/ficha_rendimiento.html", ctx)
                    pdf_bytes = html_to_pdf_bytes(html)
                    doc = st.num_documento or str(st.id)
                    safe_name = (
                        f"{st.apellido_paterno or ''}_{st.apellido_materno or ''}_{st.nombres or ''}"
                        .replace(" ", "_").replace("/", "-").strip("_")
                    )
                    fname = f"ficha_rendimiento-{doc}-{safe_name}.pdf"
                    zf.writestr(fname, pdf_bytes)
                    generated += 1
                except Exception as exc:
                    logger.exception("Error generando ficha bulk student=%s", st.id)
                    errors_list.append({
                        "dni": st.num_documento or str(st.id),
                        "name": f"{st.apellido_paterno} {st.nombres}".strip(),
                        "error": str(exc),
                    })

            # Reporte de errores
            if errors_list:
                report_lines = [
                    "FICHAS DE RENDIMIENTO — ERRORES",
                    "================================",
                    f"Total estudiantes:    {len(students)}",
                    f"Fichas generadas:     {generated}",
                    f"Estudiantes con error:{len(errors_list)}",
                    "",
                    "Detalle:",
                ]
                for e in errors_list:
                    report_lines.append(f"- {e['dni']} · {e['name']}: {e['error']}")
                zf.writestr("_ERRORES.txt", "\n".join(report_lines))

        if generated == 0:
            return Response(
                {
                    "detail": "No se pudo generar ninguna ficha.",
                    "errors": errors_list[:20],
                },
                status=500,
            )

        zip_buf.seek(0)
        suffix_parts = []
        if career_id:
            suffix_parts.append(f"carrera{career_id}")
        if semester:
            suffix_parts.append(f"ciclo{semester}")
        if period:
            suffix_parts.append(period)
        suffix = "_".join(suffix_parts) or "todos"
        fname = f"fichas_rendimiento_{suffix}.zip"
        return HttpResponse(
            zip_buf.getvalue(),
            content_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )