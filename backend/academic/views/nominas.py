"""
academic/views/nominas.py
═══════════════════════════════════════════════════════════════
Generadores de Nóminas de Matrícula (formato oficial MINEDU) y
Data Completa de Estudiantes filtrados por Carrera + Ciclo + Periodo.

Endpoints:
  GET /api/academic/reports/nominas.xlsx
       ?career_id=X&semester=N&period=2026-I[&seccion=A&turno=MAÑANA]
  GET /api/academic/reports/students-data.xlsx
       ?career_id=X&semester=N&period=2026-I

El primero genera el Excel oficial siguiendo la plantilla "NÓMINAS 2025-II.xls"
del MINEDU (cabecera con datos institucionales, tabla de estudiantes con
N° Orden, N° Matrícula, Apellidos y Nombres, Sexo, Fecha Nac, Edad, etc.,
y resumen al pie).

El segundo es un dump completo de TODOS los campos de cada estudiante
para uso interno.
"""
import io
import os
import warnings
from datetime import date, datetime

from django.conf import settings
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Silenciar warning de DataValidation extension al cargar la plantilla MINEDU
warnings.filterwarnings(
    "ignore", message=".*Data Validation extension is not supported.*"
)

# Mapa Carrera → Resolución del Plan de Estudios (catálogo MINEDU)
CAREER_RVM_MAP = {
    "EDUCACION INICIAL": "Resolución Viceministerial N° 163-2019-MINEDU",
    "EDUCACION PRIMARIA": "Resolución Viceministerial N° 204-2019-MINEDU",
    "EDUCACION FISICA": "Resolución Viceministerial N° 147-2020-MINEDU",
    "COMUNICACION": "Resolución Viceministerial N° 143-2020-MINEDU",
    "EDUCACION INICIAL INTERCULTURAL BILINGUE": "Resolución Viceministerial N° 252-2019-MINEDU",
    "EDUCACION PRIMARIA INTERCULTURAL BILINGUE": "Resolución Viceministerial N° 252-2019-MINEDU",
    "COMPUTACION E INFORMATICA": "Resolución Directoral N° 0223-2012-ED",
}


def _norm_career_key(name: str) -> str:
    import unicodedata
    s = (name or "").upper().strip()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _rvm_for_career(career_name: str) -> str:
    return CAREER_RVM_MAP.get(_norm_career_key(career_name), "")


# Mapa cyclo (int) → "CICLO I", etc.
CICLO_TXT = {i: f"CICLO {r}" for i, r in enumerate(
    ["I","II","III","IV","V","VI","VII","VIII","IX","X"], start=1)}

from students.models import Student
from academic.models import InstitutionSettings, Plan, Enrollment
from catalogs.models import Career
from catalogs.models import InstitutionSetting as CatalogInstitutionSetting


# ──────────────────────────────────────────────────────────────
#  Helpers de datos institucionales y logo
# ──────────────────────────────────────────────────────────────

def _resolve_ubigeo_names(dept_code: str, prov_code: str, dist_code: str) -> dict:
    """Convierte códigos de ubigeo (INEI) a nombres legibles."""
    out = {"region": "", "provincia": "", "distrito": ""}
    if not (dept_code or prov_code or dist_code):
        return out
    try:
        from catalogs.views.ubigeo import _load_ubigeo_pe
        ub = _load_ubigeo_pe() or {}
        dep = ub.get(str(dept_code) or "", {}) if dept_code else {}
        if dep:
            out["region"] = (dep.get("name") or "").strip()
            provs = dep.get("provinces") or {}
            prov = provs.get(str(prov_code) or "", {}) if prov_code else {}
            if prov:
                out["provincia"] = (prov.get("name") or "").strip()
                dists = prov.get("districts") or {}
                if dist_code and str(dist_code) in dists:
                    val = dists[str(dist_code)]
                    out["distrito"] = (val.get("name") if isinstance(val, dict) else str(val) or "").strip()
    except Exception:
        pass
    return out


def _get_institution_data() -> dict:
    """Combina InstitutionSettings (academic) + InstitutionSetting (catalogs JSON).

    El UI de Administración guarda TODO en `InstitutionSetting.data` (catalog
    JSON) — incluido name, ruc, address, department/province/district (códigos
    de ubigeo), bank_*, etc. Aquí preferimos esos valores y caemos a la tabla
    `InstitutionSettings` (academic) solo como fallback.

    Retorna dict con todas las llaves útiles para la nómina:
      name, ruc, address, dre, ugel, codigo_modular, denominacion,
      gestion, ds_creacion, region, provincia, distrito, director_name,
      director_resolution, logo_url
    """
    out = {
        "name": "",
        "ruc": "",
        "address": "",
        "dre": "DREJ",
        "ugel": "UGEL TARMA",
        "codigo_modular": "0609370",
        "denominacion": "IESP",
        "gestion": "Público",
        "ds_creacion": "D.S. 059-1984-ED",
        "region": "Junín",
        "provincia": "Tarma",
        "distrito": "Tarma",
        "director_name": "",
        "director_resolution": "",
        "secretary_name": "",
        "rvm": "",
        "logo_url": "",
        "phone": "",
        "email": "",
        "website": "",
    }

    # 1) InstitutionSetting (catalogs JSON) — fuente principal del UI
    cat = CatalogInstitutionSetting.objects.filter(pk=1).first()
    cat_data = cat.data if (cat and isinstance(cat.data, dict)) else {}

    def _grab(key):
        v = cat_data.get(key)
        if isinstance(v, str):
            return v.strip()
        return v or ""

    # Identidad / contacto / dirección — claves directas del UI
    for k in ("name", "ruc", "address", "phone", "email", "website",
              "dre", "ugel", "codigo_modular", "denominacion", "gestion",
              "ds_creacion", "region", "provincia", "distrito",
              "director_name", "director_resolution", "secretary_name",
              "rvm", "logo_url"):
        v = _grab(k)
        if v:
            out[k] = v

    # Autoridades y resoluciones: el UI de Institución las guarda con las
    # MISMAS claves que usan las actas (director_name / rd_encargatura /
    # resolucion_autorizacion) — la nómina salía con "—" porque buscaba
    # otras claves. Fallbacks con los datos institucionales vigentes.
    if not out["director_name"]:
        out["director_name"] = "GARCIA PORRAS, MARIA ELVIRA"
    if not out["director_resolution"]:
        out["director_resolution"] = _grab("rd_encargatura") or "R.D.R. N° 017-2026-DREJ"
    if not out["rvm"]:
        out["rvm"] = (_grab("resolucion_autorizacion")
                      or "R.D. N° 306-2016-MINEDU/VMGP/DIGEDD/DIFOID")

    # Si en el catálogo hay códigos de ubigeo (department/province/district)
    # y los nombres legibles aún no están seteados, resolverlos desde el JSON
    # de ubigeo_pe.
    needs_ub = not (cat_data.get("region") and cat_data.get("provincia") and cat_data.get("distrito"))
    if needs_ub and (cat_data.get("department") or cat_data.get("province") or cat_data.get("district")):
        names = _resolve_ubigeo_names(
            cat_data.get("department") or "",
            cat_data.get("province") or "",
            cat_data.get("district") or "",
        )
        if names["region"] and not cat_data.get("region"):
            out["region"] = names["region"]
        if names["provincia"] and not cat_data.get("provincia"):
            out["provincia"] = names["provincia"]
        if names["distrito"] and not cat_data.get("distrito"):
            out["distrito"] = names["distrito"]

    # 2) InstitutionSettings (academic) — fallback para campos que no estén en
    #    el catálogo (instalaciones más antiguas).
    inst = InstitutionSettings.objects.first()
    if inst:
        if not out["name"]:
            out["name"] = (inst.name or "").strip()
        if not out["ruc"]:
            out["ruc"] = (inst.ruc or "").strip()
        if not out["address"]:
            out["address"] = (inst.address or "").strip()
        if not out["logo_url"]:
            out["logo_url"] = (inst.logo_url or "").strip()
        if not out["phone"]:
            out["phone"] = (getattr(inst, "phone", "") or "").strip()
        if not out["email"]:
            out["email"] = (getattr(inst, "email", "") or "").strip()
        if not out["website"]:
            out["website"] = (getattr(inst, "website", "") or "").strip()

    if not out["name"]:
        out["name"] = "GUSTAVO ALLENDE LLAVERIA"

    return out


def _media_url_to_abs_path(media_url: str):
    """Convierte URL /media/... a path absoluto en disco."""
    if not media_url:
        return None
    s = str(media_url).strip()
    rel = s.split("/media/", 1)[1] if "/media/" in s else s.lstrip("/")
    return os.path.join(str(settings.MEDIA_ROOT), rel)


def _logo_abs_path(inst_data: dict):
    """Devuelve path absoluto del logo institucional, o None."""
    p = _media_url_to_abs_path(inst_data.get("logo_url") or "")
    if p and os.path.isfile(p):
        return p
    # Fallback: cualquier logo*.png en media/institution/
    inst_dir = os.path.join(str(settings.MEDIA_ROOT), "institution")
    if os.path.isdir(inst_dir):
        for fn in sorted(os.listdir(inst_dir)):
            if fn.lower().startswith("logo") and fn.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                p = os.path.join(inst_dir, fn)
                if os.path.isfile(p):
                    return p
    return None


# ──────────────────────────────────────────────────────────────
#  Helpers de query
# ──────────────────────────────────────────────────────────────

def _sistema_logo_path():
    """Logo del SISTEMA (Tarma) para el lado derecho del membrete."""
    try:
        from academic.views.acta_excel import _institution_logo_paths
        _inst, sist = _institution_logo_paths()
        return sist if sist and os.path.isfile(sist) else None
    except Exception:
        return None


def _tipos_matricula(students, period):
    """student_id → tipo de SU matrícula confirmada del período (para G/R)."""
    ids = [s.id for s in students]
    m = {}
    try:
        for e in Enrollment.objects.filter(
                student_id__in=ids, period=str(period).strip(),
                status=Enrollment.STATUS_CONFIRMED,
        ).values("student_id", "tipo_matricula"):
            m[e["student_id"]] = (e["tipo_matricula"] or "").upper()
    except Exception:
        pass
    return m


def _gp_de(tipo):
    """Columna 'Gratuito o Pagante': institución PÚBLICA → todos G (gratuito);
    matrícula de reincorporación → G/R (indicación de Secretaría)."""
    return "G/R" if tipo == "REINCORPORACION" else "G"


def _resolve_career(career_id) -> Career | None:
    if not career_id:
        return None
    try:
        return Career.objects.filter(pk=int(career_id)).first()
    except (TypeError, ValueError):
        return None


def _filter_students(career_id, semester, period, only_enrolled=True,
                     solo_regulares=False):
    """Estudiantes de una carrera + ciclo + periodo.

    Si `only_enrolled=True` (default), filtra SOLO los estudiantes que tienen
    Enrollment CONFIRMADO en ese periodo (matriculados de verdad).
    Si `only_enrolled=False`, usa solo los campos `Student.ciclo` y
    `Student.periodo` (lista de candidatos / admitidos).

    Con `solo_regulares=True` se excluyen los estudiantes de SUBSANACIÓN: en la
    Nómina de Matrícula figuran únicamente los regulares, los de subsanación
    llevan sus actas aparte (indicación de Secretaría Académica). Sí deben
    seguir apareciendo en el acta del curso, que no usa este filtro.
    """
    qs = Student.objects.select_related("plan", "plan__career", "user")

    if solo_regulares:
        qs = qs.exclude(estado_academico=Student.ESTADO_SUBSANACION)

    if career_id:
        career = _resolve_career(career_id)
        if career:
            qs = qs.filter(plan__career_id=career.id)

    if semester:
        try:
            qs = qs.filter(ciclo=int(semester))
        except (TypeError, ValueError):
            pass

    if only_enrolled and period:
        # Solo estudiantes con matrícula CONFIRMADA en el periodo
        qs = qs.filter(
            enrollments__period=str(period).strip(),
            enrollments__status=Enrollment.STATUS_CONFIRMED,
        ).distinct()
    elif period:
        qs = qs.filter(periodo=str(period).strip())

    return qs.order_by("apellido_paterno", "apellido_materno", "nombres")


def _calc_age(fecha_nac):
    if not fecha_nac:
        return ""
    today = date.today()
    return today.year - fecha_nac.year - (
        (today.month, today.day) < (fecha_nac.month, fecha_nac.day)
    )


def _full_name_apellidos_primero(st):
    parts = [
        (st.apellido_paterno or "").strip(),
        (st.apellido_materno or "").strip(),
        (st.nombres or "").strip(),
    ]
    return ", ".join([" ".join(parts[:2]).strip(), parts[2]]).strip(", ").strip()


SEMESTER_ROMAN = {
    1: "I", 2: "II", 3: "III", 4: "IV", 5: "V",
    6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X",
}


# ──────────────────────────────────────────────────────────────
#  Vista 1: NÓMINA MINEDU
# ──────────────────────────────────────────────────────────────

class NominasMatriculaXlsxView(APIView):
    """Genera la Nómina de Matrícula oficial en formato MINEDU."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        career_id = request.query_params.get("career_id")
        semester = request.query_params.get("semester")
        period = (request.query_params.get("period") or "").strip()
        seccion = (request.query_params.get("seccion") or "A").strip()
        turno = (request.query_params.get("turno") or "MAÑANA").strip()
        # ?include_admitted=1 → incluir admitidos aunque no estén matriculados
        include_admitted = str(
            request.query_params.get("include_admitted", "")
        ).lower() in ("1", "true", "yes")

        if not career_id:
            return Response({"detail": "Falta career_id"}, status=400)
        if not period:
            return Response({"detail": "Falta period (e.g. 2026-I)"}, status=400)

        career = _resolve_career(career_id)
        if not career:
            return Response({"detail": "Carrera no encontrada"}, status=404)

        students = list(_filter_students(
            career_id, semester, period, only_enrolled=not include_admitted,
            solo_regulares=True,   # los de subsanación llevan actas aparte
        ))

        if not students:
            mode = "matriculados" if not include_admitted else "registrados"
            return Response(
                {"detail": f"No hay estudiantes {mode} para "
                           f"{career.name} ciclo {semester} en {period}"},
                status=404,
            )

        inst = _get_institution_data()
        try:
            sem_int = int(semester) if semester else 0
        except (TypeError, ValueError):
            sem_int = 0
        sem_roman = SEMESTER_ROMAN.get(sem_int, "")

        wb = Workbook()
        ws = wb.active
        ws.title = f"{career.name[:25]} {sem_roman}"[:31]

        # ── Estilos ──
        bold = Font(name="Calibri", bold=True, size=10)
        bold_lg = Font(name="Calibri", bold=True, size=14)
        normal = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        thin = Side(style="thin", color="000000")
        bx = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Anchos ajustados para que cada bloque del encabezado quepa bien
        widths = {
            "A": 3.5, "B": 3.0, "C": 2.5, "D": 2.5, "E": 2.5, "F": 2.5, "G": 2.0,
            "H": 2.5, "I": 4.4, "J": 6.0, "K": 9.0, "L": 14.5, "M": 9.0,
            "N": 11.0, "O": 2.5, "P": 11.0, "Q": 14.0, "R": 9.0,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Helper para aplicar borde a un rango de celdas (incluye merges)
        def apply_border(rng):
            for row in ws[rng]:
                for cell in row:
                    cell.border = bx

        def merge_set(rng, value, font=None, alignment=None, fill=None, fmt=None):
            ws.merge_cells(rng)
            top_left = rng.split(":")[0]
            c = ws[top_left]
            c.value = value
            c.font = font or normal
            c.alignment = alignment or center
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            apply_border(rng)
            return c

        # ── Logos (institucional a la izquierda, sistema a la derecha) ──
        logo_path = _logo_abs_path(inst)
        if logo_path:
            try:
                img = XLImage(logo_path)
                img.width = 95
                img.height = 95
                ws.add_image(img, "A1")
            except Exception:
                pass
        sist_path = _sistema_logo_path()
        if sist_path:
            try:
                img2 = XLImage(sist_path)
                img2.width = 95
                img2.height = 95
                ws.add_image(img2, "Q1")
            except Exception:
                pass
        # Reservar altura para los logos (filas 1-4)
        for r in range(1, 5):
            ws.row_dimensions[r].height = 22

        # ── Título A5:R5 ──
        ws.merge_cells("A5:R5")
        c = ws["A5"]
        c.value = "NÓMINA DE MATRÍCULA"
        c.font = bold_lg
        c.alignment = center
        ws.row_dimensions[5].height = 26

        # ══════════════════════════════════════════════════════════
        # BLOQUE INSTITUCIONAL (filas 7-10) — replicando merges originales
        # ══════════════════════════════════════════════════════════
        # Fila 7-8 spanned: A7:I8 = "Nombre de la Institución"
        merge_set("A7:I8", "Nombre de la Institución", font=bold, alignment=center)
        # J7:M8 = nombre de la institución (valor)
        merge_set("J7:M8", inst["name"], font=bold, alignment=center)
        # N7:O7 = "DRE"   |   P7:R7 = nombre DRE (DREJ)
        merge_set("N7:O7", "DRE", font=bold, alignment=center)
        merge_set("P7:R7", inst["dre"], font=bold, alignment=center)
        # N8:O8 = "UGEL"  |   P8:R8 = ugel
        merge_set("N8:O8", "UGEL", font=bold, alignment=center)
        merge_set("P8:R8", inst["ugel"], font=bold, alignment=center)

        # Filas 9-10: alturas para wrap-text
        ws.row_dimensions[9].height = 30
        ws.row_dimensions[10].height = 22

        # Fila 9: cabeceras
        merge_set("A9:G9", "Código Modular", font=bold, alignment=center)
        merge_set("H9:J9", "Denominación", font=bold, alignment=center)
        merge_set("K9:K9", "Gestión", font=bold, alignment=center)
        merge_set("L9:L9", "D.S./R.M. de Creación",
                  font=bold, alignment=center)
        merge_set("M9:M9", "Dirección", font=bold, alignment=center)
        merge_set("N9:R9", inst["address"] or "—",
                  font=normal, alignment=center)

        # Fila 10: valores + provincia/distrito
        merge_set("A10:G10", inst["codigo_modular"], font=normal, alignment=center)
        merge_set("H10:J10", inst["denominacion"], font=normal, alignment=center)
        merge_set("K10:K10", inst["gestion"], font=normal, alignment=center)
        merge_set("L10:L10", inst["ds_creacion"], font=normal, alignment=center)
        merge_set("M10:M10", "Provincia", font=bold, alignment=center)
        merge_set("N10:O10", inst["provincia"], font=normal, alignment=center)
        merge_set("P10:P10", "Distrito", font=bold, alignment=center)
        merge_set("Q10:R10", inst["distrito"], font=normal, alignment=center)

        # ══════════════════════════════════════════════════════════
        # BLOQUE PROGRAMA (filas 12-13)
        # ══════════════════════════════════════════════════════════
        # Estructura idéntica a la plantilla:
        #   A:I = etiqueta, J:N = valor, P:Q = etiqueta derecha, R = valor derecho
        ws.row_dimensions[12].height = 24
        ws.row_dimensions[13].height = 22

        merge_set("A12:I12", "Programa de estudios / Turno",
                  font=bold, alignment=center)
        merge_set("J12:N12", f"{career.name} / TURNO: {turno}",
                  font=bold, alignment=center)
        merge_set("P12:Q12", "Periodo Académico",
                  font=bold, alignment=center)
        # Separar año del semestre: "2026-I" → "2026 - I"
        period_pretty = period.replace("-", " - ") if "-" in period else period
        ws["R12"] = period_pretty
        ws["R12"].font = bold
        ws["R12"].alignment = center
        ws["R12"].border = bx

        merge_set("A13:I13", "Resolución de Autorización",
                  font=bold, alignment=center)
        merge_set("J13:N13", inst.get("rvm") or "R.D.",
                  font=normal, alignment=center)
        merge_set("P13:Q13", "Ciclo - Sección",
                  font=bold, alignment=center)
        ws["R13"] = (
            f'{sem_roman} - "{seccion}"' if sem_roman else f'"{seccion}"'
        )
        ws["R13"].font = bold
        ws["R13"].alignment = center
        ws["R13"].border = bx

        # ══════════════════════════════════════════════════════════
        # BLOQUE DIRECTOR (fila 15)
        # ══════════════════════════════════════════════════════════
        ws.row_dimensions[15].height = 22
        merge_set("A15:I15", "Director (e) General", font=bold, alignment=center)
        merge_set("J15:M15", (inst["director_name"] or "").upper() or "—",
                  font=bold, alignment=center)
        merge_set("N15:P15", "R.D. de Nombramiento o Encargatura",
                  font=bold, alignment=center)
        merge_set("Q15:R15", inst["director_resolution"] or "—",
                  font=normal, alignment=center)

        # ══════════════════════════════════════════════════════════
        # CABECERA DE TABLA (fila 17) — con merges
        # ══════════════════════════════════════════════════════════
        ws.row_dimensions[17].height = 38
        merge_set("A17:B17", "N° Orden", font=bold, alignment=center, fill=hdr_fill)
        merge_set("C17:I17", "N° Matrícula\n(DNI)", font=bold, alignment=center, fill=hdr_fill)
        merge_set("J17:M17", "APELLIDOS Y NOMBRES (Por orden Alfabético)",
                  font=bold, alignment=center, fill=hdr_fill)
        merge_set("N17:O17", "Gratuito o Pagante",
                  font=bold, alignment=center, fill=hdr_fill)
        # P, Q, R individuales (se eliminó "Fecha de Matrícula" S)
        for ref, val in (
            ("P17", "Sexo H/M"),
            ("Q17", "Fecha de Nacimiento"),
            ("R17", "Edad"),
        ):
            ws[ref] = val
            ws[ref].font = bold
            ws[ref].alignment = center
            ws[ref].fill = hdr_fill
            ws[ref].border = bx

        # ══════════════════════════════════════════════════════════
        # FILAS DE DATOS — con merges por estudiante
        # ══════════════════════════════════════════════════════════
        START = 18
        hombres = mujeres = 0
        gratuitos = pagantes = 0
        today_str = date.today().strftime("%d-%m-%Y")
        tipos = _tipos_matricula(students, period)

        for i, st in enumerate(students):
            r = START + i
            ws.row_dimensions[r].height = 18

            sexo = (st.sexo or "").strip().upper()
            if sexo in ("M", "MASCULINO", "H"):
                hombres += 1
                sexo_letra = "H"
            elif sexo in ("F", "FEMENINO"):
                mujeres += 1
                sexo_letra = "M"
            else:
                sexo_letra = ""

            gratuitos += 1

            # N° Orden (A:B)
            merge_set(f"A{r}:B{r}", i + 1, font=normal, alignment=center)
            # N° Matrícula DNI (C:I)
            merge_set(f"C{r}:I{r}", st.num_documento, font=normal, alignment=center)
            # APELLIDOS Y NOMBRES (J:M) — siempre en MAYÚSCULAS
            merge_set(f"J{r}:M{r}",
                      _full_name_apellidos_primero(st).upper(),
                      font=normal, alignment=left)
            # Gratuito o Pagante (N:O) — G, o G/R si es reincorporación
            merge_set(f"N{r}:O{r}", _gp_de(tipos.get(st.id, "")),
                      font=normal, alignment=center)
            # Sexo
            ws[f"P{r}"] = sexo_letra
            ws[f"P{r}"].font = normal
            ws[f"P{r}"].alignment = center
            ws[f"P{r}"].border = bx
            # Fecha Nac
            ws[f"Q{r}"] = st.fecha_nac
            ws[f"Q{r}"].font = normal
            ws[f"Q{r}"].alignment = center
            ws[f"Q{r}"].border = bx
            if st.fecha_nac:
                ws[f"Q{r}"].number_format = "DD/MM/YYYY"
            # Edad
            ws[f"R{r}"] = _calc_age(st.fecha_nac)
            ws[f"R{r}"].font = normal
            ws[f"R{r}"].alignment = center
            ws[f"R{r}"].border = bx

        # ══════════════════════════════════════════════════════════
        # RESUMEN AL PIE
        # ══════════════════════════════════════════════════════════
        last = START + len(students)
        sum_row = last + 1
        # Cabecera "Resumen / Total"
        merge_set(f"A{sum_row}:C{sum_row}", "Resumen", font=bold, alignment=center, fill=hdr_fill)
        merge_set(f"D{sum_row}:F{sum_row}", "", font=bold, alignment=center, fill=hdr_fill)
        merge_set(f"G{sum_row}:I{sum_row}", "Total", font=bold, alignment=center, fill=hdr_fill)

        # El total siempre debe coincidir con la cantidad real de estudiantes
        # del cuadro principal (independiente de campos vacíos en sexo).
        total_alumnos = len(students)
        rows_def = [
            ("Hombres",  hombres,            total_alumnos),
            ("Mujeres",  mujeres,            None),
            ("Gratuitos", gratuitos,         total_alumnos),
            ("Pagantes", pagantes,           None),
        ]
        for k, (label, count, total) in enumerate(rows_def, start=1):
            rr = sum_row + k
            merge_set(f"A{rr}:C{rr}", label, font=bold, alignment=center)
            merge_set(f"D{rr}:F{rr}", count, font=normal, alignment=center)
            if total is not None:
                merge_set(f"G{rr}:I{rr}", total, font=normal, alignment=center)
            else:
                merge_set(f"G{rr}:I{rr}", "", font=normal, alignment=center)

        # Lugar y fecha
        loc_row = sum_row + len(rows_def) + 1
        merge_set(
            f"P{loc_row}:R{loc_row}",
            f"{inst['provincia'].upper()}, {today_str}",
            font=bold, alignment=center,
        )

        # Print settings
        try:
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "17:17"
        except Exception:
            pass

        # ── Guardar y responder ──
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        fname = (
            f"NOMINA_{career.name.replace(' ', '_')}"
            f"_{sem_roman}_{period}.xlsx"
        )
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


# ──────────────────────────────────────────────────────────────
#  Vista 1-B: NÓMINA MINEDU (PDF)
# ──────────────────────────────────────────────────────────────

class NominasMatriculaPDFView(APIView):
    """Versión PDF de la Nómina de Matrícula (mismo contenido y orden que el
    Excel, en A4 horizontal con cabecera institucional, programa, director y
    tabla alfabética de estudiantes + resumen al pie)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        career_id = request.query_params.get("career_id")
        semester = request.query_params.get("semester")
        period = (request.query_params.get("period") or "").strip()
        seccion = (request.query_params.get("seccion") or "A").strip()
        turno = (request.query_params.get("turno") or "MAÑANA").strip()
        include_admitted = str(
            request.query_params.get("include_admitted", "")
        ).lower() in ("1", "true", "yes")

        if not career_id:
            return Response({"detail": "Falta career_id"}, status=400)
        if not period:
            return Response({"detail": "Falta period (e.g. 2026-I)"}, status=400)

        career = _resolve_career(career_id)
        if not career:
            return Response({"detail": "Carrera no encontrada"}, status=404)

        students = list(_filter_students(
            career_id, semester, period, only_enrolled=not include_admitted,
            solo_regulares=True,   # los de subsanación llevan actas aparte
        ))
        if not students:
            mode = "matriculados" if not include_admitted else "registrados"
            return Response(
                {"detail": f"No hay estudiantes {mode} para "
                           f"{career.name} ciclo {semester} en {period}"},
                status=404,
            )

        inst = _get_institution_data()
        try:
            sem_int = int(semester) if semester else 0
        except (TypeError, ValueError):
            sem_int = 0
        sem_roman = SEMESTER_ROMAN.get(sem_int, "")
        period_pretty = period.replace("-", " - ") if "-" in period else period
        today_str = date.today().strftime("%d-%m-%Y")
        logo_path = _logo_abs_path(inst)

        # ── Conteos ──
        hombres = mujeres = 0
        for st in students:
            s = (st.sexo or "").strip().upper()
            if s in ("M", "MASCULINO", "H"):
                hombres += 1
            elif s in ("F", "FEMENINO"):
                mujeres += 1

        # ══════════════════════════════════════════════════════════
        # PDF formato oficial vertical (A4) — paleta clara, tablas
        # tipo "Excel" y firmas Director / Secretario / V°B° al pie.
        # ══════════════════════════════════════════════════════════
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            BaseDocTemplate, PageTemplate, Frame, Table, TableStyle,
            Paragraph, Spacer, Image, KeepTogether,
        )

        # ── Paleta clara (al estilo del Excel oficial) ──
        INK      = colors.HexColor("#1F2937")  # tinta principal
        LINE     = colors.HexColor("#1F2937")  # bordes finos
        LBL_BG   = colors.HexColor("#EFF3F8")  # fondo suave para labels
        HEAD_BG  = colors.HexColor("#DCE6F1")  # encabezados de tabla
        ROW_ALT  = colors.HexColor("#F7F9FC")  # cebra muy suave
        GREY_TXT = colors.HexColor("#475569")

        bio = io.BytesIO()
        page_w, page_h = A4
        margin_x = 1.0 * cm
        margin_top = 1.0 * cm
        margin_bottom = 1.0 * cm

        doc = BaseDocTemplate(
            bio, pagesize=A4,
            leftMargin=margin_x, rightMargin=margin_x,
            topMargin=margin_top, bottomMargin=margin_bottom,
            title=f"Nómina {career.name} {sem_roman} {period}",
            author="Sistema Académico",
        )
        frame = Frame(
            margin_x, margin_bottom,
            page_w - 2 * margin_x, page_h - margin_top - margin_bottom,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            id="content",
        )
        doc.addPageTemplates([PageTemplate(id="main", frames=[frame])])

        styles = getSampleStyleSheet()
        def PS(name, **kw):
            return ParagraphStyle(name, parent=styles["Normal"], **kw)

        sty_h1       = PS("h1", fontName="Helvetica-Bold", fontSize=14,
                          alignment=TA_CENTER, leading=18, textColor=INK)
        sty_lbl      = PS("lbl", fontName="Helvetica-Bold", fontSize=7,
                          alignment=TA_CENTER, leading=8.5, textColor=INK)
        sty_lbl_l    = PS("lbll", fontName="Helvetica-Bold", fontSize=7,
                          alignment=TA_LEFT, leading=8.5, textColor=INK)
        sty_val      = PS("val", fontName="Helvetica", fontSize=7.5,
                          alignment=TA_CENTER, leading=9, textColor=INK)
        sty_val_l    = PS("vall", fontName="Helvetica", fontSize=7.5,
                          alignment=TA_LEFT, leading=9, textColor=INK)
        sty_val_bold = PS("valb", fontName="Helvetica-Bold", fontSize=7.5,
                          alignment=TA_CENTER, leading=9, textColor=INK)
        sty_th       = PS("th", fontName="Helvetica-Bold", fontSize=7,
                          alignment=TA_CENTER, leading=8.5, textColor=INK)
        sty_td       = PS("td", fontName="Helvetica", fontSize=7,
                          alignment=TA_CENTER, leading=8.5, textColor=INK)
        sty_td_name  = PS("tdn", fontName="Helvetica-Bold", fontSize=7,
                          alignment=TA_LEFT, leading=8.5, textColor=INK)
        sty_sig      = PS("sig", fontName="Helvetica-Bold", fontSize=8,
                          alignment=TA_CENTER, leading=10, textColor=INK)
        sty_sig_sub  = PS("sigs", fontName="Helvetica", fontSize=6.5,
                          alignment=TA_CENTER, leading=8, textColor=GREY_TXT)

        story = []
        avail_w = page_w - 2 * margin_x   # ≈ 19 cm

        # ════════════════════════════════════════════════════
        # ENCABEZADO — logo + título grande "NÓMINA DE MATRÍCULA"
        # ════════════════════════════════════════════════════
        if logo_path and os.path.isfile(logo_path):
            try:
                logo_img = Image(logo_path, width=2.0 * cm, height=2.0 * cm)
                logo_img.hAlign = "CENTER"
            except Exception:
                logo_img = ""
        else:
            logo_img = ""

        # Logo del SISTEMA (Tarma) al lado derecho, como los demás documentos
        sist_path = _sistema_logo_path()
        sist_img = ""
        if sist_path:
            try:
                sist_img = Image(sist_path, width=2.0 * cm, height=2.0 * cm)
                sist_img.hAlign = "CENTER"
            except Exception:
                sist_img = ""

        hdr_tbl = Table(
            [[logo_img,
              Paragraph("NÓMINA DE MATRÍCULA", sty_h1),
              sist_img]],
            colWidths=[2.5 * cm, avail_w - 5.0 * cm, 2.5 * cm],
        )
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 0.2 * cm))

        # ════════════════════════════════════════════════════
        # BLOQUE INSTITUCIONAL  (replicando el Excel oficial)
        # ════════════════════════════════════════════════════
        # 8 columnas: lbl·val·lbl·val·lbl·val·lbl·val (cada par es 25% del ancho)
        col_w = [
            2.6*cm, 4.2*cm,   # par 1
            1.6*cm, 2.4*cm,   # par 2
            1.6*cm, 1.6*cm,   # par 3
            1.6*cm, avail_w - (2.6+4.2+1.6+2.4+1.6+1.6+1.6)*cm,  # par 4
        ]

        # Fila 1: Nombre Institución (col 0-1) | "VALOR" (col 2-5) | DREJ + valor (6-7)
        # Fila 2:                       continuation         | UGEL + valor (6-7)
        # Fila 3: Dirección spans (col 1-7) full address
        # Fila 4: Cód Modular | Denominación | Gestión | D.S.R.M. de Creación
        # Fila 5: 0609370 | IESP | Público | D.S. 059-1984-ED
        # Fila 6: (label) Provincia | Tarma | (label) Distrito | Tarma
        # — Simplificamos a 8 columnas con SPAN donde aplica.
        inst_rows = [
            # Fila 1
            [Paragraph("Nombre de la Institución", sty_lbl),
             Paragraph(f"<b>{inst['name']}</b>", sty_val), "", "", "", "",
             Paragraph("DREJ", sty_lbl), Paragraph(inst.get("dre") or "—", sty_val)],
            # Fila 2 (UGEL en columna derecha)
            [Paragraph("Dirección", sty_lbl),
             Paragraph(inst.get("address") or "—", sty_val_l), "", "", "", "",
             Paragraph("UGEL", sty_lbl), Paragraph(inst.get("ugel") or "—", sty_val)],
            # Fila 3 (encabezados de datos modulares)
            [Paragraph("Código Modular", sty_lbl),
             Paragraph("Denominación", sty_lbl), "",
             Paragraph("Gestión", sty_lbl), "",
             Paragraph("D.S.R.M. de Creación y R.D. de Revalidación", sty_lbl), "", ""],
            # Fila 4 (valores)
            [Paragraph(inst.get("codigo_modular") or "—", sty_val),
             Paragraph(inst.get("denominacion") or "—", sty_val), "",
             Paragraph(inst.get("gestion") or "—", sty_val), "",
             Paragraph(inst.get("ds_creacion") or "—", sty_val), "", ""],
            # Fila 5 (Provincia + Distrito)
            [Paragraph("Provincia", sty_lbl),
             Paragraph(inst.get("provincia") or "—", sty_val), "", "",
             Paragraph("Distrito", sty_lbl),
             Paragraph(inst.get("distrito") or "—", sty_val), "", ""],
        ]
        inst_tbl = Table(inst_rows, colWidths=col_w)
        inst_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, LINE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # SPAN: valor del nombre institución 4 columnas
            ("SPAN", (1, 0), (5, 0)),
            # SPAN: dirección 4 columnas
            ("SPAN", (1, 1), (5, 1)),
            # SPAN: D.S.R.M. de Creación 3 columnas (label fila 3)
            ("SPAN", (5, 2), (7, 2)),
            ("SPAN", (5, 3), (7, 3)),  # valor
            # SPAN denominación/gestión (cada uno 2 cols)
            ("SPAN", (1, 2), (2, 2)),
            ("SPAN", (1, 3), (2, 3)),
            ("SPAN", (3, 2), (4, 2)),
            ("SPAN", (3, 3), (4, 3)),
            # SPAN provincia (2 cols) + distrito (2 cols)
            ("SPAN", (1, 4), (3, 4)),
            ("SPAN", (5, 4), (7, 4)),
            # Backgrounds suaves en labels
            ("BACKGROUND", (0, 0), (0, 1), LBL_BG),
            ("BACKGROUND", (6, 0), (6, 1), LBL_BG),
            ("BACKGROUND", (0, 2), (-1, 2), LBL_BG),
            ("BACKGROUND", (0, 4), (0, 4), LBL_BG),
            ("BACKGROUND", (4, 4), (4, 4), LBL_BG),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(inst_tbl)

        # ════════════════════════════════════════════════════
        # BLOQUE PROGRAMA + AUTORIDADES
        # ════════════════════════════════════════════════════
        ciclo_seccion = f'{sem_roman} - "{seccion}"' if sem_roman else f'"{seccion}"'
        # Nombre del programa + RVM (heurística por carrera)
        program_label = career.name
        if inst.get("rvm"):
            # Insertar RVM corto si existe (ej. "RVM. 163-2019-MINEDU")
            rvm_short = inst["rvm"].replace("Resolución Viceministerial",
                                            "RVM.").replace("  ", " ").strip()
            program_label = f"{career.name} ({rvm_short})"

        prog_rows = [
            # Fila A: Programa | valor (4 cols) | Turno label/val | Periodo label/val
            [Paragraph("Programa de estudios / Turno", sty_lbl),
             Paragraph(f"<b>{program_label}</b>", sty_val), "", "",
             Paragraph(f"TURNO: {turno}", sty_lbl),
             Paragraph("Periodo Académico", sty_lbl),
             Paragraph(f"<b>{period_pretty}</b>", sty_val), ""],
            # Fila B: Resolución | valor (4 cols) | Ciclo-Sec | valor
            [Paragraph("Resolución de Autorización", sty_lbl),
             Paragraph(inst.get("rvm") or "R.D.", sty_val), "", "", "",
             Paragraph("Ciclo - Sección", sty_lbl),
             Paragraph(f"<b>{ciclo_seccion}</b>", sty_val), ""],
            # Fila C: Director (e) | NOMBRE | R.D. de Encargatura | resolución
            [Paragraph("Director General (e)", sty_lbl),
             Paragraph(f"<b>{(inst.get('director_name') or '—').upper()}</b>",
                       sty_val), "", "", "",
             Paragraph("R.D. de Encargatura", sty_lbl),
             Paragraph(inst.get("director_resolution") or "—", sty_val), ""],
        ]
        prog_tbl = Table(prog_rows, colWidths=col_w)
        prog_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, LINE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # SPANs
            ("SPAN", (1, 0), (3, 0)),
            ("SPAN", (6, 0), (7, 0)),
            ("SPAN", (1, 1), (4, 1)),
            ("SPAN", (6, 1), (7, 1)),
            ("SPAN", (1, 2), (4, 2)),
            ("SPAN", (6, 2), (7, 2)),
            # Backgrounds suaves
            ("BACKGROUND", (0, 0), (0, -1), LBL_BG),
            ("BACKGROUND", (4, 0), (4, 0), LBL_BG),
            ("BACKGROUND", (5, 0), (5, -1), LBL_BG),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(prog_tbl)
        story.append(Spacer(1, 0.25 * cm))

        # alias de placeholder para mantener compatibilidad con el bloque siguiente
        sty_inst_sub = sty_val  # noqa: usado más abajo en _stat_card legacy
        # ════════════════════════════════════════════════════
        # TABLA DE ESTUDIANTES  (al estilo Excel, con 20 filas mín.)
        # ════════════════════════════════════════════════════
        thead = [
            Paragraph("N°<br/>Orden", sty_th),
            Paragraph("N° Matrícula<br/>(DNI)", sty_th),
            Paragraph("APELLIDOS Y NOMBRES<br/><font size='6'>(Por orden Alfabético)</font>", sty_th),
            Paragraph("Gratuito o<br/>Pagante", sty_th),
            Paragraph("Sexo<br/>H/M", sty_th),
            Paragraph("Fecha de<br/>Nacimiento", sty_th),
            Paragraph("Edad", sty_th),
        ]
        tipos = _tipos_matricula(students, period)
        table_data = [thead]
        for i, st in enumerate(students, start=1):
            sexo = (st.sexo or "").strip().upper()
            if sexo in ("M", "MASCULINO", "H"):
                sexo_letra = "H"
            elif sexo in ("F", "FEMENINO"):
                sexo_letra = "M"
            else:
                sexo_letra = ""
            fnac = st.fecha_nac.strftime("%d/%m/%Y") if st.fecha_nac else ""
            table_data.append([
                Paragraph(str(i), sty_td),
                Paragraph(str(st.num_documento or ""), sty_td),
                Paragraph(_full_name_apellidos_primero(st).upper(), sty_td_name),
                # Institución pública: todos GRATUITOS; reincorporación → G/R
                Paragraph(_gp_de(tipos.get(st.id, "")), sty_td),
                Paragraph(sexo_letra, sty_td),
                Paragraph(fnac, sty_td),
                Paragraph(str(_calc_age(st.fecha_nac) or ""), sty_td),
            ])

        # Rellenar hasta 20 filas con vacíos (look del Excel oficial)
        MIN_ROWS = 20
        for i in range(len(students) + 1, MIN_ROWS + 1):
            table_data.append([
                Paragraph(str(i), sty_td),
                "", "", "", "", "", "",
            ])

        # Anchos
        s_col_n   = 0.9 * cm
        s_col_dni = 2.3 * cm
        s_col_gp  = 1.4 * cm
        s_col_sex = 1.0 * cm
        s_col_fn  = 2.3 * cm
        s_col_ed  = 1.0 * cm
        s_col_name = avail_w - (s_col_n + s_col_dni + s_col_gp + s_col_sex + s_col_fn + s_col_ed)
        s_col_widths = [s_col_n, s_col_dni, s_col_name, s_col_gp, s_col_sex, s_col_fn, s_col_ed]

        students_tbl = Table(table_data, colWidths=s_col_widths, repeatRows=1)
        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), HEAD_BG),
            ("TEXTCOLOR",  (0, 0), (-1, 0), INK),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
            ("ALIGN",      (2, 1), (2, -1), "LEFT"),
            ("GRID",       (0, 0), (-1, -1), 0.4, LINE),
            ("LEFTPADDING",   (0, 0), (-1, -1), 3),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
            ("TOPPADDING",    (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]
        # Cebra muy suave solo en filas con dato real
        for row_idx in range(1, len(students) + 1):
            if row_idx % 2 == 0:
                ts.append(("BACKGROUND", (0, row_idx), (-1, row_idx), ROW_ALT))
        students_tbl.setStyle(TableStyle(ts))
        story.append(students_tbl)
        story.append(Spacer(1, 0.4 * cm))

        # ════════════════════════════════════════════════════
        # RESUMEN (4 filas: Hombres / Mujeres / Gratuitos / Pagantes)
        # alineado a la izquierda, con la fecha a la derecha
        # ════════════════════════════════════════════════════
        total_alumnos = len(students)
        gratuitos = total_alumnos   # institución pública: todos gratuitos
        pagantes = 0

        sum_inner = Table([
            [Paragraph("Resumen", sty_lbl), Paragraph("Total", sty_lbl)],
            [Paragraph("Hombres",   sty_lbl_l), Paragraph(str(hombres),   sty_val)],
            [Paragraph("Mujeres",   sty_lbl_l), Paragraph(str(mujeres),   sty_val)],
            [Paragraph("Gratuitos", sty_lbl_l), Paragraph(str(gratuitos), sty_val)],
            [Paragraph("Pagantes",  sty_lbl_l), Paragraph(str(pagantes),  sty_val)],
        ], colWidths=[2.6 * cm, 1.6 * cm])
        sum_inner.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, LINE),
            ("BACKGROUND", (0, 0), (-1, 0), HEAD_BG),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))

        loc_label = (inst.get("provincia") or "").upper()
        date_para = Paragraph(
            f"<para align='right'>{loc_label}, {today_str}</para>",
            sty_val_l,
        )

        sum_row = Table(
            [[sum_inner, "", date_para]],
            colWidths=[4.2 * cm, avail_w - 4.2 * cm - 6.0 * cm, 6.0 * cm],
        )
        sum_row.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(sum_row)
        story.append(Spacer(1, 0.9 * cm))

        # ════════════════════════════════════════════════════
        # FIRMAS — Director / Secretario / V°B° DRE-UGEL
        # ════════════════════════════════════════════════════
        dir_name = (inst.get("director_name") or "").upper() or "—"
        sec_name = (inst.get("secretary_name") or "").upper() or "—"

        # Tres firmas con SEPARACIÓN real entre líneas (columnas de aire)
        gap = 1.0 * cm
        sig_col_w = (avail_w - 2 * gap) / 3

        sig_tbl = Table([
            ["", "", "", "", ""],
            [Paragraph(dir_name, sty_sig), "",
             Paragraph(sec_name, sty_sig), "",
             Paragraph("", sty_sig)],
            [Paragraph("DIRECTOR(A) GENERAL", sty_sig), "",
             Paragraph("SECRETARIO(A) ACADÉMICO", sty_sig), "",
             Paragraph("V° B° DRE / UGEL", sty_sig)],
            [Paragraph("Firma, Post Firma y Sello", sty_sig_sub), "",
             Paragraph("Firma, Post Firma y Sello", sty_sig_sub), "",
             Paragraph("Firma y Sello", sty_sig_sub)],
        ], colWidths=[sig_col_w, gap, sig_col_w, gap, sig_col_w])
        sig_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN",  (0, 0), (-1, -1), "CENTER"),
            # Línea de firma SOLO sobre las columnas de firma (no el aire)
            ("LINEABOVE", (0, 1), (0, 1), 0.8, LINE),
            ("LINEABOVE", (2, 1), (2, 1), 0.8, LINE),
            ("LINEABOVE", (4, 1), (4, 1), 0.8, LINE),
            ("TOPPADDING", (0, 1), (-1, 1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 42),  # espacio para firmar
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(sig_tbl)

        try:
            doc.build(story)
        except Exception as exc:
            return Response({"detail": f"Error generando PDF: {exc}"}, status=500)

        bio.seek(0)
        fname = (
            f"NOMINA_{career.name.replace(' ', '_')}"
            f"_{sem_roman}_{period}.pdf"
        )
        resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


# ──────────────────────────────────────────────────────────────
#  Vista 2: DATA COMPLETA
# ──────────────────────────────────────────────────────────────

class StudentsDataXlsxView(APIView):
    """Dump completo de TODOS los campos de cada estudiante filtrado."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        career_id = request.query_params.get("career_id")
        semester = request.query_params.get("semester")
        period = (request.query_params.get("period") or "").strip()
        # ?only_enrolled=1 → solo matriculados confirmados.
        # Por default trae TODOS los del ciclo+periodo (admitidos + matriculados)
        only_enrolled = str(
            request.query_params.get("only_enrolled", "")
        ).lower() in ("1", "true", "yes")

        students = list(_filter_students(
            career_id, semester, period, only_enrolled=only_enrolled,
        ))

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Estudiantes"

        bold = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                               fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="B0B0B0")
        bx = Border(left=thin, right=thin, top=thin, bottom=thin)

        cols = [
            ("N°", 5),
            ("DNI", 12),
            ("Apellido Paterno", 18),
            ("Apellido Materno", 18),
            ("Nombres", 22),
            ("Sexo", 6),
            ("Fecha Nacimiento", 14),
            ("Edad", 6),
            ("Email", 25),
            ("Celular", 14),
            ("Carrera", 30),
            ("Plan", 32),
            ("Ciclo", 8),
            ("Sección", 10),
            ("Periodo", 10),
            ("Turno", 10),
            ("Región", 14),
            ("Provincia", 14),
            ("Distrito", 14),
            ("Código Modular Origen", 14),
            ("Institución de Origen", 28),
            ("Gestión", 12),
            ("Tipo Inst.", 12),
            ("Lengua Materna", 14),
            ("Discapacidad", 12),
            ("Tipo Discapacidad", 22),
            ("Username", 14),
            ("Activo", 8),
            ("Creado", 16),
        ]
        for i, (name, w) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.font = bold
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = bx
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for i, st in enumerate(students, start=2):
            user = st.user
            row = [
                i - 1,
                st.num_documento,
                st.apellido_paterno,
                st.apellido_materno,
                st.nombres,
                st.sexo,
                st.fecha_nac,
                _calc_age(st.fecha_nac),
                st.email or (user.email if user else ""),
                st.celular,
                st.programa_carrera,
                st.plan.name if st.plan else "",
                st.ciclo,
                st.seccion,
                st.periodo,
                st.turno,
                st.region,
                st.provincia,
                st.distrito,
                st.codigo_modular,
                st.nombre_institucion,
                st.gestion,
                st.tipo,
                st.lengua,
                st.discapacidad,
                st.tipo_discapacidad,
                user.username if user else "",
                "SI" if (user and user.is_active) else "NO",
                st.created_at.strftime("%Y-%m-%d %H:%M") if st.created_at else "",
            ]
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = bx
                if j == 7 and st.fecha_nac:
                    cell.number_format = "DD/MM/YYYY"

        # Print
        try:
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(students) + 1}"
            ws.print_title_rows = "1:1"
        except Exception:
            pass

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        career = _resolve_career(career_id)
        career_part = career.name.replace(" ", "_") if career else "TODAS"
        sem_part = SEMESTER_ROMAN.get(int(semester), str(semester or "X")) \
            if semester else "X"
        fname = f"DATA_ESTUDIANTES_{career_part}_{sem_part}_{period or 'TODOS'}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp



# ──────────────────────────────────────────────────────────────
#  Vista 3: REPORTE MINEDU SIA (formato oficial)
# ──────────────────────────────────────────────────────────────

# Path a la plantilla MINEDU oficial (con catálogos -AYUDA-, -MAESTRO GENERAL-,
# -PERIODO ACADÉMICO-, -LENGUA MATERNA-, -IDENTIFICACIÓN ÉTNICA-,
# -CARRERAS-PROGRAMAS-, -CICLO-)
MINEDU_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "..", "templates_xlsx", "reporte_minedu_template.xlsx",
)


def _sex_minedu(sexo: str) -> str:
    s = (sexo or "").strip().upper()
    if s in ("M", "MASCULINO", "H"):
        return "MASCULINO"
    if s in ("F", "FEMENINO"):
        return "FEMENINO"
    return ""


def _yes_no_minedu(val) -> str:
    if val is None or val == "":
        return "NO"
    s = str(val).strip().upper()
    if s in ("SI", "SÍ", "YES", "TRUE", "1"):
        return "SÍ"
    return "NO"


class ReporteMineduXlsxView(APIView):
    """Genera el Reporte de Matrícula en formato oficial MINEDU SIA.

    Usa la plantilla `reporte_minedu_template.xlsx` (incluye todas las hojas
    de catálogo necesarias) y rellena la hoja principal con los estudiantes
    matriculados del periodo + carrera + ciclo seleccionados.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        career_id = request.query_params.get("career_id")
        semester = request.query_params.get("semester")
        period = (request.query_params.get("period") or "").strip()
        seccion = (request.query_params.get("seccion") or "A").strip().upper()
        turno = (request.query_params.get("turno") or "MAÑANA").strip().upper()
        include_admitted = str(
            request.query_params.get("include_admitted", "")
        ).lower() in ("1", "true", "yes")

        if not career_id:
            return Response({"detail": "Falta career_id"}, status=400)
        if not period:
            return Response({"detail": "Falta period (e.g. 2026-I)"}, status=400)

        career = _resolve_career(career_id)
        if not career:
            return Response({"detail": "Carrera no encontrada"}, status=404)

        students = list(_filter_students(
            career_id, semester, period, only_enrolled=not include_admitted,
        ))

        if not students:
            mode = "matriculados" if not include_admitted else "registrados"
            return Response(
                {"detail": f"No hay estudiantes {mode} para "
                           f"{career.name} ciclo {semester} en {period}"},
                status=404,
            )

        if not os.path.isfile(MINEDU_TEMPLATE_PATH):
            return Response(
                {"detail": f"Plantilla MINEDU no encontrada en "
                           f"{MINEDU_TEMPLATE_PATH}"},
                status=500,
            )

        # Cargar plantilla (preserva catálogos en otras hojas)
        wb = load_workbook(MINEDU_TEMPLATE_PATH)
        ws = wb.worksheets[0]

        try:
            ws.title = f"REPORTE DE INFORMACIÓN {period}"[:31]
        except Exception:
            pass

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        inst = _get_institution_data()
        try:
            year = period.split("-")[0]
            default_resolution = (
                inst.get("director_resolution") or
                f'RDI N° 031-{year}-DG-IESPP"GALL"T'
            )
        except Exception:
            default_resolution = inst.get("director_resolution") or ""

        plan_rvm = _rvm_for_career(career.name) or (
            inst.get("rvm") or "Resolución Viceministerial N° 163-2019-MINEDU"
        )

        try:
            ciclo_int = int(semester) if semester else 1
        except (TypeError, ValueError):
            ciclo_int = 1
        ciclo_txt = CICLO_TXT.get(ciclo_int, f"CICLO {ciclo_int}")

        for i, st in enumerate(students, start=1):
            r = i + 1
            row_values = [
                str(i),
                "MATRÍCULA REGULAR",
                default_resolution,
                career.name,
                plan_rvm,
                st.turno.upper() if st.turno else turno,
                (st.seccion or seccion).upper(),
                ciclo_txt,
                "DNI",
                int(st.num_documento) if st.num_documento and
                str(st.num_documento).isdigit() else st.num_documento,
                "PERUANA",
                (st.nombres or "").upper(),
                (st.apellido_paterno or "").upper(),
                (st.apellido_materno or "").upper(),
                _sex_minedu(st.sexo),
                st.fecha_nac,
                "",
                (st.lengua or "CASTELLANO").upper(),
                "MESTIZO",
                "",
                "",
                st.celular or "",
                "NO",
                "",
                _yes_no_minedu(st.discapacidad),
                (st.tipo_discapacidad or "").upper(),
                "",
                "SOLTERO(A)",
                "NO",
                "",
                "NO",
                "",
                "",
                "NO",
                "",
                "NO",
                "",
                "",
                "",
                "NO",
                "",
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                if col_idx == 16 and st.fecha_nac:
                    cell.number_format = "DD/MM/YYYY"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        career_part = career.name.replace(" ", "_")
        sem_part = SEMESTER_ROMAN.get(ciclo_int, str(ciclo_int))
        fname = f"REPORTE_MINEDU_{career_part}_{sem_part}_{period}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp



# ──────────────────────────────────────────────────────────────
#  Vista 4: OFICIO DE MATRICULADOS (PDF con membrete)
# ──────────────────────────────────────────────────────────────

class OficioMatriculadosPDFView(APIView):
    """Genera el oficio en formato PDF con diseño de hoja membretada
    (logo, cinta lateral, marca de agua, firma del director)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .oficio_generator import generate_oficio_pdf

        inst = _get_institution_data()
        logo_path = _logo_abs_path(inst)

        # Firma: leer signature_url del catálogo o academic.InstitutionSettings
        sig_url = ""
        cat = CatalogInstitutionSetting.objects.filter(pk=1).first()
        if cat and isinstance(cat.data, dict):
            sig_url = (cat.data.get("signature_url") or "").strip()
        if not sig_url:
            inst_a = InstitutionSettings.objects.first()
            if inst_a:
                sig_url = (inst_a.signature_url or "").strip()
        sig_path = _media_url_to_abs_path(sig_url) if sig_url else None
        # Fallback: cualquier firma*.png en media/institution/
        if not sig_path or not os.path.isfile(sig_path):
            inst_dir = os.path.join(str(settings.MEDIA_ROOT), "institution")
            if os.path.isdir(inst_dir):
                for fn in sorted(os.listdir(inst_dir)):
                    fl = fn.lower()
                    if (fl.startswith("firma") or "firma" in fl) and fl.endswith((".png",".jpg",".jpeg")):
                        cand = os.path.join(inst_dir, fn)
                        if os.path.isfile(cand):
                            sig_path = cand
                            break

        qp = request.query_params
        period = (qp.get("period") or "").strip() or f"{date.today().year}-I"

        # Construir payload combinando defaults + query params
        payload = {
            "oficio_number": (qp.get("oficio_number") or "").strip(),
            "oficio_year": (qp.get("oficio_year") or str(date.today().year)).strip(),
            "lema": (qp.get("lema") or "").strip(),
            "city_date": (qp.get("city_date") or "").strip(),
            "recipient_treatment": (qp.get("recipient_treatment") or "Señor:").strip(),
            "recipient_name": (qp.get("recipient_name") or "").strip(),
            "recipient_position": (qp.get("recipient_position") or "").strip(),
            "recipient_city": (qp.get("recipient_city") or "LIMA").strip(),
            "asunto": (qp.get("asunto") or f"REPORTE DE ESTUDIANTES MATRICULADOS {period}").strip(),
            "atencion_name": (qp.get("atencion_name") or "").strip(),
            "atencion_position": (qp.get("atencion_position") or "").strip(),
            "period": period,
        }

        # Permitir override de body via query param multilinea separado por "||"
        body_raw = (qp.get("body") or "").strip()
        if body_raw:
            payload["body_paragraphs"] = [
                p.strip() for p in body_raw.split("||") if p.strip()
            ]

        try:
            pdf_bytes = generate_oficio_pdf(
                payload=payload,
                inst_data=inst,
                logo_path=logo_path,
                signature_path=sig_path,
            )
        except Exception as exc:
            return Response(
                {"detail": f"Error generando PDF: {exc}"},
                status=500,
            )

        oficio_num = payload["oficio_number"] or "SN"
        fname = f"Oficio_{oficio_num}_{payload['oficio_year']}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp
