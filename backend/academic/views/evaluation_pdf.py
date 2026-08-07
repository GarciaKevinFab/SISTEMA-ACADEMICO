"""
Evaluación — versiones PDF de los documentos y reportes de méritos.

  PDF (Chromium via pdf_render.html_to_pdf_bytes):
    GET sections/<id>/acta-area.pdf              Acta de Evaluación de Área (docente/admin)
    GET admin/evaluation/actas.pdf               Acta consolidada
    GET admin/evaluation/rendimiento.pdf         Reporte de rendimiento
    GET admin/evaluation/primeros-lugares        (?fmt=pdf|xlsx) top 3 por carrera+ciclo
    GET admin/evaluation/tercio-quinto           (?tipo=tercio|quinto&fmt=) promoción 10mo
    GET admin/evaluation/constancias-beca.zip    constancias PDF (promedio >= 17)

  Excel complementarios:
    GET admin/evaluation/boletas.xlsx            detalle alumno×curso del período
    GET admin/evaluation/fichas.xlsx             resumen alumno×término (todos los períodos)
"""
import html as html_mod
import zipfile
from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from academic.models import Section, AcademicGradeRecord, PlanCourse
from students.models import Student
from academic.pdf_render import html_to_pdf_bytes

from .evaluation import (
    _require_grades_admin, _sections_for, _bundle_map, _section_eval_row,
    _entry_for, _final_of,
)
from .acta_excel import (
    _roster, _acta_area_inst, _roman, _cualitativa_de_vigesimal,
    _section_header_info, _xlsx_response,
)

import logging
logger = logging.getLogger(__name__)


def _esc(v):
    return html_mod.escape(str(v if v is not None else ""))


def _logo_datauris():
    try:
        from .kardex_helpers import _get_institution_media_datauris
        logo, logo2, _sig = _get_institution_media_datauris()
        return logo or "", logo2 or ""
    except Exception:
        return "", ""


def _pdf_shell(titulo, cuerpo, landscape=False):
    """Marco HTML común: membrete con logos + título + contenido."""
    logo, logo2 = _logo_datauris()
    inst = _acta_area_inst()
    size = "A4 landscape" if landscape else "A4"
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  @page {{ size: {size}; margin: 12mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #111; margin: 0; }}
  .head {{ display: flex; align-items: center; gap: 10px; border-bottom: 2px solid #1F4E79; padding-bottom: 6px; }}
  .head img {{ height: 52px; }}
  .head .tit {{ flex: 1; text-align: center; }}
  .head .tit h1 {{ font-size: 14px; margin: 0; color: #1F4E79; }}
  .head .tit p {{ margin: 2px 0 0; font-size: 9px; color: #444; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 8px; }}
  th, td {{ border: 1px solid #555; padding: 3px 5px; }}
  th {{ background: #1F4E79; color: #fff; font-size: 9px; }}
  td {{ font-size: 9.5px; }}
  .c {{ text-align: center; }}
  h2 {{ font-size: 11px; color: #1F4E79; margin: 12px 0 2px; }}
  .grupo {{ page-break-inside: avoid; }}
  .salto {{ page-break-before: always; }}
  .firma {{ margin-top: 34px; text-align: center; }}
  .firma .linea {{ display: inline-block; border-top: 1px solid #111; padding-top: 3px; min-width: 230px; font-size: 9px; }}
</style></head><body>
<div class="head">
  {f'<img src="{logo}">' if logo else ''}
  <div class="tit">
    <h1>{_esc(inst['nombre'])}</h1>
    <p>{_esc(titulo)}</p>
  </div>
  {f'<img src="{logo2}">' if logo2 else ''}
</div>
{cuerpo}
</body></html>"""


# ══════════════════════════════════════════════════════════════
# Promedios ponderados
# ══════════════════════════════════════════════════════════════

def _creditos_de(rec):
    pc = rec.plan_course
    if pc and getattr(pc, "credits", 0):
        return int(pc.credits)
    c = getattr(rec.course, "credits", 0)
    return int(c or 0)


def _promedios_por_alumno(term=None, student_ids=None):
    """{student_id: promedio ponderado} para un término (o todos si term=None)."""
    qs = AcademicGradeRecord.objects.select_related("plan_course", "course")
    if term:
        qs = qs.filter(term=term)
    if student_ids is not None:
        qs = qs.filter(student_id__in=list(student_ids))
    acc = {}
    for rec in qs:
        try:
            g = float(rec.final_grade)
        except (TypeError, ValueError):
            continue
        cr = _creditos_de(rec) or 1
        p, c = acc.get(rec.student_id, (0.0, 0))
        acc[rec.student_id] = (p + g * cr, c + cr)
    return {sid: round(p / c, 2) for sid, (p, c) in acc.items() if c}


def _nombre(st):
    return (f"{st.apellido_paterno or ''} {st.apellido_materno or ''}, "
            f"{st.nombres or ''}").strip(", ").strip().upper()


def _filtrar_students(period, career_id=None, semester=None, anio=None):
    qs = (Student.objects.filter(grade_records__term=period)
          .select_related("plan", "plan__career").distinct())
    if career_id:
        try:
            qs = qs.filter(plan__career_id=int(career_id))
        except (TypeError, ValueError):
            pass
    if semester:
        try:
            qs = qs.filter(ciclo=int(semester))
        except (TypeError, ValueError):
            pass
    if anio:
        try:
            n = int(anio)
            qs = qs.filter(ciclo__in=[2 * n - 1, 2 * n])
        except (TypeError, ValueError):
            pass
    return list(qs)


# ══════════════════════════════════════════════════════════════
# 1. Acta de Evaluación de Área — PDF (docente y admin)
# ══════════════════════════════════════════════════════════════

def _acta_area_pdf_bytes(sec):
    from academic.models import SectionGrades
    curso, codigo, docente = _section_header_info(sec)
    pc = sec.plan_course
    career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "").upper()
    creditos = getattr(pc, "credits", 0) or 0
    ciclo = _roman(pc.semester if pc else "")
    inst = _acta_area_inst()

    bundle = SectionGrades.objects.filter(section=sec).first()
    grades = (bundle.grades or {}) if bundle else {}
    alumnos = _roster(sec)
    if not alumnos:
        return None, f"La sección no tiene alumnos matriculados en {sec.period}."

    def _entry(st):
        for key in (st["id"], st.get("pk")):
            if key is not None and isinstance(grades.get(str(key)), dict):
                return grades[str(key)]
        return {}

    def _final(entry):
        if (entry.get("status") or "").upper() == "DPI":
            return 0
        for k in ("final_grade", "PROMEDIO_FINAL", "FINAL"):
            try:
                f = float(entry.get(k))
                if 0 <= f <= 20:
                    return int(round(f))
            except (TypeError, ValueError):
                continue
        return None

    filas = []
    finales = []
    for n, st in enumerate(alumnos, 1):
        e = _entry(st)
        f = _final(e)
        finales.append(f)
        lic = st.get("estado") == "LICENCIA"
        calif = "LICENCIA" if lic else (f if f is not None else "")
        filas.append(
            f"<tr><td class='c'>{n}</td><td class='c'>{_esc(st['dni'])}</td>"
            f"<td>{_esc(st['nombre'])}</td>"
            f"<td class='c'>{calif}</td>"
            f"<td class='c'>{creditos if f is not None else ''}</td>"
            f"<td class='c'>{round(f * creditos) if f is not None else ''}</td>"
            f"<td class='c'>{_esc(_cualitativa_de_vigesimal(f)) if f is not None and not lic else ''}</td></tr>"
        )

    aprobados = sum(1 for f in finales if f is not None and f >= 11)
    desaprob = sum(1 for f in finales if f is not None and f < 11)
    n_dpi = sum(1 for st in alumnos if (_entry(st).get("status") or "").upper() == "DPI")

    cuerpo = f"""
<h2 style="text-align:center">ACTA DE EVALUACIÓN DE ÁREA {_esc(curso.upper())}</h2>
<table style="margin-top:4px">
  <tr><td><b>Programa de estudios</b></td><td>{_esc(career)} / TURNO: MAÑANA</td>
      <td><b>Periodo Académico</b></td><td class='c'>{_esc(sec.period)}</td></tr>
  <tr><td><b>Director General (e)</b></td><td>{_esc(inst['director'])}</td>
      <td><b>Ciclo - Sección</b></td><td class='c'>{_esc(ciclo)} - "{_esc(sec.label or 'A')}"</td></tr>
  <tr><td><b>Docente</b></td><td>{_esc(docente.upper())}</td>
      <td><b>R.D. de Encargatura</b></td><td class='c'>{_esc(inst['rd_encargatura'])}</td></tr>
</table>
<table>
  <tr><th>N°<br>Orden</th><th>N° Matrícula<br>(DNI)</th><th>APELLIDOS Y NOMBRES<br>(Por orden Alfabético)</th>
      <th>Calificativo</th><th>Crédito</th><th>Puntaje</th><th>Calificación<br>Cualitativa</th></tr>
  {''.join(filas)}
</table>
<table style="width:45%">
  <tr><th colspan="2">Resumen</th></tr>
  <tr><td>Matriculados</td><td class='c'>{len(alumnos)}</td></tr>
  <tr><td>Aprobados</td><td class='c'>{aprobados}</td></tr>
  <tr><td>Desaprobados</td><td class='c'>{desaprob}</td></tr>
  <tr><td>Con Licencia</td><td class='c'>{sum(1 for st in alumnos if st.get('estado') == 'LICENCIA')}</td></tr>
  <tr><td>Límite de Inasistencia</td><td class='c'>{n_dpi}</td></tr>
</table>
<div class="firma"><span class="linea">Firma del Docente: {_esc(docente.upper())}</span></div>
"""
    html = _pdf_shell(f"ACTA DE EVALUACIÓN DE ÁREA — {sec.period}", cuerpo)
    return html_to_pdf_bytes(html), f"ACTA_AREA_{codigo or 'CURSO'}_{sec.label or 'A'}_{sec.period}.pdf"


class SectionActaAreaPdfView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id: int):
        from .teachers import _grades_section_access_denied
        sec = get_object_or_404(
            Section.objects.select_related(
                "plan_course__course", "plan_course__plan__career", "teacher__user"),
            id=section_id)
        if err := _grades_section_access_denied(request, sec):
            return err
        pdf, fname_or_err = _acta_area_pdf_bytes(sec)
        if pdf is None:
            return Response({"detail": fname_or_err}, status=400)
        return HttpResponse(pdf, content_type="application/pdf",
                            headers={"Content-Disposition": f'attachment; filename="{fname_or_err}"'})


# ══════════════════════════════════════════════════════════════
# 1b. Horario del docente — PDF con sus datos personales y foto
# ══════════════════════════════════════════════════════════════

class TeacherSelfSchedulePdfView(APIView):
    """GET /api/academic/teachers/me/horario.pdf — horario semanal del
    docente logueado, con sus datos personales (foto, grado, contacto)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    DIAS = {1: "LUNES", 2: "MARTES", 3: "MIÉRCOLES", 4: "JUEVES",
            5: "VIERNES", 6: "SÁBADO", 7: "DOMINGO"}

    def get(self, request):
        from academic.models import Teacher as AcademicTeacher, Section, SectionScheduleSlot
        from catalogs.models import Teacher as CatalogTeacher, Period

        teacher = AcademicTeacher.objects.filter(user=request.user).first()
        if not teacher:
            return Response({"detail": "No se encontró perfil de docente."}, status=404)

        # Período solicitado (permite descargar horarios de períodos pasados)
        period_code = (request.query_params.get("period") or "").strip().upper()
        if not period_code:
            per = Period.objects.filter(is_active=True).first()
            period_code = (per.code or "").strip() if per else ""

        sections = Section.objects.filter(teacher=teacher).select_related(
            "plan_course__course", "plan_course__plan__career", "classroom")
        if period_code:
            sections = sections.filter(period=period_code)
        sections = list(sections)
        if not sections:
            return Response(
                {"detail": f"No tienes secciones asignadas en {period_code or 'el período actual'}."},
                status=404)

        # ── Datos personales (catalogs.Teacher: perfil editable) ──
        ct = CatalogTeacher.objects.filter(user=request.user).first()
        grado = dict(CatalogTeacher.GRADOS_ACADEMICOS).get(
            getattr(ct, "grado_academico", "") or "", "")
        celular = getattr(ct, "phone", "") or ""
        correo = getattr(ct, "email", "") or getattr(request.user, "email", "")
        fecha_nac = ""
        if getattr(ct, "fecha_nac", None):
            fecha_nac = ct.fecha_nac.strftime("%d/%m/%Y")
        condicion = dict(CatalogTeacher.CONDICIONES).get(
            getattr(ct, "condicion_laboral", "") or "", "")
        rd_num = getattr(ct, "rd_nombramiento", "") or ""
        rd_fecha = ""
        if getattr(ct, "rd_fecha", None):
            rd_fecha = ct.rd_fecha.strftime("%d/%m/%Y")
        rd_txt = " ".join(filter(None, [
            rd_num, f"({rd_fecha})" if rd_fecha else ""])) or "—"
        foto_uri = ""
        try:
            if ct and ct.photo:
                from .utils import _file_to_data_uri
                foto_uri = _file_to_data_uri(ct.photo.path) or ""
        except Exception:
            pass

        full_name = (getattr(request.user, "full_name", "")
                     or request.user.username or "").upper()
        dni = getattr(ct, "document", "") or request.user.username or ""

        # ── Horario semanal ──
        slots = (SectionScheduleSlot.objects
                 .filter(section__in=sections)
                 .select_related("section__plan_course__course",
                                 "section__plan_course__plan__career",
                                 "section__classroom")
                 .order_by("weekday", "start"))
        filas = []
        for sl in slots:
            sec = sl.section
            pc = sec.plan_course
            career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "")
            filas.append(
                f"<tr><td class='c'><b>{self.DIAS.get(int(sl.weekday), sl.weekday)}</b></td>"
                f"<td class='c'>{str(sl.start)[:5]} – {str(sl.end)[:5]}</td>"
                f"<td>{_esc(pc.effective_name if pc else '')}</td>"
                f"<td>{_esc(career)}</td>"
                f"<td class='c'>{pc.semester if pc else ''} - \"{_esc(sec.label or 'A')}\"</td>"
                f"<td class='c'>{_esc(sec.classroom.code if sec.classroom else '—')}</td></tr>")

        foto_html = (f'<img src="{foto_uri}" style="height:88px;width:74px;'
                     f'object-fit:cover;border:1px solid #999;border-radius:4px">'
                     if foto_uri else "")
        cursos_html = "".join(
            f"<li>{_esc(s.plan_course.effective_name if s.plan_course else '')} "
            f"(Ciclo {s.plan_course.semester if s.plan_course else '?'} · Sec. {_esc(s.label or 'A')})</li>"
            for s in sections)

        cuerpo = f"""
<h2 style="text-align:center">HORARIO DEL DOCENTE — {_esc(period_code or 'PERÍODO ACTUAL')}</h2>
<table style="margin-top:6px"><tr>
  <td style="width:86px;text-align:center;vertical-align:top;border:none">{foto_html}</td>
  <td style="border:none">
    <table>
      <tr><td style="width:32%"><b>Apellidos y Nombres</b></td><td>{_esc(full_name)}</td></tr>
      <tr><td><b>DNI</b></td><td>{_esc(dni)}</td></tr>
      <tr><td><b>Grado académico</b></td><td>{_esc(grado or '—')}</td></tr>
      <tr><td><b>Condición laboral</b></td><td>{_esc(condicion or '—')}</td></tr>
      <tr><td><b>R.D. de Nombramiento / Contrato</b></td><td>{_esc(rd_txt)}</td></tr>
      <tr><td><b>Fecha de nacimiento</b></td><td>{_esc(fecha_nac or '—')}</td></tr>
      <tr><td><b>Celular</b></td><td>{_esc(celular or '—')}</td></tr>
      <tr><td><b>Correo institucional</b></td><td>{_esc(correo or '—')}</td></tr>
    </table>
  </td>
</tr></table>
<table>
  <tr><th>Día</th><th>Hora</th><th>Curso / Módulo</th><th>Programa</th><th>Ciclo - Sec.</th><th>Aula</th></tr>
  {''.join(filas) if filas else '<tr><td colspan="6" class="c">Sin horario registrado para el período</td></tr>'}
</table>
<h2>Cursos a cargo ({len(sections)})</h2>
<ul style="font-size:10px;margin:4px 0 0 14px">{cursos_html or '<li>—</li>'}</ul>
<div class="firma"><span class="linea">{_esc(full_name)}<br>DOCENTE</span></div>
"""
        html = _pdf_shell(f"HORARIO DE CLASES — {period_code}", cuerpo)
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="horario-docente-{period_code or "actual"}.pdf"'})


# ══════════════════════════════════════════════════════════════
# 2. Acta consolidada / Reporte de rendimiento — PDF
# ══════════════════════════════════════════════════════════════

def _params(request):
    q = request.query_params
    return ((q.get("period") or "").strip().upper(), q.get("career_id"),
            q.get("semester"), q.get("anio"))


class EvaluationConsolidadaPdfView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        sections = list(_sections_for(period, career_id, semester, anio=anio))
        if not sections:
            return Response({"detail": f"No hay secciones en {period}"}, status=404)
        bmap = _bundle_map(sections)

        groups = {}
        for sec in sections:
            pc = sec.plan_course
            career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "?")
            key = (career, (pc.semester if pc else 0) or 0, sec.label or "A")
            groups.setdefault(key, []).append(sec)

        bloques = []
        primero = True
        for (career, sem, label), secs in sorted(groups.items()):
            rows = [_section_eval_row(s, bmap) for s in secs]
            by_id = {}
            for rw in rows:
                for st in rw["_students"]:
                    by_id[st.id] = st
            if not by_id:
                continue
            students = sorted(by_id.values(), key=_nombre)
            heads = "".join(
                f"<th>{_esc((rw['course_name'] or '').upper()[:30])}<br>"
                f"<span style='font-weight:normal'>Cr. {getattr(secs[i].plan_course, 'credits', '') or ''}</span></th>"
                for i, rw in enumerate(rows))
            filas = []
            for n, st in enumerate(students, 1):
                celdas, puntaje, cred = [], 0, 0
                for i, rw in enumerate(rows):
                    f = _final_of(_entry_for(rw["_grades"], st))
                    cr = getattr(secs[i].plan_course, "credits", 0) or 0
                    if f is not None:
                        puntaje += round(float(f) * cr)
                        cred += cr
                        celdas.append(f"<td class='c'>{round(float(f))}</td>")
                    else:
                        celdas.append("<td class='c'></td>")
                prom = round(puntaje / cred, 2) if cred else ""
                filas.append(
                    f"<tr><td class='c'>{n}</td><td class='c'>{_esc(st.num_documento or '')}</td>"
                    f"<td>{_esc(_nombre(st))}</td>{''.join(celdas)}"
                    f"<td class='c'>{puntaje if cred else ''}</td><td class='c'>{cred or ''}</td>"
                    f"<td class='c'><b>{prom}</b></td></tr>")
            bloques.append(f"""
<div class="grupo{'' if primero else ' salto'}">
<h2>{_esc(career.upper())} — Ciclo {_esc(_roman(sem))} · Sección "{_esc(label)}" · {_esc(period)}</h2>
<table>
  <tr><th>N°</th><th>DNI</th><th>Apellidos y Nombres</th>{heads}
      <th>Puntaje</th><th>Créd.</th><th>Prom.<br>Pond.</th></tr>
  {''.join(filas)}
</table>
</div>""")
            primero = False

        if not bloques:
            return Response({"detail": "Sin alumnos para el filtro"}, status=404)
        html = _pdf_shell(f"ACTA CONSOLIDADA DE EVALUACIÓN — {period}", "".join(bloques), landscape=True)
        return HttpResponse(html_to_pdf_bytes(html), content_type="application/pdf",
                            headers={"Content-Disposition":
                                     f'attachment; filename="acta-consolidada-{period}.pdf"'})


class EvaluationRendimientoPdfView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        sections = list(_sections_for(period, career_id, semester,
                                      request.query_params.get("section_id"), anio=anio))
        if not sections:
            return Response({"detail": "Sin secciones para el filtro"}, status=404)
        bmap = _bundle_map(sections)

        filas = []
        for sec in sections:
            row = _section_eval_row(sec, bmap)
            finals = list(row["_finals"].values())
            aprob = sum(1 for f in finals if f >= 11)
            desap = sum(1 for f in finals if f < 11)
            prom = round(sum(finals) / len(finals), 1) if finals else ""
            filas.append(
                f"<tr><td>{_esc(row['career_name'])}</td><td class='c'>{row['semester'] or ''}</td>"
                f"<td>{_esc(row['course_name'])}</td><td class='c'>{_esc(row['label'])}</td>"
                f"<td>{_esc(row['teacher_name'])}</td><td class='c'>{row['n_students']}</td>"
                f"<td class='c'>{row['n_loaded']}</td><td class='c'>{aprob}</td>"
                f"<td class='c'>{desap}</td><td class='c'><b>{prom}</b></td>"
                f"<td class='c'>{'CERRADA' if row['submitted'] else 'ABIERTA'}</td></tr>")
        cuerpo = f"""
<h2>Resumen por curso — {_esc(period)}</h2>
<table>
  <tr><th>Carrera</th><th>Ciclo</th><th>Curso</th><th>Sec.</th><th>Docente</th>
      <th>Matric.</th><th>Con notas</th><th>Aprob.</th><th>Desaprob.</th><th>Promedio</th><th>Acta</th></tr>
  {''.join(filas)}
</table>"""
        html = _pdf_shell(f"REPORTE DE RENDIMIENTO ACADÉMICO — {period}", cuerpo, landscape=True)
        return HttpResponse(html_to_pdf_bytes(html), content_type="application/pdf",
                            headers={"Content-Disposition":
                                     f'attachment; filename="rendimiento-{period}.pdf"'})


# ══════════════════════════════════════════════════════════════
# 2b. Reporte de ASISTENCIA (PDF / Excel) con filtros
# ══════════════════════════════════════════════════════════════

DPI_UMBRAL = 0.30


def _asistencia_por_seccion(sections):
    """[(row_seccion, resumen, [alumnos])] con % de inasistencia por alumno."""
    from academic.models import AttendanceSession, AttendanceRow
    from .evaluation import _bundle_map

    bmap = _bundle_map(sections)
    out = []
    for sec in sections:
        row = _section_eval_row(sec, bmap)
        sess_ids = list(AttendanceSession.objects
                        .filter(section=sec).values_list("id", flat=True))
        n_sess = len(sess_ids)
        faltas, tardanzas = {}, {}
        if sess_ids:
            for r in (AttendanceRow.objects
                      .filter(session_id__in=sess_ids)
                      .values_list("student_id", "status")):
                sid, status = r[0], (r[1] or "").upper()
                if status == "ABSENT":
                    faltas[sid] = faltas.get(sid, 0) + 1
                elif status == "LATE":
                    tardanzas[sid] = tardanzas.get(sid, 0) + 1

        alumnos, en_riesgo = [], 0
        for st in sorted(row["_students"], key=_nombre):
            keys = [k for k in (getattr(st, "user_id", None), st.id) if k is not None]
            f = sum(faltas.get(k, 0) for k in keys)
            t = sum(tardanzas.get(k, 0) for k in keys)
            pct = round(f / n_sess * 100) if n_sess else 0
            riesgo = n_sess > 0 and (f / n_sess) > DPI_UMBRAL
            if riesgo:
                en_riesgo += 1
            alumnos.append({"st": st, "faltas": f, "tardanzas": t,
                            "pct": pct, "riesgo": riesgo})
        resumen = {"sesiones": n_sess, "en_riesgo": en_riesgo,
                   "n_students": row["n_students"]}
        out.append((row, resumen, alumnos))
    return out


class EvaluationAsistenciaReporteView(APIView):
    """
    GET /api/academic/admin/evaluation/asistencia.(pdf|xlsx)
        ?period=&career_id=&semester=&anio=&section_id=
    Reporte de asistencia: por curso (sesiones, alumnos en riesgo DPI) y
    detalle por alumno con faltas, tardanzas y % de inasistencia.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, fmt="pdf"):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        sections = list(_sections_for(period, career_id, semester,
                                      request.query_params.get("section_id"), anio=anio))
        if not sections:
            return Response({"detail": "Sin secciones para el filtro"}, status=404)
        data = _asistencia_por_seccion(sections)

        if fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumen por curso"
            head_fill = PatternFill("solid", start_color="1F4E79")
            head_font = Font(bold=True, color="FFFFFF")
            ws["A1"] = f"REPORTE DE ASISTENCIA — {period}"
            ws["A1"].font = Font(bold=True, size=13)
            for c, h in enumerate(["CARRERA", "CICLO", "CURSO", "SEC.", "DOCENTE",
                                   "ALUMNOS", "SESIONES", "EN RIESGO DPI (>30%)"], 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.fill = head_fill
                cell.font = head_font
            ws2 = wb.create_sheet("Detalle por alumno")
            for c, h in enumerate(["CARRERA", "CICLO", "CURSO", "SEC.", "DNI", "ALUMNO",
                                   "SESIONES", "FALTAS", "TARDANZAS", "% INASISTENCIA",
                                   "ESTADO"], 1):
                cell = ws2.cell(row=1, column=c, value=h)
                cell.fill = head_fill
                cell.font = head_font
            r1, r2 = 4, 2
            for row, res, alumnos in data:
                for c, v in enumerate([row["career_name"], row["semester"], row["course_name"],
                                       row["label"], row["teacher_name"], res["n_students"],
                                       res["sesiones"], res["en_riesgo"]], 1):
                    ws.cell(row=r1, column=c, value=v)
                r1 += 1
                for a in alumnos:
                    st = a["st"]
                    for c, v in enumerate([row["career_name"], row["semester"], row["course_name"],
                                           row["label"], st.num_documento or "", _nombre(st),
                                           res["sesiones"], a["faltas"], a["tardanzas"],
                                           a["pct"], "RIESGO DPI" if a["riesgo"] else "OK"], 1):
                        ws2.cell(row=r2, column=c, value=v)
                    r2 += 1
            for col, w in zip("ABCDEFGH", (24, 6, 38, 5, 28, 9, 9, 18)):
                ws.column_dimensions[col].width = w
            for col, w in zip("ABCDEFGHIJK", (24, 6, 34, 5, 11, 38, 9, 8, 10, 13, 12)):
                ws2.column_dimensions[col].width = w
            return _xlsx_response(wb, f"asistencia-{period}.xlsx")

        bloques = []
        for row, res, alumnos in data:
            filas = "".join(
                f"<tr><td class='c'>{i}</td><td class='c'>{_esc(a['st'].num_documento or '')}</td>"
                f"<td>{_esc(_nombre(a['st']))}</td><td class='c'>{res['sesiones']}</td>"
                f"<td class='c'>{a['faltas']}</td><td class='c'>{a['tardanzas']}</td>"
                f"<td class='c'><b>{a['pct']}%</b></td>"
                f"<td class='c'>{'RIESGO DPI' if a['riesgo'] else ''}</td></tr>"
                for i, a in enumerate(alumnos, 1))
            bloques.append(f"""
<div class="grupo">
<h2>{_esc(row['career_name'].upper())} — Ciclo {row['semester'] or '?'} · {_esc(row['course_name'])} (Sec. {_esc(row['label'])})</h2>
<p style="font-size:9px;margin:0 0 2px">Docente: {_esc(row['teacher_name'] or '—')} ·
Sesiones registradas: <b>{res['sesiones']}</b> · Alumnos: {res['n_students']} ·
En riesgo DPI (&gt;30%): <b>{res['en_riesgo']}</b></p>
<table><tr><th>N°</th><th>DNI</th><th>Apellidos y Nombres</th><th>Sesiones</th>
<th>Faltas</th><th>Tardanzas</th><th>% Inasist.</th><th>Estado</th></tr>{filas}</table>
</div>""")
        html = _pdf_shell(f"REPORTE DE ASISTENCIA — {period}", "".join(bloques))
        return HttpResponse(html_to_pdf_bytes(html), content_type="application/pdf",
                            headers={"Content-Disposition":
                                     f'attachment; filename="asistencia-{period}.pdf"'})


# ══════════════════════════════════════════════════════════════
# 2b-bis. Registro MENSUAL de asistencia de una sección — PDF
#         + estadística acumulada de todo el ciclo (gráficos)
# ══════════════════════════════════════════════════════════════

# Estados de asistencia con su marca, etiqueta y color para los gráficos
EST_ASIS = [
    ("PRESENT", "P", "Asistió",     "#15803D"),
    ("LATE",    "T", "Tardanza",    "#B45309"),
    ("ABSENT",  "F", "Faltó",       "#B91C1C"),
    ("EXCUSED", "J", "Justificó",   "#1D4ED8"),
    ("HOLIDAY", "0", "Feriado",     "#64748B"),
]
EST_KEYS = [e[0] for e in EST_ASIS]
MESES_COR = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Set", "Oct", "Nov", "Dic"]


def _stats_asistencia_ciclo(sec, alumnos, key_of):
    """Estadística de TODAS las sesiones de la sección (el ciclo completo).

    Devuelve (n_sesiones, n_lectivas, totales, por_alumno, por_mes) donde
      n_sesiones = sesiones registradas en la sección
      n_lectivas = sesiones con clase efectiva (se descartan las que solo
                   tienen feriados y las vacías) → denominador del % oficial
      totales    = {estado: cantidad}
      por_alumno = {roster_id: {estado: cantidad}}
      por_mes    = {(año, mes): {estado: cantidad, "lectivas": n}} por fecha
    """
    from academic.models import AttendanceSession, AttendanceRow

    sesiones = list(AttendanceSession.objects.filter(section=sec).order_by("date"))
    fecha_de = {s.id: s.date for s in sesiones}
    n_sess = len(sesiones)

    totales = {e: 0 for e in EST_KEYS}
    por_alumno = {str(st["id"]): {e: 0 for e in EST_KEYS} for st in alumnos}
    por_mes = {}
    estados_de_sesion = {}

    if sesiones:
        for sid_row, status, sess_id in (
                AttendanceRow.objects
                .filter(session_id__in=list(fecha_de.keys()))
                .values_list("student_id", "status", "session_id")):
            est = (status or "").upper()
            if est not in totales:
                continue
            estados_de_sesion.setdefault(sess_id, set()).add(est)
            rid = key_of.get(sid_row)
            if rid is None or str(rid) not in por_alumno:
                continue
            por_alumno[str(rid)][est] += 1
            totales[est] += 1
            f = fecha_de.get(sess_id)
            if f:
                por_mes.setdefault((f.year, f.month),
                                   {**{e: 0 for e in EST_KEYS}, "lectivas": 0})[est] += 1

    # sesión lectiva = tiene al menos una marca distinta de feriado
    lectivas = {sid for sid, ests in estados_de_sesion.items() if ests - {"HOLIDAY"}}
    for sid in lectivas:
        f = fecha_de.get(sid)
        if f and (f.year, f.month) in por_mes:
            por_mes[(f.year, f.month)]["lectivas"] += 1

    return n_sess, len(lectivas), totales, por_alumno, dict(sorted(por_mes.items()))


def _barra_apilada(totales):
    """Barra apilada horizontal con la distribución de marcas del ciclo."""
    total = sum(totales.values())
    if not total:
        return "<p class='nodata'>Sin marcas registradas.</p>"
    tramos = []
    for est, marca, label, color in EST_ASIS:
        n = totales.get(est, 0)
        if not n:
            continue
        pct = n / total * 100
        texto = f"{marca} {pct:.1f}%" if pct >= 7 else (marca if pct >= 3 else "")
        tramos.append(
            f"<span style='width:{pct:.4f}%;background:{color}' "
            f"title='{_esc(label)}: {n}'>{texto}</span>")
    return f"<div class='stack'>{''.join(tramos)}</div>"


def _grafico_por_mes(por_mes, umbral_pct=30):
    """Columnas verticales: % de inasistencia por mes, con línea del umbral.

    El eje Y se escala al máximo observado (mínimo 50%) para que las barras
    sean legibles sin dejar de mostrar la línea del umbral en su sitio.
    """
    if not por_mes:
        return "<p class='nodata'>Sin sesiones registradas en el ciclo.</p>"

    datos = []
    for (_y, m), d in por_mes.items():
        lect = d.get("lectivas", 0)
        # promedio de faltas por sesión lectiva, en %
        alumnos_marcados = max(1, sum(d[e] for e in ("PRESENT", "LATE", "ABSENT", "EXCUSED")))
        pct = d["ABSENT"] / alumnos_marcados * 100
        datos.append((m, pct, lect))

    tope = max([umbral_pct * 5 / 3] + [p for _m, p, _e in datos])
    tope = min(100.0, max(50.0, tope * 1.15))

    cols = []
    for m, pct, lect in datos:
        color = "#B91C1C" if pct > umbral_pct else "#1F6E43"
        cols.append(
            f"<div class='col'>"
            f"<div class='bar' style='height:{pct / tope * 100:.2f}%;background:{color}'></div>"
            f"<span class='lbl'>{MESES_COR[m]}<br><b>{pct:.0f}%</b>"
            f"<br><span style='color:#666'>{lect} ses.</span></span></div>")

    return f"""
<div class='plotwrap'>
  <div class='plot'>
    <div class='eje' style='bottom:100%'><i>{tope:.0f}%</i></div>
    <div class='ref' style='bottom:{umbral_pct / tope * 100:.2f}%'><i>{umbral_pct}%</i></div>
    {''.join(cols)}
  </div>
</div>
<p class='mini'>Cada columna es el % de inasistencia del aula en ese mes
(«ses.» = sesiones lectivas). La línea roja es el límite del {umbral_pct}%.</p>"""

class SectionAttendanceMonthPdfView(APIView):
    """
    GET /api/academic/sections/<id>/attendance/mes.pdf?month=YYYY-MM
    Registro de asistencia mensual (grilla alumnos × días) tal como lo ve el
    docente, para imprimir/archivar.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id: int):
        from calendar import monthrange
        from datetime import date as date_cls
        from academic.models import AttendanceSession, AttendanceRow
        from .attendance import _section_access_denied
        from .acta_excel import _roster, _section_header_info, _roman

        if err := _section_access_denied(request, section_id):
            return err
        sec = get_object_or_404(
            Section.objects.select_related(
                "plan_course__course", "plan_course__plan__career", "teacher__user"),
            id=section_id)

        raw = (request.query_params.get("month") or "").strip()
        try:
            y, m = (int(x) for x in raw.split("-")[:2])
            assert 1 <= m <= 12
        except Exception:
            hoy = date_cls.today()
            y, m = hoy.year, hoy.month
        n_days = monthrange(y, m)[1]
        LETRAS = ["L", "M", "X", "J", "V", "S", "D"]
        MARK = {"PRESENT": "P", "LATE": "T", "ABSENT": "F",
                "EXCUSED": "J", "HOLIDAY": "0"}

        alumnos = _roster(sec)
        if not alumnos:
            return Response({"detail": "La sección no tiene alumnos matriculados."}, status=400)

        key_of = {}
        for st in alumnos:
            key_of[st["id"]] = st["id"]
            if st.get("pk") is not None:
                key_of[st["pk"]] = st["id"]

        marks, cerrados = {}, set()
        for sess in (AttendanceSession.objects
                     .filter(section=sec, date__year=y, date__month=m)
                     .prefetch_related("rows")):
            if sess.closed:
                cerrados.add(sess.date.day)
            for row in sess.rows.all():
                rid = key_of.get(row.student_id)
                if rid is None:
                    continue
                mk = MARK.get((row.status or "").upper())
                if mk:
                    marks.setdefault(str(rid), {})[sess.date.day] = mk

        dias = [{"d": d, "letra": LETRAS[date_cls(y, m, d).weekday()],
                 "wknd": date_cls(y, m, d).weekday() >= 5} for d in range(1, n_days + 1)]

        heads = "".join(
            f"<th class='{'wk' if d['wknd'] else ''}'>{d['d']}<br>"
            f"<span style='font-weight:normal'>{d['letra']}</span>"
            f"{'<br>🔒' if d['d'] in cerrados else ''}</th>" for d in dias)

        filas = []
        for i, st in enumerate(alumnos, 1):
            celdas = []
            cuenta = {m: 0 for _e, m, _l, _c in EST_ASIS}
            for d in dias:
                v = marks.get(str(st["id"]), {}).get(d["d"], "")
                if v in cuenta:
                    cuenta[v] += 1
                celdas.append(f"<td class='{'wk' if d['wknd'] else ''}'>{v}</td>")
            efectivas = cuenta["P"] + cuenta["T"] + cuenta["F"] + cuenta["J"]
            pct_mes = round(cuenta["F"] / efectivas * 100) if efectivas else 0
            filas.append(
                f"<tr><td class='c'>{i}</td><td>{_esc(st['nombre'])}</td>"
                f"{''.join(celdas)}"
                f"<td class='c'>{cuenta['P'] or ''}</td><td class='c'>{cuenta['T'] or ''}</td>"
                f"<td class='c'><b>{cuenta['F'] or ''}</b></td><td class='c'>{cuenta['J'] or ''}</td>"
                f"<td class='c'>{pct_mes}%</td></tr>")

        # ── Estadística acumulada de TODO el ciclo ──
        (n_ciclo, n_lectivas, tot_ciclo,
         por_alumno, por_mes) = _stats_asistencia_ciclo(sec, alumnos, key_of)
        UMBRAL = int(DPI_UMBRAL * 100)
        base = n_lectivas or 0        # denominador único del % (sesiones lectivas)

        stat_filas, n_retirados, n_riesgo, n_licencia = [], 0, 0, 0
        for i, st in enumerate(alumnos, 1):
            d = por_alumno.get(str(st["id"]), {e: 0 for e in EST_KEYS})
            lic = (st.get("estado") or "").upper() == "LICENCIA"
            marcadas = sum(d[e] for e in ("PRESENT", "LATE", "ABSENT", "EXCUSED"))
            sin_marca = max(0, base - marcadas)
            pct = (d["ABSENT"] / base * 100) if base else 0
            asis_pct = ((d["PRESENT"] + d["LATE"]) / base * 100) if base else 0
            if lic:
                n_licencia += 1
                estado_txt, estado_cls = "CON LICENCIA", "lic"
            elif base and pct > UMBRAL:
                n_retirados += 1
                estado_txt, estado_cls = f"RETIRADO POR INASISTENCIA (&gt;{UMBRAL}%)", "ret"
            elif base and pct >= UMBRAL * 2 / 3:
                n_riesgo += 1
                estado_txt, estado_cls = "EN RIESGO", "rie"
            else:
                estado_txt, estado_cls = "HABILITADO", "ok"
            barra = (f"<div class='minibar'><i style='width:{min(100.0, pct):.2f}%;"
                     f"background:{'#B91C1C' if pct > UMBRAL else '#1F6E43'}'></i></div>")
            stat_filas.append(
                f"<tr class='{estado_cls}'><td class='c'>{i}</td>"
                f"<td class='c'>{_esc(st.get('dni') or '')}</td><td>{_esc(st['nombre'])}</td>"
                f"<td class='c'>{base}</td>"
                f"<td class='c'>{d['PRESENT']}</td><td class='c'>{d['LATE']}</td>"
                f"<td class='c'><b>{d['ABSENT']}</b></td><td class='c'>{d['EXCUSED']}</td>"
                f"<td class='c'>{sin_marca or ''}</td>"
                f"<td class='c'>{asis_pct:.0f}%</td>"
                f"<td class='c'><b>{pct:.0f}%</b></td>"
                f"<td class='barcell'>{barra}</td>"
                f"<td class='c est'>{estado_txt}</td></tr>")

        tot_marcas = sum(tot_ciclo.values())
        tot_efectivas = sum(tot_ciclo[e] for e in ("PRESENT", "LATE", "ABSENT", "EXCUSED"))
        pct_asis_sec = ((tot_ciclo["PRESENT"] + tot_ciclo["LATE"]) / tot_efectivas * 100
                        if tot_efectivas else 0)
        kpis = "".join(
            f"<div class='kpi'><b>{v}</b><span>{t}</span></div>"
            for t, v in [
                ("Matriculados", len(alumnos)),
                ("Sesiones lectivas", n_lectivas),
                ("% de asistencia", f"{pct_asis_sec:.1f}%"),
                (f"Retirados (&gt;{UMBRAL}%)", n_retirados),
                ("En riesgo", n_riesgo),
                ("Con licencia", n_licencia),
            ])
        leyenda_tot = " · ".join(
            f"<b style='color:{color}'>{marca}</b> {label}: {tot_ciclo.get(est, 0)}"
            f" ({tot_ciclo.get(est, 0) / tot_marcas * 100:.1f}%)" if tot_marcas else
            f"<b style='color:{color}'>{marca}</b> {label}: 0"
            for est, marca, label, color in EST_ASIS)

        estadistica = f"""
<div class="salto"></div>
<h2 style="text-align:center">ESTADÍSTICA DE ASISTENCIA DEL CICLO — {_esc(sec.period)}</h2>
<div class="kpis">{kpis}</div>

<h2>Distribución de marcas en todo el ciclo</h2>
{_barra_apilada(tot_ciclo)}
<p class="mini">{leyenda_tot}</p>

<h2>Evolución del % de inasistencia por mes</h2>
{_grafico_por_mes(por_mes, UMBRAL)}

<h2>Detalle por estudiante (acumulado del ciclo)</h2>
<table class="stats">
  <tr><th>N°</th><th>DNI</th><th style="text-align:left">APELLIDOS Y NOMBRES</th>
      <th>Sesiones<br>lectivas</th><th>Asistió<br>(P)</th><th>Tardanza<br>(T)</th>
      <th>Faltó<br>(F)</th><th>Justificó<br>(J)</th><th>Sin<br>marca</th>
      <th>%<br>Asist.</th><th>%<br>Inasist.</th>
      <th style="width:80px">Inasistencia (línea = {UMBRAL}%)</th><th>Condición</th></tr>
  {''.join(stat_filas)}
</table>
<p style="font-size:8px;margin-top:5px">
<b>NOTA:</b> el % se calcula sobre las <b>{n_lectivas} sesiones lectivas</b> del ciclo
(de {n_ciclo} sesiones registradas se descartan las de feriado y las vacías).
% de asistencia = (P + T) ÷ sesiones lectivas; % de inasistencia = F ÷ sesiones lectivas.
Se considera <b>RETIRADO POR INASISTENCIA</b> (desaprobado por inasistencia — DPI) al
estudiante con más del {UMBRAL}% de faltas en el ciclo, conforme a la
RVM N° 277-2019-MINEDU; <b>EN RIESGO</b> desde el {UMBRAL * 2 // 3}%.
Los estudiantes con licencia no entran en el cómputo. La columna
«Sin marca» son sesiones lectivas en las que no se registró marca para el estudiante.</p>
"""

        curso, _codigo, docente = _section_header_info(sec)
        pc = sec.plan_course
        career = (pc.plan.career.name if pc and pc.plan and pc.plan.career else "")
        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
                 "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]

        cuerpo = f"""
<style>
  th, td {{ padding: 1px 2px; font-size: 7.5px; text-align: center; }}
  table.grid td:nth-child(2) {{ text-align: left; font-size: 7.5px; white-space: nowrap; }}
  th.wk, td.wk {{ background: #C7CDD4; }}
  th {{ background: #1F6E43; }}
  .nodata {{ font-size: 9px; color: #666; font-style: italic; margin: 4px 0; }}
  .mini {{ font-size: 7.5px; margin: 3px 0 0; color: #333; }}

  /* KPIs */
  .kpis {{ display: flex; gap: 6px; margin-top: 6px; }}
  .kpi {{ flex: 1; border: 1px solid #1F4E79; border-radius: 3px; padding: 4px 2px;
          text-align: center; background: #F3F7FB; }}
  .kpi b {{ display: block; font-size: 14px; color: #1F4E79; line-height: 1.1; }}
  .kpi span {{ font-size: 6.8px; color: #333; text-transform: uppercase; }}

  /* Barra apilada de distribución */
  .stack {{ display: flex; width: 100%; height: 20px; border: 1px solid #555;
            border-radius: 2px; overflow: hidden; }}
  .stack span {{ font-size: 7px; color: #fff; font-weight: bold; text-align: center;
                 line-height: 20px; overflow: hidden; white-space: nowrap; }}

  /* Columnas por mes */
  .plotwrap {{ padding-left: 4px; }}
  .plot {{ position: relative; height: 90px; border-bottom: 1px solid #555;
           border-left: 1px solid #555; display: flex; align-items: flex-end;
           gap: 10px; padding: 0 10px; margin-bottom: 16px; }}
  .plot .col {{ position: relative; width: 26px; height: 100%;
                display: flex; align-items: flex-end; }}
  .plot .bar {{ width: 100%; border-radius: 2px 2px 0 0; min-height: 1px; }}
  .plot .lbl {{ position: absolute; top: 100%; left: -6px; right: -6px;
                text-align: center; font-size: 6.5px; line-height: 1.15; }}
  .plot .ref {{ position: absolute; left: 0; right: 0; border-top: 1px dashed #B91C1C; }}
  .plot .ref i {{ position: absolute; right: 1px; top: -8px; font-size: 6px;
                  color: #B91C1C; font-style: normal; font-weight: bold; }}
  .plot .eje {{ position: absolute; left: 0; right: 0; border-top: 1px dotted #B9C2CC; }}
  .plot .eje i {{ position: absolute; left: -20px; top: -4px; font-size: 6px;
                  color: #667; font-style: normal; }}

  /* Tabla de estadística por alumno */
  table.stats td, table.stats th {{ font-size: 7px; }}
  table.stats td:nth-child(3) {{ text-align: left; }}
  table.stats td.est {{ font-size: 6.2px; font-weight: bold; white-space: nowrap; }}
  .barcell {{ padding: 2px 3px !important; }}
  .minibar {{ position: relative; height: 7px; background: #E5E7EB;
              border: 1px solid #9CA3AF; }}
  .minibar i {{ display: block; height: 100%; }}
  .minibar::after {{ content: ''; position: absolute; left: {int(DPI_UMBRAL * 100)}%;
                     top: -1px; bottom: -1px; border-left: 1px dashed #B91C1C; }}
  tr.ret td {{ background: #FEE2E2; }}
  tr.ret td.est {{ color: #B91C1C; }}
  tr.rie td {{ background: #FEF3C7; }}
  tr.rie td.est {{ color: #B45309; }}
  tr.lic td {{ background: #EDE9FE; }}
  tr.lic td.est {{ color: #6D28D9; }}
  tr.ok td.est {{ color: #15803D; }}
</style>
<h2 style="text-align:center">REGISTRO DE ASISTENCIA — {_esc(meses[m])} {y}</h2>
<table style="margin-top:4px">
  <tr><td style="width:16%"><b>Programa</b></td><td>{_esc(career.upper())}</td>
      <td style="width:12%"><b>Período</b></td><td class='c'>{_esc(sec.period)}</td></tr>
  <tr><td><b>Curso / Módulo</b></td><td>{_esc(curso)}</td>
      <td><b>Ciclo - Sec.</b></td><td class='c'>{_esc(_roman(pc.semester if pc else ''))} - "{_esc(sec.label or 'A')}"</td></tr>
  <tr><td><b>Docente</b></td><td colspan="3">{_esc(docente.upper())}</td></tr>
</table>
<table class="grid">
  <tr><th>N°</th><th style="text-align:left">APELLIDOS Y NOMBRES</th>{heads}
      <th>P</th><th>T</th><th>F</th><th>J</th><th>% Inas.</th></tr>
  {''.join(filas)}
</table>
<p style="font-size:8px;margin-top:5px">
<b>LEYENDA:</b> P = Presente · T = Tardanza · F = Falta · J = Justificado · 0 = Feriado ·
🔒 Día cerrado · Celdas grises: sábado y domingo ·
Las últimas columnas resumen el mes ({_esc(meses[m])} {y}).</p>
{estadistica}
<div class="firma"><span class="linea">{_esc(docente.upper())}<br>DOCENTE</span></div>
"""
        html = _pdf_shell(f"REGISTRO DE ASISTENCIA MENSUAL — {sec.period}", cuerpo, landscape=True)
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="asistencia-{y}-{m:02d}-sec{sec.id}.pdf"'})


# ══════════════════════════════════════════════════════════════
# 2c. Documentos del PROPIO ALUMNO (boleta y asistencia)
# ══════════════════════════════════════════════════════════════

class StudentSelfBoletaPdfView(APIView):
    """GET /api/academic/student/me/boleta.pdf?period=2026-I
    Boleta de calificaciones del alumno logueado (sus notas del período)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .dashboard_student import _student_of, _current_period

        st = _student_of(request.user)
        if not st:
            return Response({"detail": "No se encontró perfil de estudiante."}, status=404)
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            per = _current_period()
            period = (getattr(per, "code", None) or "").strip()
        if not period:
            return Response({"detail": "No hay período activo"}, status=400)

        recs = list(AcademicGradeRecord.objects
                    .filter(student=st, term=period)
                    .select_related("course", "plan_course")
                    .order_by("course__name"))
        if not recs:
            return Response(
                {"detail": f"Aún no hay notas procesadas para {period}. "
                           "Tus notas aparecerán cuando Secretaría Académica procese el período."},
                status=404)

        filas, ptos, creds = [], 0, 0
        for i, rec in enumerate(recs, 1):
            try:
                g = round(float(rec.final_grade))
            except (TypeError, ValueError):
                g = None
            cr = _creditos_de(rec)
            ptj = (g or 0) * cr
            if g is not None:
                ptos += ptj
                creds += cr
            filas.append(
                f"<tr><td class='c'>{i}</td><td>{_esc(rec.course.name)}</td>"
                f"<td class='c'>{cr}</td><td class='c'><b>{g if g is not None else ''}</b></td>"
                f"<td class='c'>{_esc(_cualitativa_de_vigesimal(g)) if g is not None else ''}</td>"
                f"<td class='c'>{ptj if g is not None else ''}</td></tr>")
        prom = round(ptos / creds, 2) if creds else ""
        career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")

        cuerpo = f"""
<h2 style="text-align:center">BOLETA DE CALIFICACIONES</h2>
<table style="margin-top:4px">
  <tr><td style="width:22%"><b>Estudiante</b></td><td>{_esc(_nombre(st))}</td>
      <td style="width:18%"><b>Período</b></td><td class='c'>{_esc(period)}</td></tr>
  <tr><td><b>N° de Matrícula (DNI)</b></td><td>{_esc(st.num_documento or '')}</td>
      <td><b>Ciclo</b></td><td class='c'>{_esc(st.ciclo or '')}</td></tr>
  <tr><td><b>Programa de Estudios</b></td><td colspan="3">{_esc(career.upper())}</td></tr>
</table>
<table>
  <tr><th>N°</th><th>Curso / Módulo</th><th>Créd.</th><th>Calificación</th>
      <th>Calificación cualitativa</th><th>Puntaje</th></tr>
  {''.join(filas)}
  <tr><td colspan="5" style="text-align:right"><b>PROMEDIO PONDERADO DEL PERÍODO</b></td>
      <td class='c'><b>{prom}</b></td></tr>
</table>
<div class="firma"><span class="linea">SECRETARIO(A) ACADÉMICO(A)</span></div>
"""
        html = _pdf_shell(f"BOLETA DE CALIFICACIONES — {period}", cuerpo)
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="boleta-{st.num_documento or st.id}-{period}.pdf"'})


class StudentSelfAsistenciaPdfView(APIView):
    """GET /api/academic/student/me/asistencia.pdf?period=2026-I
    Reporte de asistencia del alumno logueado, por curso."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .dashboard_student import _student_of, _current_period
        from academic.models import (
            EnrollmentItem, Enrollment, Section, AttendanceSession, AttendanceRow)

        st = _student_of(request.user)
        if not st:
            return Response({"detail": "No se encontró perfil de estudiante."}, status=404)
        period = (request.query_params.get("period") or "").strip().upper()
        if not period:
            per = _current_period()
            period = (getattr(per, "code", None) or "").strip()
        if not period:
            return Response({"detail": "No hay período activo"}, status=400)

        keys = [k for k in (st.user_id, st.id) if k is not None]
        items = (EnrollmentItem.objects
                 .filter(enrollment__student=st, enrollment__period=period)
                 .exclude(enrollment__status=Enrollment.STATUS_CANCELLED)
                 .select_related("plan_course__course", "section"))
        filas = []
        for item in items:
            pc = item.plan_course
            if not pc:
                continue
            sec = item.section or (Section.objects
                                   .filter(plan_course=pc, period=period)
                                   .order_by("label", "id").first())
            n_sess = f = t = 0
            if sec:
                sess_ids = list(AttendanceSession.objects
                                .filter(section=sec).values_list("id", flat=True))
                n_sess = len(sess_ids)
                if sess_ids:
                    for status in (AttendanceRow.objects
                                   .filter(session_id__in=sess_ids, student_id__in=keys)
                                   .values_list("status", flat=True)):
                        s = (status or "").upper()
                        if s == "ABSENT":
                            f += 1
                        elif s == "LATE":
                            t += 1
            pct = round(f / n_sess * 100) if n_sess else 0
            riesgo = n_sess > 0 and (f / n_sess) > DPI_UMBRAL
            filas.append((pc.effective_name, n_sess, f, t, pct, riesgo))

        if not filas:
            return Response({"detail": f"No tienes cursos matriculados en {period}"}, status=404)
        filas.sort(key=lambda x: x[0])
        html_filas = "".join(
            f"<tr><td class='c'>{i}</td><td>{_esc(n)}</td><td class='c'>{s}</td>"
            f"<td class='c'>{fa}</td><td class='c'>{ta}</td><td class='c'><b>{p}%</b></td>"
            f"<td class='c'>{'RIESGO DPI' if r else 'OK'}</td></tr>"
            for i, (n, s, fa, ta, p, r) in enumerate(filas, 1))
        career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")

        cuerpo = f"""
<h2 style="text-align:center">REPORTE DE ASISTENCIA DEL ESTUDIANTE</h2>
<table style="margin-top:4px">
  <tr><td style="width:22%"><b>Estudiante</b></td><td>{_esc(_nombre(st))}</td>
      <td style="width:18%"><b>Período</b></td><td class='c'>{_esc(period)}</td></tr>
  <tr><td><b>N° de Matrícula (DNI)</b></td><td>{_esc(st.num_documento or '')}</td>
      <td><b>Ciclo</b></td><td class='c'>{_esc(st.ciclo or '')}</td></tr>
  <tr><td><b>Programa de Estudios</b></td><td colspan="3">{_esc(career.upper())}</td></tr>
</table>
<table>
  <tr><th>N°</th><th>Curso / Módulo</th><th>Sesiones</th><th>Faltas</th>
      <th>Tardanzas</th><th>% Inasistencia</th><th>Estado</th></tr>
  {html_filas}
</table>
<p style="font-size:9px;margin-top:6px">
Se considera <b>riesgo de DPI</b> (Desaprobado Por Inasistencia) cuando las faltas
superan el 30% de las sesiones registradas del curso.</p>
<div class="firma"><span class="linea">SECRETARIO(A) ACADÉMICO(A)</span></div>
"""
        html = _pdf_shell(f"REPORTE DE ASISTENCIA — {period}", cuerpo)
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="asistencia-{st.num_documento or st.id}-{period}.pdf"'})


# ══════════════════════════════════════════════════════════════
# 3. Méritos: primeros lugares / tercio y quinto superior / becas
# ══════════════════════════════════════════════════════════════

class EvaluationPrimerosLugaresView(APIView):
    """Top 3 por especialidad (carrera) + aula (ciclo) según promedio
    ponderado del período. ?fmt=pdf|xlsx"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        top_n = 3
        students = _filtrar_students(period, career_id, semester, anio)
        if not students:
            return Response({"detail": f"No hay alumnos con notas procesadas en {period}"}, status=404)
        proms = _promedios_por_alumno(term=period, student_ids=[s.id for s in students])

        groups = {}
        for st in students:
            if st.id not in proms:
                continue
            career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "—")
            groups.setdefault((career, st.ciclo or 0), []).append(st)

        resultado = []   # (carrera, ciclo, puesto, st, prom)
        for (career, ciclo), sts in sorted(groups.items()):
            ranked = sorted(sts, key=lambda s: (-proms[s.id], _nombre(s)))[:top_n]
            for i, st in enumerate(ranked, 1):
                resultado.append((career, ciclo, i, st, proms[st.id]))

        if (request.query_params.get("fmt") or "").lower() == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Primeros lugares"
            ws["A1"] = f"PRIMEROS LUGARES POR ESPECIALIDAD Y AULA — {period}"
            ws["A1"].font = Font(bold=True, size=12)
            heads = ["CARRERA", "CICLO", "PUESTO", "DNI", "APELLIDOS Y NOMBRES", "PROMEDIO"]
            for c, h in enumerate(heads, 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1F4E79")
            for r, (career, ciclo, puesto, st, prom) in enumerate(resultado, 4):
                for c, v in enumerate([career, ciclo, puesto, st.num_documento or "",
                                       _nombre(st), prom], 1):
                    ws.cell(row=r, column=c, value=v)
            for col, w in zip("ABCDEF", (30, 7, 8, 12, 42, 10)):
                ws.column_dimensions[col].width = w
            return _xlsx_response(wb, f"primeros-lugares-{period}.xlsx")

        filas, last_key = [], None
        for career, ciclo, puesto, st, prom in resultado:
            key = (career, ciclo)
            sep = "" if key == last_key else f"<tr><td colspan='5' style='background:#E8EEF7'><b>{_esc(career.upper())} — Ciclo {ciclo}</b></td></tr>"
            last_key = key
            filas.append(sep +
                f"<tr><td class='c'><b>{puesto}°</b></td><td class='c'>{_esc(st.num_documento or '')}</td>"
                f"<td>{_esc(_nombre(st))}</td><td class='c'><b>{prom}</b></td>"
                f"<td class='c'>{_esc(_cualitativa_de_vigesimal(prom))}</td></tr>")
        cuerpo = f"""
<h2>Primeros lugares por especialidad y aula — {_esc(period)}</h2>
<table><tr><th>Puesto</th><th>DNI</th><th>Apellidos y Nombres</th><th>Promedio</th><th>Calificación</th></tr>
{''.join(filas)}</table>
<div class="firma"><span class="linea">DIRECTOR(A) GENERAL</span></div>"""
        html = _pdf_shell(f"PRIMEROS LUGARES — {period}", cuerpo)
        return HttpResponse(html_to_pdf_bytes(html), content_type="application/pdf",
                            headers={"Content-Disposition":
                                     f'attachment; filename="primeros-lugares-{period}.pdf"'})


class EvaluationTercioQuintoView(APIView):
    """Tercio/Quinto superior de la promoción que culmina el 10mo ciclo
    (todas las especialidades), por promedio ponderado de toda la carrera.
    ?tipo=tercio|quinto&fmt=pdf|xlsx"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        tipo = (request.query_params.get("tipo") or "tercio").lower()
        divisor = 5 if tipo == "quinto" else 3
        titulo = "QUINTO SUPERIOR" if divisor == 5 else "TERCIO SUPERIOR"

        students = list(Student.objects.filter(ciclo=10, grade_records__isnull=False)
                        .select_related("plan", "plan__career").distinct())
        if not students:
            return Response({"detail": "No hay alumnos de 10mo ciclo con notas"}, status=404)
        proms = _promedios_por_alumno(student_ids=[s.id for s in students])
        ranked = sorted((s for s in students if s.id in proms),
                        key=lambda s: (-proms[s.id], _nombre(s)))
        corte = max(1, len(ranked) // divisor)
        seleccion = ranked[:corte]

        if (request.query_params.get("fmt") or "").lower() == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = titulo[:31]
            ws["A1"] = f"{titulo} — PROMOCIÓN (10mo ciclo, todas las especialidades)"
            ws["A1"].font = Font(bold=True, size=12)
            ws["A2"] = f"Total promoción: {len(ranked)} · {titulo.lower()}: {corte}"
            heads = ["PUESTO", "DNI", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "PROMEDIO GENERAL"]
            for c, h in enumerate(heads, 1):
                cell = ws.cell(row=4, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1F4E79")
            for r, st in enumerate(seleccion, 5):
                career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")
                for c, v in enumerate([r - 4, st.num_documento or "", _nombre(st),
                                       career, proms[st.id]], 1):
                    ws.cell(row=r, column=c, value=v)
            for col, w in zip("ABCDE", (8, 12, 42, 30, 16)):
                ws.column_dimensions[col].width = w
            return _xlsx_response(wb, f"{tipo}-superior-promocion.xlsx")

        filas = "".join(
            f"<tr><td class='c'><b>{i}°</b></td><td class='c'>{_esc(st.num_documento or '')}</td>"
            f"<td>{_esc(_nombre(st))}</td>"
            f"<td>{_esc(st.plan.career.name if st.plan_id and st.plan and st.plan.career else '')}</td>"
            f"<td class='c'><b>{proms[st.id]}</b></td></tr>"
            for i, st in enumerate(seleccion, 1))
        cuerpo = f"""
<h2>{titulo} de la promoción — 10mo ciclo (todas las especialidades)</h2>
<p style="font-size:9px">Total de la promoción: {len(ranked)} estudiantes · {titulo.title()}: {corte} estudiantes
(promedio ponderado de toda la formación).</p>
<table><tr><th>Puesto</th><th>DNI</th><th>Apellidos y Nombres</th><th>Especialidad</th><th>Promedio</th></tr>
{filas}</table>
<div class="firma"><span class="linea">DIRECTOR(A) GENERAL</span></div>"""
        html = _pdf_shell(titulo, cuerpo)
        return HttpResponse(html_to_pdf_bytes(html), content_type="application/pdf",
                            headers={"Content-Disposition":
                                     f'attachment; filename="{tipo}-superior-promocion.pdf"'})


class EvaluationConstanciasBecaZipView(APIView):
    """Constancias de beca (PDF) para alumnos con promedio del período >= 17,
    por aula. ?period=&career_id=&semester=&anio=&min_avg=17"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        try:
            min_avg = float(request.query_params.get("min_avg", 17))
        except (TypeError, ValueError):
            min_avg = 17.0

        students = _filtrar_students(period, career_id, semester, anio)
        proms = _promedios_por_alumno(term=period, student_ids=[s.id for s in students])
        becarios = sorted((s for s in students if proms.get(s.id, 0) >= min_avg),
                          key=lambda s: (-proms[s.id], _nombre(s)))
        if not becarios:
            return Response(
                {"detail": f"Ningún alumno alcanza promedio >= {min_avg:g} en {period}"},
                status=404)

        inst = _acta_area_inst()
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for st in becarios:
                career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")
                cuerpo = f"""
<div style="text-align:center; margin-top:30px">
  <h2 style="font-size:16px">CONSTANCIA DE BECA</h2>
</div>
<p style="font-size:11px; line-height:1.9; text-align:justify; margin-top:18px">
El (La) Director(a) General del {_esc(inst['nombre'])}, que suscribe, hace constar que:
</p>
<p style="text-align:center; font-size:13px"><b>{_esc(_nombre(st))}</b><br>
<span style="font-size:10px">DNI N° {_esc(st.num_documento or '')} — Programa de Estudios de {_esc(career.upper())}
— Ciclo {_esc(st.ciclo or '')}</span></p>
<p style="font-size:11px; line-height:1.9; text-align:justify">
ha obtenido un <b>promedio ponderado de {proms[st.id]:.2f}</b> en el período académico
<b>{_esc(period)}</b>, alcanzando un rendimiento igual o superior a {min_avg:g},
por lo que se hace acreedor(a) a la <b>BECA DE ESTUDIOS</b> conforme a las
disposiciones institucionales vigentes.
</p>
<p style="font-size:11px; line-height:1.9; text-align:justify">
Se expide la presente constancia a solicitud del (de la) interesado(a) para los fines
que estime conveniente.
</p>
<p style="font-size:10px; text-align:right; margin-top:24px">Tarma, ____ de ____________ de {_esc(period.split('-')[0])}</p>
<div class="firma" style="margin-top:60px"><span class="linea">{_esc(inst['director'])}<br>DIRECTOR(A) GENERAL</span></div>
"""
                html = _pdf_shell(f"CONSTANCIA DE BECA — {period}", cuerpo)
                try:
                    pdf = html_to_pdf_bytes(html)
                except Exception:
                    logger.exception("Error constancia beca student=%s", st.id)
                    continue
                doc = st.num_documento or str(st.id)
                safe = _nombre(st).replace(" ", "_").replace(",", "").replace("/", "-")
                zf.writestr(f"constancia-beca-{period}-{doc}-{safe}.pdf", pdf)

        zip_buf.seek(0)
        return HttpResponse(
            zip_buf.getvalue(), content_type="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="constancias-beca-{period}.zip"'})


# ══════════════════════════════════════════════════════════════
# 4. Excel complementarios (Boleta / Ficha en tabla)
# ══════════════════════════════════════════════════════════════

class EvaluationBoletasXlsxView(APIView):
    """Detalle alumno×curso del período con promedio ponderado (versión Excel
    de las boletas)."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        students = _filtrar_students(period, career_id, semester, anio)
        if not students:
            return Response({"detail": f"No hay alumnos con notas en {period}"}, status=404)
        proms = _promedios_por_alumno(term=period, student_ids=[s.id for s in students])

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletas (detalle)"
        ws["A1"] = f"BOLETAS DE INFORMACIÓN (DETALLE) — {period}"
        ws["A1"].font = Font(bold=True, size=12)
        heads = ["DNI", "APELLIDOS Y NOMBRES", "CARRERA", "CICLO", "CURSO",
                 "CRÉD.", "NOTA", "CALIFICACIÓN", "PROMEDIO PONDERADO"]
        for c, h in enumerate(heads, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F4E79")
        r = 4
        for st in sorted(students, key=_nombre):
            recs = (AcademicGradeRecord.objects
                    .filter(student=st, term=period)
                    .select_related("course", "plan_course").order_by("course__name"))
            career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")
            for rec in recs:
                try:
                    g = round(float(rec.final_grade))
                except (TypeError, ValueError):
                    g = ""
                vals = [st.num_documento or "", _nombre(st), career, st.ciclo or "",
                        rec.course.name, _creditos_de(rec), g,
                        _cualitativa_de_vigesimal(g) if g != "" else "",
                        proms.get(st.id, "")]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=r, column=c, value=v)
                r += 1
        for col, w in zip("ABCDEFGHI", (12, 40, 26, 7, 45, 7, 7, 15, 12)):
            ws.column_dimensions[col].width = w
        return _xlsx_response(wb, f"boletas-detalle-{period}.xlsx")


class EvaluationFichasXlsxView(APIView):
    """Resumen alumno×término (todos los períodos): versión Excel de la ficha
    de rendimiento."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        period, career_id, semester, anio = _params(request)
        students = _filtrar_students(period, career_id, semester, anio) if period else []
        if not students:
            return Response({"detail": "No hay alumnos con notas para el filtro"}, status=404)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Fichas (resumen)"
        ws["A1"] = "FICHA DE RENDIMIENTO (RESUMEN POR PERÍODO)"
        ws["A1"].font = Font(bold=True, size=12)
        heads = ["DNI", "APELLIDOS Y NOMBRES", "CARRERA", "PERÍODO",
                 "CURSOS", "CRÉDITOS", "PROMEDIO PONDERADO", "CALIFICACIÓN"]
        for c, h in enumerate(heads, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F4E79")
        r = 4
        for st in sorted(students, key=_nombre):
            career = (st.plan.career.name if st.plan_id and st.plan and st.plan.career else "")
            por_term = {}
            for rec in (AcademicGradeRecord.objects.filter(student=st)
                        .select_related("plan_course", "course")):
                try:
                    g = float(rec.final_grade)
                except (TypeError, ValueError):
                    continue
                cr = _creditos_de(rec) or 1
                cur = por_term.setdefault(rec.term, [0.0, 0, 0])
                cur[0] += g * cr
                cur[1] += cr
                cur[2] += 1
            for term in sorted(por_term):
                pts, cr, n = por_term[term]
                prom = round(pts / cr, 2) if cr else ""
                vals = [st.num_documento or "", _nombre(st), career, term, n, cr, prom,
                        _cualitativa_de_vigesimal(prom) if prom != "" else ""]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=r, column=c, value=v)
                r += 1
        for col, w in zip("ABCDEFGH", (12, 40, 26, 10, 8, 9, 12, 15)):
            ws.column_dimensions[col].width = w
        return _xlsx_response(wb, "fichas-resumen.xlsx")
