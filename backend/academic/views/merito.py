"""
Orden de mérito por promedio ponderado del período.

  ADMIN (Evaluación → Boletas y reportes):
    GET /academic/admin/evaluation/merito.xlsx?period=&career_id=&semester=
    GET /academic/admin/evaluation/merito.pdf?…
      · con carrera + ciclo  → mérito del AULA
      · con carrera sola     → mérito de la ESPECIALIDAD
      · sin filtros          → mérito de TODO EL INSTITUTO

  ESTUDIANTE (autoservicio, solo sus datos):
    GET /academic/student/merito?period=
      → su promedio y su puesto en aula / especialidad / instituto,
        y si alcanza el DESCUENTO del 25% de la matrícula (promedio ≥ 17
        del aula, política institucional).
    GET /academic/student/merito/constancia.pdf?period=
      → constancia que ACREDITA el mérito para el descuento (solo si ≥ 17).

El puesto usa ranking de competencia (empates comparten puesto: 1,2,2,4).
El "ciclo cursado" del alumno sale de los cursos de su kárdex del período,
no de Student.ciclo (que ya avanzó al ciclo siguiente al promover).
"""
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from rest_framework_simplejwt.authentication import JWTAuthentication

from academic.models import AcademicGradeRecord
from students.models import Student
from students.name_utils import clave_orden

from .evaluation_pdf import (
    _filtrar_students, _promedios_por_alumno, _nombre, _pdf_shell,
    _require_grades_admin, _esc, _roman,
)
from ..pdf_render import html_to_pdf_bytes

UMBRAL_DESCUENTO = 17.0


def _ciclo_cursado(student_id, period):
    """Ciclo que el alumno CURSÓ en el período (máximo semestre del kárdex)."""
    sems = (AcademicGradeRecord.objects
            .filter(student_id=student_id, term=period,
                    plan_course__semester__isnull=False)
            .values_list("plan_course__semester", flat=True))
    return max(sems) if sems else None


def _ranking(students, proms):
    """[(puesto, student, promedio)] — mayor promedio primero; los empates
    comparten puesto (1, 2, 2, 4)."""
    con_prom = [(st, proms[st.id]) for st in students if st.id in proms]
    con_prom.sort(key=lambda t: (-t[1], clave_orden(_nombre(t[0]))))
    out, puesto_prev, prom_prev = [], 0, None
    for i, (st, prom) in enumerate(con_prom, 1):
        puesto = puesto_prev if prom == prom_prev else i
        out.append((puesto, st, prom))
        puesto_prev, prom_prev = puesto, prom
    return out


def _puesto_de(ranking, student_id):
    """(puesto, total) del alumno dentro de un ranking, o (None, total)."""
    for puesto, st, _prom in ranking:
        if st.id == student_id:
            return puesto, len(ranking)
    return None, len(ranking)


def _fecha_emision():
    return timezone.localtime(timezone.now()).strftime(
        "Documento generado el %d/%m/%Y a las %H:%M")


# ══════════════════════════════════════════════════════════════
# ADMIN — Orden de mérito (aula / especialidad / instituto)
# ══════════════════════════════════════════════════════════════

class EvaluationMeritoView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    fmt = "pdf"   # sobreescrito por urls.py

    def get(self, request):
        if err := _require_grades_admin(request):
            return err
        q = request.query_params
        period = (q.get("period") or "").strip().upper()
        if not period:
            return Response({"detail": "period es requerido"}, status=400)
        career_id, semester = q.get("career_id"), q.get("semester")

        students = _filtrar_students(period, career_id, semester,
                                     q.get("anio"))
        if not students:
            return Response(
                {"detail": f"No hay alumnos con notas procesadas en {period} "
                           "con los filtros elegidos."}, status=404)
        proms = _promedios_por_alumno(term=period,
                                      student_ids=[s.id for s in students])
        ranking = _ranking(students, proms)
        if not ranking:
            return Response({"detail": "Ningún alumno tiene promedio."},
                            status=404)

        if career_id and semester:
            carrera = next((s.plan.career.name for s in students
                            if s.plan and s.plan.career), "")
            alcance = f"AULA — {carrera.upper()} CICLO {_roman(semester)}"
        elif career_id:
            carrera = next((s.plan.career.name for s in students
                            if s.plan and s.plan.career), "")
            alcance = f"ESPECIALIDAD — {carrera.upper()}"
        else:
            alcance = "TODO EL INSTITUTO"

        filas = []
        for puesto, st, prom in ranking:
            carrera_st = (st.plan.career.name
                          if st.plan and st.plan.career else "")
            ciclo = _ciclo_cursado(st.id, period)
            filas.append([puesto, st.num_documento or "", _nombre(st),
                          carrera_st.upper(), _roman(ciclo) if ciclo else "",
                          f"{prom:.2f}",
                          "SÍ" if prom >= UMBRAL_DESCUENTO else ""])

        titulo = f"ORDEN DE MÉRITO {alcance} — {period}"
        if self.fmt == "xlsx":
            return self._xlsx(titulo, period, filas)
        return self._pdf(titulo, filas)

    HEADERS = ["Puesto", "DNI", "Apellidos y Nombres", "Programa de Estudios",
               "Ciclo", "Promedio\nPonderado", f"Beca 25%\n(≥ {UMBRAL_DESCUENTO:g})"]

    def _pdf(self, titulo, filas):
        def _td(i, v):
            attr = "" if i == 2 else " class='c'"
            return f"<td{attr}>{_esc(str(v))}</td>"

        cuerpo_filas = "".join(
            "<tr>" + "".join(_td(i, v) for i, v in enumerate(fila)) + "</tr>"
            for fila in filas)
        heads = "".join(f"<th>{_esc(h)}</th>"
                        for h in [h.replace(chr(10), " ") for h in self.HEADERS])
        cuerpo = f"""
<table>
  <thead><tr>{heads}</tr></thead>
  <tbody>{cuerpo_filas}</tbody>
</table>
<p style="margin-top:6px; font-size:8px; color:#777; text-align:right">
  Total: {len(filas)} estudiante(s) · {_esc(_fecha_emision())}</p>
"""
        html = _pdf_shell(titulo, cuerpo)
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     'attachment; filename="orden-merito.pdf"'})

    def _xlsx(self, titulo, period, filas):
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from .acta_excel import _acta_area_inst

        wb = Workbook()
        ws = wb.active
        ws.title = "ORDEN DE MÉRITO"
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        n = len(self.HEADERS)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        t = ws.cell(row=1, column=1, value=_acta_area_inst()["nombre"])
        t.font = Font(bold=True, size=12)
        t.alignment = center
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        s = ws.cell(row=2, column=1, value=titulo)
        s.font = Font(bold=True, size=10)
        s.alignment = center

        H = 4
        for c, h in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=H, column=c, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", start_color="E7E6E6")
            cell.alignment = center
            cell.border = border
        for r, fila in enumerate(filas, H + 1):
            for c, v in enumerate(fila, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(size=9)
                cell.border = border
                if c != 3:
                    cell.alignment = Alignment(horizontal="center")
        for c, w in enumerate((8, 12, 40, 26, 7, 11, 10), 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{H + 1}"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.getvalue(),
            content_type=("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     f'attachment; filename="orden-merito-{period}.xlsx"'})


# ══════════════════════════════════════════════════════════════
# ESTUDIANTE — mi puesto y mi constancia
# ══════════════════════════════════════════════════════════════

def _periodo_por_defecto():
    from catalogs.models import Period
    p = Period.objects.filter(is_active=True).first()
    if not p:
        return ""
    return (p.code or f"{p.year}-{p.term}").strip().upper()


def _periodos_candidatos(student, explicito=""):
    """Períodos donde buscar el mérito del alumno, en orden de preferencia.

    Con `?period=` explícito, solo ese. Si no: el período activo primero y
    después TODOS los términos del kárdex del alumno, del más reciente al
    más antiguo. Así el panel no muere con "No hay período activo" (pasó en
    producción) ni queda vacío cuando el período recién empieza y las notas
    procesadas son las del anterior.
    """
    if explicito:
        return [explicito]
    from .utils import _term_sort_key
    out = []
    activo = _periodo_por_defecto()
    if activo:
        out.append(activo)
    terms = (AcademicGradeRecord.objects.filter(student=student)
             .values_list("term", flat=True).distinct())
    propios = sorted({(t or "").strip().upper() for t in terms if t},
                     key=_term_sort_key, reverse=True)
    out.extend(t for t in propios if t not in out)
    return out


def _merito_del_estudiante(student, period):
    """Puesto del alumno en aula/especialidad/instituto. None si no tiene
    notas procesadas en el período."""
    prom_propio = _promedios_por_alumno(
        term=period, student_ids=[student.id]).get(student.id)
    if prom_propio is None:
        return None

    ciclo = _ciclo_cursado(student.id, period)
    career_id = student.plan.career_id if student.plan_id and student.plan else None

    niveles = {}
    for clave, (cid, sem) in {
        "aula": (career_id, ciclo),
        "especialidad": (career_id, None),
        "instituto": (None, None),
    }.items():
        cohorte = _filtrar_students(period, cid, sem)
        proms = _promedios_por_alumno(term=period,
                                      student_ids=[s.id for s in cohorte])
        puesto, total = _puesto_de(_ranking(cohorte, proms), student.id)
        niveles[clave] = {"puesto": puesto, "total": total}

    return {
        "period": period,
        "promedio": prom_propio,
        "ciclo_cursado": ciclo,
        "carrera": (student.plan.career.name
                    if student.plan_id and student.plan and student.plan.career
                    else (student.programa_carrera or "")),
        "umbral_descuento": UMBRAL_DESCUENTO,
        "alcanza_descuento": prom_propio >= UMBRAL_DESCUENTO,
        **niveles,
    }


class StudentMeritoView(APIView):
    """Mérito del alumno autenticado — solo lee SUS datos."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        st = Student.objects.select_related("plan", "plan__career")\
            .filter(user=request.user).first()
        if not st:
            return Response({"detail": "No se encontró perfil de estudiante."},
                            status=404)
        explicito = (request.query_params.get("period") or "").strip().upper()
        for period in _periodos_candidatos(st, explicito):
            data = _merito_del_estudiante(st, period)
            if data is not None:
                return Response(data)
        return Response(
            {"detail": "Aún no tienes notas procesadas en ningún período. "
                       "Tu puesto aparecerá cuando Secretaría Académica "
                       "procese las calificaciones."}, status=404)


class StudentMeritoConstanciaView(APIView):
    """Constancia PDF que acredita el mérito para el descuento del 25%."""
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .acta_excel import _acta_area_inst

        st = Student.objects.select_related("plan", "plan__career")\
            .filter(user=request.user).first()
        if not st:
            return Response({"detail": "No se encontró perfil de estudiante."},
                            status=404)
        explicito = (request.query_params.get("period") or "").strip().upper()
        data = None
        for period in _periodos_candidatos(st, explicito):
            data = _merito_del_estudiante(st, period)
            if data is not None:
                break
        if not data:
            return Response(
                {"detail": "No tienes notas procesadas en ningún período."},
                status=404)
        period = data["period"]
        if not data["alcanza_descuento"]:
            return Response(
                {"detail": f"Tu promedio de {period} es {data['promedio']:.2f}: "
                           f"la constancia se emite desde "
                           f"{UMBRAL_DESCUENTO:g} (descuento del 25%)."},
                status=409)

        inst = _acta_area_inst()
        aula = data["aula"]
        anio = period.split("-")[0]
        cuerpo = f"""
<div style="text-align:center; margin-top:30px">
  <h2 style="font-size:16px">CONSTANCIA DE MÉRITO ACADÉMICO</h2>
</div>
<p style="font-size:11px; line-height:1.9; text-align:justify; margin-top:18px">
El (La) que suscribe, Secretario(a) Académico(a) del {_esc(inst['nombre'])},
hace constar que:
</p>
<p style="text-align:center; font-size:13px"><b>{_esc(_nombre(st))}</b><br>
<span style="font-size:10px">DNI N° {_esc(st.num_documento or '')} —
Programa de Estudios de {_esc((data['carrera'] or '').upper())} —
Ciclo {_esc(_roman(data['ciclo_cursado']) if data['ciclo_cursado'] else '')}</span></p>
<p style="font-size:11px; line-height:1.9; text-align:justify">
ha obtenido un <b>promedio ponderado de {data['promedio']:.2f}</b> en el período
académico <b>{_esc(period)}</b>, ocupando el
<b>puesto {aula['puesto']} de {aula['total']}</b> estudiantes de su aula,
con un rendimiento igual o superior a {UMBRAL_DESCUENTO:g}, por lo que
<b>ACREDITA el derecho al descuento del 25% del pago de matrícula</b>
conforme a las disposiciones institucionales vigentes.
</p>
<p style="font-size:11px; line-height:1.9; text-align:justify">
Se expide la presente constancia a solicitud del (de la) interesado(a) para
los fines que estime conveniente.
</p>
<p style="font-size:10px; text-align:right; margin-top:24px">
Tarma, ____ de ____________ de {_esc(anio)}</p>
<div class="firma" style="margin-top:60px"><span class="linea">
SECRETARIO(A) ACADÉMICO(A)<br>
<span style="font-size:8px; font-weight:normal">Firma, Post Firma y Sello</span></span></div>
<p style="margin-top:16px; font-size:8px; color:#777; text-align:center">
{_esc(_fecha_emision())}</p>
"""
        html = _pdf_shell(f"CONSTANCIA DE MÉRITO — {period}", cuerpo)
        dni = st.num_documento or st.id
        return HttpResponse(
            html_to_pdf_bytes(html), content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="constancia-merito-{dni}-{period}.pdf"'})
