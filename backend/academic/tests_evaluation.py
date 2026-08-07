"""
Tests del Centro de Evaluación y del acta oficial.
    python manage.py test academic.tests_evaluation -v 2

Con DUMP_SAMPLES=<dir> además guarda los Excel generados en <dir>
(para revisión visual de los formatos).
"""
import os
from io import BytesIO
from types import SimpleNamespace


def _dump(nombre, contenido):
    d = os.environ.get("DUMP_SAMPLES")
    if d:
        with open(os.path.join(d, nombre), "wb") as fh:
            fh.write(contenido)

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate
from openpyxl import Workbook, load_workbook

from catalogs.models import Career
from students.models import Student
from academic.models import (
    Course, Plan, PlanCourse, Section, Teacher, Enrollment, EnrollmentItem,
    SectionGrades, AcademicGradeRecord, AcademicPeriod,
)
from academic.views import acta_excel as AE
from academic.views.evaluation import (
    EvaluationStateView, EvaluationProcessView, EvaluationActaConsolidadaView,
    EvaluationSectionsView, _section_eval_row, _entry_for, _final_of,
)
from academic.views.teachers import (
    _check_grades_window, AdminGradesOverviewView,
    _calc_escala_0_5, _calc_promedio_final_0_20, _calif_curso,
)

User = get_user_model()
PERIOD = "2026-I"


class BaseEvalTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user("admin_test", "admin_test@t.pe", "x")
        cls.admin.is_staff = True
        cls.admin.save()

        cls.tuser = User.objects.create_user(
            "21065570", "docente@t.pe", "x", full_name="GOYAS BALDOCEDA, ANA MARIA")
        cls.teacher = Teacher.objects.create(user=cls.tuser)

        cls.career = Career.objects.create(name="EDUCACIÓN INICIAL", code="EI")
        cls.plan = Plan.objects.create(career=cls.career, name="Plan 2020")
        cls.c1 = Course.objects.create(code="DP1", name="Desarrollo Personal I")
        cls.c2 = Course.objects.create(code="PM1", name="Resolución de Problemas Matemáticos I")
        cls.pc1 = PlanCourse.objects.create(plan=cls.plan, course=cls.c1, semester=1, credits=3)
        cls.pc2 = PlanCourse.objects.create(plan=cls.plan, course=cls.c2, semester=1, credits=5)
        cls.sec1 = Section.objects.create(plan_course=cls.pc1, teacher=cls.teacher,
                                          period=PERIOD, label="A")
        cls.sec2 = Section.objects.create(plan_course=cls.pc2, teacher=cls.teacher,
                                          period=PERIOD, label="A")

        # st1 con cuenta de usuario (clave del acta = user_id); st2 sin cuenta (clave = pk)
        su = User.objects.create_user("60634719", "st1@t.pe", "x",
                                      full_name="ATAPOMA ROQUE, SHEYLA")
        cls.st1 = Student.objects.create(user=su, num_documento="60634719",
                                         nombres="SHEYLA", apellido_paterno="ATAPOMA",
                                         apellido_materno="ROQUE")
        cls.st2 = Student.objects.create(num_documento="61576678", nombres="DAYANIRA",
                                         apellido_paterno="BALDEON", apellido_materno="CRUZ")
        for st in (cls.st1, cls.st2):
            e = Enrollment.objects.create(student=st, period=PERIOD,
                                          status=Enrollment.STATUS_CONFIRMED)
            EnrollmentItem.objects.create(enrollment=e, plan_course=cls.pc1,
                                          section=cls.sec1, credits=3)
            EnrollmentItem.objects.create(enrollment=e, plan_course=cls.pc2,
                                          section=cls.sec2, credits=5)

        cls.factory = APIRequestFactory()

    # helpers
    def _get(self, view_cls, url, params=None):
        req = self.factory.get(url, params or {})
        force_authenticate(req, user=self.admin)
        return view_cls.as_view()(req)

    def _post(self, view_cls, url, body):
        req = self.factory.post(url, body, format="json")
        force_authenticate(req, user=self.admin)
        return view_cls.as_view()(req)


class ActaOficialTests(BaseEvalTest):
    def test_generacion_acta_oficial(self):
        wb = Workbook()
        ws = wb.active
        n = AE._write_grades_sheet(ws, self.sec1)
        buf = BytesIO()
        wb.save(buf)
        _dump("muestra_acta_calificacion.xlsx", buf.getvalue())
        self.assertEqual(n, 2)
        self.assertEqual(ws["A2"].value, f"SECCION_ID: {self.sec1.id}")
        self.assertIn("ACTA_OFICIAL", str(ws["A3"].value))
        # cabecera
        self.assertEqual(ws.cell(row=6, column=3).value, "EDUCACIÓN INICIAL")
        self.assertEqual(ws.cell(row=6, column=25).value, PERIOD)
        self.assertIn('I - "A"', str(ws.cell(row=8, column=25).value))
        # alumnos ordenados alfabéticamente desde la fila 16
        self.assertEqual(ws.cell(row=16, column=2).value, "60634719")   # ATAPOMA
        self.assertEqual(ws.cell(row=17, column=2).value, "61576678")   # BALDEON
        # fórmulas vivas
        self.assertTrue(str(ws.cell(row=16, column=22).value).startswith("=IF(COUNTA(D16:H16"))
        self.assertTrue(str(ws.cell(row=16, column=26).value).startswith("=IF(Y16="))
        # resumen (2 alumnos → última fila 17, resumen en 17+3=20)
        self.assertEqual(ws.cell(row=20, column=2).value, "Resumen")
        self.assertEqual(ws.cell(row=21, column=2).value, "Matriculados")
        self.assertEqual(ws.cell(row=21, column=4).value, 2)

    def test_import_acta_oficial_roundtrip(self):
        wb = Workbook()
        ws = wb.active
        AE._write_grades_sheet(ws, self.sec1)
        # Marcar st1 (fila 16): C1=L (col G), C2=P (col L), C3=D (col T) + recomendación
        ws.cell(row=16, column=7, value="X")
        ws.cell(row=16, column=12, value="x")     # minúscula también vale
        ws.cell(row=16, column=20, value="X")
        ws.cell(row=16, column=9, value="Buen avance")

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        ws2 = load_workbook(buf).active

        req = SimpleNamespace(user=self.admin)
        res = AE._import_grades_ws(ws2, req)
        self.assertEqual(res["errores"], [])
        self.assertEqual(res["importados"], 1)

        bundle = SectionGrades.objects.get(section=self.sec1)
        key = str(self.st1.user_id)      # alumno con cuenta → clave user_id
        self.assertIn(key, bundle.grades)
        entry = bundle.grades[key]
        self.assertEqual(entry["C1_LEVEL"], "L")
        self.assertEqual(entry["C2_LEVEL"], "P")
        self.assertEqual(entry["C3_LEVEL"], "D")
        self.assertEqual(entry["C1_REC"], "Buen avance")
        # L,P,D → 4,3,5 → escala 4.0 → promedio ((4-1)/4)*20 = 15
        self.assertEqual(entry["PROMEDIO_FINAL"], 15)
        self.assertEqual(entry["ESTADO"], "Logrado")

    def test_import_marcas_numericas(self):
        """El docente puede marcar con la puntuación (4/1/5) en vez de X,
        como en el modelo oficial (ejemplo del documento → 3.3 → 12)."""
        wb = Workbook()
        ws = wb.active
        AE._write_grades_sheet(ws, self.sec1)
        # st1 (fila 16): C1=L→"4" (col G), C2=PI→"1" (col J), C3=D→"5" (col T)
        ws.cell(row=16, column=7, value=4)
        ws.cell(row=16, column=10, value=1)
        ws.cell(row=16, column=20, value=5)
        res = AE._import_grades_ws(ws, SimpleNamespace(user=self.admin))
        self.assertEqual(res["errores"], [])
        self.assertEqual(res["importados"], 1)
        entry = SectionGrades.objects.get(section=self.sec1).grades[str(self.st1.user_id)]
        self.assertEqual(entry["C1_LEVEL"], "L")
        self.assertEqual(entry["C2_LEVEL"], "PI")
        self.assertEqual(entry["C3_LEVEL"], "D")
        self.assertEqual(entry["ESCALA_0_5"], 3.3)
        self.assertEqual(entry["PROMEDIO_FINAL"], 12)
        self.assertEqual(entry["ESTADO"], "En proceso")

    def test_import_valor_invalido_en_columna(self):
        """Un número que no corresponde a la columna se rechaza con mensaje claro."""
        wb = Workbook()
        ws = wb.active
        AE._write_grades_sheet(ws, self.sec1)
        ws.cell(row=16, column=7, value=2)   # "2" en la columna L (val 4) → inválido
        res = AE._import_grades_ws(ws, SimpleNamespace(user=self.admin))
        self.assertEqual(res["importados"], 0)
        self.assertTrue(any("no válido" in e for e in res["errores"]))

    def test_import_doble_marca_es_error(self):
        wb = Workbook()
        ws = wb.active
        AE._write_grades_sheet(ws, self.sec1)
        ws.cell(row=16, column=4, value="X")
        ws.cell(row=16, column=7, value="X")   # dos niveles en Comp 1
        res = AE._import_grades_ws(ws, SimpleNamespace(user=self.admin))
        self.assertEqual(res["importados"], 0)
        self.assertTrue(any("más de un nivel" in e for e in res["errores"]))

    def test_import_formato_legado_sigue_funcionando(self):
        wb = Workbook()
        ws = wb.active
        ws["A2"] = f"SECCION_ID: {self.sec1.id}"
        headers = ["N°", "DNI", "APELLIDOS Y NOMBRES"] + AE.GRADE_COLS
        for c, h in enumerate(headers, 1):
            ws.cell(row=7, column=c, value=h)
        ws.append([])  # noop
        row = [1, "61576678", "BALDEON CRUZ, DAYANIRA", "P", "", "L", "", "I", ""]
        for c, v in enumerate(row, 1):
            ws.cell(row=8, column=c, value=v)
        res = AE._import_grades_ws(ws, SimpleNamespace(user=self.admin))
        self.assertEqual(res["errores"], [])
        self.assertEqual(res["importados"], 1)
        bundle = SectionGrades.objects.get(section=self.sec1)
        self.assertIn(str(self.st2.id), bundle.grades)   # sin user → clave pk


class CompetenciasDecimalesTests(TestCase):
    """Observaciones de los docentes: decimales por competencia con rango,
    y cursos con 1 o 2 competencias (Inglés)."""

    def _norm(self, payload):
        from academic.views.teachers import _normalize_acta_student_payload
        return _normalize_acta_student_payload(payload)

    def test_decimales_dentro_del_rango(self):
        out, errs = self._norm({"C1": 4.5, "C2": "3.7", "C3": 5})
        self.assertEqual(errs, [])
        self.assertEqual(out["C1"], 4.5)
        self.assertEqual(out["C1_LEVEL"], "L")     # nivel deducido del valor
        self.assertEqual(out["C2_LEVEL"], "P")
        self.assertEqual(out["C3_LEVEL"], "D")
        # (4.5+3.7+5)/3 = 4.4 → vigesimal 17 → Logrado
        self.assertEqual(out["ESCALA_0_5"], 4.4)
        self.assertEqual(out["PROMEDIO_FINAL"], 17)
        self.assertEqual(out["ESTADO"], "Logrado")

    def test_valor_fuera_del_rango_del_nivel_se_rechaza(self):
        # Caso del docente: marca L (4.0-4.9) y escribe 5.1 → rechazado
        out, errs = self._norm({"C1_LEVEL": "L", "C1": 5.1})
        self.assertIsNone(out)
        self.assertTrue(errs)
        # Dentro de 1-5 pero fuera del nivel marcado
        out2, errs2 = self._norm({"C1_LEVEL": "L", "C1": 3.9})
        self.assertIsNone(out2)
        self.assertTrue(any("no corresponde al nivel L" in e for e in errs2), errs2)
        self.assertTrue(any("4 – 4.9" in e for e in errs2), errs2)
        # 5.0 en D sí es válido (rango exacto)
        out3, errs3 = self._norm({"C1_LEVEL": "D", "C1": 5.0})
        self.assertEqual(errs3, [])
        self.assertEqual(out3["C1"], 5.0)

    def test_valor_fuera_de_1_a_5(self):
        out, errs = self._norm({"C1": 7})
        self.assertIsNone(out)
        self.assertTrue(any("fuera del rango permitido" in e for e in errs), errs)

    def test_curso_con_una_sola_competencia(self):
        """Inglés: una competencia → calcula igual sin exigir las otras."""
        out, errs = self._norm({"C1": 4.2})
        self.assertEqual(errs, [])
        self.assertEqual(out["ESCALA_0_5"], 4.2)
        self.assertEqual(out["PROMEDIO_FINAL"], 16)
        self.assertEqual(out["ESTADO"], "Logrado")
        self.assertEqual(out["N_COMPETENCIAS"], 1)
        self.assertEqual(out["C2"], "")

    def test_dos_competencias(self):
        out, errs = self._norm({"C1": 3.0, "C2": 4.0})
        self.assertEqual(errs, [])
        self.assertEqual(out["ESCALA_0_5"], 3.5)
        self.assertEqual(out["N_COMPETENCIAS"], 2)

    def test_solo_nivel_sin_valor_usa_minimo(self):
        out, errs = self._norm({"C1_LEVEL": "P", "C2_LEVEL": "L", "C3_LEVEL": "D"})
        self.assertEqual(errs, [])
        self.assertEqual([out["C1"], out["C2"], out["C3"]], [3.0, 4.0, 5.0])
        self.assertEqual(out["ESCALA_0_5"], 4.0)

    def test_sin_ninguna_competencia_es_error(self):
        out, errs = self._norm({"C1_REC": "solo un comentario"})
        self.assertIsNone(out)
        self.assertIn("Debe registrar al menos una competencia", errs)


class ConversionOficialTests(TestCase):
    """Tabla oficial RVM 123-2022 pág. 28 (escala 1-5 → vigesimal 1-20)."""

    def test_ejemplo_del_documento(self):
        # C1=4 (L), C2=1 (PI), C3=5 (D) → escala 3.3 → 12 → "En proceso"
        esc = _calc_escala_0_5(4, 1, 5)
        self.assertEqual(esc, 3.3)
        self.assertEqual(_calc_promedio_final_0_20(esc), 12)
        self.assertEqual(_calif_curso(esc), "En proceso")

    def test_bandas_oficiales(self):
        casos = [
            (1.0, 1), (1.3, 2), (1.5, 3), (1.9, 5),
            (2.0, 6), (2.5, 8), (2.9, 10),
            (3.0, 11), (3.2, 11), (3.3, 12), (3.5, 12), (3.6, 13), (3.8, 14), (3.9, 14),
            (4.0, 15), (4.1, 15), (4.3, 16), (4.5, 17), (4.7, 18), (4.9, 19),
            (5.0, 20),
        ]
        for escala, esperado in casos:
            self.assertEqual(_calc_promedio_final_0_20(escala), esperado,
                             f"escala {escala} debería dar {esperado}")

    def test_redondeo_a_favor(self):
        # (3+3+4)/3 = 3.333… → 3.3 (trunca al decimal más cercano)
        self.assertEqual(_calc_escala_0_5(3, 3, 4), 3.3)
        # (4+5+5)/3 = 4.666… → 4.7 (0.05 a favor del estudiante)
        self.assertEqual(_calc_escala_0_5(4, 5, 5), 4.7)
        # (5+5+5)/3 = 5.0
        self.assertEqual(_calc_escala_0_5(5, 5, 5), 5.0)

    def test_calif_cualitativa(self):
        self.assertEqual(_calif_curso(5.0), "Destacado")
        self.assertEqual(_calif_curso(4.0), "Logrado")
        self.assertEqual(_calif_curso(3.0), "En proceso")
        self.assertEqual(_calif_curso(2.0), "Inicio")
        self.assertEqual(_calif_curso(1.5), "Previo al inicio")


class ActaAreaTests(BaseEvalTest):
    def test_acta_area_descarga(self):
        from academic.views.acta_excel import SectionActaAreaView
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {
                str(self.st1.user_id): {"PROMEDIO_FINAL": 15},
                str(self.st2.id): {"PROMEDIO_FINAL": 10},
            }},
        )
        req = self.factory.get("/x")
        force_authenticate(req, user=self.admin)
        res = SectionActaAreaView.as_view()(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 200)
        _dump("muestra_acta_area.xlsx", res.content)

        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        self.assertIn("ACTA DE EVALUACIÓN DE ÁREA", str(ws.cell(row=2, column=1).value))
        # alumnos desde la fila 15: ATAPOMA → calificativo 15, crédito 3, puntaje 45, Logrado
        self.assertEqual(ws.cell(row=15, column=2).value, "60634719")
        self.assertEqual(ws.cell(row=15, column=5).value, 15)
        self.assertEqual(ws.cell(row=15, column=6).value, 3)
        self.assertEqual(ws.cell(row=15, column=7).value, 45)
        self.assertEqual(ws.cell(row=15, column=8).value, "Logrado")
        # BALDEON → 10 → Inicio
        self.assertEqual(ws.cell(row=16, column=5).value, 10)
        self.assertEqual(ws.cell(row=16, column=8).value, "Inicio")
        # La leyenda fue retirada del acta de área (pedido del 01-08-2026)
        self.assertIsNone(ws.cell(row=3, column=10).value)
        # Director y RD de encargatura completados
        textos = [str(ws.cell(row=r, column=c).value or "")
                  for r in range(1, 14) for c in range(1, 9)]
        self.assertTrue(any("GARCIA PORRAS" in t for t in textos))
        self.assertTrue(any("017-2026" in t for t in textos))

    def test_acta_area_docente_propio_y_ajeno(self):
        from academic.views.acta_excel import SectionActaAreaView
        # El docente de la sección puede descargarla
        req = self.factory.get("/x")
        force_authenticate(req, user=self.tuser)
        res = SectionActaAreaView.as_view()(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 200)
        # Otro docente NO
        otro = User.objects.create_user("otro_doc", "otro@t.pe", "x")
        req2 = self.factory.get("/x")
        force_authenticate(req2, user=otro)
        res2 = SectionActaAreaView.as_view()(req2, section_id=self.sec1.id)
        self.assertEqual(res2.status_code, 403)


class ReportesEvaluacionTests(BaseEvalTest):
    def _notas(self):
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {
                str(self.st1.user_id): {"PROMEDIO_FINAL": 15},
                str(self.st2.id): {"PROMEDIO_FINAL": 10},
            }},
        )

    def test_actas_area_zip(self):
        from academic.views.evaluation import EvaluationActasAreaZipView
        self._notas()
        res = self._get(EvaluationActasAreaZipView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        import zipfile as zf
        z = zf.ZipFile(BytesIO(res.content))
        # 2 secciones con alumnos → 2 actas
        xlsx = [n for n in z.namelist() if n.endswith(".xlsx")]
        self.assertEqual(len(xlsx), 2)

    def test_actas_area_zip_filtro_ciclo_sin_resultado(self):
        from academic.views.evaluation import EvaluationActasAreaZipView
        res = self._get(EvaluationActasAreaZipView, "/x", {"period": PERIOD, "semester": 9})
        self.assertEqual(res.status_code, 404)

    def test_reporte_rendimiento(self):
        from academic.views.evaluation import EvaluationReporteRendimientoView
        self._notas()
        res = self._get(EvaluationReporteRendimientoView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        _dump("muestra_reporte_rendimiento.xlsx", res.content)
        wb = load_workbook(BytesIO(res.content))
        self.assertEqual(wb.sheetnames, ["Resumen por curso", "Detalle por alumno"])
        ws = wb["Resumen por curso"]
        # sec1: 2 matriculados, 2 con notas, 1 aprobado, 1 desaprobado
        fila = [ws.cell(row=4, column=c).value for c in range(1, 13)]
        self.assertEqual(fila[5], 2)   # matriculados
        self.assertEqual(fila[6], 2)   # con notas
        self.assertEqual(fila[7], 1)   # aprobados
        self.assertEqual(fila[8], 1)   # desaprobados
        self.assertEqual(fila[10], 12.5)  # promedio (15+10)/2
        ws2 = wb["Detalle por alumno"]
        self.assertEqual(ws2.cell(row=2, column=7).value, 15)
        self.assertEqual(ws2.cell(row=2, column=8).value, "Logrado")

    def test_reporte_rendimiento_filtro_curso(self):
        from academic.views.evaluation import EvaluationReporteRendimientoView
        self._notas()
        res = self._get(EvaluationReporteRendimientoView, "/x",
                        {"period": PERIOD, "section_id": self.sec1.id})
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        ws = wb["Resumen por curso"]
        # solo 1 sección en el resumen (fila 4 con datos, fila 5 vacía)
        self.assertIsNotNone(ws.cell(row=4, column=1).value)
        self.assertIsNone(ws.cell(row=5, column=1).value)

    def test_consolidada_filtro_ciclo(self):
        self._notas()
        res = self._get(EvaluationActaConsolidadaView, "/x",
                        {"period": PERIOD, "semester": 1})
        self.assertEqual(res.status_code, 200)
        res2 = self._get(EvaluationActaConsolidadaView, "/x",
                         {"period": PERIOD, "semester": 9})
        self.assertEqual(res2.status_code, 404)


class BoletaAlumnoTests(BaseEvalTest):
    def test_cursos_del_alumno_con_notas_y_asistencia(self):
        from academic.views.dashboard_student import student_courses_detail
        from academic.models import AttendanceSession, AttendanceRow
        import datetime as dt

        # Nota viva en el acta de sec1 + asistencia con 2 faltas de 4 sesiones
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {str(self.st1.user_id): {
                "C1": 4, "C2": 3, "C3": 5, "PROMEDIO_FINAL": 15,
                "C1_LEVEL": "L", "C2_LEVEL": "P", "C3_LEVEL": "D",
            }}})
        for i, status in enumerate(["PRESENT", "ABSENT", "ABSENT", "LATE"], 1):
            s = AttendanceSession.objects.create(
                section=self.sec1, date=dt.date(2026, 6, i))
            AttendanceRow.objects.create(
                session=s, student_id=self.st1.user_id, status=status)

        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.st1.user)
        res = student_courses_detail(req)
        self.assertEqual(res.status_code, 200, res.data)
        cursos = {c["code"]: c for c in res.data["cursos"]}
        c1 = cursos["DP1"]
        self.assertEqual(c1["promedio"], 15)
        self.assertEqual(c1["calificacion"], "Logrado")
        self.assertEqual(c1["c1"], 4)
        self.assertEqual(c1["sesiones"], 4)
        self.assertEqual(c1["faltas"], 2)
        self.assertEqual(c1["tardanzas"], 1)
        self.assertEqual(c1["pct_faltas"], 50)
        self.assertTrue(c1["en_riesgo"])
        # curso 2 sin notas ni sesiones
        c2 = cursos["PM1"]
        self.assertIsNone(c2["promedio"])
        self.assertEqual(c2["faltas"], 0)
        # PGA ponderado: solo curso 1 con nota (3 créditos) → 15
        self.assertEqual(res.data["pga"], 15.0)

    def test_kardex_manda_sobre_acta(self):
        from academic.views.dashboard_student import student_courses_detail
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {str(self.st1.user_id): {"PROMEDIO_FINAL": 12}}})
        AcademicGradeRecord.objects.create(
            student=self.st1, course=self.c1, plan_course=self.pc1,
            term=PERIOD, final_grade=14)
        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.st1.user)
        res = student_courses_detail(req)
        c1 = next(c for c in res.data["cursos"] if c["code"] == "DP1")
        self.assertEqual(c1["promedio"], 14)   # kárdex procesado manda
        self.assertTrue(c1["procesado"])


class KardexAccesoTests(BaseEvalTest):
    """El kárdex de otro alumno NO debe ser accesible por un estudiante."""

    def setUp(self):
        super().setUp()
        # Alumna con cuenta que intentará espiar a otra
        self.espia_user = User.objects.create_user("espia", "espia@t.pe", "x")
        self.espia = Student.objects.create(
            user=self.espia_user, num_documento="70000001",
            nombres="GRAYSI", apellido_paterno="ESPIA")
        try:
            from acl.models import Role, UserRole
            rol, _ = Role.objects.get_or_create(name="STUDENT")
            UserRole.objects.get_or_create(user=self.espia_user, role=rol)
        except Exception:
            pass

    def _get(self, view_cls, user, student_id, **params):
        req = self.factory.get("/x", params)
        force_authenticate(req, user=user)
        return view_cls.as_view()(req, student_id=student_id)

    def test_alumno_no_ve_kardex_ajeno(self):
        from academic.views.kardex import KardexView
        # por DNI del otro alumno (el caso reportado)
        res = self._get(KardexView, self.espia_user, self.st1.num_documento)
        self.assertEqual(res.status_code, 403)
        self.assertIn("tu propio kárdex", res.data["detail"])
        # por id interno también
        res2 = self._get(KardexView, self.espia_user, self.st1.id)
        self.assertEqual(res2.status_code, 403)

    def test_alumno_si_ve_el_suyo(self):
        from academic.views.kardex import KardexView
        res = self._get(KardexView, self.espia_user, self.espia.num_documento)
        self.assertEqual(res.status_code, 200)
        res2 = self._get(KardexView, self.espia_user, self.espia.id)
        self.assertEqual(res2.status_code, 200)

    def test_pdfs_ajenos_bloqueados(self):
        from academic.views.kardex import (
            KardexRecordNotasPDFView, KardexBoletaPeriodoPDFView,
            KardexExportXlsxView, KardexFichaRendimientoPDFView,
        )
        for view in (KardexRecordNotasPDFView, KardexBoletaPeriodoPDFView,
                     KardexExportXlsxView, KardexFichaRendimientoPDFView):
            res = self._get(view, self.espia_user, self.st1.num_documento,
                            period=PERIOD)
            self.assertEqual(res.status_code, 403, f"{view.__name__} no bloqueó")

    def test_admin_y_docente_si_consultan(self):
        from academic.views.kardex import KardexView
        self.assertEqual(
            self._get(KardexView, self.admin, self.st1.num_documento).status_code, 200)
        self.assertEqual(
            self._get(KardexView, self.tuser, self.st1.num_documento).status_code, 200)


class VentanaNotasTests(BaseEvalTest):
    """La ventana de carga habilita/bloquea al docente con mensaje claro."""

    def _periodo(self, **kw):
        from datetime import date
        per, _ = AcademicPeriod.objects.update_or_create(
            code=PERIOD,
            defaults={"start": date(2026, 3, 1), "end": date(2026, 7, 31), **kw},
        )
        return per

    def _window(self, user=None):
        from academic.views.teachers import SectionGradesWindowView
        req = self.factory.get("/x")
        force_authenticate(req, user=user or self.tuser)
        return SectionGradesWindowView.as_view()(req, section_id=self.sec1.id)

    def test_abierta_sin_fechas(self):
        self._periodo()
        res = self._window()
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data["window_state"], "OPEN")
        self.assertTrue(res.data["can_edit"])

    def test_aun_no_abre(self):
        from django.utils import timezone
        import datetime as dt
        ahora = timezone.now()
        self._periodo(grades_start=ahora + dt.timedelta(days=3),
                      grades_end=ahora + dt.timedelta(days=10))
        res = self._window()
        self.assertEqual(res.data["window_state"], "NOT_YET")
        self.assertFalse(res.data["can_edit"])
        self.assertIn("aún no está habilitado", res.data["message"])

    def test_fecha_vencida_bloquea_y_avisa(self):
        from django.utils import timezone
        from academic.views.teachers import GradesSaveView
        import datetime as dt
        ahora = timezone.now()
        self._periodo(grades_start=ahora - dt.timedelta(days=10),
                      grades_end=ahora - dt.timedelta(days=1))

        res = self._window()
        self.assertEqual(res.data["window_state"], "EXPIRED")
        self.assertFalse(res.data["can_edit"])
        self.assertIn("se cerró", res.data["message"])

        # El docente no puede guardar: 423 con el mismo mensaje
        req = self.factory.post("/x", {
            "section_id": self.sec1.id,
            "grades": {str(self.st1.user_id): {
                "C1_LEVEL": "L", "C2_LEVEL": "L", "C3_LEVEL": "L"}},
        }, format="json")
        force_authenticate(req, user=self.tuser)
        r2 = GradesSaveView.as_view()(req)
        self.assertEqual(r2.status_code, 423)
        self.assertIn("se cerró", r2.data["detail"])
        self.assertEqual(r2.data["window_state"], "EXPIRED")

    def test_periodo_cerrado_bloquea(self):
        self._periodo(grades_state=AcademicPeriod.GRADES_CLOSED)
        res = self._window()
        self.assertEqual(res.data["window_state"], "CLOSED")
        self.assertFalse(res.data["can_edit"])
        self.assertIn("CERRADO", res.data["message"])

    def test_admin_puede_editar_pese_al_cierre(self):
        self._periodo(grades_state=AcademicPeriod.GRADES_CLOSED)
        res = self._window(user=self.admin)
        self.assertEqual(res.data["window_state"], "CLOSED")
        self.assertTrue(res.data["can_edit"])        # admin sí
        self.assertTrue(res.data["admin_override"])


class HorarioDocenteTests(BaseEvalTest):
    def test_periodos_del_docente(self):
        from academic.views.teachers import TeacherSelfPeriodsView
        # Sección extra en un período pasado
        Section.objects.create(plan_course=self.pc1, teacher=self.teacher,
                               period="2025-II", label="A")
        req = self.factory.get("/x")
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfPeriodsView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        codes = [p["code"] for p in res.data["periods"]]
        self.assertEqual(codes, [PERIOD, "2025-II"])       # descendente
        self.assertEqual(res.data["periods"][0]["sections"], 2)

    def test_horario_pdf_periodo_pasado(self):
        from academic.views.evaluation_pdf import TeacherSelfSchedulePdfView
        from academic.models import SectionScheduleSlot
        import datetime as dt
        sec_old = Section.objects.create(plan_course=self.pc1, teacher=self.teacher,
                                         period="2025-II", label="B")
        SectionScheduleSlot.objects.create(section=sec_old, weekday=3,
                                           start=dt.time(15, 0), end=dt.time(17, 0))
        req = self.factory.get("/x", {"period": "2025-II"})
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfSchedulePdfView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        from pypdf import PdfReader
        txt = "".join((p.extract_text() or "") for p in PdfReader(BytesIO(res.content)).pages)
        self.assertIn("2025-II", txt)
        self.assertIn("MIÉRCOLES", txt)
        self.assertIn("15:00", txt)

    def test_horario_periodo_sin_carga_404(self):
        from academic.views.evaluation_pdf import TeacherSelfSchedulePdfView
        req = self.factory.get("/x", {"period": "2019-I"})
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfSchedulePdfView.as_view()(req)
        self.assertEqual(res.status_code, 404)


class DocumentosAlumnoTests(BaseEvalTest):
    def test_boleta_propia_pdf(self):
        from academic.views.evaluation_pdf import StudentSelfBoletaPdfView
        AcademicGradeRecord.objects.create(
            student=self.st1, course=self.c1, plan_course=self.pc1,
            term=PERIOD, final_grade=16)
        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.st1.user)
        res = StudentSelfBoletaPdfView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "application/pdf")
        from pypdf import PdfReader
        txt = "".join((p.extract_text() or "") for p in PdfReader(BytesIO(res.content)).pages)
        self.assertIn("BOLETA DE CALIFICACIONES", txt)
        self.assertIn("ATAPOMA", txt.upper())
        self.assertIn("Logrado", txt)

    def test_boleta_sin_notas_404(self):
        from academic.views.evaluation_pdf import StudentSelfBoletaPdfView
        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.st1.user)
        res = StudentSelfBoletaPdfView.as_view()(req)
        self.assertEqual(res.status_code, 404)

    def test_asistencia_propia_pdf(self):
        from academic.views.evaluation_pdf import StudentSelfAsistenciaPdfView
        from academic.models import AttendanceSession, AttendanceRow
        import datetime as dt
        for i, status in enumerate(["PRESENT", "ABSENT", "ABSENT"], 1):
            s = AttendanceSession.objects.create(section=self.sec1, date=dt.date(2026, 6, i))
            AttendanceRow.objects.create(session=s, student_id=self.st1.user_id, status=status)
        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.st1.user)
        res = StudentSelfAsistenciaPdfView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        from pypdf import PdfReader
        txt = "".join((p.extract_text() or "") for p in PdfReader(BytesIO(res.content)).pages)
        self.assertIn("REPORTE DE ASISTENCIA", txt)
        self.assertIn("RIESGO DPI", txt)   # 2 de 3 faltas = 67%


class ReporteAsistenciaTests(BaseEvalTest):
    def test_reporte_asistencia_xlsx(self):
        from academic.views.evaluation_pdf import EvaluationAsistenciaReporteView
        from academic.models import AttendanceSession, AttendanceRow
        import datetime as dt

        # 4 sesiones; st1 con 2 faltas (50% → riesgo DPI), st2 sin faltas
        for i, (s1, s2) in enumerate(
                [("PRESENT", "PRESENT"), ("ABSENT", "PRESENT"),
                 ("ABSENT", "PRESENT"), ("LATE", "PRESENT")], 1):
            sess = AttendanceSession.objects.create(
                section=self.sec1, date=dt.date(2026, 6, i))
            AttendanceRow.objects.create(session=sess, student_id=self.st1.user_id, status=s1)
            AttendanceRow.objects.create(session=sess, student_id=self.st2.id, status=s2)

        req = self.factory.get("/x", {"period": PERIOD, "section_id": self.sec1.id})
        force_authenticate(req, user=self.admin)
        res = EvaluationAsistenciaReporteView.as_view()(req, fmt="xlsx")
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        self.assertEqual(wb.sheetnames, ["Resumen por curso", "Detalle por alumno"])
        ws = wb["Resumen por curso"]
        self.assertEqual(ws.cell(row=4, column=7).value, 4)   # sesiones
        self.assertEqual(ws.cell(row=4, column=8).value, 1)   # en riesgo DPI
        ws2 = wb["Detalle por alumno"]
        fila = {ws2.cell(row=r, column=5).value: r for r in (2, 3)}
        r1 = fila["60634719"]                                  # ATAPOMA (st1)
        self.assertEqual(ws2.cell(row=r1, column=8).value, 2)   # faltas
        self.assertEqual(ws2.cell(row=r1, column=9).value, 1)   # tardanzas
        self.assertEqual(ws2.cell(row=r1, column=10).value, 50) # % inasistencia
        self.assertEqual(ws2.cell(row=r1, column=11).value, "RIESGO DPI")


class AutoedicionAlumnoTests(BaseEvalTest):
    """El alumno solo edita contacto; su ciclo se actualiza al matricularse."""

    def test_patch_solo_acepta_contacto(self):
        from students.serializers import StudentMeUpdateSerializer
        Student.objects.filter(id=self.st1.id).update(ciclo=2)
        st = Student.objects.get(id=self.st1.id)
        ser = StudentMeUpdateSerializer(st, data={
            "email": "nuevo@correo.pe", "celular": "999888777",
            # intentos de cambiar datos bloqueados:
            "nombres": "HACKER", "numDocumento": "11111111", "ciclo": 10,
            "apellidoPaterno": "OTRO", "periodo": "2030-I",
        }, partial=True)
        self.assertTrue(ser.is_valid(), ser.errors)
        st = ser.save()
        st.refresh_from_db()
        self.assertEqual(st.email, "nuevo@correo.pe")
        self.assertEqual(st.celular, "999888777")
        # Todo lo demás intacto
        self.assertEqual(st.nombres, "SHEYLA")
        self.assertEqual(st.num_documento, "60634719")
        self.assertEqual(st.apellido_paterno, "ATAPOMA")
        self.assertEqual(st.ciclo, 2)

    def test_ciclo_sube_al_confirmar_matricula(self):
        from academic.views.enrollment import EnrollmentCommitView
        from academic.models import PlanCourse

        # Cursos de ciclo 3 para la matrícula nueva
        c3 = Course.objects.create(code="C3", name="Curso Ciclo 3")
        pc3 = PlanCourse.objects.create(plan=self.plan, course=c3, semester=3, credits=4)
        Student.objects.filter(id=self.st1.id).update(ciclo=1, plan=self.plan)

        # limpiar matrícula previa del período
        Enrollment.objects.filter(student=self.st1, period=PERIOD).delete()

        req = self.factory.post("/x", {
            "student_id": self.st1.id,
            "academic_period": PERIOD,
            "plan_course_ids": [self.pc1.id, pc3.id],
        }, format="json")
        force_authenticate(req, user=self.admin)
        res = EnrollmentCommitView.as_view()(req)
        self.assertEqual(res.status_code, 200, getattr(res, "data", None))

        self.st1.refresh_from_db()
        self.assertEqual(self.st1.ciclo, 3)        # semestre más alto matriculado
        self.assertEqual(self.st1.periodo, PERIOD)


class DetalleCursoAlumnoTests(BaseEvalTest):
    def test_detalle_completo(self):
        from academic.views.dashboard_student import student_course_detail
        from academic.models import AttendanceSession, AttendanceRow
        import datetime as dt

        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {str(self.st1.user_id): {
                "C1": 4, "C1_LEVEL": "L", "C1_REC": "Buen trabajo",
                "C2": 3, "C2_LEVEL": "P", "C3": 5, "C3_LEVEL": "D",
                "PROMEDIO_FINAL": 15, "ESTADO": "Logrado",
            }}})
        s = AttendanceSession.objects.create(section=self.sec1, date=dt.date(2026, 6, 2))
        AttendanceRow.objects.create(session=s, student_id=self.st1.user_id, status="ABSENT")
        AcademicGradeRecord.objects.create(
            student=self.st1, course=self.c1, plan_course=self.pc1,
            term="2025-II", final_grade=13)

        req = self.factory.get("/x")
        force_authenticate(req, user=self.st1.user)
        res = student_course_detail(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data["acta"]["c1"], 4)
        self.assertEqual(res.data["acta"]["c1_rec"], "Buen trabajo")
        self.assertEqual(res.data["acta"]["promedio"], 15)
        self.assertEqual(res.data["sesiones"][0]["mark"], "F")
        self.assertEqual(res.data["historial"][0]["term"], "2025-II")
        self.assertEqual(res.data["historial"][0]["final"], 13)

    def test_no_matriculado_403(self):
        from academic.views.dashboard_student import student_course_detail
        # st2 no tiene cuenta de usuario; creamos un alumno ajeno con cuenta
        u = User.objects.create_user("ajeno_st", "ajeno@t.pe", "x")
        Student.objects.create(user=u, num_documento="99999999", nombres="AJENO")
        req = self.factory.get("/x")
        force_authenticate(req, user=u)
        res = student_course_detail(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 403)


class PerfilDocenteTests(BaseEvalTest):
    def test_get_y_actualizar_perfil(self):
        from academic.views.teachers import TeacherSelfProfileView
        from catalogs.models import Teacher as CatalogTeacher

        req = self.factory.get("/x")
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfProfileView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data["full_name"], "GOYAS BALDOCEDA, ANA MARIA")

        req2 = self.factory.put("/x", {
            "fecha_nac": "1980-05-12",
            "grado_academico": "MAGISTER",
            "celular": "987654321",
            "email_institucional": "agoyas@iesppallende.edu.pe",
        }, format="json")
        force_authenticate(req2, user=self.tuser)
        res2 = TeacherSelfProfileView.as_view()(req2)
        self.assertEqual(res2.status_code, 200, res2.data)
        ct = CatalogTeacher.objects.get(user=self.tuser)
        self.assertEqual(str(ct.fecha_nac), "1980-05-12")
        self.assertEqual(ct.grado_academico, "MAGISTER")
        self.assertEqual(ct.phone, "987654321")
        self.assertEqual(ct.email, "agoyas@iesppallende.edu.pe")
        self.assertEqual(res2.data["grado_academico_label"], "Magister (a)")

    def test_rd_nombramiento(self):
        from academic.views.teachers import TeacherSelfProfileView
        from catalogs.models import Teacher as CatalogTeacher

        req = self.factory.put("/x", {
            "condicion_laboral": "NOMBRADO",
            "rd_nombramiento": "R.D. N° 0123-2026-DREJ",
            "rd_fecha": "2026-03-15",
        }, format="json")
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfProfileView.as_view()(req)
        self.assertEqual(res.status_code, 200, res.data)
        ct = CatalogTeacher.objects.get(user=self.tuser)
        self.assertEqual(ct.condicion_laboral, "NOMBRADO")
        self.assertEqual(ct.rd_nombramiento, "R.D. N° 0123-2026-DREJ")
        self.assertEqual(str(ct.rd_fecha), "2026-03-15")
        self.assertEqual(res.data["condicion_laboral_label"], "Nombrado (a)")

        # Condición inválida → 400
        req2 = self.factory.put("/x", {"condicion_laboral": "PRACTICANTE"}, format="json")
        force_authenticate(req2, user=self.tuser)
        self.assertEqual(TeacherSelfProfileView.as_view()(req2).status_code, 400)

    def test_rd_sale_en_horario_pdf(self):
        from academic.views.teachers import TeacherSelfProfileView
        from academic.views.evaluation_pdf import TeacherSelfSchedulePdfView
        from academic.models import SectionScheduleSlot
        import datetime as dt

        req = self.factory.put("/x", {
            "condicion_laboral": "CONTRATADO",
            "rd_nombramiento": "R.D. N° 0456-2026-DREJ",
            "rd_fecha": "2026-04-01",
        }, format="json")
        force_authenticate(req, user=self.tuser)
        TeacherSelfProfileView.as_view()(req)

        SectionScheduleSlot.objects.create(section=self.sec1, weekday=2,
                                           start=dt.time(10, 0), end=dt.time(12, 0))
        req2 = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req2, user=self.tuser)
        res = TeacherSelfSchedulePdfView.as_view()(req2)
        self.assertEqual(res.status_code, 200)
        from pypdf import PdfReader
        txt = "".join((p.extract_text() or "") for p in PdfReader(BytesIO(res.content)).pages)
        self.assertIn("0456-2026", txt)
        self.assertIn("Contratado", txt)
        self.assertIn("01/04/2026", txt)

    def test_grado_invalido_400(self):
        from academic.views.teachers import TeacherSelfProfileView
        req = self.factory.put("/x", {"grado_academico": "INGENIERO"}, format="json")
        force_authenticate(req, user=self.tuser)
        res = TeacherSelfProfileView.as_view()(req)
        self.assertEqual(res.status_code, 400)


class AsistenciaMensualTests(BaseEvalTest):
    def test_grabar_y_leer_mes(self):
        from academic.views.attendance import AttendanceMonthView
        from academic.models import AttendanceSession, AttendanceRow

        # Grabar: st1 presente el 3 y tardanza el 4; st2 falta el 3
        req = self.factory.post("/x", {
            "month": "2026-06",
            "marks": {
                str(self.st1.user_id): {"3": "P", "4": "T"},
                str(self.st2.id): {"3": "F"},
            },
        }, format="json")
        force_authenticate(req, user=self.tuser)   # docente de la sección
        res = AttendanceMonthView.as_view()(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data["saved_days"], 2)
        self.assertEqual(AttendanceSession.objects.filter(section=self.sec1).count(), 2)
        self.assertEqual(AttendanceRow.objects.count(), 3)

        # Leer el mes: marcas de vuelta con las claves del roster
        req2 = self.factory.get("/x", {"month": "2026-06"})
        force_authenticate(req2, user=self.tuser)
        res2 = AttendanceMonthView.as_view()(req2, section_id=self.sec1.id)
        self.assertEqual(res2.status_code, 200)
        marcas = res2.data["marks"]
        self.assertEqual(marcas[str(self.st1.user_id)]["3"], "P")
        self.assertEqual(marcas[str(self.st1.user_id)]["4"], "T")
        self.assertEqual(marcas[str(self.st2.id)]["3"], "F")
        # junio 2026: día 6 es sábado
        dia6 = next(d for d in res2.data["days"] if d["day"] == 6)
        self.assertTrue(dia6["weekend"])

        # Día cerrado no se modifica
        AttendanceSession.objects.filter(
            section=self.sec1, date="2026-06-03").update(closed=True)
        req3 = self.factory.post("/x", {
            "month": "2026-06",
            "marks": {str(self.st1.user_id): {"3": "F"}},
        }, format="json")
        force_authenticate(req3, user=self.tuser)
        res3 = AttendanceMonthView.as_view()(req3, section_id=self.sec1.id)
        self.assertEqual(res3.data["closed_days"], [3])
        row = AttendanceRow.objects.get(
            session__date="2026-06-03", student_id=self.st1.user_id)
        self.assertEqual(row.status, "PRESENT")   # sigue P, no cambió

        # Vaciar el día 4 (days explícito) → sesión eliminada
        req4 = self.factory.post("/x", {
            "month": "2026-06", "marks": {}, "days": [4],
        }, format="json")
        force_authenticate(req4, user=self.tuser)
        res4 = AttendanceMonthView.as_view()(req4, section_id=self.sec1.id)
        self.assertEqual(res4.status_code, 200)
        self.assertFalse(AttendanceSession.objects.filter(
            section=self.sec1, date="2026-06-04").exists())

    def test_docente_ajeno_403(self):
        from academic.views.attendance import AttendanceMonthView
        otro = User.objects.create_user("otro_doc2", "otro2@t.pe", "x")
        req = self.factory.get("/x", {"month": "2026-06"})
        force_authenticate(req, user=otro)
        res = AttendanceMonthView.as_view()(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 403)


class MeritosTests(BaseEvalTest):
    def _kardex(self):
        AcademicGradeRecord.objects.create(
            student=self.st1, course=self.c1, plan_course=self.pc1,
            term=PERIOD, final_grade=18)
        AcademicGradeRecord.objects.create(
            student=self.st2, course=self.c1, plan_course=self.pc1,
            term=PERIOD, final_grade=12)

    def test_primeros_lugares_xlsx(self):
        from academic.views.evaluation_pdf import EvaluationPrimerosLugaresView
        self._kardex()
        res = self._get(EvaluationPrimerosLugaresView, "/x",
                        {"period": PERIOD, "fmt": "xlsx"})
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        # st1 primero (18), st2 segundo (12)
        self.assertEqual(ws.cell(row=4, column=3).value, 1)      # puesto
        self.assertEqual(ws.cell(row=4, column=4).value, "60634719")
        self.assertEqual(ws.cell(row=4, column=6).value, 18.0)

    def test_tercio_superior_xlsx(self):
        from academic.views.evaluation_pdf import EvaluationTercioQuintoView
        self._kardex()
        Student.objects.filter(id__in=[self.st1.id, self.st2.id]).update(ciclo=10)
        res = self._get(EvaluationTercioQuintoView, "/x",
                        {"tipo": "tercio", "fmt": "xlsx"})
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        # 2 en promoción → tercio = max(1, 2//3) = 1 → solo st1 (18)
        self.assertEqual(ws.cell(row=5, column=2).value, "60634719")
        self.assertIsNone(ws.cell(row=6, column=2).value)

    def test_boletas_xlsx_detalle(self):
        from academic.views.evaluation_pdf import EvaluationBoletasXlsxView
        self._kardex()
        res = self._get(EvaluationBoletasXlsxView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        # fila 4: primer alumno A-Z (ATAPOMA=st1), nota 18, cualitativa Logrado
        self.assertEqual(ws.cell(row=4, column=1).value, "60634719")
        self.assertEqual(ws.cell(row=4, column=7).value, 18)
        self.assertEqual(ws.cell(row=4, column=8).value, "Logrado")


class ProcesamientoTests(BaseEvalTest):
    def _cargar_notas(self):
        """Acta de sec1 con las DOS claves posibles (user_id y pk)."""
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {
                str(self.st1.user_id): {"C1_LEVEL": "L", "C2_LEVEL": "L", "C3_LEVEL": "L",
                                        "PROMEDIO_FINAL": 15, "ESTADO": "Logrado"},
                str(self.st2.id): {"C1_LEVEL": "I", "C2_LEVEL": "I", "C3_LEVEL": "I",
                                   "PROMEDIO_FINAL": 5, "ESTADO": "En proceso"},
            }},
        )

    def test_section_eval_row_cuenta_ambas_claves(self):
        self._cargar_notas()
        row = _section_eval_row(self.sec1)
        self.assertEqual(row["n_students"], 2)
        self.assertEqual(row["n_loaded"], 2)
        self.assertEqual(row["_finals"][self.st1.id], 15)
        self.assertEqual(row["_finals"][self.st2.id], 5)

    def test_overview_admin_cuenta_promedio_final(self):
        self._cargar_notas()
        res = self._get(AdminGradesOverviewView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        sec_row = next(s for s in res.data["sections"] if s["section_id"] == self.sec1.id)
        self.assertEqual(sec_row["n_loaded"], 2)
        self.assertEqual(sec_row["n_failed"], 1)   # st2 con 5 < 11

    def test_procesar_al_kardex(self):
        self._cargar_notas()
        res = self._post(EvaluationProcessView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data["processed"], 1)   # sec1
        self.assertEqual(res.data["skipped"], 1)     # sec2 sin notas

        r1 = AcademicGradeRecord.objects.get(student=self.st1, course=self.c1, term=PERIOD)
        self.assertEqual(float(r1.final_grade), 15.0)
        r2 = AcademicGradeRecord.objects.get(student=self.st2, course=self.c1, term=PERIOD)
        self.assertEqual(float(r2.final_grade), 5.0)
        self.assertTrue(SectionGrades.objects.get(section=self.sec1).submitted)

        # re-procesar actualiza sin duplicar
        res2 = self._post(EvaluationProcessView, "/x", {"period": PERIOD})
        self.assertEqual(res2.status_code, 200)
        self.assertEqual(AcademicGradeRecord.objects.filter(
            student=self.st1, course=self.c1, term=PERIOD).count(), 1)

    def test_sections_endpoint(self):
        self._cargar_notas()
        res = self._get(EvaluationSectionsView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data["sections"]), 2)


class HorarioExportTests(BaseEvalTest):
    def test_horario_resuelve_seccion_por_fallback(self):
        """EnrollmentItems sin sección: el horario debe resolver la sección
        del curso por plan_course + período (docente/aula/horario)."""
        from academic.views.enrollment import ScheduleExportPDFView
        from academic.models import SectionScheduleSlot
        import datetime as dt

        # Horario en la sección del curso 1
        SectionScheduleSlot.objects.create(
            section=self.sec1, weekday=1,
            start=dt.time(8, 0), end=dt.time(10, 0))
        # Matrícula de st1 SIN sección en los items (caso real del bug)
        EnrollmentItem.objects.filter(
            enrollment__student=self.st1).update(section=None)

        req = self.factory.get("/x", {"academic_period": PERIOD,
                                      "student_id": self.st1.id})
        force_authenticate(req, user=self.admin)
        res = ScheduleExportPDFView.as_view()(req)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "application/pdf")
        # El PDF debe contener el horario resuelto (día y hora), no "Sin horario"
        from pypdf import PdfReader
        texto = "".join((p.extract_text() or "") for p in PdfReader(BytesIO(res.content)).pages)
        self.assertIn("Lunes", texto)
        self.assertIn("08:00", texto)
        self.assertIn("GOYAS", texto.upper())   # docente resuelto por fallback

    def test_padron_filtro_ciclo_y_anio(self):
        from academic.views.enrollment import StudentsOverviewView
        Student.objects.filter(id=self.st1.id).update(ciclo=3)
        Student.objects.filter(id=self.st2.id).update(ciclo=1)
        # rol STUDENT requerido por la vista
        try:
            from acl.models import Role, UserRole
            rol, _ = Role.objects.get_or_create(name="STUDENT")
            for st in (self.st1, self.st2):
                if st.user_id:
                    UserRole.objects.get_or_create(user_id=st.user_id, role=rol)
        except Exception:
            pass

        res = self._get(StudentsOverviewView, "/x",
                        {"academic_period": PERIOD, "ciclo": 3})
        self.assertEqual(res.status_code, 200)
        dnis = [s["dni"] for s in res.data["students"] if s.get("dni")]
        self.assertIn("60634719", dnis)      # st1 (ciclo 3)
        self.assertNotIn("61576678", dnis)   # st2 (ciclo 1)

        # Año 2 = ciclos 3-4 → incluye a st1, no a st2
        res2 = self._get(StudentsOverviewView, "/x",
                         {"academic_period": PERIOD, "anio": 2})
        self.assertEqual(res2.status_code, 200)
        dnis2 = [s["dni"] for s in res2.data["students"] if s.get("dni")]
        self.assertIn("60634719", dnis2)
        self.assertNotIn("61576678", dnis2)


class EstadoAcademicoTests(BaseEvalTest):
    """Licencia/Reincorporación/Traslado/Subsanación con RD y bloqueo en actas."""

    def _licenciar(self, st):
        Student.objects.filter(id=st.id).update(
            estado_academico="LICENCIA", estado_rd="R.D. N° 045-2026-DG")

    def test_guardar_acta_bloquea_licencia(self):
        from academic.views.teachers import GradesSaveView
        self._licenciar(self.st1)
        payload = {
            str(self.st1.user_id): {"C1_LEVEL": "L", "C2_LEVEL": "L", "C3_LEVEL": "L"},
            str(self.st2.id): {"C1_LEVEL": "P", "C2_LEVEL": "P", "C3_LEVEL": "P"},
        }
        res = self._post(GradesSaveView, "/x",
                         {"section_id": self.sec1.id, "grades": payload})
        self.assertEqual(res.status_code, 200, res.data)
        self.assertIn("LICENCIA", res.data.get("message", ""))
        grades = SectionGrades.objects.get(section=self.sec1).grades
        self.assertNotIn(str(self.st1.user_id), grades)   # licencia fuera
        self.assertIn(str(self.st2.id), grades)           # el resto sí

    def test_import_excel_bloquea_licencia(self):
        self._licenciar(self.st1)
        wb = Workbook()
        ws = wb.active
        AE._write_grades_sheet(ws, self.sec1)
        # La plantilla debe traer LICENCIA en el comentario del alumno (fila 16)
        self.assertIn("LICENCIA", str(ws.cell(row=16, column=9).value))
        # Intentar marcarle niveles → error al importar
        ws.cell(row=16, column=7, value="X")
        res = AE._import_grades_ws(ws, SimpleNamespace(user=self.admin))
        self.assertTrue(any("LICENCIA" in e for e in res["errores"]))
        self.assertEqual(res["importados"], 0)
        # Re-importar la plantilla intacta NO genera errores por la licencia
        wb2 = Workbook()
        ws2 = wb2.active
        AE._write_grades_sheet(ws2, self.sec1)
        res2 = AE._import_grades_ws(ws2, SimpleNamespace(user=self.admin))
        self.assertEqual(res2["errores"], [])

    def test_roster_expone_estado(self):
        from academic.views.teachers import SectionStudentsView
        self._licenciar(self.st1)
        req = self.factory.get("/x")
        force_authenticate(req, user=self.admin)
        res = SectionStudentsView.as_view()(req, section_id=self.sec1.id)
        self.assertEqual(res.status_code, 200)
        por_dni = {s["num_documento"]: s for s in res.data["students"]}
        self.assertEqual(por_dni["60634719"]["estado_academico"], "LICENCIA")
        self.assertEqual(por_dni["60634719"]["estado_rd"], "R.D. N° 045-2026-DG")
        self.assertEqual(por_dni["61576678"]["estado_academico"], "")

    def test_patch_estado_valida_rd(self):
        from students.serializers import StudentUpdateSerializer
        # Estado especial sin RD → inválido
        ser = StudentUpdateSerializer(self.st1, data={"estadoAcademico": "TRASLADO"}, partial=True)
        self.assertFalse(ser.is_valid())
        self.assertIn("estadoRd", ser.errors)
        # Con RD → válido
        ser2 = StudentUpdateSerializer(
            self.st1, data={"estadoAcademico": "TRASLADO", "estadoRd": "RD 99-2026"}, partial=True)
        self.assertTrue(ser2.is_valid(), ser2.errors)
        st = ser2.save()
        self.assertEqual(st.estado_academico, "TRASLADO")
        # Volver a normal limpia la RD
        ser3 = StudentUpdateSerializer(st, data={"estadoAcademico": ""}, partial=True)
        self.assertTrue(ser3.is_valid(), ser3.errors)
        st = ser3.save()
        self.assertEqual(st.estado_academico, "")
        self.assertEqual(st.estado_rd, "")

    def test_padron_incluye_estado(self):
        from academic.views.enrollment import StudentsOverviewView
        self._licenciar(self.st1)
        try:
            from acl.models import Role, UserRole
            rol, _ = Role.objects.get_or_create(name="STUDENT")
            for st in (self.st1, self.st2):
                if st.user_id:
                    UserRole.objects.get_or_create(user_id=st.user_id, role=rol)
        except Exception:
            pass
        res = self._get(StudentsOverviewView, "/x", {"academic_period": PERIOD})
        self.assertEqual(res.status_code, 200)
        fila = next(s for s in res.data["students"] if s["dni"] == "60634719")
        self.assertEqual(fila["estado_academico"], "LICENCIA")
        self.assertEqual(fila["estado_rd"], "R.D. N° 045-2026-DG")


class TeacherDashboardTests(BaseEvalTest):
    def test_metricas_reales(self):
        from academic.views.dashboard_teacher import teacher_dashboard
        from catalogs.models import Period as CatalogPeriod
        CatalogPeriod.objects.create(code=PERIOD, year=2026, term="I",
                                     label=PERIOD, is_active=True)
        # Quitar la sección de los items (caso real): debe contar por fallback
        EnrollmentItem.objects.all().update(section=None)
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {str(self.st1.user_id): {"PROMEDIO_FINAL": 15}}})

        req = self.factory.get("/x")
        force_authenticate(req, user=self.tuser)
        res = teacher_dashboard(req)
        self.assertEqual(res.status_code, 200)
        d = res.data
        self.assertEqual(d["total_sections"], 2)
        self.assertEqual(d["total_students"], 4)      # 2 alumnos × 2 secciones (fallback)
        self.assertEqual(d["grades_pending"], 1)      # sec2 sin notas
        self.assertEqual(d["acts_pending"], 2)        # ninguna acta cerrada
        self.assertEqual(len(d["sections"]), 2)
        sec1_row = next(s for s in d["sections"] if s["id"] == self.sec1.id)
        self.assertEqual(sec1_row["students"], 2)
        self.assertEqual(sec1_row["grades_loaded"], 1)
        self.assertEqual(sec1_row["grades_pct"], 50)
        self.assertFalse(sec1_row["submitted"])


class EstadoPeriodoTests(BaseEvalTest):
    def test_abrir_cerrar_periodo(self):
        # cerrar crea el AcademicPeriod si no existe
        res = self._post(EvaluationStateView, "/x", {"period": PERIOD, "action": "close"})
        self.assertEqual(res.status_code, 200, res.data)
        per = AcademicPeriod.objects.get(code=PERIOD)
        self.assertEqual(per.grades_state, AcademicPeriod.GRADES_CLOSED)
        self.assertIsNotNone(per.grades_closed_at)

        # con el período cerrado, el docente queda bloqueado y el admin no
        ok_doc, err = _check_grades_window(self.sec1, self.tuser)
        self.assertFalse(ok_doc)
        self.assertIn("CERRADO", err["detail"])
        ok_admin, _ = _check_grades_window(self.sec1, self.admin)
        self.assertTrue(ok_admin)

        res = self._post(EvaluationStateView, "/x", {"period": PERIOD, "action": "open"})
        self.assertEqual(res.status_code, 200)
        per.refresh_from_db()
        self.assertEqual(per.grades_state, AcademicPeriod.GRADES_OPEN)
        ok_doc, _ = _check_grades_window(self.sec1, self.tuser)
        self.assertTrue(ok_doc)

    def test_estado_get(self):
        res = self._get(EvaluationStateView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data["stats"]["sections"], 2)
        self.assertEqual(res.data["stats"]["students"], 4)   # 2 alumnos × 2 secciones

    def test_no_admin_recibe_403(self):
        req = self.factory.get("/x", {"period": PERIOD})
        force_authenticate(req, user=self.tuser)   # docente, no admin
        res = EvaluationStateView.as_view()(req)
        self.assertEqual(res.status_code, 403)


class ConsolidadaTests(BaseEvalTest):
    def test_acta_consolidada_formato_oficial(self):
        SectionGrades.objects.update_or_create(
            section=self.sec1,
            defaults={"grades": {
                str(self.st1.user_id): {"PROMEDIO_FINAL": 15},
                str(self.st2.id): {"PROMEDIO_FINAL": 10},
            }},
        )
        SectionGrades.objects.update_or_create(
            section=self.sec2,
            defaults={"grades": {str(self.st1.user_id): {"PROMEDIO_FINAL": 18}}},
        )
        res = self._get(EvaluationActaConsolidadaView, "/x", {"period": PERIOD})
        self.assertEqual(res.status_code, 200)
        _dump("muestra_acta_consolidada.xlsx", res.content)

        wb = load_workbook(BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 1)     # un grupo: EDUCACIÓN INICIAL I A
        ws = wb[wb.sheetnames[0]]
        self.assertIn("ACTA CONSOLIDADA", str(ws.cell(row=2, column=1).value))
        # tabla: alumnos desde fila 17 (H=11 + 6)
        r0 = 17
        self.assertEqual(ws.cell(row=r0, column=2).value, "60634719")       # ATAPOMA
        # curso 1 (cols D,E,F): C=L, CS=15, PTJ=45
        self.assertEqual(ws.cell(row=r0, column=4).value, "L")
        self.assertEqual(ws.cell(row=r0, column=5).value, 15)
        self.assertEqual(ws.cell(row=r0, column=6).value, 45)
        # curso 2 (cols G,H,I): CS=18, créditos 5 → PTJ=90
        self.assertEqual(ws.cell(row=r0, column=8).value, 18)
        self.assertEqual(ws.cell(row=r0, column=9).value, 90)
        # totales (tot_c=10): puntaje 135, créditos 8, promedio 16.88, calificación L
        self.assertEqual(ws.cell(row=r0, column=10).value, 135)
        self.assertEqual(ws.cell(row=r0, column=11).value, 8)
        self.assertAlmostEqual(ws.cell(row=r0, column=12).value, 16.88, places=2)
        self.assertEqual(ws.cell(row=r0, column=13).value, "L")
        # st2: solo curso 1 con nota 10 → EP... no: 10 → I (6-10). Verificar
        r1 = r0 + 1
        self.assertEqual(ws.cell(row=r1, column=4).value, "I")
        self.assertEqual(ws.cell(row=r1, column=11).value, 3)
        # firmas: docente del curso 1
        self.assertTrue(any(
            "GOYAS" in str(ws.cell(row=r, column=3).value or "")
            for r in range(r0 + 2, r0 + 12)
        ))
