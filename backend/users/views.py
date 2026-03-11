# backend/users/views.py
from django.contrib.auth import get_user_model
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.utils.crypto import get_random_string

from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from acl.models import UserRole, Role
from .serializers import UserSerializer, UserCreateSerializer, UserUpdateSerializer
from django.db import transaction
from django.http import HttpResponse

from django.db.models.deletion import ProtectedError
from django.db import IntegrityError

import io

from academic.models import (
    AcademicGradeRecord,
    AcademicProcess,
    ProcessFile,
    AttendanceRow,
    SectionGrades,
)

from students.models import Student

User = get_user_model()


def _qs_users_by_role_name(role_name: str):
    role_name = (role_name or "").strip()
    if not role_name:
        return User.objects.none()

    role_ids = list(Role.objects.filter(name__iexact=role_name).values_list("id", flat=True))
    if not role_ids:
        return User.objects.none()

    user_ids = (
        UserRole.objects
        .filter(role_id__in=role_ids)
        .values_list("user_id", flat=True)
        .distinct()
    )
    return User.objects.filter(id__in=user_ids)


# ---------- permisos ----------
class IsAdminOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in ("GET", "HEAD", "OPTIONS"):
            return request.user and request.user.is_authenticated
        return request.user and request.user.is_authenticated and request.user.is_staff


def _has_role(user, role_name: str) -> bool:
    if not user or not user.is_authenticated:
        return False

    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True

    try:
        if hasattr(user, "roles") and user.roles.filter(name__iexact=role_name).exists():
            return True
    except Exception:
        pass

    try:
        return UserRole.objects.filter(user=user, role__name__iexact=role_name).exists()
    except Exception:
        return False


def _require_staff(request):
    # ✅ STAFF o ADMIN_ACADEMIC
    if _has_role(request.user, "ADMIN_ACADEMIC"):
        return None
    return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)


def _split_full_name(full_name: str, fallback: str = ""):
    full = (full_name or "").strip()
    if not full:
        full = (fallback or "").strip()

    if not full:
        return "", ""

    parts = [p for p in full.split() if p]
    if len(parts) == 1:
        return parts[0], ""
    if len(parts) >= 3:
        nombres = " ".join(parts[:-2])
        apellidos = " ".join(parts[-2:])
        return nombres, apellidos
    return parts[0], parts[1]


def _ensure_student_for_user(user: User):
    if hasattr(user, "student_profile"):
        return None

    username = getattr(user, "username", "") or ""
    email = getattr(user, "email", "") or ""
    full_name = getattr(user, "full_name", "") or username

    nombres, apellidos = _split_full_name(full_name, fallback=username)
    temp_doc = username[:12] if username else "TMP-" + get_random_string(9).upper()

    ap_parts = apellidos.split() if apellidos else []
    ap_pat = ap_parts[0] if len(ap_parts) >= 1 else ""
    ap_mat = " ".join(ap_parts[1:]) if len(ap_parts) >= 2 else ""

    st = Student.objects.create(
        user=user,
        num_documento=temp_doc,
        nombres=nombres or username,
        apellido_paterno=ap_pat,
        apellido_materno=ap_mat,
        email=email,
    )
    return st


# ===================== PAGINACIÓN SIMPLE =====================
def _int_param(request, key, default):
    raw = request.query_params.get(key, None)
    if raw is None or raw == "":
        return default
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default


def _paginate_queryset(request, qs, default_page_size=10, max_page_size=100):
    page = _int_param(request, "page", 1)
    page_size = _int_param(request, "page_size", default_page_size)

    # saneo
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = default_page_size
    if page_size > max_page_size:
        page_size = max_page_size

    total = qs.count()
    start = (page - 1) * page_size
    end = start + page_size

    items = qs[start:end]

    # next/previous como números (simple para front)
    next_page = page + 1 if end < total else None
    prev_page = page - 1 if page > 1 else None

    return {
        "count": total,
        "page": page,
        "page_size": page_size,
        "next": next_page,
        "previous": prev_page,
        "items": items,
    }


# ---------- AUTH ----------
@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def change_password(request):
    """
    Cambia la contraseña del usuario logueado (self-service).
    POST /api/auth/change-password
    body: { current_password, new_password }
    """
    u = request.user
    current_password = (request.data.get("current_password") or "").strip()
    new_password = (request.data.get("new_password") or "").strip()

    if not current_password or not new_password:
        return Response({"detail": "Faltan campos."}, status=status.HTTP_400_BAD_REQUEST)

    if not u.check_password(current_password):
        return Response({"detail": "Contraseña actual incorrecta."}, status=status.HTTP_400_BAD_REQUEST)

    # política Django (si tienes AUTH_PASSWORD_VALIDATORS)
    try:
        validate_password(new_password, user=u)
    except ValidationError as e:
        return Response({"detail": " ".join(e.messages)}, status=status.HTTP_400_BAD_REQUEST)

    u.set_password(new_password)
    if hasattr(u, "must_change_password"):
        u.must_change_password = False
        u.save(update_fields=["password", "must_change_password"])
    else:
        u.save(update_fields=["password"])

    return Response({"status": "password_changed"})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def auth_me(request):
    u = request.user

    # roles seguros
    try:
        roles = list(u.roles.values_list("name", flat=True))
    except Exception:
        roles = []

    # permissions seguros (NO rompe auth/me)
    perm_codes = []
    try:
        perm_codes = list(u.roles.values_list("permissions__code", flat=True).distinct())
    except Exception:
        perm_codes = []

    # student_profile seguro
    student_id = None
    try:
        student_id = u.student_profile.id
    except Exception:
        student_id = None

    return Response({
        "id": u.id,
        "username": u.username,
        "email": getattr(u, "email", ""),
        "full_name": getattr(u, "full_name", ""),
        "is_active": u.is_active,
        "is_staff": u.is_staff,
        "is_superuser": u.is_superuser,
        "roles": roles,
        "permissions": perm_codes,
        "student_id": student_id,
        "must_change_password": getattr(u, "must_change_password", False),
    })


# ✅ ---------- USERS (COLLECTION) ----------
@api_view(["GET", "POST"])
@permission_classes([permissions.IsAuthenticated])
def users_collection(request):
    """
    GET  /api/users?q=...&page=1&page_size=10
    POST /api/users      (solo staff)
    """
    # ----- GET (PAGINADO) -----
    if request.method == "GET":
        q = (request.query_params.get("q") or "").strip()

        qs = User.objects.all().order_by("id")
        if q:
            qs = qs.filter(
                Q(username__icontains=q) |
                Q(email__icontains=q) |
                Q(full_name__icontains=q)
            )

        pag = _paginate_queryset(request, qs, default_page_size=10, max_page_size=100)
        ser = UserSerializer(pag["items"], many=True)

        return Response({
            "count": pag["count"],
            "page": pag["page"],
            "page_size": pag["page_size"],
            "next": pag["next"],
            "previous": pag["previous"],
            "results": ser.data,
        })

    # ----- POST -----
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    roles = request.data.get("roles", [])
    ser = UserCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    user = ser.save(is_active=True)

    # roles opcionales
    if isinstance(roles, list) and hasattr(user, "roles"):
        role_objs = []
        for name in roles:
            r, _ = Role.objects.get_or_create(name=str(name).strip())
            role_objs.append(r)
        user.roles.set(role_objs)

        role_names = [r.name.upper() for r in role_objs]
        if "STUDENT" in role_names:
            _ensure_student_for_user(user)

    return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)


# ✅ ---------- USERS (DETAIL) ----------
@api_view(["PUT", "PATCH", "DELETE"])
@permission_classes([permissions.IsAuthenticated])
def users_detail(request, pk: int):
    """
    PUT/PATCH/DELETE /api/users/<id>   (solo staff)
    DELETE = eliminación REAL del usuario (si DB lo permite). NO purge.
    """
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    # ✅ DELETE (eliminación real, con manejo de errores para no dar 500)
    if request.method == "DELETE":
        if request.user.id == pk:
            return Response({"detail": "No puedes eliminarte a ti mismo."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Desvincular Teacher records antes de eliminar (evita ProtectedError)
                try:
                    from catalogs.models import Teacher as CatalogTeacher
                    CatalogTeacher.objects.filter(user_id=pk).update(user=None)
                except Exception:
                    pass
                try:
                    from academic.models import Teacher as AcademicTeacher
                    AcademicTeacher.objects.filter(user_id=pk).update(user=None)
                except Exception:
                    pass

                user.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        except ProtectedError as e:
            sample = [str(x) for x in list(e.protected_objects)[:10]]
            return Response({
                "detail": "No se puede eliminar este usuario porque tiene registros relacionados protegidos. "
                          "Si es un estudiante y quieres borrar también su data académica, usa el botón 'Purge estudiantes'.",
                "protected_sample": sample,
            }, status=status.HTTP_409_CONFLICT)

        except IntegrityError:
            return Response({
                "detail": "No se puede eliminar este usuario por integridad referencial en la base de datos. "
                          "Usa el botón 'Purge estudiantes' si corresponde."
            }, status=status.HTTP_409_CONFLICT)

        except Exception as ex:
            return Response({
                "detail": "Error inesperado al eliminar el usuario.",
                "error": str(ex)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # PUT/PATCH
    partial = (request.method == "PATCH")
    ser = UserUpdateSerializer(user, data=request.data, partial=partial)
    ser.is_valid(raise_exception=True)
    user = ser.save()
    return Response(UserSerializer(user).data)


# ✅ ---------- BÚSQUEDA (ALIAS PAGINADO) ----------
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def users_search(request):
    """
    GET /api/users/search?q=...&page=1&page_size=10
    Alias paginado del listado.
    """
    q = (request.query_params.get("q") or "").strip()

    qs = User.objects.all().order_by("id")
    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q) |
            Q(full_name__icontains=q)
        )

    pag = _paginate_queryset(request, qs, default_page_size=10, max_page_size=100)
    ser = UserSerializer(pag["items"], many=True)

    return Response({
        "count": pag["count"],
        "page": pag["page"],
        "page_size": pag["page_size"],
        "next": pag["next"],
        "previous": pag["previous"],
        "results": ser.data,
    })


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_deactivate(request, pk: int):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    user.is_active = False
    user.save(update_fields=["is_active"])
    return Response({"status": "deactivated"})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_activate(request, pk: int):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    user.is_active = True
    user.save(update_fields=["is_active"])
    return Response({"status": "activated"})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_reset_password(request, pk: int):
    """
    Admin reset: genera temporal + must_change_password=True
    """
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    tmp = get_random_string(10)
    user.set_password(tmp)
    if hasattr(user, "must_change_password"):
        user.must_change_password = True
        user.save(update_fields=["password", "must_change_password"])
    else:
        user.save(update_fields=["password"])

    return Response({"status": "password_reset", "temporary_password": tmp})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_assign_roles(request, pk: int):
    """
    POST /api/users/<id>/roles
    body: { roles: ["ADMIN", "STUDENT"] }
    """
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    roles = request.data.get("roles", [])
    if not isinstance(roles, list):
        return Response({"detail": "roles must be a list"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    if not hasattr(user, "roles"):
        return Response({"detail": "Este modelo User no tiene roles."}, status=status.HTTP_400_BAD_REQUEST)

    role_objs = []
    for name in roles:
        r, _ = Role.objects.get_or_create(name=str(name).strip())
        role_objs.append(r)

    user.roles.set(role_objs)

    # ✅ si se asigna STUDENT, crear Student automáticamente
    role_names = [r.name.upper() for r in role_objs]
    if "STUDENT" in role_names:
        _ensure_student_for_user(user)

    return Response({"status": "roles_assigned", "roles": [r.name for r in role_objs]})


# ✅ ---------- ADMIN SET PASSWORD (OPCIÓN B) ----------
@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_set_password(request, pk: int):
    """
    Admin set password (NO requiere actual).
    POST /api/users/<id>/set-password
    body: { new_password, confirm_password }
    - must_change_password=True
    """
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "No existe."}, status=status.HTTP_404_NOT_FOUND)

    new_password = (request.data.get("new_password") or "").strip()
    confirm_password = (request.data.get("confirm_password") or "").strip()

    errors = {}
    if not new_password:
        errors["new_password"] = "La nueva contraseña es obligatoria."
    if not confirm_password:
        errors["confirm_password"] = "Confirma la nueva contraseña."
    if errors:
        return Response({"errors": errors}, status=status.HTTP_400_BAD_REQUEST)

    if new_password != confirm_password:
        return Response({"errors": {"confirm_password": "La confirmación no coincide."}}, status=status.HTTP_400_BAD_REQUEST)

    try:
        validate_password(new_password, user=user)
    except ValidationError as e:
        return Response({"errors": {"new_password": list(e.messages)}}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)

    if hasattr(user, "must_change_password"):
        user.must_change_password = True
        user.save(update_fields=["password", "must_change_password"])
    else:
        user.save(update_fields=["password"])

    return Response({"status": "password_set", "must_change_password": getattr(user, "must_change_password", False)})


def _purge_one_student(user_obj: User):
    """
    Borra todo lo de un estudiante (eliminación total):
    - AcademicGradeRecord por student_profile
    - Enrollment + EnrollmentItem + EnrollmentPayment (por student)
    - AcademicProcess + ProcessFile (por student_id = Student.id)
    - AttendanceRow (por student_id = user.id)
    - SectionGrades.grades: elimina llave str(user.id)
    - Student profile
    - Teacher records (catalogs + academic) si existen
    - UserRole
    - User
    Retorna conteos.
    """
    counts = {
        "grade_records": 0,
        "enrollments": 0,
        "enrollment_payments": 0,
        "processes": 0,
        "process_files": 0,
        "attendance_rows": 0,
        "section_grades_touched": 0,
        "student_profiles": 0,
        "teachers_unlinked": 0,
        "user_roles": 0,
        "users": 0,
    }

    uid = int(user_obj.id)

    st = None
    try:
        st = user_obj.student_profile
    except Exception:
        st = None

    # ── 1. Grade records ──
    if st:
        counts["grade_records"] = AcademicGradeRecord.objects.filter(student=st).delete()[0]

    # ── 2. Enrollments (y sus items / payments) ──
    if st:
        try:
            from academic.models import Enrollment, EnrollmentPayment
            counts["enrollment_payments"] = EnrollmentPayment.objects.filter(student=st).delete()[0]
            counts["enrollments"] = Enrollment.objects.filter(student=st).delete()[0]
        except Exception:
            pass

    # ── 3. Processes + files ──
    if st:
        proc_ids = list(AcademicProcess.objects.filter(student_id=st.id).values_list("id", flat=True))
        if proc_ids:
            counts["process_files"] = ProcessFile.objects.filter(process_id__in=proc_ids).delete()[0]
        counts["processes"] = AcademicProcess.objects.filter(student_id=st.id).delete()[0]

    # ── 4. Attendance rows ──
    counts["attendance_rows"] = AttendanceRow.objects.filter(student_id=uid).delete()[0]

    # ── 5. Section grades (JSON key) ──
    key = str(uid)
    touched = 0
    try:
        bundles = SectionGrades.objects.filter(grades__has_key=key)  # noqa
    except Exception:
        bundles = SectionGrades.objects.all()

    for b in bundles:
        g = b.grades or {}
        if key in g:
            g.pop(key, None)
            b.grades = g
            b.save(update_fields=["grades", "updated_at"])
            touched += 1

    counts["section_grades_touched"] = touched

    # ── 6. Student profile ──
    if st:
        counts["student_profiles"] = Student.objects.filter(id=st.id).delete()[0]

    # ── 7. Desvincular Teacher records (PROTECT → SET_NULL ya aplicado,
    #        pero por seguridad desvinculamos explícitamente) ──
    try:
        from catalogs.models import Teacher as CatalogTeacher
        n = CatalogTeacher.objects.filter(user_id=uid).update(user=None)
        counts["teachers_unlinked"] += n
    except Exception:
        pass
    try:
        from academic.models import Teacher as AcademicTeacher
        n = AcademicTeacher.objects.filter(user_id=uid).update(user=None)
        counts["teachers_unlinked"] += n
    except Exception:
        pass

    # ── 8. Desvincular Applicant (admission) ──
    try:
        from admission.models import Applicant
        Applicant.objects.filter(user_id=uid).update(user=None)
    except Exception:
        pass

    # ── 9. UserRole ──
    counts["user_roles"] = UserRole.objects.filter(user_id=uid).delete()[0]

    # ── 10. User (final) ──
    counts["users"] = User.objects.filter(id=uid).delete()[0]

    return counts


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_purge(request):
    """
    POST /api/users/purge

    body:
    {
      "mode": "role" | "ids",
      "role": "STUDENT",          # si mode=role
      "user_ids": [1,2,3],        # si mode=ids
      "dry_run": true|false
    }

    - Solo STAFF o ADMIN_ACADEMIC
    - Por seguridad: NO permite borrarte a ti mismo
    """
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    body = request.data or {}
    mode = (body.get("mode") or "").strip().lower()
    dry_run = bool(body.get("dry_run", False))

    if mode not in ("role", "ids"):
        return Response({"detail": "mode inválido. Usa 'role' o 'ids'."}, status=status.HTTP_400_BAD_REQUEST)

    if mode == "role":
        role = (body.get("role") or "").strip()
        if not role:
            return Response({"detail": "role es requerido cuando mode=role"}, status=status.HTTP_400_BAD_REQUEST)
        qs = _qs_users_by_role_name(role)
    else:
        user_ids = body.get("user_ids") or []
        if not isinstance(user_ids, list) or not user_ids:
            return Response({"detail": "user_ids debe ser lista no vacía cuando mode=ids"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user_ids = [int(x) for x in user_ids]
        except Exception:
            return Response({"detail": "user_ids inválidos"}, status=status.HTTP_400_BAD_REQUEST)
        qs = User.objects.filter(id__in=user_ids)

    qs = qs.exclude(id=request.user.id)

    safe_targets = list(qs)

    if not safe_targets:
        return Response({
            "dry_run": dry_run,
            "targets": 0,
            "message": "No hay usuarios para purgar (o intentaste borrar tu propio usuario)."
        })

    if dry_run:
        preview = []
        for u in safe_targets[:200]:
            preview.append({
                "id": u.id,
                "username": getattr(u, "username", ""),
                "email": getattr(u, "email", ""),
                "full_name": getattr(u, "full_name", ""),
            })
        return Response({
            "dry_run": True,
            "targets": len(safe_targets),
            "sample": preview,
            "message": "Vista previa. Enviar dry_run=false para ejecutar."
        })

    totals = {
        "grade_records": 0,
        "enrollments": 0,
        "enrollment_payments": 0,
        "processes": 0,
        "process_files": 0,
        "attendance_rows": 0,
        "section_grades_touched": 0,
        "student_profiles": 0,
        "teachers_unlinked": 0,
        "user_roles": 0,
        "users": 0,
    }

    errors = []

    with transaction.atomic():
        for u in safe_targets:
            try:
                c = _purge_one_student(u)
                for k in totals:
                    totals[k] += int(c.get(k, 0) or 0)
            except Exception as ex:
                errors.append({
                    "user_id": u.id,
                    "username": getattr(u, "username", ""),
                    "error": str(ex),
                })

    if errors and totals["users"] == 0:
        return Response({
            "dry_run": False,
            "targets": len(safe_targets),
            "deleted": totals,
            "errors": errors,
            "message": f"Purge falló para {len(errors)} usuario(s). Revisa los errores."
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({
        "dry_run": False,
        "targets": len(safe_targets),
        "deleted": totals,
        "errors": errors if errors else None,
        "message": f"Purge ejecutado. {totals['users']} usuario(s) eliminado(s)."
    })


def _get_institution_info():
    """Obtiene datos de la institución para encabezados de reportes."""
    import os as _os
    from django.conf import settings as _settings

    info = {"name": "", "ruc": "", "address": "", "phone": "", "email": "", "logo_path": ""}
    try:
        from catalogs.models import InstitutionSetting
        obj = InstitutionSetting.objects.filter(pk=1).first()
        if obj and obj.data:
            d = obj.data
            info["name"] = d.get("name") or d.get("nombre") or ""
            info["ruc"] = d.get("ruc") or ""
            info["address"] = d.get("address") or d.get("direccion") or ""
            info["phone"] = d.get("phone") or d.get("telefono") or ""
            info["email"] = d.get("email") or ""
            # Resolver logo a ruta de archivo
            logo_url = d.get("logo_url") or ""
            if logo_url and "/media/" in logo_url:
                rel = logo_url.split("/media/")[-1]
                fpath = _os.path.join(_settings.MEDIA_ROOT, rel)
                if _os.path.exists(fpath):
                    info["logo_path"] = fpath
    except Exception:
        pass
    if not info["name"]:
        try:
            from academic.models import InstitutionSettings
            inst = InstitutionSettings.objects.first()
            if inst:
                info["name"] = inst.name or ""
                info["ruc"] = getattr(inst, "ruc", "") or ""
                info["address"] = getattr(inst, "address", "") or ""
                info["phone"] = getattr(inst, "phone", "") or ""
                info["email"] = getattr(inst, "email", "") or ""
        except Exception:
            pass
    # Fallback: buscar logo en static
    if not info["logo_path"]:
        for lp in [
            _os.path.join(_settings.BASE_DIR, "media", "institution"),
            _os.path.join(_settings.BASE_DIR, "static", "img"),
            _os.path.join(_settings.BASE_DIR, "staticfiles", "img"),
        ]:
            if _os.path.isdir(lp):
                for f in _os.listdir(lp):
                    if "logo" in f.lower() and f.lower().endswith((".png", ".jpg", ".jpeg")):
                        info["logo_path"] = _os.path.join(lp, f)
                        break
            if info["logo_path"]:
                break
    return info


def _build_credentials_excel(rows, role_name, inst):
    """Genera un Excel con formato profesional para credenciales."""
    import openpyxl as _xl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = _xl.Workbook()
    ws = wb.active
    ws.title = f"Credenciales {role_name}"

    # --- Colores y estilos ---
    DARK_BLUE = "1F4E79"
    LIGHT_BLUE = "D6E4F0"
    WHITE = "FFFFFF"

    fill_header = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    fill_subheader = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    font_title = Font(name="Arial", size=13, bold=True, color=WHITE)
    font_subtitle = Font(name="Arial", size=10, color=WHITE)
    font_col_header = Font(name="Arial", size=10, bold=True, color=WHITE)
    font_data = Font(name="Arial", size=10)
    font_note = Font(name="Arial", size=9, italic=True, color="666666")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Anchos de columna
    col_widths = {"A": 18, "B": 38, "C": 18, "D": 18}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    # --- Logo (si existe) ---
    logo_path = inst.get("logo_path") or ""
    has_logo = False
    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XlImage
            img = XlImage(logo_path)
            img.width = 60
            img.height = 60
            img.anchor = "A1"
            ws.add_image(img)
            has_logo = True
        except Exception:
            pass

    # --- Fila 1: Nombre de institución ---
    if has_logo:
        ws.merge_cells("B1:D1")
        cell_title = ws["B1"]
    else:
        ws.merge_cells("A1:D1")
        cell_title = ws["A1"]
    cell_title.value = inst.get("name") or "INSTITUCIÓN EDUCATIVA"
    cell_title.font = font_title
    cell_title.fill = fill_header
    cell_title.alignment = align_center
    # Pintar fondo de A1 si tiene logo
    if has_logo:
        ws["A1"].fill = fill_header
    ws.row_dimensions[1].height = 32

    # --- Fila 2: Subtítulo ---
    if has_logo:
        ws.merge_cells("B2:D2")
        cell_sub = ws["B2"]
        ws["A2"].fill = fill_header
    else:
        ws.merge_cells("A2:D2")
        cell_sub = ws["A2"]
    role_label = "DOCENTES" if role_name == "TEACHER" else "ESTUDIANTES"
    cell_sub.value = f"CREDENCIALES DE ACCESO - {role_label}"
    cell_sub.font = font_subtitle
    cell_sub.fill = fill_header
    cell_sub.alignment = align_center
    ws.row_dimensions[2].height = 22

    # --- Fila 3: Info adicional (merge A3:D3) ---
    ws.merge_cells("A3:D3")
    parts = []
    if inst.get("address"):
        parts.append(inst["address"])
    if inst.get("phone"):
        parts.append(f"Tel: {inst['phone']}")
    if inst.get("email"):
        parts.append(inst["email"])
    cell_info = ws["A3"]
    cell_info.value = " | ".join(parts) if parts else ""
    cell_info.font = Font(name="Arial", size=9, color="888888")
    cell_info.alignment = align_center
    ws.row_dimensions[3].height = 18

    # --- Fila 4: Espacio ---
    ws.row_dimensions[4].height = 8

    # --- Fila 5: Headers de columna ---
    header_row = 5
    col_headers = ["DOCUMENTO", "NOMBRE COMPLETO", "USUARIO", "CONTRASEÑA"]
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_col_header
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 24

    # --- Filas de datos ---
    for i, r in enumerate(rows):
        row_num = header_row + 1 + i
        is_alt = i % 2 == 1
        values = [r["documento"], r["nombre"], r["usuario"], r["contraseña"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = align_left
            if is_alt:
                cell.fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # --- Fila final: Nota ---
    note_row = header_row + len(rows) + 2
    ws.merge_cells(f"A{note_row}:D{note_row}")
    cell_note = ws.cell(row=note_row, column=1)
    cell_note.value = (
        f"Total: {len(rows)} {role_label.lower()} | "
        "Las contraseñas son temporales. El usuario deberá cambiarla al iniciar sesión."
    )
    cell_note.font = font_note
    cell_note.alignment = align_center

    # --- Configurar impresión ---
    ws.print_title_rows = "1:5"
    ws.sheet_properties.pageSetUpPr = _xl.worksheet.properties.PageSetupProperties(fitToPage=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def users_bulk_credentials(request):
    """
    Regenera contraseñas temporales para todos los usuarios activos de un rol
    y devuelve un Excel descargable con las credenciales.
    Body: { "role": "STUDENT" | "TEACHER" }
    """
    import traceback as _tb

    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    role_name = (request.data.get("role") or "").strip().upper()
    if role_name not in ("STUDENT", "TEACHER"):
        return Response(
            {"detail": "role debe ser STUDENT o TEACHER"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        rows = []

        if role_name == "TEACHER":
            from catalogs.models import Teacher
            teachers = Teacher.objects.filter(
                user__isnull=False, user__is_active=True
            ).select_related("user").order_by("full_name", "user__username")

            if not teachers.exists():
                return Response(
                    {"detail": "No hay docentes con usuario activo."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            for t in teachers:
                u = t.user
                tmp = get_random_string(10)
                u.set_password(tmp)
                u.must_change_password = True
                u.save(update_fields=["password", "must_change_password"])
                rows.append({
                    "documento": t.document or u.username,
                    "nombre": u.full_name or t.full_name or "",
                    "usuario": u.username,
                    "contraseña": tmp,
                })

        else:  # STUDENT
            from students.models import Student
            students = list(
                Student.objects.filter(
                    user__isnull=False, user__is_active=True
                ).select_related("user").order_by("apellido_paterno", "apellido_materno", "nombres")
            )

            if not students:
                return Response(
                    {"detail": "No hay estudiantes con usuario activo."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Generar contraseñas en lote para mejor rendimiento
            users_to_update = []
            for st in students:
                u = st.user
                tmp = get_random_string(10)
                u.set_password(tmp)
                u.must_change_password = True
                users_to_update.append(u)
                rows.append({
                    "documento": st.num_documento or u.username,
                    "nombre": u.full_name or (
                        f"{st.apellido_paterno or ''} {st.apellido_materno or ''} {st.nombres or ''}".strip()
                    ),
                    "usuario": u.username,
                    "contraseña": tmp,
                })

            # bulk_update para no hacer 1092 queries individuales
            User.objects.bulk_update(users_to_update, ["password", "must_change_password"], batch_size=200)

        inst = _get_institution_info()
        excel_bytes = _build_credentials_excel(rows, role_name, inst)

        filename = f"credenciales_{role_name.lower()}.xlsx"
        resp = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    except Exception as exc:
        return Response(
            {"detail": f"Error generando credenciales: {exc}", "trace": _tb.format_exc()},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
