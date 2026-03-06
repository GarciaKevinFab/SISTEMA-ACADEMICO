"""
research/views.py
Módulo de Investigación – IESPP Gustavo Allende Llavería
"""
from django.http import HttpResponse
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, FloatField
from django.db.models.functions import Cast
from django.conf import settings
import os
import traceback
from datetime import datetime
from io import BytesIO

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError

from .models import *
from .serializers import *

import logging
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════
# Decorator: captura errores y devuelve JSON útil
# ══════════════════════════════════════════════

def safe_view(fn):
    """Envuelve una vista DRF para que nunca devuelva 500 sin contexto."""
    from functools import wraps
    @wraps(fn)
    def wrapper(request, *args, **kwargs):
        try:
            return fn(request, *args, **kwargs)
        except ValidationError:
            raise  # DRF ya maneja esto
        except Exception as e:
            logger.error(
                "Error in %s: %s\n%s", fn.__name__, str(e), traceback.format_exc()
            )
            return Response(
                {"detail": f"Error interno: {str(e)}", "_view": fn.__name__},
                status=500,
            )
    return wrapper


# ══════════════════════════════════════════════
# Dashboard – stats para la vista principal
# ══════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def dashboard_stats(request):
    """Estadísticas para el dashboard de investigación."""
    from django.utils import timezone
    today = timezone.now().date()

    total = Project.objects.count()
    active = Project.objects.filter(status__in=["RUNNING", "APPROVED", "IN_REVIEW"]).count()
    completed = Project.objects.filter(status="FINISHED").count()
    draft = Project.objects.filter(status="DRAFT").count()

    # Entregables pendientes
    pending_deliverables = Deliverable.objects.filter(
        status__in=["PENDING", ""]
    ).count()

    # Entregables vencidos (con fecha pasada y no aprobados)
    overdue_deliverables = 0
    try:
        overdue_deliverables = Deliverable.objects.filter(
            due_date__lt=today,
            status__in=["PENDING", ""],
        ).count()
    except Exception:
        pass

    # Advisors activos
    total_advisors = Advisor.objects.filter(is_active=True).count()

    # Líneas de investigación
    total_lines = ResearchLine.objects.count()

    # Convocatorias abiertas
    open_calls = 0
    try:
        open_calls = Call.objects.filter(end_date__gte=today).count()
    except Exception:
        pass

    # Promedio de evaluación (robusto)
    avg_score = 0.0
    try:
        avg_score = _safe_avg_rubric_total(Project.objects.all())
    except Exception:
        pass

    # Proyectos por estado
    by_status = []
    try:
        by_status = list(
            Project.objects.values("status")
            .order_by()
            .annotate(count=Count("id"))
        )
    except Exception:
        pass

    # Últimos proyectos modificados
    recent = []
    try:
        recent = list(
            Project.objects.select_related("line", "advisor")
            .order_by("-updated_at")[:5]
            .values("id", "title", "status", "updated_at")
        )
    except Exception:
        pass

    return Response({
        "total_projects": total,
        "active_projects": active,
        "completed_projects": completed,
        "draft_projects": draft,
        "pending_deliverables": pending_deliverables,
        "overdue_deliverables": overdue_deliverables,
        "total_advisors": total_advisors,
        "total_lines": total_lines,
        "open_calls": open_calls,
        "avg_score": round(avg_score, 2),
        "by_status": by_status,
        "recent_projects": recent,
    })


# ══════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════

def _deliverable_to_fe(d: Deliverable):
    meta = d.meta or {}
    return {
        "id": d.id,
        "name": d.name,
        "description": d.description,
        "due_date": d.due_date,
        "status": d.status,
        "link": meta.get("link", ""),
        "updated_at": d.updated_at,
        "file_url": d.file.url if d.file else None,
    }


def _budget_item_to_fe(b: BudgetItem):
    meta = b.meta or {}
    return {
        "id": b.id,
        "date": meta.get("date"),
        "category": b.category,
        "item": b.concept,
        "planned": float(b.amount or 0),
        "executed": float(b.executed or 0),
        "doc_type": meta.get("doc_type", ""),
        "doc_number": meta.get("doc_number", ""),
        "receipt_url": b.receipt.url if b.receipt else None,
    }


def _fe_to_budget_payload(payload: dict):
    meta = {}
    if payload.get("date"):
        meta["date"] = payload["date"]
    if payload.get("doc_type"):
        meta["doc_type"] = payload["doc_type"]
    if payload.get("doc_number"):
        meta["doc_number"] = payload["doc_number"]
    return {
        "category": payload.get("category") or "OTHER",
        "concept": payload.get("item") or "",
        "amount": payload.get("planned", 0) or 0,
        "executed": payload.get("executed", 0) or 0,
        "meta": meta,
    }


def _schedule_item_to_fe(s: ScheduleItem):
    m = s.meta or {}
    due = s.end or s.start
    return {
        "id": s.id,
        "title": s.name,
        "due_date": due,
        "responsible": m.get("responsible", ""),
        "status": m.get("status", "PLANNED"),
        "progress": s.progress or 0,
    }


def _fe_to_schedule_model(it: dict):
    meta = {}
    if it.get("responsible"):
        meta["responsible"] = it["responsible"]
    if it.get("status"):
        meta["status"] = it["status"]
    due = it.get("due_date")
    return {
        "name": it.get("title") or "",
        "start": due or it.get("start") or it.get("end"),
        "end": due or it.get("end") or it.get("start"),
        "progress": int(it.get("progress") or 0),
        "meta": meta,
    }


def _normalize_date(v):
    """
    Devuelve 'YYYY-MM-DD' o None.
    Prioriza formato latinoamericano DD/MM/YYYY sobre MM/DD/YYYY.
    """
    if not v:
        return None
    s = str(v).strip()[:10]
    # Ya es ISO
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    # DD/MM/YYYY (formato PE)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            continue
    return s


def _ensure_media_tmp():
    tmpdir = os.path.join(settings.MEDIA_ROOT, "tmp")
    os.makedirs(tmpdir, exist_ok=True)
    return tmpdir


def _write_stub_pdf(abs_path: str, title="Reporte", subtitle=""):
    """Genera un PDF con ReportLab o fallback minimal."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        c = canvas.Canvas(abs_path, pagesize=A4)
        w, h = A4
        c.setFont("Helvetica-Bold", 16)
        c.drawString(72, h - 72, title)
        c.setFont("Helvetica", 12)
        if subtitle:
            c.drawString(72, h - 100, subtitle)
        c.drawString(72, h - 130, "Documento generado automáticamente.")
        c.save()
    except Exception:
        minimal_pdf = (
            b"%PDF-1.4\n1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n"
            b"2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n"
            b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>endobj\n"
            b"4 0 obj<< /Length 44 >>stream\nBT /F1 16 Tf 72 720 Td (Reporte) Tj ET\n"
            b"endstream\nendobj\n"
            b"5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n"
            b"xref\n0 6\n"
            b"0000000000 65535 f \n0000000010 00000 n \n0000000061 00000 n \n"
            b"0000000116 00000 n \n0000000270 00000 n \n0000000370 00000 n \n"
            b"trailer<< /Size 6 /Root 1 0 R >>\nstartxref\n450\n%%EOF\n"
        )
        with open(abs_path, "wb") as f:
            f.write(minimal_pdf)


# ══════════════════════════════════════════════
# Catálogos – Líneas de investigación
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def catalog_lines(request):
    if request.method == "GET":
        qs = ResearchLine.objects.all()
        return Response(LineSerializer(qs, many=True).data)
    ser = LineSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save()
    return Response(LineSerializer(obj).data, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def catalog_line_detail(request, id: int):
    try:
        obj = ResearchLine.objects.get(pk=id)
    except ResearchLine.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        ser = LineSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(LineSerializer(obj).data)
    obj.delete()
    return Response(status=204)


# ══════════════════════════════════════════════
# Catálogos – Asesores
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def catalog_advisors(request):
    if request.method == "GET":
        return Response(AdvisorSerializer(Advisor.objects.all(), many=True).data)
    ser = AdvisorSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save()
    return Response(AdvisorSerializer(obj).data, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def catalog_advisor_detail(request, id: int):
    try:
        obj = Advisor.objects.get(pk=id)
    except Advisor.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        ser = AdvisorSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(AdvisorSerializer(obj).data)
    obj.delete()
    return Response(status=204)


# ══════════════════════════════════════════════
# Proyectos – CRUD + cambio de estado
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def projects_collection(request):
    if request.method == "GET":
        qs = Project.objects.select_related("line", "advisor").all()
        status_f = request.query_params.get("status")
        line_id = request.query_params.get("line_id")
        if status_f:
            qs = qs.filter(status=status_f)
        if line_id:
            qs = qs.filter(line_id=line_id)
        try:
            data = ProjectSerializer(qs[:200], many=True).data
        except Exception as e:
            logger.error("ProjectSerializer error: %s\n%s", e, traceback.format_exc())
            # Fallback: devolver lista mínima
            data = list(qs[:200].values(
                "id", "title", "status", "start_date", "end_date",
                "line_id", "advisor_id", "meta", "created_at", "updated_at",
            ))
        return Response(data)

    try:
        payload = request.data.copy() if hasattr(request.data, "copy") else dict(request.data)

        # Normalizar FKs
        if "line_id" in payload and "line" not in payload and payload.get("line_id") not in (None, "", "null"):
            payload["line"] = payload["line_id"]
        if "advisor_id" in payload and "advisor" not in payload and payload.get("advisor_id") not in (None, "", "null"):
            payload["advisor"] = payload["advisor_id"]

        for k in ("start_date", "end_date"):
            if payload.get(k):
                payload[k] = _normalize_date(payload[k])

        if payload.get("budget") in (None, "", "null"):
            payload["budget"] = 0
        payload.setdefault("status", "DRAFT")

        ser = ProjectSerializer(data=payload)
        ser.is_valid(raise_exception=True)

        advisors_ids = payload.get("advisors_ids") or payload.get("advisors") or []

        with transaction.atomic():
            try:
                obj = ser.save(created_by=request.user)
            except TypeError:
                obj = ser.save()

            # M2M co-asesores
            if advisors_ids and hasattr(obj, "advisors"):
                if isinstance(advisors_ids, str):
                    advisors_ids = [x for x in advisors_ids.split(",") if x.strip()]
                try:
                    obj.advisors.set(list(map(int, advisors_ids)))
                except (ValueError, TypeError):
                    pass

        return Response(ProjectSerializer(obj).data, status=201)

    except ValidationError as e:
        return Response({"detail": "Validation error", "errors": e.detail}, status=400)
    except IntegrityError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:
        tb = traceback.format_exc(limit=2)
        return Response({"detail": f"{e}", "hint": tb}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def project_detail(request, id: int):
    try:
        obj = Project.objects.select_related("line", "advisor").get(pk=id)
    except Project.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(ProjectSerializer(obj).data)

    if request.method == "PATCH":
        payload = request.data.copy() if hasattr(request.data, "copy") else dict(request.data)

        if "line_id" in payload and payload.get("line_id") not in (None, "", "null"):
            payload["line"] = payload.pop("line_id")
        if "advisor_id" in payload and payload.get("advisor_id") not in (None, "", "null"):
            payload["advisor"] = payload.pop("advisor_id")

        for k in ("start_date", "end_date"):
            if payload.get(k):
                payload[k] = _normalize_date(payload[k])

        if "budget" in payload and payload.get("budget") in (None, "", "null"):
            payload["budget"] = 0

        ser = ProjectSerializer(obj, data=payload, partial=True)
        try:
            ser.is_valid(raise_exception=True)
            with transaction.atomic():
                try:
                    obj = ser.save(updated_by=request.user)
                except TypeError:
                    obj = ser.save()

                advisors_ids = payload.get("advisors_ids") or payload.get("advisors")
                if advisors_ids is not None and hasattr(obj, "advisors"):
                    if isinstance(advisors_ids, str):
                        advisors_ids = [x for x in advisors_ids.split(",") if x.strip()]
                    try:
                        obj.advisors.set(list(map(int, advisors_ids)))
                    except (ValueError, TypeError):
                        pass
        except ValidationError as e:
            return Response({"detail": "Validation error", "errors": e.detail}, status=400)

        return Response(ProjectSerializer(obj).data)

    # DELETE
    obj.delete()
    return Response(status=204)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def project_change_status(request, id: int):
    try:
        obj = Project.objects.get(pk=id)
    except Project.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    st = request.data.get("status")
    if not st:
        return Response({"detail": "status requerido"}, status=400)
    obj.status = st
    obj.save(update_fields=["status"])
    return Response({"ok": True, "id": obj.id, "status": obj.status})


# ══════════════════════════════════════════════
# Cronograma
# ══════════════════════════════════════════════
@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def schedule_list(request, projectId: int):
    items = ScheduleItem.objects.filter(project_id=projectId)
    return Response([_schedule_item_to_fe(s) for s in items])


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def schedule_bulk(request, projectId: int):
    items = (request.data or {}).get("items", [])
    with transaction.atomic():
        ScheduleItem.objects.filter(project_id=projectId).delete()
        objs = [
            ScheduleItem(project_id=projectId, **_fe_to_schedule_model(it))
            for it in items
        ]
        if objs:
            ScheduleItem.objects.bulk_create(objs)
    return Response({"ok": True, "count": len(objs)})


# ══════════════════════════════════════════════
# Entregables
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def deliverables_collection(request, projectId: int):
    if request.method == "GET":
        rows = Deliverable.objects.filter(project_id=projectId)
        return Response([_deliverable_to_fe(d) for d in rows])
    payload = request.data.copy()
    meta = payload.get("meta") or {}
    if payload.get("link"):
        meta["link"] = payload["link"]
    ser = DeliverableSerializer(data={**payload, "meta": meta})
    ser.is_valid(raise_exception=True)
    obj = ser.save(project_id=projectId)
    return Response(_deliverable_to_fe(obj), status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def deliverable_update(request, deliverableId: int):
    try:
        d = Deliverable.objects.get(pk=deliverableId)
    except Deliverable.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "DELETE":
        d.delete()
        return Response(status=204)
    payload = request.data.copy()
    meta = d.meta or {}
    if "link" in payload:
        if payload["link"] in (None, ""):
            meta.pop("link", None)
        else:
            meta["link"] = payload["link"]
    payload["meta"] = meta
    ser = DeliverableSerializer(d, data=payload, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(_deliverable_to_fe(d))


# ══════════════════════════════════════════════
# Evaluaciones de proyecto
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def evaluations_collection(request, projectId: int):
    if request.method == "GET":
        qs = Evaluation.objects.filter(project_id=projectId)
        return Response(EvaluationSerializer(qs, many=True).data)
    payload = request.data or {}
    ev = Evaluation.objects.create(
        project_id=projectId,
        evaluator=request.user if hasattr(request, "user") else None,
        rubric=payload,
    )
    return Response(EvaluationSerializer(ev).data, status=201)


# ══════════════════════════════════════════════
# Equipo
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def team_collection(request, projectId: int):
    if request.method == "GET":
        qs = TeamMember.objects.filter(project_id=projectId)
        return Response(TeamMemberSerializer(qs, many=True).data)
    ser = TeamMemberSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save(project_id=projectId)
    return Response(TeamMemberSerializer(obj).data, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def team_member_detail(request, projectId: int, memberId: int):
    try:
        m = TeamMember.objects.get(pk=memberId, project_id=projectId)
    except TeamMember.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        ser = TeamMemberSerializer(m, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(TeamMemberSerializer(m).data)
    m.delete()
    return Response(status=204)


# ══════════════════════════════════════════════
# Presupuesto
# ══════════════════════════════════════════════
@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_list(request, projectId: int):
    items = BudgetItem.objects.filter(project_id=projectId)
    data = [_budget_item_to_fe(b) for b in items]
    summary = {
        "planned": sum(x["planned"] for x in data),
        "executed": sum(x["executed"] for x in data),
    }
    return Response({"items": data, "summary": summary})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_create_item(request, projectId: int):
    mapped = _fe_to_budget_payload(request.data or {})
    ser = BudgetItemSerializer(data=mapped)
    ser.is_valid(raise_exception=True)
    obj = ser.save(project_id=projectId)
    return Response(_budget_item_to_fe(obj), status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_item_detail(request, projectId: int, itemId: int):
    try:
        it = BudgetItem.objects.get(pk=itemId, project_id=projectId)
    except BudgetItem.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        mapped = _fe_to_budget_payload(request.data or {})
        meta = it.meta or {}
        meta.update(mapped.get("meta") or {})
        mapped["meta"] = meta
        ser = BudgetItemSerializer(it, data=mapped, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(_budget_item_to_fe(it))
    it.delete()
    return Response(status=204)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_upload_receipt(request, projectId: int, itemId: int):
    try:
        it = BudgetItem.objects.get(pk=itemId, project_id=projectId)
    except BudgetItem.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "file requerido"}, status=400)
    it.receipt = f
    it.save(update_fields=["receipt"])
    return Response({
        "ok": True,
        "id": it.id,
        "receipt_url": it.receipt.url if it.receipt else None,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_export_pdf_stub(request, projectId: int):
    tmpdir = _ensure_media_tmp()
    filename = f"budget-{projectId}.pdf"
    abs_path = os.path.join(tmpdir, filename)
    _write_stub_pdf(abs_path, title=f"Presupuesto del Proyecto #{projectId}")
    return Response({"success": True, "downloadUrl": f"/media/tmp/{filename}"})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def budget_export_xlsx(request, projectId: int):
    """
    FIX: Genera un XLSX real con openpyxl (o fallback CSV).
    """
    items = BudgetItem.objects.filter(project_id=projectId)
    rows = [_budget_item_to_fe(b) for b in items]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presupuesto"

        headers = ["Categoría", "Concepto", "Planificado", "Ejecutado", "Tipo Doc", "Nro Doc", "Fecha"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=r["category"])
            ws.cell(row=row_idx, column=2, value=r["item"])
            ws.cell(row=row_idx, column=3, value=r["planned"])
            ws.cell(row=row_idx, column=4, value=r["executed"])
            ws.cell(row=row_idx, column=5, value=r["doc_type"])
            ws.cell(row=row_idx, column=6, value=r["doc_number"])
            ws.cell(row=row_idx, column=7, value=r["date"] or "")

        # Totales
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(r["planned"] for r in rows)).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=sum(r["executed"] for r in rows)).font = Font(bold=True)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="project_{projectId}_budget.xlsx"'
        return resp

    except ImportError:
        # Fallback: CSV
        import csv
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="project_{projectId}_budget.csv"'
        writer = csv.writer(resp)
        writer.writerow(["Categoría", "Concepto", "Planificado", "Ejecutado", "Tipo Doc", "Nro Doc", "Fecha"])
        for r in rows:
            writer.writerow([r["category"], r["item"], r["planned"], r["executed"], r["doc_type"], r["doc_number"], r["date"] or ""])
        return resp


# ══════════════════════════════════════════════
# Ética & Propiedad Intelectual
# ══════════════════════════════════════════════
@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def ethics_ip_get(request, projectId: int):
    obj, _ = EthicsIP.objects.get_or_create(project_id=projectId)
    return Response(EthicsIPSerializer(obj).data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
@safe_view
def ethics_set(request, projectId: int):
    obj, _ = EthicsIP.objects.get_or_create(project_id=projectId)
    obj.ethics = request.data or {}
    obj.save(update_fields=["ethics"])
    return Response(EthicsIPSerializer(obj).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def ethics_upload_doc(request, projectId: int):
    obj, _ = EthicsIP.objects.get_or_create(project_id=projectId)
    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "file requerido"}, status=400)
    obj.ethics_doc = f
    obj.save(update_fields=["ethics_doc"])
    return Response({"ok": True, "url": obj.ethics_doc.url if obj.ethics_doc else None})


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
@safe_view
def ip_set(request, projectId: int):
    obj, _ = EthicsIP.objects.get_or_create(project_id=projectId)
    obj.ip = request.data or {}
    obj.save(update_fields=["ip"])
    return Response(EthicsIPSerializer(obj).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def ip_upload_doc(request, projectId: int):
    obj, _ = EthicsIP.objects.get_or_create(project_id=projectId)
    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "file requerido"}, status=400)
    obj.ip_doc = f
    obj.save(update_fields=["ip_doc"])
    return Response({"ok": True, "url": obj.ip_doc.url if obj.ip_doc else None})


# ══════════════════════════════════════════════
# Publicaciones
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def publications_collection(request, projectId: int):
    if request.method == "GET":
        qs = Publication.objects.filter(project_id=projectId)
        return Response(PublicationSerializer(qs, many=True).data)
    ser = PublicationSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save(project_id=projectId)
    return Response(PublicationSerializer(obj).data, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def publication_detail(request, projectId: int, pubId: int):
    try:
        p = Publication.objects.get(pk=pubId, project_id=projectId)
    except Publication.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        ser = PublicationSerializer(p, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(PublicationSerializer(p).data)
    p.delete()
    return Response(status=204)


# ══════════════════════════════════════════════
# Convocatorias
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def calls_collection(request):
    if request.method == "GET":
        return Response(CallSerializer(Call.objects.all(), many=True).data)
    ser = CallSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save()
    return Response(CallSerializer(obj).data, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
@safe_view
def call_detail(request, id: int):
    try:
        c = Call.objects.get(pk=id)
    except Call.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    if request.method == "PATCH":
        ser = CallSerializer(c, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(CallSerializer(c).data)
    c.delete()
    return Response(status=204)


# ══════════════════════════════════════════════
# Propuestas
# ══════════════════════════════════════════════
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@safe_view
def proposals_collection(request, callId: int):
    if request.method == "GET":
        qs = Proposal.objects.filter(call_id=callId)
        return Response(ProposalSerializer(qs, many=True).data)
    ser = ProposalSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    obj = ser.save(call_id=callId)
    return Response(ProposalSerializer(obj).data, status=201)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def proposal_submit(request, callId: int, proposalId: int):
    try:
        p = Proposal.objects.get(pk=proposalId, call_id=callId)
    except Proposal.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    p.status = "SUBMITTED"
    p.save(update_fields=["status"])
    return Response({"ok": True, "status": p.status})


# ══════════════════════════════════════════════
# Revisiones de propuestas
# ══════════════════════════════════════════════
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def review_assign(request, callId: int, proposalId: int):
    reviewer_id = request.data.get("reviewer_id")
    if not reviewer_id:
        return Response({"detail": "reviewer_id requerido"}, status=400)
    try:
        p = Proposal.objects.get(pk=proposalId, call_id=callId)
    except Proposal.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    ProposalReview.objects.create(proposal=p, reviewer_id=reviewer_id, rubric={})
    return Response({"ok": True})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def review_rubric_get(request, callId: int, proposalId: int):
    """
    FIX: Devuelve la rúbrica del Call (rubric_template) + la última review si existe.
    Antes solo devolvía la última review sin la plantilla de rúbrica.
    """
    try:
        p = Proposal.objects.get(pk=proposalId, call_id=callId)
    except Proposal.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    call = p.call
    last_review = p.reviews.order_by("-id").first()

    data = {
        "rubric_template": call.rubric_template or [],
        "review": ProposalReviewSerializer(last_review).data if last_review else None,
    }
    return Response(data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@safe_view
def review_save(request, callId: int, proposalId: int):
    try:
        p = Proposal.objects.get(pk=proposalId, call_id=callId)
    except Proposal.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    payload = request.data or {}
    pr = ProposalReview.objects.create(
        proposal=p,
        reviewer_id=payload.get("reviewer_id") or (request.user.id if hasattr(request, "user") else 0),
        rubric=payload,
    )
    p.status = "REVIEWED"
    p.save(update_fields=["status"])
    return Response(ProposalReviewSerializer(pr).data, status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def calls_ranking(request, callId: int):
    rows = []
    for pr in Proposal.objects.filter(call_id=callId):
        totals = []
        for r in pr.reviews.all():
            try:
                totals.append(float(r.rubric.get("total", 0)))
            except (TypeError, ValueError, AttributeError):
                pass
        avg = sum(totals) / len(totals) if totals else 0.0
        rows.append({
            "proposal_id": pr.id,
            "title": pr.title,
            "avg_total": round(avg, 2),
            "review_count": len(totals),
        })
    rows.sort(key=lambda x: x["avg_total"], reverse=True)
    return Response(rows)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def calls_ranking_export(request, callId: int):
    tmpdir = _ensure_media_tmp()
    filename = f"call-{callId}-ranking.pdf"
    abs_path = os.path.join(tmpdir, filename)
    _write_stub_pdf(abs_path, title=f"Ranking – Convocatoria #{callId}")
    return Response({"success": True, "downloadUrl": f"/media/tmp/{filename}"})


# ══════════════════════════════════════════════
# Reportes
# ══════════════════════════════════════════════
def _safe_avg_rubric_total(qs_projects):
    """Promedio robusto de rubric.total (SQL con fallback Python)."""
    base = Evaluation.objects.filter(project__in=qs_projects)
    try:
        val = (
            base
            .exclude(rubric__total__isnull=True)
            .exclude(rubric__total="")
            .aggregate(a=Avg(Cast("rubric__total", FloatField())))
            .get("a")
        )
        return float(val or 0)
    except Exception:
        totals = list(base.values_list("rubric__total", flat=True))
        nums = []
        for t in totals:
            try:
                nums.append(float(t))
            except (TypeError, ValueError):
                pass
        return (sum(nums) / len(nums)) if nums else 0.0


@api_view(["GET"])
@permission_classes([IsAuthenticated])
@safe_view
def reports_summary(request):
    year = request.query_params.get("year")
    status_f = request.query_params.get("status")

    qs = Project.objects.all()
    if status_f:
        qs = qs.filter(status=status_f)
    if year:
        try:
            year_int = int(year)
            qs = qs.filter(start_date__year=year_int)
        except (ValueError, TypeError):
            pass

    total_projects = qs.count()

    total_advisors = 0
    try:
        total_advisors = Advisor.objects.filter(
            projects_as_main__in=qs
        ).distinct().count()
    except Exception:
        pass

    total_deliverables = 0
    try:
        total_deliverables = Deliverable.objects.filter(project__in=qs).count()
    except Exception:
        pass

    avg_score = 0.0
    try:
        avg_score = _safe_avg_rubric_total(qs)
    except Exception:
        pass

    by_status = []
    try:
        by_status = list(qs.values("status").order_by().annotate(count=Count("id")))
    except Exception:
        pass

    data = {
        "total_projects": total_projects,
        "total_advisors": total_advisors,
        "total_deliverables": total_deliverables,
        "avg_score": round(avg_score, 2),
        "by_status": by_status,
    }
    return Response(data)


@api_view(["POST", "GET"])
@permission_classes([IsAuthenticated])
@safe_view
def reports_summary_export_stub(request):
    year = request.query_params.get("year") or (request.data or {}).get("year") or ""
    suffix = f"-{year}" if year else ""
    filename = f"research-summary{suffix}.pdf"
    tmpdir = _ensure_media_tmp()
    abs_path = os.path.join(tmpdir, filename)
    _write_stub_pdf(
        abs_path,
        title="Reporte de Investigación",
        subtitle=f"Año: {year}" if year else "Resumen general",
    )
    return Response({"success": True, "downloadUrl": f"/media/tmp/{filename}"})