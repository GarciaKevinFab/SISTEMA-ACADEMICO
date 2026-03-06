"""
admission/views_reports.py

Vistas para Reportes y Dashboard de Admision

FIX:
  1. reports_admission_xlsx genera Excel REAL con openpyxl
     (antes devolvia bytes vacios b"")
  2. _get_app_field() busca en data.profile / data.school / applicant FK
     con aliases para soportar ambos wizards (espanol e ingles)
  3. Fallback a CSV si openpyxl no esta instalado
  4. select_related + prefetch_related para evitar N+1
"""
import io
import csv
from datetime import timedelta
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from admission.models import AdmissionCall, Application, ApplicationPreference


# ══════════════════════════════════════════════════════════════
# Dashboard
# ══════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admission_dashboard(request):
    """Dashboard con estadisticas generales"""
    calls_open = AdmissionCall.objects.filter(published=False).count()
    total_applications = Application.objects.count()

    by_career = (
        Application.objects.values("career_name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    by_career = [
        {"name": r["career_name"] or "Sin carrera", "value": r["count"]}
        for r in by_career
    ]

    return Response({
        "total_applications": total_applications,
        "calls_open": calls_open,
        "by_career": by_career,
        "trend": [],
    })


# ══════════════════════════════════════════════════════════════
# HELPER: extraer datos del JSON anidado de Application.data
# ══════════════════════════════════════════════════════════════

# Aliases: el PublicApplicationWizard puede usar nombres en ingles,
# pero el reporte MINEDU necesita nombres en espanol
_ALIASES = {
    "nombres":          ["nombres", "first_names", "names"],
    "apellido_paterno": ["apellido_paterno", "last_name_father"],
    "apellido_materno": ["apellido_materno", "last_name_mother"],
    "dni":              ["dni", "document_number", "numero_documento_identidad"],
    "sexo":             ["sexo", "sex"],
    "fecha_nacimiento": ["fecha_nacimiento", "birth_date"],
    "nacionalidad":     ["nacionalidad", "nationality"],
    "email":            ["email"],
    "phone":            ["phone", "mobile", "telefono"],
    "direccion":        ["direccion", "direccion_domicilio", "address"],
    "lengua_materna":   ["lengua_materna", "mother_tongue"],
    "estado_civil":     ["estado_civil"],
    "autoidentificacion_etnica": ["autoidentificacion_etnica", "ethnic_identity"],
    "discapacidad":     ["discapacidad"],
    "tipo_discapacidad": ["tipo_discapacidad"],
    "colegio_procedencia": ["colegio_procedencia", "school_name"],
    "anio_egreso":      ["anio_egreso", "promotion_year", "anio_finalizo_estudios_secundarios"],
    "modalidad_admision": ["modalidad_admision"],
}


def _get_app_field(app, key, fallback=""):
    """
    Busca un campo en la estructura anidada de Application:
      1. data.profile[key]   <- donde public_apply guarda campos MINEDU
      2. data.school[key]    <- datos escolares
      3. data[key]           <- nivel raiz (compatibilidad)
      4. applicant.{key}     <- FK al modelo Applicant
      5. app.{key}           <- campo directo del modelo Application

    Usa aliases para soportar ambos wizards (espanol e ingles).
    """
    data = app.data if isinstance(getattr(app, "data", None), dict) else {}
    profile = data.get("profile") or {}
    school = data.get("school") or {}
    applicant = app.applicant

    # Lista de keys a buscar (con aliases)
    keys_to_try = _ALIASES.get(key, [key])

    for k in keys_to_try:
        # 1) profile
        val = profile.get(k)
        if val not in (None, ""):
            return str(val)
        # 2) school
        val = school.get(k)
        if val not in (None, ""):
            return str(val)
        # 3) data raiz (solo valores simples)
        val = data.get(k)
        if val not in (None, "") and not isinstance(val, (dict, list)):
            return str(val)

    # 4) applicant FK (solo key original)
    if applicant and hasattr(applicant, key):
        val = getattr(applicant, key, None)
        if val not in (None, ""):
            return str(val)

    # 5) app directo (solo key original)
    if hasattr(app, key):
        val = getattr(app, key, None)
        if val not in (None, "") and not isinstance(val, (dict, list)):
            return str(val)

    return fallback


# ══════════════════════════════════════════════════════════════
# Columnas MINEDU para el reporte Excel
# ══════════════════════════════════════════════════════════════

REPORT_COLUMNS = [
    ("N",                     5),
    ("DNI",                  12),
    ("Apellido Paterno",     18),
    ("Apellido Materno",     18),
    ("Nombres",              20),
    ("Sexo",                 12),
    ("Fecha Nacimiento",     16),
    ("Nacionalidad",         14),
    ("Email",                28),
    ("Telefono",             14),
    ("Direccion",            30),
    ("Estado Civil",         12),
    ("Lengua Materna",       14),
    ("Autoidentif. Etnica",  18),
    ("Discapacidad",         12),
    ("Tipo Discapacidad",    18),
    ("Colegio Procedencia",  25),
    ("Anio Egreso",          12),
    ("Programa de Estudios", 28),
    ("Modalidad Admision",   22),
    ("Convocatoria",         25),
    ("Estado",               14),
    ("Fecha Registro",       18),
]


def _get_filtered_qs(request):
    """Queryset de Applications filtrado por call_id y career_id."""
    call_id = request.query_params.get("call_id")
    career_id = request.query_params.get("career_id")

    qs = (
        Application.objects
        .select_related("applicant", "call")
        .prefetch_related("preferences__career")
        .order_by("id")
    )
    if call_id:
        qs = qs.filter(call_id=call_id)
    if career_id:
        qs = qs.filter(preferences__career_id=career_id).distinct()

    return qs


def _build_row(idx, app):
    """Construye una fila de datos para el reporte."""
    # DNI
    dni = _get_app_field(app, "dni")
    if not dni and app.applicant:
        dni = app.applicant.dni or ""

    # Nombre separado
    ap_pat = _get_app_field(app, "apellido_paterno")
    ap_mat = _get_app_field(app, "apellido_materno")
    nombres = _get_app_field(app, "nombres")

    # Email y telefono (fallback a applicant)
    email = _get_app_field(app, "email")
    if not email and app.applicant:
        email = app.applicant.email or ""
    phone = _get_app_field(app, "phone")
    if not phone and app.applicant:
        phone = app.applicant.phone or ""

    # Carrera (primera preferencia o career_name)
    career_name = app.career_name or ""
    if not career_name:
        try:
            first_pref = app.preferences.order_by("rank").first()
            if first_pref and first_pref.career:
                career_name = first_pref.career.name
        except Exception:
            pass

    # Fecha de registro
    created = ""
    if hasattr(app, "created_at") and app.created_at:
        created = app.created_at.strftime("%d/%m/%Y %H:%M")

    # Convocatoria
    call_title = ""
    if app.call:
        call_title = getattr(app.call, "title", "") or getattr(app.call, "name", "") or ""

    return [
        idx,
        dni,
        ap_pat,
        ap_mat,
        nombres,
        _get_app_field(app, "sexo"),
        _get_app_field(app, "fecha_nacimiento"),
        _get_app_field(app, "nacionalidad"),
        email,
        phone,
        _get_app_field(app, "direccion"),
        _get_app_field(app, "estado_civil"),
        _get_app_field(app, "lengua_materna"),
        _get_app_field(app, "autoidentificacion_etnica"),
        _get_app_field(app, "discapacidad"),
        _get_app_field(app, "tipo_discapacidad"),
        _get_app_field(app, "colegio_procedencia"),
        _get_app_field(app, "anio_egreso"),
        career_name,
        _get_app_field(app, "modalidad_admision"),
        call_title,
        app.status or "",
        created,
    ]


# ══════════════════════════════════════════════════════════════
# Endpoint: Reporte Excel
# ══════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_admission_xlsx(request):
    """
    Reporte general de admision en formato Excel (.xlsx).
    Con fallback a CSV si openpyxl no esta instalado.
    """
    qs = _get_filtered_qs(request)
    headers = [col[0] for col in REPORT_COLUMNS]

    # Construir filas
    rows_data = []
    for idx, app in enumerate(qs, start=1):
        rows_data.append(_build_row(idx, app))

    # ── Intentar generar Excel con openpyxl ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Postulantes Admision"

        # ── Estilos ──
        hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="B0B0B0")
        thin_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side,
        )
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center", wrap_text=False)
        # Filas alternadas para legibilidad
        stripe_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ── Cabeceras ──
        for col_idx, (col_name, col_width) in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 30

        # ── Datos con filas alternadas ──
        for row_idx, row_data in enumerate(rows_data, start=2):
            row_fill = stripe_fill if (row_idx % 2 == 0) else white_fill
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                cell.fill = row_fill

        # ── Freeze panes (cabecera fija al hacer scroll) ──
        ws.freeze_panes = "A2"

        # ── Tabla con formato + auto-filter (compatibilidad cross-version) ──
        last_col = get_column_letter(len(REPORT_COLUMNS))
        last_row = len(rows_data) + 1
        table_added = False

        if rows_data:
            try:
                tab = Table(
                    displayName="PostulantesAdmision",
                    ref=f"A1:{last_col}{last_row}",
                )
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)
                table_added = True
            except Exception:
                pass

        # Auto-filter solo si no se pudo crear la Tabla (evitar conflicto)
        if not table_added:
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # ── Print settings para impresión ──
        try:
            from openpyxl.worksheet.properties import PageSetupProperties
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "1:1"  # Repetir cabecera en cada página
        except Exception:
            pass  # No critico, el archivo funciona sin esto

        # ── Guardar en buffer ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        content = buf.getvalue()
        buf.close()

        # Validar que el archivo no este vacio
        if len(content) < 200:
            raise ValueError("Archivo generado demasiado pequeno, posible error")

        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_admision.xlsx"'
        response["Content-Length"] = len(content)
        return response

    except ImportError:
        # openpyxl no instalado → fallback a CSV
        pass
    except Exception as exc:
        # Error generando Excel → fallback a CSV
        import logging
        logging.getLogger(__name__).warning("Error generando XLSX, fallback CSV: %s", exc)

    # ── Fallback: CSV ──
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="reporte_admision.csv"'
    response.write("\ufeff")  # BOM para Excel

    writer = csv.writer(response)
    writer.writerow(headers)
    for row_data in rows_data:
        writer.writerow(row_data)

    return response


# ══════════════════════════════════════════════════════════════
# Otros reportes (stubs)
# ══════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_admission_summary(request):
    """Resumen simple de admision"""
    total = Application.objects.count()
    return Response({"total_applications": total})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_ranking_xlsx(request):
    """Ranking de postulantes (TODO)"""
    return Response({"detail": "No implementado aun"}, status=501)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_vacants_vs_xlsx(request):
    """Vacantes vs postulantes (TODO)"""
    return Response({"detail": "No implementado aun"}, status=501)