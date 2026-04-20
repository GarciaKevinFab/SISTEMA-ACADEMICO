"""
admission/views/ingresantes_import.py

Importador de ingresantes POR FASES desde Excel del proceso de admisión.

FASE 1 — Examen general (Promedio Primera Fase / APROBADO|REPROBADO)
  → crea Applicant + Application, status = PHASE1_PASSED / PHASE1_FAILED,
    guarda nota como EvaluationScore(phase="WRITTEN"). NO crea Users.

FASE 2 — Resultados finales (Promedio Final / ALCANZÓ VACANTE | NO)
  → actualiza Application, status = ADMITTED / NOT_ADMITTED, guarda nota
    como EvaluationScore(phase="FINAL"). SOLO ADMITTED crea Student +
    User + credencial temporal.

La fase se detecta automáticamente del encabezado.
"""
import io
import logging
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from admission.models import (
    AdmissionCall, Applicant, Application, ApplicationPreference,
    EvaluationScore,
)
from students.models import Student
from students.views import _create_user_for_student
from catalogs.models import Career

try:
    from catalogs.helpers import match_career_robust, career_base_name
except ImportError:
    def match_career_robust(name):
        return Career.objects.filter(name__iexact=(name or "").strip()).first()
    def career_base_name(full_name):
        import re
        s = (full_name or "").strip().upper()
        s = re.sub(r"\s*\([^)]*\)\s*", "", s).strip()
        return s


logger = logging.getLogger("admission.ingresantes")

# ── Columnas esperadas del Excel (índices 0-based) ──
COL_CARRERA = 0
COL_DNI = 1
COL_AP_PAT = 2
COL_AP_MAT = 3
COL_NOMBRES = 4
COL_PROMEDIO = 5
COL_CONDICION = 6
COL_FECHA = 7
COL_DISCAPACIDAD = 8


def _norm_upper(s) -> str:
    s = (str(s) if s is not None else "").upper()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _is_admitted(condicion: str) -> bool:
    """Fase 2: alcanzó vacante."""
    s = _norm_upper(condicion)
    return ("ALCANZO" in s and "VACANTE" in s) or "INGRESANTE" in s or "ADMITIDO" in s


def _is_phase1_approved(condicion: str) -> bool:
    """Fase 1: aprobado."""
    s = _norm_upper(condicion)
    return "APROBADO" in s and "REPROBADO" not in s


def _to_decimal(v):
    """Convierte a Decimal, retorna None si no es número."""
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _detect_phase_from_header(header_row) -> int:
    """
    Detecta la fase del Excel por el encabezado.
    Retorna 1 (primera fase), 2 (segunda/final), o 0 si no se pudo detectar.
    """
    if not header_row:
        return 0
    joined = " ".join(_norm_upper(c) for c in header_row if c)
    if "PROMEDIO PRIMERA FASE" in joined:
        return 1
    if "PROMEDIO FINAL" in joined or "ALCANZO VACANTE" in joined or "DISCAPACIDAD" in joined:
        return 2
    return 0


def _get_or_create_default_call():
    """Crea/recupera una convocatoria por default para el año actual."""
    year = datetime.now().year
    title = f"ADMISIÓN {year} - I"
    period = f"{year}-I"
    call = AdmissionCall.objects.filter(period=period).order_by("-id").first()
    if not call:
        call = AdmissionCall.objects.filter(title__iexact=title).first()
    if not call:
        call = AdmissionCall.objects.create(
            title=title,
            period=period,
            published=True,
            meta={"source": "auto-ingresantes-import"},
        )
    return call


def _clean_str(v, default="") -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _get_or_create_applicant(dni, nombres, ap_pat, ap_mat, email=""):
    """Busca Applicant por DNI, crea si no existe."""
    full_name = " ".join(x for x in [nombres, ap_pat, ap_mat] if x).strip()
    app, created = Applicant.objects.get_or_create(
        dni=dni,
        defaults={
            "names": full_name[:120],
            "email": email or "",
            "phone": "",
        },
    )
    if not created:
        # Actualizar nombre si cambió
        if full_name and app.names != full_name:
            app.names = full_name[:120]
            app.save(update_fields=["names"])
    return app, created


def _get_or_create_application(applicant, call, career, career_name_raw,
                               modalidad="ORDINARIO", status="CREATED",
                               discapacidad_si=False,
                               nombres="", ap_pat="", ap_mat=""):
    """Busca Application por (applicant, call), crea si no existe.
    Actualiza el status al valor pedido."""
    app = Application.objects.filter(applicant=applicant, call=call).first()
    created = False
    career_display = (career.name if career else career_name_raw) or ""

    # Campos MINEDU separados para que la UI los muestre correctamente
    base_profile = {
        "nombres": nombres or (applicant.names.split(" ")[0] if applicant.names else ""),
        "apellido_paterno": ap_pat or "",
        "apellido_materno": ap_mat or "",
        "dni": applicant.dni,
        "modalidad_admision": modalidad,
    }
    if discapacidad_si:
        base_profile["discapacidad"] = "SI"

    if not app:
        app = Application.objects.create(
            applicant=applicant,
            call=call,
            career_name=career_display,
            status=status,
            data={"profile": base_profile, "source": "IMPORT_INGRESANTES"},
        )
        created = True
    else:
        app.status = status
        if career and not app.career_name:
            app.career_name = career.name
        # Fusionar profile existente con los campos MINEDU separados
        data = app.data if isinstance(app.data, dict) else {}
        profile = data.get("profile") or {}
        for k, v in base_profile.items():
            if v and not profile.get(k):
                profile[k] = v
        if discapacidad_si:
            profile["discapacidad"] = "SI"
        data["profile"] = profile
        app.data = data
        app.save(update_fields=["status", "career_name", "data"])

    # Asegurar preferencia con la carrera
    if career:
        ApplicationPreference.objects.get_or_create(
            application=app, career=career, defaults={"rank": 1}
        )
    return app, created


def _upsert_score(application, phase, promedio):
    """Crea o actualiza EvaluationScore para una fase."""
    if promedio is None:
        return
    EvaluationScore.objects.update_or_create(
        application=application,
        phase=phase,
        defaults={"total": promedio, "rubric": {"promedio": str(promedio)}},
    )


def _upsert_student(dni, nombres, ap_pat, ap_mat, career_name, discapacidad_si):
    """Crea o actualiza Student. Retorna (student, created)."""
    st = Student.objects.filter(num_documento=dni).first()
    created = False
    discap = "SI" if discapacidad_si else "NO"

    if not st:
        st = Student.objects.create(
            num_documento=dni,
            nombres=nombres,
            apellido_paterno=ap_pat,
            apellido_materno=ap_mat,
            programa_carrera=career_name,
            discapacidad=discap,
        )
        created = True
    else:
        dirty = []
        if nombres and st.nombres != nombres:
            st.nombres = nombres
            dirty.append("nombres")
        if ap_pat and st.apellido_paterno != ap_pat:
            st.apellido_paterno = ap_pat
            dirty.append("apellido_paterno")
        if ap_mat and st.apellido_materno != ap_mat:
            st.apellido_materno = ap_mat
            dirty.append("apellido_materno")
        if career_name and st.programa_carrera != career_name:
            st.programa_carrera = career_name
            dirty.append("programa_carrera")
        if st.discapacidad != discap:
            st.discapacidad = discap
            dirty.append("discapacidad")
        if dirty:
            st.save(update_fields=dirty)
    return st, created


def _generate_temp_password(length=10):
    """Contraseña temporal: 8 alfanuméricos + 1 dígito + '!'."""
    from django.utils.crypto import get_random_string
    import string
    body = get_random_string(length - 2, allowed_chars=string.ascii_letters + string.digits)
    tail = get_random_string(1, allowed_chars=string.digits) + "!"
    return body + tail


def _ensure_user_with_temp_password(student):
    """Crea un User con username=DNI y contraseña temporal aleatoria.
    Si ya tenía usuario, RESETEA su contraseña a una nueva temporal.
    Retorna (user, temp_password, was_created)."""
    from django.contrib.auth import get_user_model
    from acl.models import Role, UserRole
    User = get_user_model()

    if student.user_id:
        # Ya tiene usuario → resetear contraseña y asegurar rol STUDENT
        user = student.user
        temp_password = _generate_temp_password()
        user.set_password(temp_password)
        user.is_active = True
        user.save(update_fields=["password", "is_active"])

        student_role = Role.objects.filter(name__iexact="STUDENT").first()
        if student_role:
            UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id)

        return user, temp_password, False

    username = student.num_documento
    k = 1
    uname = username
    while User.objects.filter(username=uname).exists():
        k += 1
        uname = f"{username}-{k}"

    full_name = " ".join(x for x in [student.nombres, student.apellido_paterno, student.apellido_materno] if x).strip()
    user = User(username=uname, is_active=True, is_staff=False)

    # Email sintético si no hay
    try:
        User._meta.get_field("email")
        user.email = (student.email or "").strip().lower() or f"{uname}@no-email.local"
    except Exception:
        pass

    if hasattr(user, "full_name"):
        user.full_name = full_name
    elif hasattr(user, "first_name"):
        user.first_name = full_name[:150]

    # Contraseña temporal aleatoria
    temp_password = _generate_temp_password()
    user.set_password(temp_password)
    user.save()

    # Asignar rol STUDENT
    student_role = Role.objects.filter(name__iexact="STUDENT").first()
    if student_role:
        UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id)

    # Vincular al estudiante
    student.user = user
    student.save(update_fields=["user"])

    return user, temp_password, True


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def ingresantes_import(request):
    """
    POST /admission/ingresantes/import
    Body (multipart):
      - file: archivo .xlsx
      - phase: "1", "2" o "auto" (default "auto" — detecta del encabezado)
      - call_id: id de la convocatoria (opcional — auto si no viene)
      - modalidad: modalidad de admisión (opcional, default ORDINARIO)
      - dry_run: "1" para simular sin guardar (opcional)
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Falta archivo 'file'"}, status=400)

    call_id = request.data.get("call_id")
    call = None
    if call_id:
        try:
            call = AdmissionCall.objects.get(pk=call_id)
        except AdmissionCall.DoesNotExist:
            return Response({"detail": "Convocatoria no encontrada"}, status=404)
    else:
        call = _get_or_create_default_call()

    phase_param = _clean_str(request.data.get("phase"), "auto").lower()
    modalidad = _clean_str(request.data.get("modalidad"), "ORDINARIO")
    dry_run = str(request.data.get("dry_run", "")).lower() in ("1", "true", "yes")

    # Leer workbook
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as exc:
        return Response({"detail": f"No se pudo leer el Excel: {exc}"}, status=400)

    # Detectar fase y encabezado
    rows = []
    current_career_name = None
    header_seen = False
    detected_phase = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not header_seen:
            if row and any("CARRERA" in _norm_upper(c) for c in row if c):
                header_seen = True
                detected_phase = _detect_phase_from_header(row)
                continue
            continue

        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        carrera_raw = row[COL_CARRERA] if len(row) > COL_CARRERA else None
        if carrera_raw and str(carrera_raw).strip():
            current_career_name = str(carrera_raw).strip()

        dni_raw = row[COL_DNI] if len(row) > COL_DNI else None
        if not dni_raw:
            continue
        dni = str(dni_raw).strip().replace(".0", "")
        if not dni.isdigit():
            continue

        rows.append({
            "row_idx": row_idx,
            "carrera_raw": current_career_name,
            "dni": dni,
            "ap_pat": _clean_str(row[COL_AP_PAT] if len(row) > COL_AP_PAT else ""),
            "ap_mat": _clean_str(row[COL_AP_MAT] if len(row) > COL_AP_MAT else ""),
            "nombres": _clean_str(row[COL_NOMBRES] if len(row) > COL_NOMBRES else ""),
            "promedio": row[COL_PROMEDIO] if len(row) > COL_PROMEDIO else None,
            "condicion": _clean_str(row[COL_CONDICION] if len(row) > COL_CONDICION else ""),
            "discapacidad_si": _clean_str(row[COL_DISCAPACIDAD] if len(row) > COL_DISCAPACIDAD else "").upper() == "SI",
        })

    # Resolver fase efectiva
    if phase_param in ("1", "2"):
        phase = int(phase_param)
    else:
        phase = detected_phase
    if phase not in (1, 2):
        return Response({
            "detail": "No se pudo detectar la fase del Excel. "
                      "Envía phase=1 o phase=2 manualmente."
        }, status=400)

    # Procesar filas
    credentials = []
    errors = []
    counts = {
        "total_rows": len(rows),
        "phase": phase,
        "phase1_approved": 0,
        "phase1_failed": 0,
        "admitted": 0,
        "not_admitted": 0,
        "created_students": 0,
        "updated_students": 0,
        "created_users": 0,
        "reset_users": 0,
        "created_applicants": 0,
        "created_applications": 0,
        "updated_applications": 0,
        "scores_saved": 0,
        "errors": 0,
        "call_id": call.id,
        "call_title": call.title,
    }

    career_cache = {}

    for r in rows:
        try:
            # Determinar el estado según fase + condición
            if phase == 1:
                if _is_phase1_approved(r["condicion"]):
                    status = "PHASE1_PASSED"
                    counts["phase1_approved"] += 1
                else:
                    status = "PHASE1_FAILED"
                    counts["phase1_failed"] += 1
                score_phase = "WRITTEN"
            else:  # phase == 2
                if _is_admitted(r["condicion"]):
                    status = "ADMITTED"
                    counts["admitted"] += 1
                else:
                    status = "NOT_ADMITTED"
                    counts["not_admitted"] += 1
                score_phase = "FINAL"

            # Match de carrera
            carrera_raw = r["carrera_raw"] or ""
            if carrera_raw not in career_cache:
                career_cache[carrera_raw] = match_career_robust(carrera_raw)
            career = career_cache[carrera_raw]
            career_display = (career.name if career else career_base_name(carrera_raw)) or carrera_raw

            if dry_run:
                continue

            with transaction.atomic():
                # 1. Applicant
                applicant, ap_created = _get_or_create_applicant(
                    dni=r["dni"],
                    nombres=r["nombres"],
                    ap_pat=r["ap_pat"],
                    ap_mat=r["ap_mat"],
                )
                if ap_created:
                    counts["created_applicants"] += 1

                # 2. Application con status correcto según fase
                app, app_created = _get_or_create_application(
                    applicant=applicant,
                    call=call,
                    career=career,
                    career_name_raw=career_display,
                    modalidad=modalidad,
                    status=status,
                    discapacidad_si=r["discapacidad_si"],
                    nombres=r["nombres"],
                    ap_pat=r["ap_pat"],
                    ap_mat=r["ap_mat"],
                )
                if app_created:
                    counts["created_applications"] += 1
                else:
                    counts["updated_applications"] += 1

                # 3. Guardar score en EvaluationScore
                promedio = _to_decimal(r["promedio"])
                if promedio is not None:
                    _upsert_score(app, score_phase, promedio)
                    counts["scores_saved"] += 1

                # 4. Solo fase 2 + ADMITTED → Student + User + credencial
                if phase == 2 and status == "ADMITTED":
                    student, st_created = _upsert_student(
                        dni=r["dni"],
                        nombres=r["nombres"],
                        ap_pat=r["ap_pat"],
                        ap_mat=r["ap_mat"],
                        career_name=career_display,
                        discapacidad_si=r["discapacidad_si"],
                    )
                    if st_created:
                        counts["created_students"] += 1
                    else:
                        counts["updated_students"] += 1

                    user, password, user_created = _ensure_user_with_temp_password(student)
                    if user_created:
                        counts["created_users"] += 1
                    else:
                        counts["reset_users"] += 1

                    credentials.append({
                        "dni": r["dni"],
                        "nombres": f"{r['ap_pat']} {r['ap_mat']} {r['nombres']}".strip(),
                        "carrera": career_display,
                        "username": user.username,
                        "password": password,
                        "is_new": user_created,
                    })
        except Exception as exc:
            logger.exception("Error importando fila %s: %s", r["row_idx"], exc)
            counts["errors"] += 1
            errors.append({
                "row": r["row_idx"],
                "dni": r["dni"],
                "reason": str(exc),
            })

    return Response({
        "summary": counts,
        "credentials": credentials,
        "errors": errors,
        "dry_run": dry_run,
        "detected_phase": detected_phase,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def ingresantes_regenerate_credentials(request):
    """
    POST /admission/ingresantes/regenerate-credentials

    Regenera las credenciales (resetea contraseñas temporales) de los
    estudiantes cuyos DNIs se envíen, y retorna el archivo con usuarios
    y contraseñas nuevas.

    Body (JSON):
      {
        "dnis": ["12345678", "87654321", ...]   # lista de DNIs (opcional)
        "call_id": 23                            # o una convocatoria entera
      }

    Si se envía `call_id` sin `dnis`, procesa TODOS los estudiantes con
    Application ADMITTED en esa convocatoria.
    """
    data = request.data or {}
    dnis = data.get("dnis") or []
    call_id = data.get("call_id")

    if not dnis and not call_id:
        return Response(
            {"detail": "Envía 'dnis' (lista) o 'call_id' (convocatoria)"},
            status=400,
        )

    # Normalizar DNIs
    if isinstance(dnis, str):
        dnis = [d.strip() for d in dnis.replace(",", "\n").split("\n") if d.strip()]
    dnis = [str(d).strip() for d in dnis if str(d).strip()]

    # Si viene call_id, obtener los DNIs de los admitidos de esa convocatoria
    if call_id and not dnis:
        admitted_apps = (
            Application.objects
            .filter(call_id=call_id, status="ADMITTED")
            .select_related("applicant")
        )
        dnis = list({app.applicant.dni for app in admitted_apps if app.applicant and app.applicant.dni})

    if not dnis:
        return Response({"detail": "No hay DNIs para procesar"}, status=400)

    credentials = []
    errors = []
    counts = {
        "total_requested": len(dnis),
        "found": 0,
        "created_users": 0,
        "reset_users": 0,
        "not_found": 0,
        "errors": 0,
    }

    counts["created_students"] = 0

    for dni in dnis:
        try:
            student = Student.objects.filter(num_documento=dni).first()

            # Si no hay Student, intentar crearlo desde Applicant + Application ADMITTED
            if not student:
                applicant = Applicant.objects.filter(dni=dni).first()
                if not applicant:
                    counts["not_found"] += 1
                    errors.append({"dni": dni, "reason": "No existe como postulante ni estudiante"})
                    continue

                # Buscar la Application más reciente (preferencia: ADMITTED)
                app = (
                    Application.objects
                    .filter(applicant=applicant)
                    .select_related("call")
                    .prefetch_related("preferences__career")
                    .order_by("-id")
                    .first()
                )
                # Extraer nombres
                full = applicant.names or ""
                parts = full.split()
                if len(parts) >= 3:
                    ap_pat, ap_mat = parts[0], parts[1]
                    nombres_only = " ".join(parts[2:])
                elif len(parts) == 2:
                    ap_pat, ap_mat, nombres_only = parts[0], "", parts[1]
                else:
                    ap_pat, ap_mat, nombres_only = "", "", full

                # Carrera del Application
                career_name = ""
                if app:
                    career_name = app.career_name or ""
                    if not career_name:
                        fp = app.preferences.order_by("rank").first()
                        if fp and fp.career:
                            career_name = fp.career.name

                # Discapacidad del profile JSON
                discap_flag = False
                if app and isinstance(app.data, dict):
                    prof = app.data.get("profile") or {}
                    dv = str(prof.get("discapacidad", "")).strip().upper()
                    discap_flag = dv in ("SI", "YES", "TRUE", "1")

                student, _ = _upsert_student(
                    dni=dni,
                    nombres=nombres_only,
                    ap_pat=ap_pat,
                    ap_mat=ap_mat,
                    career_name=career_name,
                    discapacidad_si=discap_flag,
                )
                counts["created_students"] += 1

                # Marcar Application como ADMITTED si no lo estaba (ya que el admin está
                # regenerando credenciales → se asume ingresante)
                if app and app.status != "ADMITTED":
                    app.status = "ADMITTED"
                    app.save(update_fields=["status"])

            counts["found"] += 1
            career_display = student.programa_carrera or ""

            user, password, was_created = _ensure_user_with_temp_password(student)
            if was_created:
                counts["created_users"] += 1
            else:
                counts["reset_users"] += 1

            full_name = " ".join(x for x in [
                student.apellido_paterno, student.apellido_materno, student.nombres
            ] if x).strip()

            credentials.append({
                "dni": dni,
                "nombres": full_name,
                "carrera": career_display,
                "username": user.username,
                "password": password,
                "is_new": was_created,
            })
        except Exception as exc:
            logger.exception("Error regenerando credencial DNI %s: %s", dni, exc)
            counts["errors"] += 1
            errors.append({"dni": dni, "reason": str(exc)})

    return Response({
        "summary": counts,
        "credentials": credentials,
        "errors": errors,
    })
