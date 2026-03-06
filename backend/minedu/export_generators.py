"""
export_generators.py — Generadores de documentos SIA para IESPP

Genera archivos Excel (XLSX) y CSV con formato oficial del SIA / MINEDU
para uso institucional, archivo y referencia durante la carga manual al SIA.

Documentos soportados:
  1. ENROLLMENT   — Nómina de Matrícula
  2. FICHA        — Ficha de Matrícula (individual)
  3. BOLETA       — Boleta de Notas (individual)
  4. ACTA         — Acta Consolidada de Evaluación
  5. REPORTE      — Reporte de Información del Sistema (Kardex)
  6. REGISTRO_AUX — Registro Auxiliar de Evaluación
  7. CERTIFICADO  — Certificado de Estudios

Referencia normativa:
  - RVM N° 123-2022-MINEDU (evaluación formativa, escala vigesimal referencial)
  - RVM N° 039-2023-MINEDU (modificatoria)
  - Sistema SIA: sia.pedagogicos.pe

Campos marcados con # AJUSTAR: pueden necesitar adaptación a los nombres
reales de tus modelos Django.
"""

import csv
import io
from datetime import date, datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle,
)
from openpyxl.utils import get_column_letter
from django.db.models import Q

# ── Modelos ──
from catalogs.models import Career
from academic.models import (
    Plan,
    PlanCourse,         # AJUSTAR: nombre del modelo curso↔plan
    Enrollment,
    EnrollmentItem,     # AJUSTAR: nombre del modelo item/detalle de matrícula
    AcademicGradeRecord,
    InstitutionSettings,
)
from students.models import Student
from minedu.models import MineduCatalogMapping


# ═══════════════════════════════════════════════════════════
# Constantes y helpers
# ═══════════════════════════════════════════════════════════

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_BOTTOM = Border(bottom=_THIN)

_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_SUBHEADER = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_FILL_LIGHT = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

_FONT_TITLE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_FONT_SUBTITLE = Font(name="Arial", size=10, bold=True)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_NORMAL = Font(name="Arial", size=9)
_FONT_SMALL = Font(name="Arial", size=8)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _nota_letra(nota):
    """
    Convierte nota vigesimal a escala descriptiva (RVM 123-2022-MINEDU).
      P: Previo al inicio (< 11)
      I: Inicio (11-12)
      E: En proceso (13-14)
      L: Logrado (15-17)
      D: Destacado (18-20)
    """
    if nota is None:
        return ""
    nota = int(nota)
    if nota < 11:
        return "P"
    elif nota <= 12:
        return "I"
    elif nota <= 14:
        return "E"
    elif nota <= 17:
        return "L"
    else:
        return "D"


def _nota_estado(nota):
    """Devuelve condición: Aprobado / Desaprobado."""
    if nota is None:
        return ""
    return "Aprobado" if int(nota) >= 11 else "Desaprobado"


def _get_institution():
    """Datos de la institución."""
    try:
        inst = InstitutionSettings.objects.first()
        if inst:
            return {
                "name": inst.name or "IESPP",
                "ruc": getattr(inst, "ruc", "") or "",
                "address": getattr(inst, "address", "") or "",
                "dre": getattr(inst, "dre", "") or "",
                "ugel": getattr(inst, "ugel", "") or "",
                "codigo_modular": getattr(inst, "codigo_modular", "") or "",
                "director": getattr(inst, "director_name", "") or "",
                "secretario": getattr(inst, "academic_secretary_name", "") or "",
            }
    except Exception:
        pass
    return {
        "name": "IESPP", "ruc": "", "address": "", "dre": "", "ugel": "",
        "codigo_modular": "", "director": "", "secretario": "",
    }


def _get_minedu_code(catalog_type, local_id):
    """Busca código MINEDU mapeado para un registro local."""
    try:
        m = MineduCatalogMapping.objects.filter(
            type=catalog_type, local_id=local_id
        ).first()
        return m.minedu_code if m and m.minedu_code else ""
    except Exception:
        return ""


def _parse_period(period_code):
    """
    '2024-I' → (2024, 'I')
    '2024-II' → (2024, 'II')
    """
    parts = period_code.split("-", 1)
    year = int(parts[0])
    period = parts[1] if len(parts) > 1 else "I"
    return year, period


def _sia_header_rows(ws, title, period_code, inst, start_row=1, max_col=10):
    """
    Escribe encabezado estilo SIA en la hoja.
    Retorna la fila siguiente disponible.
    """
    year, period = _parse_period(period_code)

    # Fila 1: Título institucional
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=max_col
    )
    cell = ws.cell(row=start_row, column=1)
    cell.value = f"MINISTERIO DE EDUCACIÓN - {inst['dre'] or 'DRE'}".upper()
    cell.font = _FONT_TITLE
    cell.fill = _FILL_HEADER
    cell.alignment = _ALIGN_CENTER

    # Fila 2: Nombre institución
    r2 = start_row + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=max_col)
    cell = ws.cell(row=r2, column=1)
    cell.value = inst["name"].upper()
    cell.font = _FONT_TITLE
    cell.fill = _FILL_HEADER
    cell.alignment = _ALIGN_CENTER

    # Fila 3: Título del documento
    r3 = start_row + 2
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=max_col)
    cell = ws.cell(row=r3, column=1)
    cell.value = title.upper()
    cell.font = Font(name="Arial", size=12, bold=True)
    cell.alignment = _ALIGN_CENTER

    # Fila 4: Período
    r4 = start_row + 3
    ws.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=max_col)
    cell = ws.cell(row=r4, column=1)
    cell.value = f"Período Académico {year} - {period}"
    cell.font = _FONT_SUBTITLE
    cell.alignment = _ALIGN_CENTER

    # Fila 5: Código modular y datos
    r5 = start_row + 4
    ws.cell(row=r5, column=1).value = f"Código Modular: {inst['codigo_modular']}"
    ws.cell(row=r5, column=1).font = _FONT_SMALL
    ws.cell(row=r5, column=max_col - 2).value = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
    ws.cell(row=r5, column=max_col - 2).font = _FONT_SMALL

    return start_row + 6


def _signature_rows(ws, inst, row, max_col):
    """Escribe bloque de firmas: Director, Secretario Académico, Especialista DRE."""
    row += 2  # espacio
    sig_cols = [2, max_col // 2, max_col - 2]
    titles = [
        ("_" * 30, inst["director"] or "Director General", "Director General"),
        ("_" * 30, inst["secretario"] or "Secretario(a) Académico(a)", "Secretario(a) Académico(a)"),
        ("_" * 30, "Especialista DRE", "Especialista DRE"),
    ]
    for col_idx, (line, name, cargo) in zip(sig_cols, titles):
        ws.cell(row=row, column=col_idx).value = line
        ws.cell(row=row, column=col_idx).font = _FONT_SMALL
        ws.cell(row=row, column=col_idx).alignment = _ALIGN_CENTER
        ws.cell(row=row + 1, column=col_idx).value = name
        ws.cell(row=row + 1, column=col_idx).font = _FONT_SMALL
        ws.cell(row=row + 1, column=col_idx).alignment = _ALIGN_CENTER
        ws.cell(row=row + 2, column=col_idx).value = cargo
        ws.cell(row=row + 2, column=col_idx).font = _FONT_SMALL
        ws.cell(row=row + 2, column=col_idx).alignment = _ALIGN_CENTER

    return row + 4


def _apply_table_style(ws, header_row, data_start, data_end, max_col):
    """Aplica bordes y estilos de tabla al rango de datos."""
    # Header
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_SUBHEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_ALL

    # Data rows
    for r in range(data_start, data_end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_ALL
            cell.font = _FONT_NORMAL
            if c == 1:
                cell.alignment = _ALIGN_CENTER
            # Zebra striping
            if (r - data_start) % 2 == 1:
                cell.fill = _FILL_LIGHT


def _workbook_to_bytes(wb):
    """Serializa workbook a bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _get_enrollments(period_code):
    """Obtiene matrículas confirmadas del período."""
    year, period = _parse_period(period_code)
    return (
        Enrollment.objects
        .filter(
            academic_year=year,
            academic_period=period,     # AJUSTAR: campo de período en tu modelo
            status="CONFIRMED",
        )
        .select_related("student", "plan", "plan__career")
        .order_by("student__apellido_paterno", "student__apellido_materno", "student__nombres")
    )


def _student_fullname(student):
    """Nombre completo: Apellido Paterno Apellido Materno, Nombres."""
    ap = (student.apellido_paterno or "").strip()
    am = (student.apellido_materno or "").strip()
    n = (student.nombres or "").strip()
    return f"{ap} {am}, {n}".strip(", ")


def _get_grades_for_enrollment(enrollment):
    """
    Retorna lista de dicts con la info de cada curso + nota del estudiante.
    AJUSTAR: según tus modelos reales (EnrollmentItem, AcademicGradeRecord, etc.)
    """
    items = []
    try:
        enrollment_items = EnrollmentItem.objects.filter(
            enrollment=enrollment
        ).select_related("plan_course", "plan_course__course")

        for ei in enrollment_items:
            pc = ei.plan_course
            course = pc.course if pc else None
            course_name = course.name if course else "Sin nombre"
            course_code = course.code if course else ""
            credits = getattr(pc, "credits", 0) or 0
            hours = getattr(pc, "hours_per_week", 0) or getattr(pc, "hours", 0) or 0

            # Buscar nota
            grade_record = AcademicGradeRecord.objects.filter(
                student=enrollment.student,
                # AJUSTAR: filtro correcto para vincular nota al item
                plan_course=pc,
                academic_year=enrollment.academic_year,
            ).first()

            nota = None
            if grade_record:
                nota = grade_record.final_grade

            puntaje = (int(nota) * credits) if (nota is not None and credits) else 0

            items.append({
                "code": course_code,
                "name": course_name,
                "credits": credits,
                "hours": hours,
                "nota": nota,
                "nota_int": int(nota) if nota is not None else None,
                "nota_letra": _nota_letra(nota),
                "estado": _nota_estado(nota),
                "puntaje": puntaje,
                "is_subsanacion": getattr(ei, "is_subsanacion", False),
            })
    except Exception as exc:
        # Si EnrollmentItem no existe, intenta con PlanCourse directo
        try:
            plan = enrollment.plan
            if plan:
                plan_courses = PlanCourse.objects.filter(
                    plan=plan
                ).select_related("course")
                for pc in plan_courses:
                    course = pc.course
                    credits = getattr(pc, "credits", 0) or 0
                    hours = getattr(pc, "hours_per_week", 0) or getattr(pc, "hours", 0) or 0

                    grade_record = AcademicGradeRecord.objects.filter(
                        student=enrollment.student,
                        plan_course=pc,
                        academic_year=enrollment.academic_year,
                    ).first()

                    nota = grade_record.final_grade if grade_record else None
                    puntaje = (int(nota) * credits) if (nota is not None and credits) else 0

                    items.append({
                        "code": course.code if course else "",
                        "name": course.name if course else "",
                        "credits": credits,
                        "hours": hours,
                        "nota": nota,
                        "nota_int": int(nota) if nota is not None else None,
                        "nota_letra": _nota_letra(nota),
                        "estado": _nota_estado(nota),
                        "puntaje": puntaje,
                        "is_subsanacion": False,
                    })
        except Exception:
            pass

    return items


# ═══════════════════════════════════════════════════════════
# 1. NÓMINA DE MATRÍCULA
# ═══════════════════════════════════════════════════════════

def _generate_enrollment_xlsx(period_code):
    """
    Nómina de Matrícula: padrón general de estudiantes matriculados.
    Columnas: N°, Apellidos y Nombres, DNI, Carrera, Plan, Ciclo, Estado
    """
    inst = _get_institution()
    enrollments = _get_enrollments(period_code)
    total = enrollments.count()

    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina de Matrícula"
    max_col = 8

    # Anchos
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    row = _sia_header_rows(ws, "Nómina de Matrícula", period_code, inst, max_col=max_col)

    # Encabezado tabla
    headers = ["N°", "Apellidos y Nombres", "N° Documento", "Programa de Estudios",
               "Plan de Estudios", "Ciclo", "Cód. MINEDU", "Condición"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c).value = h
    header_row = row
    row += 1

    data_start = row
    for idx, enr in enumerate(enrollments, 1):
        s = enr.student
        career = enr.plan.career if enr.plan else None
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = _student_fullname(s)
        ws.cell(row=row, column=3).value = s.num_documento or ""
        ws.cell(row=row, column=4).value = career.name if career else ""
        ws.cell(row=row, column=5).value = enr.plan.name if enr.plan else ""
        ws.cell(row=row, column=6).value = getattr(enr, "cycle", "") or getattr(enr, "semester", "") or ""
        ws.cell(row=row, column=7).value = _get_minedu_code("STUDENT", s.id)
        ws.cell(row=row, column=8).value = enr.status or "CONFIRMADO"
        row += 1

    _apply_table_style(ws, header_row, data_start, row - 1, max_col)

    # Total
    ws.cell(row=row + 1, column=1).value = f"Total de estudiantes matriculados: {total}"
    ws.cell(row=row + 1, column=1).font = _FONT_SUBTITLE

    _signature_rows(ws, inst, row + 3, max_col)

    filename = f"nomina_matricula_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_enrollment_csv(period_code):
    enrollments = _get_enrollments(period_code)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["N°", "Apellidos y Nombres", "N° Documento", "Programa",
                     "Plan", "Ciclo", "Cód. MINEDU", "Condición"])
    for idx, enr in enumerate(enrollments, 1):
        s = enr.student
        career = enr.plan.career if enr.plan else None
        writer.writerow([
            idx, _student_fullname(s), s.num_documento or "",
            career.name if career else "", enr.plan.name if enr.plan else "",
            getattr(enr, "cycle", "") or "", _get_minedu_code("STUDENT", s.id),
            enr.status or "",
        ])
    content = buf.getvalue().encode("utf-8-sig")
    return f"nomina_matricula_{period_code}.csv", content, enrollments.count()


# ═══════════════════════════════════════════════════════════
# 2. FICHA DE MATRÍCULA (individual, una hoja por alumno)
# ═══════════════════════════════════════════════════════════

def _generate_ficha_xlsx(period_code):
    """
    Ficha de Matrícula individual: datos del alumno + listado de cursos
    matriculados con horas y créditos + sección subsanación.
    """
    inst = _get_institution()
    enrollments = _get_enrollments(period_code)
    total = enrollments.count()

    wb = Workbook()
    # Eliminar hoja default si hay alumnos
    if total > 0 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for enr in enrollments:
        s = enr.student
        career = enr.plan.career if enr.plan else None
        safe_name = _student_fullname(s)[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        max_col = 8

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12

        row = _sia_header_rows(ws, "Ficha de Matrícula", period_code, inst, max_col=max_col)

        # Datos del alumno
        data_pairs = [
            ("Estudiante:", _student_fullname(s)),
            ("N° Documento:", s.num_documento or ""),
            ("Programa:", career.name if career else ""),
            ("Plan:", enr.plan.name if enr.plan else ""),
            ("Ciclo:", getattr(enr, "cycle", "") or getattr(enr, "semester", "") or ""),
            ("Cód. MINEDU:", _get_minedu_code("STUDENT", s.id)),
        ]
        for label, val in data_pairs:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _FONT_HEADER
            ws.cell(row=row, column=2).value = val
            ws.cell(row=row, column=2).font = _FONT_NORMAL
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        row += 1  # espacio

        # Tabla de cursos
        headers = ["N°", "Código", "Área / Curso", "Tipo", "Horas", "Créditos", "Condición", "Observación"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c).value = h
        header_row = row
        row += 1

        items = _get_grades_for_enrollment(enr)
        regular_items = [i for i in items if not i.get("is_subsanacion")]
        subsanacion_items = [i for i in items if i.get("is_subsanacion")]

        data_start = row
        total_hours = 0
        total_credits = 0
        for idx, item in enumerate(regular_items, 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = item["code"]
            ws.cell(row=row, column=3).value = item["name"]
            ws.cell(row=row, column=4).value = "OB"  # AJUSTAR: tipo (OB/EL/OP)
            ws.cell(row=row, column=5).value = item["hours"]
            ws.cell(row=row, column=6).value = item["credits"]
            ws.cell(row=row, column=7).value = "Matriculado"
            ws.cell(row=row, column=8).value = ""
            total_hours += item["hours"]
            total_credits += item["credits"]
            row += 1

        _apply_table_style(ws, header_row, data_start, row - 1, max_col)

        # Totales
        ws.cell(row=row, column=4).value = "TOTAL"
        ws.cell(row=row, column=4).font = _FONT_HEADER
        ws.cell(row=row, column=5).value = total_hours
        ws.cell(row=row, column=5).font = _FONT_HEADER
        ws.cell(row=row, column=6).value = total_credits
        ws.cell(row=row, column=6).font = _FONT_HEADER
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = _BORDER_ALL
        row += 2

        # Sección Subsanación
        ws.cell(row=row, column=1).value = "CURSOS DE SUBSANACIÓN"
        ws.cell(row=row, column=1).font = _FONT_SUBTITLE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        row += 1

        sub_headers = ["N°", "Código", "Curso de Subsanación", "", "Horas", "Créditos", "Período", ""]
        for c, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=c).value = h
        sub_header_row = row
        row += 1

        if subsanacion_items:
            sub_start = row
            for idx, item in enumerate(subsanacion_items, 1):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = item["code"]
                ws.cell(row=row, column=3).value = item["name"]
                ws.cell(row=row, column=5).value = item["hours"]
                ws.cell(row=row, column=6).value = item["credits"]
                ws.cell(row=row, column=7).value = period_code
                row += 1
            _apply_table_style(ws, sub_header_row, sub_start, row - 1, max_col)
        else:
            # Filas vacías
            for _ in range(2):
                for c in range(1, max_col + 1):
                    ws.cell(row=row, column=c).border = _BORDER_ALL
                row += 1
            _apply_table_style(ws, sub_header_row, row - 2, row - 1, max_col)

        _signature_rows(ws, inst, row + 1, max_col)

    if total == 0:
        ws = wb.active
        ws.cell(row=1, column=1).value = "No hay matrículas confirmadas para este período."

    filename = f"fichas_matricula_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_ficha_csv(period_code):
    enrollments = _get_enrollments(period_code)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Estudiante", "DNI", "Programa", "Plan", "Ciclo",
                     "Código Curso", "Curso", "Horas", "Créditos", "Subsanación"])
    count = 0
    for enr in enrollments:
        s = enr.student
        career = enr.plan.career if enr.plan else None
        items = _get_grades_for_enrollment(enr)
        for item in items:
            writer.writerow([
                _student_fullname(s), s.num_documento or "",
                career.name if career else "", enr.plan.name if enr.plan else "",
                getattr(enr, "cycle", "") or "",
                item["code"], item["name"], item["hours"], item["credits"],
                "Sí" if item.get("is_subsanacion") else "No",
            ])
        count += 1
    content = buf.getvalue().encode("utf-8-sig")
    return f"fichas_matricula_{period_code}.csv", content, count


# ═══════════════════════════════════════════════════════════
# 3. BOLETA DE NOTAS (individual, una hoja por alumno)
# ═══════════════════════════════════════════════════════════

def _generate_boleta_xlsx(period_code):
    """
    Boleta de Notas: notas individuales con puntaje y promedio ponderado.
    Puntaje = Nota × Créditos
    Promedio Ponderado = ΣPuntajes / ΣCréditos
    """
    inst = _get_institution()
    enrollments = _get_enrollments(period_code)
    total = enrollments.count()

    wb = Workbook()
    if total > 0 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for enr in enrollments:
        s = enr.student
        career = enr.plan.career if enr.plan else None
        safe_name = _student_fullname(s)[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        max_col = 10

        col_widths = [5, 12, 30, 10, 8, 8, 10, 10, 12, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _sia_header_rows(ws, "Boleta de Notas", period_code, inst, max_col=max_col)

        # Datos alumno
        info = [
            ("Estudiante:", _student_fullname(s)),
            ("N° Documento:", s.num_documento or ""),
            ("Programa:", career.name if career else ""),
            ("Ciclo:", getattr(enr, "cycle", "") or ""),
        ]
        for label, val in info:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _FONT_HEADER
            ws.cell(row=row, column=3).value = val
            ws.cell(row=row, column=3).font = _FONT_NORMAL
            row += 1
        row += 1

        # Tabla de notas
        headers = ["N°", "Código", "Área / Curso", "Tipo", "Horas", "Créd.",
                   "Calificación", "Nivel", "Puntaje", "Condición"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c).value = h
        header_row = row
        row += 1

        items = _get_grades_for_enrollment(enr)
        regular_items = [i for i in items if not i.get("is_subsanacion")]
        subsanacion_items = [i for i in items if i.get("is_subsanacion")]

        data_start = row
        sum_puntaje = 0
        sum_creditos = 0
        for idx, item in enumerate(regular_items, 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = item["code"]
            ws.cell(row=row, column=3).value = item["name"]
            ws.cell(row=row, column=4).value = "OB"  # AJUSTAR
            ws.cell(row=row, column=5).value = item["hours"]
            ws.cell(row=row, column=6).value = item["credits"]
            ws.cell(row=row, column=7).value = item["nota_int"] if item["nota_int"] is not None else ""
            ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
            ws.cell(row=row, column=8).value = item["nota_letra"]
            ws.cell(row=row, column=8).alignment = _ALIGN_CENTER
            ws.cell(row=row, column=9).value = item["puntaje"]
            ws.cell(row=row, column=9).alignment = _ALIGN_CENTER
            ws.cell(row=row, column=10).value = item["estado"]
            sum_puntaje += item["puntaje"]
            sum_creditos += item["credits"]
            row += 1

        _apply_table_style(ws, header_row, data_start, row - 1, max_col)

        # Totales y promedio
        promedio = round(sum_puntaje / sum_creditos, 2) if sum_creditos > 0 else 0
        ws.cell(row=row, column=5).value = "TOTAL"
        ws.cell(row=row, column=5).font = _FONT_HEADER
        ws.cell(row=row, column=6).value = sum_creditos
        ws.cell(row=row, column=6).font = _FONT_HEADER
        ws.cell(row=row, column=8).value = "Σ Puntaje:"
        ws.cell(row=row, column=8).font = _FONT_HEADER
        ws.cell(row=row, column=9).value = sum_puntaje
        ws.cell(row=row, column=9).font = _FONT_HEADER
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = _BORDER_ALL
        row += 1

        ws.cell(row=row, column=8).value = "Promedio Ponderado:"
        ws.cell(row=row, column=8).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=9).value = promedio
        ws.cell(row=row, column=9).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=10).value = _nota_letra(round(promedio))
        ws.cell(row=row, column=10).font = Font(name="Arial", size=10, bold=True)
        row += 2

        # Sección subsanación
        ws.cell(row=row, column=1).value = "CURSOS DE SUBSANACIÓN"
        ws.cell(row=row, column=1).font = _FONT_SUBTITLE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        row += 1

        sub_headers = ["N°", "Código", "Curso", "", "Horas", "Créd.",
                       "Calificación", "Nivel", "Puntaje", "Condición"]
        for c, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=c).value = h
        sub_header = row
        row += 1

        if subsanacion_items:
            sub_start = row
            for idx, item in enumerate(subsanacion_items, 1):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = item["code"]
                ws.cell(row=row, column=3).value = item["name"]
                ws.cell(row=row, column=5).value = item["hours"]
                ws.cell(row=row, column=6).value = item["credits"]
                ws.cell(row=row, column=7).value = item["nota_int"] if item["nota_int"] is not None else ""
                ws.cell(row=row, column=8).value = item["nota_letra"]
                ws.cell(row=row, column=9).value = item["puntaje"]
                ws.cell(row=row, column=10).value = item["estado"]
                row += 1
            _apply_table_style(ws, sub_header, sub_start, row - 1, max_col)
        else:
            for _ in range(2):
                for c in range(1, max_col + 1):
                    ws.cell(row=row, column=c).border = _BORDER_ALL
                row += 1
            _apply_table_style(ws, sub_header, row - 2, row - 1, max_col)

        _signature_rows(ws, inst, row + 1, max_col)

    if total == 0:
        ws = wb.active
        ws.cell(row=1, column=1).value = "No hay matrículas para este período."

    filename = f"boletas_notas_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_boleta_csv(period_code):
    enrollments = _get_enrollments(period_code)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Estudiante", "DNI", "Programa", "Código Curso", "Curso",
                     "Créditos", "Nota", "Nivel", "Puntaje", "Condición", "Promedio Ponderado"])
    count = 0
    for enr in enrollments:
        s = enr.student
        career = enr.plan.career if enr.plan else None
        items = _get_grades_for_enrollment(enr)
        sum_p = sum(i["puntaje"] for i in items)
        sum_c = sum(i["credits"] for i in items)
        prom = round(sum_p / sum_c, 2) if sum_c > 0 else 0
        for item in items:
            writer.writerow([
                _student_fullname(s), s.num_documento or "",
                career.name if career else "",
                item["code"], item["name"], item["credits"],
                item["nota_int"] if item["nota_int"] is not None else "",
                item["nota_letra"], item["puntaje"], item["estado"], prom,
            ])
        count += 1
    content = buf.getvalue().encode("utf-8-sig")
    return f"boletas_notas_{period_code}.csv", content, count


# ═══════════════════════════════════════════════════════════
# 4. ACTA CONSOLIDADA DE EVALUACIÓN
# ═══════════════════════════════════════════════════════════

def _generate_acta_xlsx(period_code):
    """
    Acta Consolidada de Evaluación: tabla alumnos × asignaturas.
    Cada asignatura tiene 3 subcolumnas: C (calificación), CS (condición), PTJ (puntaje).
    Al final: Σ créditos, Σ puntaje, promedio ponderado.
    """
    inst = _get_institution()
    enrollments = list(_get_enrollments(period_code))
    total = len(enrollments)

    if total == 0:
        wb = Workbook()
        wb.active.cell(row=1, column=1).value = "No hay matrículas para este período."
        return f"acta_consolidada_{period_code}.xlsx", _workbook_to_bytes(wb), 0

    # Recopilar cursos del plan (asumimos todos tienen el mismo plan)
    first_plan = enrollments[0].plan if enrollments else None
    courses = []
    if first_plan:
        plan_courses = PlanCourse.objects.filter(
            plan=first_plan
        ).select_related("course").order_by("order", "id")  # AJUSTAR: campo de orden
        for pc in plan_courses:
            courses.append({
                "id": pc.id,
                "code": pc.course.code if pc.course else "",
                "name": pc.course.name if pc.course else "",
                "credits": getattr(pc, "credits", 0) or 0,
            })

    num_courses = len(courses)
    # Columnas: N° | Apellidos y Nombres | DNI | [C,CS,PTJ]×cursos | Créd.Total | Ptje.Total | Promedio
    fixed_cols = 3
    course_cols = num_courses * 3
    summary_cols = 3  # créditos total, puntaje total, promedio
    max_col = fixed_cols + course_cols + summary_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Acta Consolidada"

    # Anchos
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12

    row = _sia_header_rows(ws, "Acta Consolidada de Evaluación", period_code, inst, max_col=max_col)

    # Info de carrera/plan
    if first_plan and first_plan.career:
        ws.cell(row=row, column=1).value = f"Programa: {first_plan.career.name}"
        ws.cell(row=row, column=1).font = _FONT_HEADER
        ws.cell(row=row, column=5).value = f"Plan: {first_plan.name}"
        ws.cell(row=row, column=5).font = _FONT_HEADER
        row += 1

    # Fila de nombres de curso (merge 3 columnas cada uno)
    course_name_row = row
    for i, course in enumerate(courses):
        col_start = fixed_cols + 1 + i * 3
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_start + 2
        )
        cell = ws.cell(row=row, column=col_start)
        cell.value = course["name"][:20]
        cell.font = _FONT_SMALL
        cell.alignment = _ALIGN_CENTER
        cell.fill = _FILL_SUBHEADER
        cell.border = _BORDER_ALL
        # Set width
        for j in range(3):
            ws.column_dimensions[get_column_letter(col_start + j)].width = 5

    # Summary headers
    sum_col = fixed_cols + course_cols + 1
    ws.cell(row=row, column=sum_col).value = "Créd."
    ws.cell(row=row, column=sum_col).font = _FONT_SMALL
    ws.cell(row=row, column=sum_col).alignment = _ALIGN_CENTER
    ws.cell(row=row, column=sum_col + 1).value = "Ptje."
    ws.cell(row=row, column=sum_col + 1).font = _FONT_SMALL
    ws.cell(row=row, column=sum_col + 1).alignment = _ALIGN_CENTER
    ws.cell(row=row, column=sum_col + 2).value = "Prom."
    ws.cell(row=row, column=sum_col + 2).font = _FONT_SMALL
    ws.cell(row=row, column=sum_col + 2).alignment = _ALIGN_CENTER
    for c in range(sum_col, sum_col + 3):
        ws.cell(row=row, column=c).fill = _FILL_SUBHEADER
        ws.cell(row=row, column=c).border = _BORDER_ALL
        ws.column_dimensions[get_column_letter(c)].width = 7

    row += 1

    # Subheader: N° | Apellidos | DNI | C CS PTJ × n | totals
    headers_fixed = ["N°", "Apellidos y Nombres", "DNI"]
    for c, h in enumerate(headers_fixed, 1):
        ws.cell(row=row, column=c).value = h
    for i in range(num_courses):
        col_start = fixed_cols + 1 + i * 3
        ws.cell(row=row, column=col_start).value = "C"
        ws.cell(row=row, column=col_start + 1).value = "CS"
        ws.cell(row=row, column=col_start + 2).value = "PTJ"
    header_row = row
    # Apply header styling
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _FONT_SMALL
        cell.fill = _FILL_SUBHEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_ALL
    row += 1

    # Fila de créditos
    ws.cell(row=row, column=1).value = ""
    ws.cell(row=row, column=2).value = "Créditos →"
    ws.cell(row=row, column=2).font = _FONT_SMALL
    for i, course in enumerate(courses):
        col_start = fixed_cols + 1 + i * 3
        ws.cell(row=row, column=col_start + 2).value = course["credits"]
        ws.cell(row=row, column=col_start + 2).font = _FONT_SMALL
        ws.cell(row=row, column=col_start + 2).alignment = _ALIGN_CENTER
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = _BORDER_ALL
    row += 1

    # Data rows
    data_start = row
    for idx, enr in enumerate(enrollments, 1):
        s = enr.student
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = _student_fullname(s)
        ws.cell(row=row, column=3).value = s.num_documento or ""

        sum_puntaje = 0
        sum_creditos = 0

        for i, course in enumerate(courses):
            col_start = fixed_cols + 1 + i * 3
            grade = AcademicGradeRecord.objects.filter(
                student=s,
                plan_course_id=course["id"],
                academic_year=enr.academic_year,
            ).first()

            nota = grade.final_grade if grade else None
            nota_int = int(nota) if nota is not None else None
            credits = course["credits"]
            puntaje = (nota_int * credits) if nota_int is not None else 0

            # C = calificación
            ws.cell(row=row, column=col_start).value = nota_int if nota_int is not None else ""
            ws.cell(row=row, column=col_start).alignment = _ALIGN_CENTER
            # CS = condición (A/D)
            if nota_int is not None:
                ws.cell(row=row, column=col_start + 1).value = "A" if nota_int >= 11 else "D"
            ws.cell(row=row, column=col_start + 1).alignment = _ALIGN_CENTER
            # PTJ = puntaje
            ws.cell(row=row, column=col_start + 2).value = puntaje if nota_int is not None else ""
            ws.cell(row=row, column=col_start + 2).alignment = _ALIGN_CENTER

            sum_puntaje += puntaje
            sum_creditos += credits

        # Summary cols
        promedio = round(sum_puntaje / sum_creditos, 2) if sum_creditos > 0 else 0
        ws.cell(row=row, column=sum_col).value = sum_creditos
        ws.cell(row=row, column=sum_col).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=sum_col + 1).value = sum_puntaje
        ws.cell(row=row, column=sum_col + 1).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=sum_col + 2).value = promedio
        ws.cell(row=row, column=sum_col + 2).alignment = _ALIGN_CENTER

        # Borders
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = _BORDER_ALL
            ws.cell(row=row, column=c).font = _FONT_SMALL
            # Zebra
            if idx % 2 == 0:
                ws.cell(row=row, column=c).fill = _FILL_LIGHT

        row += 1

    _signature_rows(ws, inst, row + 2, max_col)

    filename = f"acta_consolidada_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_acta_csv(period_code):
    enrollments = list(_get_enrollments(period_code))
    if not enrollments:
        buf = io.StringIO()
        csv.writer(buf).writerow(["Sin datos"])
        return f"acta_consolidada_{period_code}.csv", buf.getvalue().encode("utf-8-sig"), 0

    first_plan = enrollments[0].plan
    plan_courses = list(PlanCourse.objects.filter(
        plan=first_plan
    ).select_related("course").order_by("order", "id")) if first_plan else []

    buf = io.StringIO()
    writer = csv.writer(buf)

    header = ["N°", "Estudiante", "DNI"]
    for pc in plan_courses:
        name = pc.course.name if pc.course else ""
        header += [f"{name} C", f"{name} CS", f"{name} PTJ"]
    header += ["Créditos", "Puntaje", "Promedio"]
    writer.writerow(header)

    for idx, enr in enumerate(enrollments, 1):
        s = enr.student
        row_data = [idx, _student_fullname(s), s.num_documento or ""]
        sum_p = sum_c = 0
        for pc in plan_courses:
            grade = AcademicGradeRecord.objects.filter(
                student=s, plan_course=pc, academic_year=enr.academic_year
            ).first()
            nota = int(grade.final_grade) if grade and grade.final_grade is not None else None
            credits = getattr(pc, "credits", 0) or 0
            puntaje = (nota * credits) if nota is not None else 0
            cond = ("A" if nota >= 11 else "D") if nota is not None else ""
            row_data += [nota or "", cond, puntaje]
            sum_p += puntaje
            sum_c += credits
        prom = round(sum_p / sum_c, 2) if sum_c > 0 else 0
        row_data += [sum_c, sum_p, prom]
        writer.writerow(row_data)

    content = buf.getvalue().encode("utf-8-sig")
    return f"acta_consolidada_{period_code}.csv", content, len(enrollments)


# ═══════════════════════════════════════════════════════════
# 5. REPORTE DE INFORMACIÓN (KARDEX)
# ═══════════════════════════════════════════════════════════

def _generate_reporte_xlsx(period_code):
    """
    Kardex: historial académico completo del estudiante,
    agrupado por período con totales por período y general.
    """
    inst = _get_institution()
    enrollments = _get_enrollments(period_code)
    total = enrollments.count()

    # Recopilar TODOS los períodos de cada estudiante
    student_ids = set(enrollments.values_list("student_id", flat=True))

    wb = Workbook()
    if student_ids and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for student_id in student_ids:
        student = Student.objects.get(pk=student_id)
        safe_name = _student_fullname(student)[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        max_col = 10

        col_widths = [5, 12, 30, 8, 8, 8, 8, 10, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _sia_header_rows(ws, "Reporte de Información del Sistema", period_code, inst, max_col=max_col)

        # Datos del alumno
        info = [
            ("Estudiante:", _student_fullname(student)),
            ("N° Documento:", student.num_documento or ""),
            ("Fecha Nac.:", student.fecha_nac.strftime("%d/%m/%Y") if student.fecha_nac else ""),
            ("Cód. MINEDU:", _get_minedu_code("STUDENT", student.id)),
        ]
        for label, val in info:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _FONT_HEADER
            ws.cell(row=row, column=3).value = val
            ws.cell(row=row, column=3).font = _FONT_NORMAL
            row += 1
        row += 1

        # Obtener TODAS las matrículas del alumno
        all_enrollments = (
            Enrollment.objects
            .filter(student=student, status="CONFIRMED")
            .select_related("plan", "plan__career")
            .order_by("academic_year", "academic_period")
        )

        grand_puntaje = 0
        grand_creditos = 0

        for enr in all_enrollments:
            career = enr.plan.career if enr.plan else None
            per = f"{enr.academic_year}-{enr.academic_period}"

            # Título período
            ws.cell(row=row, column=1).value = f"Período: {per}"
            ws.cell(row=row, column=1).font = _FONT_SUBTITLE
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=6).value = f"Programa: {career.name if career else ''}"
            ws.cell(row=row, column=6).font = _FONT_SMALL
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=max_col)
            row += 1

            headers = ["N°", "Código", "Área / Curso", "Tipo", "Horas", "Créd.",
                       "Calificación", "Nivel", "Puntaje", "Condición"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c).value = h
            h_row = row
            row += 1

            items = _get_grades_for_enrollment(enr)
            d_start = row
            per_puntaje = 0
            per_creditos = 0
            for idx, item in enumerate(items, 1):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = item["code"]
                ws.cell(row=row, column=3).value = item["name"]
                ws.cell(row=row, column=4).value = "OB"
                ws.cell(row=row, column=5).value = item["hours"]
                ws.cell(row=row, column=6).value = item["credits"]
                ws.cell(row=row, column=7).value = item["nota_int"] if item["nota_int"] is not None else ""
                ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
                ws.cell(row=row, column=8).value = item["nota_letra"]
                ws.cell(row=row, column=8).alignment = _ALIGN_CENTER
                ws.cell(row=row, column=9).value = item["puntaje"]
                ws.cell(row=row, column=9).alignment = _ALIGN_CENTER
                ws.cell(row=row, column=10).value = item["estado"]
                per_puntaje += item["puntaje"]
                per_creditos += item["credits"]
                row += 1

            if items:
                _apply_table_style(ws, h_row, d_start, row - 1, max_col)

            # Totales período
            per_promedio = round(per_puntaje / per_creditos, 2) if per_creditos > 0 else 0
            ws.cell(row=row, column=5).value = "Subtotal:"
            ws.cell(row=row, column=5).font = _FONT_HEADER
            ws.cell(row=row, column=6).value = per_creditos
            ws.cell(row=row, column=9).value = per_puntaje
            ws.cell(row=row, column=10).value = f"Prom: {per_promedio}"
            ws.cell(row=row, column=10).font = _FONT_HEADER
            row += 2

            grand_puntaje += per_puntaje
            grand_creditos += per_creditos

        # Totales generales
        grand_promedio = round(grand_puntaje / grand_creditos, 2) if grand_creditos > 0 else 0
        ws.cell(row=row, column=1).value = "RESUMEN GENERAL"
        ws.cell(row=row, column=1).font = Font(name="Arial", size=11, bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        row += 1
        ws.cell(row=row, column=1).value = f"Total Créditos: {grand_creditos}"
        ws.cell(row=row, column=1).font = _FONT_SUBTITLE
        ws.cell(row=row, column=4).value = f"Total Puntaje: {grand_puntaje}"
        ws.cell(row=row, column=4).font = _FONT_SUBTITLE
        ws.cell(row=row, column=7).value = f"Promedio Ponderado Acumulado: {grand_promedio}"
        ws.cell(row=row, column=7).font = Font(name="Arial", size=10, bold=True)
        row += 1

        _signature_rows(ws, inst, row + 1, max_col)

    if not student_ids:
        ws = wb.active
        ws.cell(row=1, column=1).value = "No hay datos."

    filename = f"reporte_kardex_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_reporte_csv(period_code):
    enrollments = _get_enrollments(period_code)
    student_ids = set(enrollments.values_list("student_id", flat=True))

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Estudiante", "DNI", "Período", "Código", "Curso",
                     "Créditos", "Nota", "Nivel", "Puntaje", "Condición"])
    for sid in student_ids:
        student = Student.objects.get(pk=sid)
        all_enr = Enrollment.objects.filter(
            student=student, status="CONFIRMED"
        ).select_related("plan").order_by("academic_year", "academic_period")
        for enr in all_enr:
            per = f"{enr.academic_year}-{enr.academic_period}"
            items = _get_grades_for_enrollment(enr)
            for item in items:
                writer.writerow([
                    _student_fullname(student), student.num_documento or "", per,
                    item["code"], item["name"], item["credits"],
                    item["nota_int"] if item["nota_int"] is not None else "",
                    item["nota_letra"], item["puntaje"], item["estado"],
                ])

    content = buf.getvalue().encode("utf-8-sig")
    return f"reporte_kardex_{period_code}.csv", content, len(student_ids)


# ═══════════════════════════════════════════════════════════
# 6. REGISTRO AUXILIAR DE EVALUACIÓN (nuevo)
# ═══════════════════════════════════════════════════════════

def _generate_registro_aux_xlsx(period_code):
    """
    Registro Auxiliar: tabla por curso, con columnas para notas parciales
    (Unidad 1, 2, 3, 4) + Nota Final.
    Genera una hoja por cada curso del plan.

    NOTA: Si tu modelo no tiene notas parciales (solo final_grade),
    las columnas de unidades quedarán vacías. El docente las llena a mano.
    """
    inst = _get_institution()
    year, period = _parse_period(period_code)

    # Obtener plan del período
    enrollments = list(_get_enrollments(period_code))
    if not enrollments:
        wb = Workbook()
        wb.active.cell(row=1, column=1).value = "No hay matrículas."
        return f"registro_auxiliar_{period_code}.xlsx", _workbook_to_bytes(wb), 0

    first_plan = enrollments[0].plan
    plan_courses = PlanCourse.objects.filter(
        plan=first_plan
    ).select_related("course").order_by("order", "id") if first_plan else []

    students = [enr.student for enr in enrollments]
    total = len(students)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for pc in plan_courses:
        course = pc.course
        safe_name = (course.name if course else "Curso")[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        max_col = 10

        col_widths = [5, 30, 12, 8, 8, 8, 8, 10, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _sia_header_rows(ws, "Registro Auxiliar de Evaluación", period_code, inst, max_col=max_col)

        # Info del curso
        ws.cell(row=row, column=1).value = f"Curso: {course.name if course else ''}"
        ws.cell(row=row, column=1).font = _FONT_SUBTITLE
        ws.cell(row=row, column=6).value = f"Créditos: {getattr(pc, 'credits', 0)}"
        ws.cell(row=row, column=6).font = _FONT_HEADER
        row += 1
        ws.cell(row=row, column=1).value = f"Docente: ___________________________"
        ws.cell(row=row, column=1).font = _FONT_NORMAL
        row += 2

        # Headers
        headers = ["N°", "Apellidos y Nombres", "DNI",
                   "Unidad 1", "Unidad 2", "Unidad 3", "Unidad 4",
                   "Nota Final", "Nivel", "Condición"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c).value = h
        h_row = row
        row += 1

        d_start = row
        for idx, student in enumerate(students, 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = _student_fullname(student)
            ws.cell(row=row, column=3).value = student.num_documento or ""

            # Notas parciales: buscar si el modelo las tiene
            # AJUSTAR: si tienes campos u1, u2, u3, u4 en AcademicGradeRecord
            grade = AcademicGradeRecord.objects.filter(
                student=student,
                plan_course=pc,
                academic_year=year,
            ).first()

            # Unidades (vacías si no hay datos parciales)
            for u_col in [4, 5, 6, 7]:
                ws.cell(row=row, column=u_col).value = ""
                ws.cell(row=row, column=u_col).alignment = _ALIGN_CENTER

            # Nota final
            nota = grade.final_grade if grade else None
            ws.cell(row=row, column=8).value = int(nota) if nota is not None else ""
            ws.cell(row=row, column=8).alignment = _ALIGN_CENTER
            ws.cell(row=row, column=9).value = _nota_letra(nota)
            ws.cell(row=row, column=9).alignment = _ALIGN_CENTER
            ws.cell(row=row, column=10).value = _nota_estado(nota)
            row += 1

        _apply_table_style(ws, h_row, d_start, row - 1, max_col)

        # Firma del docente
        row += 3
        ws.cell(row=row, column=2).value = "_" * 30
        ws.cell(row=row, column=2).alignment = _ALIGN_CENTER
        ws.cell(row=row + 1, column=2).value = "Firma del Docente"
        ws.cell(row=row + 1, column=2).font = _FONT_SMALL
        ws.cell(row=row + 1, column=2).alignment = _ALIGN_CENTER

    filename = f"registro_auxiliar_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_registro_aux_csv(period_code):
    enrollments = list(_get_enrollments(period_code))
    if not enrollments:
        buf = io.StringIO()
        csv.writer(buf).writerow(["Sin datos"])
        return f"registro_auxiliar_{period_code}.csv", buf.getvalue().encode("utf-8-sig"), 0

    first_plan = enrollments[0].plan
    plan_courses = PlanCourse.objects.filter(
        plan=first_plan
    ).select_related("course").order_by("order", "id") if first_plan else []

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Curso", "N°", "Estudiante", "DNI",
                     "Unidad 1", "Unidad 2", "Unidad 3", "Unidad 4",
                     "Nota Final", "Nivel", "Condición"])
    for pc in plan_courses:
        course = pc.course
        for idx, enr in enumerate(enrollments, 1):
            grade = AcademicGradeRecord.objects.filter(
                student=enr.student, plan_course=pc, academic_year=enr.academic_year
            ).first()
            nota = int(grade.final_grade) if grade and grade.final_grade is not None else None
            writer.writerow([
                course.name if course else "", idx,
                _student_fullname(enr.student), enr.student.num_documento or "",
                "", "", "", "",
                nota if nota is not None else "", _nota_letra(nota), _nota_estado(nota),
            ])

    content = buf.getvalue().encode("utf-8-sig")
    return f"registro_auxiliar_{period_code}.csv", content, len(enrollments)


# ═══════════════════════════════════════════════════════════
# 7. CERTIFICADO DE ESTUDIOS (nuevo)
# ═══════════════════════════════════════════════════════════

def _generate_certificado_xlsx(period_code):
    """
    Certificado de Estudios: resumen por período con promedios,
    apto para impresión como documento oficial.
    """
    inst = _get_institution()
    enrollments = _get_enrollments(period_code)
    student_ids = set(enrollments.values_list("student_id", flat=True))
    total = len(student_ids)

    wb = Workbook()
    if total > 0 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for student_id in student_ids:
        student = Student.objects.get(pk=student_id)
        safe_name = _student_fullname(student)[:28].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        max_col = 8

        col_widths = [5, 15, 30, 10, 10, 10, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _sia_header_rows(ws, "Certificado de Estudios", period_code, inst, max_col=max_col)

        # Datos del alumno
        info = [
            ("Estudiante:", _student_fullname(student)),
            ("N° Documento:", student.num_documento or ""),
            ("Fecha Nac.:", student.fecha_nac.strftime("%d/%m/%Y") if student.fecha_nac else ""),
        ]
        for label, val in info:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _FONT_HEADER
            ws.cell(row=row, column=3).value = val
            ws.cell(row=row, column=3).font = _FONT_NORMAL
            row += 1
        row += 1

        # Obtener matrícula del período actual + carrera
        current_enr = enrollments.filter(student=student).first()
        career = current_enr.plan.career if current_enr and current_enr.plan else None

        if career:
            ws.cell(row=row, column=1).value = f"Programa de Estudios: {career.name}"
            ws.cell(row=row, column=1).font = _FONT_SUBTITLE
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            row += 2

        # Historial por período
        all_enr = (
            Enrollment.objects
            .filter(student=student, status="CONFIRMED")
            .select_related("plan", "plan__career")
            .order_by("academic_year", "academic_period")
        )

        grand_credits = 0
        grand_puntaje = 0
        period_summaries = []

        for enr in all_enr:
            per_code = f"{enr.academic_year}-{enr.academic_period}"
            items = _get_grades_for_enrollment(enr)
            per_credits = sum(i["credits"] for i in items)
            per_puntaje = sum(i["puntaje"] for i in items)
            per_prom = round(per_puntaje / per_credits, 2) if per_credits > 0 else 0
            grand_credits += per_credits
            grand_puntaje += per_puntaje
            period_summaries.append((per_code, per_credits, per_puntaje, per_prom, items))

        for per_code, per_credits, per_puntaje, per_prom, items in period_summaries:
            ws.cell(row=row, column=1).value = f"Período: {per_code}"
            ws.cell(row=row, column=1).font = _FONT_SUBTITLE
            row += 1

            headers = ["N°", "Código", "Área / Curso", "Horas", "Créd.",
                       "Calificación", "Nivel", "Condición"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c).value = h
            h_row = row
            row += 1

            d_start = row
            for idx, item in enumerate(items, 1):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = item["code"]
                ws.cell(row=row, column=3).value = item["name"]
                ws.cell(row=row, column=4).value = item["hours"]
                ws.cell(row=row, column=5).value = item["credits"]
                ws.cell(row=row, column=6).value = item["nota_int"] if item["nota_int"] is not None else ""
                ws.cell(row=row, column=6).alignment = _ALIGN_CENTER
                ws.cell(row=row, column=7).value = item["nota_letra"]
                ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
                ws.cell(row=row, column=8).value = item["estado"]
                row += 1

            if items:
                _apply_table_style(ws, h_row, d_start, row - 1, max_col)

            # Promedio del período
            ws.cell(row=row, column=5).value = f"Créd: {per_credits}"
            ws.cell(row=row, column=5).font = _FONT_SMALL
            ws.cell(row=row, column=7).value = f"Prom: {per_prom}"
            ws.cell(row=row, column=7).font = _FONT_HEADER
            row += 2

        # Resumen general
        grand_prom = round(grand_puntaje / grand_credits, 2) if grand_credits > 0 else 0
        ws.cell(row=row, column=1).value = "PROMEDIO PONDERADO ACUMULADO"
        ws.cell(row=row, column=1).font = Font(name="Arial", size=11, bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=6).value = grand_prom
        ws.cell(row=row, column=6).font = Font(name="Arial", size=12, bold=True)
        ws.cell(row=row, column=6).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=7).value = _nota_letra(round(grand_prom))
        ws.cell(row=row, column=7).font = Font(name="Arial", size=11, bold=True)
        ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
        row += 1

        _signature_rows(ws, inst, row + 1, max_col)

    if not student_ids:
        ws = wb.active
        ws.cell(row=1, column=1).value = "No hay datos."

    filename = f"certificado_estudios_{period_code}.xlsx"
    return filename, _workbook_to_bytes(wb), total


def _generate_certificado_csv(period_code):
    enrollments = _get_enrollments(period_code)
    student_ids = set(enrollments.values_list("student_id", flat=True))

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Estudiante", "DNI", "Período", "Código", "Curso",
                     "Créditos", "Nota", "Nivel", "Condición"])
    for sid in student_ids:
        student = Student.objects.get(pk=sid)
        all_enr = Enrollment.objects.filter(
            student=student, status="CONFIRMED"
        ).select_related("plan").order_by("academic_year", "academic_period")
        for enr in all_enr:
            per = f"{enr.academic_year}-{enr.academic_period}"
            items = _get_grades_for_enrollment(enr)
            for item in items:
                writer.writerow([
                    _student_fullname(student), student.num_documento or "", per,
                    item["code"], item["name"], item["credits"],
                    item["nota_int"] if item["nota_int"] is not None else "",
                    item["nota_letra"], item["estado"],
                ])

    content = buf.getvalue().encode("utf-8-sig")
    return f"certificado_estudios_{period_code}.csv", content, len(student_ids)


# ═══════════════════════════════════════════════════════════
# DISPATCHER — punto de entrada principal
# ═══════════════════════════════════════════════════════════

# Mapa: (data_type, format) → función generadora
# Cada función retorna: (filename, bytes_content, total_records)
GENERATORS = {
    # ── 5 documentos SIA principales ──
    ("ENROLLMENT", "XLSX"):   _generate_enrollment_xlsx,
    ("ENROLLMENT", "CSV"):    _generate_enrollment_csv,
    ("FICHA", "XLSX"):        _generate_ficha_xlsx,
    ("FICHA", "CSV"):         _generate_ficha_csv,
    ("BOLETA", "XLSX"):       _generate_boleta_xlsx,
    ("BOLETA", "CSV"):        _generate_boleta_csv,
    ("ACTA", "XLSX"):         _generate_acta_xlsx,
    ("ACTA", "CSV"):          _generate_acta_csv,
    ("REPORTE", "XLSX"):      _generate_reporte_xlsx,
    ("REPORTE", "CSV"):       _generate_reporte_csv,
    # ── 2 documentos extras ──
    ("REGISTRO_AUX", "XLSX"): _generate_registro_aux_xlsx,
    ("REGISTRO_AUX", "CSV"):  _generate_registro_aux_csv,
    ("CERTIFICADO", "XLSX"):  _generate_certificado_xlsx,
    ("CERTIFICADO", "CSV"):   _generate_certificado_csv,
}

# Aliases legacy → canónico
_LEGACY_TYPE_MAP = {
    "FICHA_MATRICULA": "FICHA",
    "GRADES":          "ACTA",
    "STUDENTS":        "REPORTE",
}


def generate_export(data_type, export_format, period_code):
    """
    Punto de entrada principal. Llamado desde views.py.

    Args:
        data_type: str — tipo de documento (ENROLLMENT, FICHA, BOLETA, etc.)
        export_format: str — formato (XLSX, CSV)
        period_code: str — período académico (e.g. "2024-I")

    Returns:
        tuple: (filename, bytes_content, total_records)

    Raises:
        ValueError: si el tipo o formato no es soportado
    """
    # Normalizar legacy types
    data_type = _LEGACY_TYPE_MAP.get(data_type, data_type)
    export_format = export_format.upper()

    key = (data_type, export_format)
    generator = GENERATORS.get(key)

    if generator is None:
        supported = sorted(set(k[0] for k in GENERATORS.keys()))
        raise ValueError(
            f"Tipo '{data_type}' con formato '{export_format}' no soportado. "
            f"Tipos disponibles: {', '.join(supported)}. "
            f"Formatos: XLSX, CSV."
        )

    return generator(period_code)