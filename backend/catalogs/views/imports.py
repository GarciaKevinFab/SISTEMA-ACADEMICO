"""
Import Engine - Procesamiento asíncrono de archivos
IMPLEMENTACIÓN COMPLETA con toda la lógica de imports
"""
import io
import re
import threading
from datetime import date
from typing import List, Optional
from django.db import transaction, close_old_connections, models, IntegrityError
from django.core.files.base import ContentFile
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string
from openpyxl import load_workbook
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from catalogs.models import ImportJob, Period, Career
from students.models import Student
from academic.models import Plan, Course, PlanCourse
from acl.models import Role, UserRole
from .utils import (
    _require_staff, _norm, _norm_key, _to_int, _to_float,
    _clean_text, _clean_spaces, _normalize_course_name, _pc_key,
    _parse_period_code, _ensure_period, _slug_code,
    _build_norm_index, _find_by_norm_cached,
    _make_unique_career_code, _ensure_career_cached,
    _read_rows,
)

# Import helpers de catalogs
try:
    from catalogs.helpers import (
        match_career_robust,
        pick_plan_for_student,
        match_plan_course_for_grade,
    )
except ImportError:
    def match_career_robust(name):
        return Career.objects.filter(name__iexact=name).first()

    def pick_plan_for_student(career, period_code):
        qs = Plan.objects.filter(career=career).order_by("-start_year", "-id")
        return qs.first()

    def match_plan_course_for_grade(student, course, plan_id=None):
        pid = plan_id or getattr(student, "plan_id", None)
        if not pid:
            return None
        return PlanCourse.objects.filter(plan_id=pid, course=course).first()

# Import AcademicGradeRecord
try:
    from academic.models import AcademicGradeRecord
except ImportError:
    AcademicGradeRecord = None

User = get_user_model()


# ═══════════════════════════════════════════════════════════════
# READERS ESPECÍFICOS
# ═══════════════════════════════════════════════════════════════

def _is_electivos_name(name: str) -> bool:
    nk = _norm_key(name or "")
    return nk in ("electivos", "electivo", "elective", "electives")


def _normalize_semester_value(v) -> str:
    s = "" if v is None else str(v)
    s = re.sub(r"\s+", "", s, flags=re.UNICODE)
    import unicodedata
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower()


def _is_semester_label(v) -> Optional[int]:
    key = _normalize_semester_value(v)
    if not key:
        return None

    sem_map = {
        "primero": 1, "segundo": 2, "tercero": 3, "cuarto": 4, "quinto": 5,
        "sexto": 6, "septimo": 7, "octavo": 8, "noveno": 9, "decimo": 10,
        "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6, "7": 7, "8": 8, "9": 9, "10": 10,
        "i": 1, "ii": 2, "iii": 3, "iv": 4, "v": 5, "vi": 6, "vii": 7, "viii": 8, "ix": 9, "x": 10,
    }

    if key in sem_map:
        return sem_map[key]

    if key.startswith("semestre"):
        tail = key.replace("semestre", "")
        return sem_map.get(tail)

    return None


def _extract_year_from_text(v):
    if v is None:
        return None
    try:
        if isinstance(v, (int, float)):
            y = int(v)
            if 2000 <= y <= 2100:
                return y
    except Exception:
        pass
    s = str(v)
    m = re.search(r"(20\d{2})", s)
    if m:
        y = int(m.group(1))
        if 2000 <= y <= 2100:
            return y
    return None


def _extract_year_range_from_text(v):
    if v is None:
        return (None, None)
    s = str(v)
    m = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 2000 <= a <= 2100 and 2000 <= b <= 2100 and a <= b:
            return (a, b)
    y = _extract_year_from_text(v)
    return (y, y) if y else (None, None)


def _sheet_plan_year_range(ws, sheet_name: str = "", filename: str = ""):
    try:
        for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
            for cell in row:
                a, b = _extract_year_range_from_text(cell)
                if a and b:
                    return a, b
    except Exception:
        pass
    a, b = _extract_year_range_from_text(sheet_name or "")
    if a and b:
        return a, b
    a, b = _extract_year_range_from_text(filename or "")
    if a and b:
        return a, b
    y = date.today().year
    return y, y


def _career_from_sheet_name(sheet: str) -> str:
    raw = (sheet or "").strip()
    raw = re.sub(r"\b20\d{2}\b", "", raw).strip()
    raw = re.sub(r"\s+", " ", raw).strip()
    key = _norm(raw)

    MAP = {
        "inicial": "EDUCACIÓN INICIAL",
        "educacion inicial": "EDUCACIÓN INICIAL",
        "ed inicial": "EDUCACIÓN INICIAL",
        "ed. inicial": "EDUCACIÓN INICIAL",
        "primaria": "EDUCACIÓN PRIMARIA",
        "educacion primaria": "EDUCACIÓN PRIMARIA",
        "ed primaria": "EDUCACIÓN PRIMARIA",
        "ed. primaria": "EDUCACIÓN PRIMARIA",
        "ed. fisica": "EDUCACIÓN FÍSICA",
        "ed fisica": "EDUCACIÓN FÍSICA",
        "educacion fisica": "EDUCACIÓN FÍSICA",
        "fisica": "EDUCACIÓN FÍSICA",
        "comunicacion": "COMUNICACIÓN",
        "educacion secundaria comunicacion": "EDUCACIÓN SECUNDARIA, ESPECIALIDAD: COMUNICACIÓN",
        "sec comunicacion": "EDUCACIÓN SECUNDARIA, ESPECIALIDAD: COMUNICACIÓN",
        "secundaria comunicacion": "EDUCACIÓN SECUNDARIA, ESPECIALIDAD: COMUNICACIÓN",
        "computacion": "COMPUTACIÓN E INFORMÁTICA",
        "computacion e informatica": "COMPUTACIÓN E INFORMÁTICA",
        "informatica": "COMPUTACIÓN E INFORMÁTICA",
    }
    return MAP.get(key, raw.upper())


def _read_study_plan_xlsx(file):
    filename = getattr(file, "name", "") or ""
    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    out = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        career_name = _career_from_sheet_name(sheet)
        if not career_name:
            continue

        plan_start, plan_end = _sheet_plan_year_range(ws, sheet, filename)
        current_sem = None
        course_col = None
        cred_col = None
        hours_col = None
        header_found = False
        empty_streak = 0

        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_norm = [(_norm(x) if x is not None else "") for x in row]

            is_header_row = (
                any(("areas" in h) or ("asignaturas" in h) for h in row_norm) and
                any(("cred" in h) for h in row_norm)
            )

            if is_header_row:
                course_col = None
                cred_col = None
                hours_col = None

                for j, h in enumerate(row_norm):
                    if ("areas" in h) or ("asignaturas" in h):
                        course_col = j
                        break

                for j, h in enumerate(row_norm):
                    if ("cred" in h) or ("credito" in h) or ("creditos" in h):
                        cred_col = j
                        break

                for j, h in enumerate(row_norm):
                    if ("hora" in h) or ("horas" in h) or (h == "nota"):
                        hours_col = j
                        break

                header_found = course_col is not None and cred_col is not None
                empty_streak = 0
                continue

            if not header_found:
                continue

            found_sem = None
            for cell in row:
                sem = _is_semester_label(cell)
                if sem:
                    found_sem = sem
                    break

            if found_sem:
                current_sem = int(found_sem)
                empty_streak = 0

            if not current_sem:
                continue

            if course_col is None or course_col >= len(row):
                continue

            course_name = "" if row[course_col] is None else str(row[course_col]).strip()

            if not course_name:
                if all((c is None or str(c).strip() == "") for c in row):
                    empty_streak += 1
                else:
                    empty_streak = 0
                if empty_streak >= 250:
                    break
                continue

            empty_streak = 0

            bad = _norm(course_name)
            if bad in ("sem", "n", "n°", "no", "numero", "areas/asignaturas",
                      "areas / asignaturas", "asignaturas", "total", "totales"):
                continue

            credits = 0
            if cred_col is not None and cred_col < len(row):
                credits = int(_to_float(row[cred_col], 0) or 0)

            hours = 0
            if hours_col is not None and hours_col < len(row):
                hours = _to_int(row[hours_col], 0) or 0

            out.append({
                "__row__": ridx,
                "career_name": career_name,
                "plan_start_year": int(plan_start),
                "plan_end_year": int(plan_end),
                "semester": int(current_sem),
                "course_name": course_name,
                "credits": int(credits),
                "hours": int(hours),
            })

    return out


def _read_calificaciones_xlsx(file):
    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [(_norm(h) if h is not None else "") for h in rows[0]]

    def idx_match(*names):
        wanted = {_norm(x) for x in names}
        for i, h in enumerate(headers):
            if h in wanted:
                return i
        return None

    i_doc = idx_match("NUMERO_DOCUMENTO", "NUM DOCUMENTO", "NUM_DOCUMENTO", "DNI")
    i_periodo = idx_match("PERIODO", "PERIOD")
    i_programa = idx_match("PROGRAMA", "PROGRAMA / CARRERA", "PROGRAMA_CARRERA")
    i_ciclo = idx_match("CICLO")
    i_curso = idx_match("CURSO", "ASIGNATURA")
    i_nota = idx_match("PROMEDIO_VIGESIMAL", "PROMEDIO", "FINAL", "NOTA", "PROM_FINAL")

    out = []
    for ridx, r in enumerate(rows[1:], start=2):
        def get(i):
            return r[i] if i is not None and i < len(r) else None

        doc = "" if get(i_doc) is None else str(get(i_doc)).strip()
        doc = re.sub(r"\.0$", "", doc)
        if doc.isdigit() and len(doc) < 8:
            doc = doc.zfill(8)

        periodo = "" if get(i_periodo) is None else str(get(i_periodo)).strip()
        programa = "" if get(i_programa) is None else str(get(i_programa)).strip()
        ciclo = _to_int(get(i_ciclo), None)
        curso = "" if get(i_curso) is None else str(get(i_curso)).strip()
        nota = _to_float(get(i_nota), None)

        if not doc or not periodo or not curso or nota is None:
            continue

        out.append({
            "__row__": ridx,
            "doc": doc,
            "periodo": periodo,
            "programa": programa,
            "ciclo": ciclo,
            "curso": curso,
            "nota": nota,
        })

    return out


# ═══════════════════════════════════════════════════════════════
# LECTOR XLSX PARA TRASLADOS (alumno + notas en un solo Excel)
# ═══════════════════════════════════════════════════════════════

LEVEL_TO_NUM_IMPORT = {"PI": 1, "I": 2, "P": 3, "L": 4, "D": 5}

def _calc_escala_import(c1, c2, c3):
    if c1 is None or c2 is None or c3 is None:
        return None
    return round(((c1 + c2 + c3) / 3.0), 1)

def _calc_promedio_from_escala(escala):
    if escala is None:
        return None
    return int(round(((escala - 1.0) / 4.0) * 20.0))


def _read_traslados_xlsx(file):
    """
    Lee un Excel de traslados: cada fila = 1 nota de 1 alumno.
    Un alumno puede tener múltiples filas (una por curso/período).

    Columnas esperadas:
      NUM_DOCUMENTO, NOMBRES, APELLIDO_PATERNO, APELLIDO_MATERNO, SEXO,
      PROGRAMA_CARRERA, CICLO, EMAIL, CELULAR,
      PERIODO, CURSO, C1_LEVEL, C2_LEVEL, C3_LEVEL, PROMEDIO_FINAL
    """
    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [(_norm(h) if h is not None else "") for h in rows[0]]

    def idx_match(*names):
        wanted = {_norm(x) for x in names}
        for i, h in enumerate(headers):
            if h in wanted:
                return i
        return None

    # Datos del alumno
    i_doc = idx_match("NUMERO_DOCUMENTO", "NUM DOCUMENTO", "NUM_DOCUMENTO", "DNI")
    i_nombres = idx_match("NOMBRES")
    i_ap_pat = idx_match("APELLIDO_PATERNO", "AP_PATERNO")
    i_ap_mat = idx_match("APELLIDO_MATERNO", "AP_MATERNO")
    i_sexo = idx_match("SEXO")
    i_programa = idx_match("PROGRAMA", "PROGRAMA_CARRERA", "PROGRAMA / CARRERA")
    i_ciclo = idx_match("CICLO")
    i_email = idx_match("EMAIL", "CORREO")
    i_celular = idx_match("CELULAR", "TELEFONO")

    # Datos de la nota
    i_periodo = idx_match("PERIODO", "PERIOD")
    i_curso = idx_match("CURSO", "ASIGNATURA")
    i_c1_level = idx_match("C1_LEVEL", "C1_NIVEL", "NIVEL_C1")
    i_c2_level = idx_match("C2_LEVEL", "C2_NIVEL", "NIVEL_C2")
    i_c3_level = idx_match("C3_LEVEL", "C3_NIVEL", "NIVEL_C3")
    i_promedio = idx_match("PROMEDIO_FINAL", "PROMEDIO_VIGESIMAL", "PROMEDIO", "NOTA", "FINAL")

    out = []
    for ridx, r in enumerate(rows[1:], start=2):
        def get(i):
            return r[i] if i is not None and i < len(r) else None

        doc = "" if get(i_doc) is None else str(get(i_doc)).strip()
        doc = re.sub(r"\.0$", "", doc)
        if doc.isdigit() and len(doc) < 8:
            doc = doc.zfill(8)

        nombres = "" if get(i_nombres) is None else str(get(i_nombres)).strip()
        ap_pat = "" if get(i_ap_pat) is None else str(get(i_ap_pat)).strip()
        ap_mat = "" if get(i_ap_mat) is None else str(get(i_ap_mat)).strip()
        sexo = "" if get(i_sexo) is None else str(get(i_sexo)).strip()
        programa = "" if get(i_programa) is None else str(get(i_programa)).strip()
        ciclo = _to_int(get(i_ciclo), None)
        email = "" if get(i_email) is None else str(get(i_email)).strip().lower()
        celular = "" if get(i_celular) is None else str(get(i_celular)).strip()

        periodo = "" if get(i_periodo) is None else str(get(i_periodo)).strip()
        curso = "" if get(i_curso) is None else str(get(i_curso)).strip()

        c1_lv = "" if get(i_c1_level) is None else str(get(i_c1_level)).strip().upper()
        c2_lv = "" if get(i_c2_level) is None else str(get(i_c2_level)).strip().upper()
        c3_lv = "" if get(i_c3_level) is None else str(get(i_c3_level)).strip().upper()
        promedio = _to_float(get(i_promedio), None)

        if not doc or not periodo or not curso:
            continue

        # Auto-calcular promedio desde componentes si no viene
        components = {}
        if c1_lv and c2_lv and c3_lv:
            c1 = LEVEL_TO_NUM_IMPORT.get(c1_lv)
            c2 = LEVEL_TO_NUM_IMPORT.get(c2_lv)
            c3 = LEVEL_TO_NUM_IMPORT.get(c3_lv)
            if c1 and c2 and c3:
                escala = _calc_escala_import(c1, c2, c3)
                prom_calc = _calc_promedio_from_escala(escala)
                estado = "Logrado" if prom_calc >= 11 else "En proceso"
                components = {
                    "C1": c1, "C2": c2, "C3": c3,
                    "C1_LEVEL": c1_lv, "C2_LEVEL": c2_lv, "C3_LEVEL": c3_lv,
                    "ESCALA_0_5": escala,
                    "PROMEDIO_FINAL": prom_calc,
                    "ESTADO": estado,
                }
                if promedio is None:
                    promedio = prom_calc

        if promedio is None:
            continue

        out.append({
            "__row__": ridx,
            "doc": doc,
            "nombres": nombres,
            "ap_pat": ap_pat,
            "ap_mat": ap_mat,
            "sexo": sexo,
            "programa": programa,
            "ciclo": ciclo,
            "email": email,
            "celular": celular,
            "periodo": periodo,
            "curso": curso,
            "promedio": promedio,
            "components": components,
        })

    return out


# ═══════════════════════════════════════════════════════════════
# HELPERS PARA GRADO/TÍTULO
# ═══════════════════════════════════════════════════════════════

def _get_default_grado_titulo_type():
    """Obtiene el GradoTituloType marcado como default, o None."""
    try:
        from graduates.models import GradoTituloType
        return GradoTituloType.objects.filter(is_default=True, is_active=True).first()
    except (ImportError, Exception):
        return None


def _derive_grado_titulo_fallback(especialidad: str) -> str:
    """Derivación legacy si no hay catálogo configurado."""
    import unicodedata

    esp = (especialidad or "").strip().upper()
    if not esp:
        return "PROFESOR(A) EN EDUCACIÓN"

    esp_norm = unicodedata.normalize("NFKD", esp).encode("ascii", "ignore").decode("ascii")

    if "COMUNICACION" in esp_norm:
        return "PROFESOR(A) EN EDUCACIÓN SECUNDARIA, ESPECIALIDAD: COMUNICACIÓN"
    if "COMPUTACION" in esp_norm or "INFORMATICA" in esp_norm:
        return "PROFESOR(A) EN COMPUTACIÓN E INFORMÁTICA"

    return f"PROFESOR(A) EN {esp}"


# ═══════════════════════════════════════════════════════════════
# IMPORT ENGINE WORKER
# ═══════════════════════════════════════════════════════════════

def _run_import_job(job_id: int, raw: bytes, safe_name: str, type: str, mapping: dict):
    close_old_connections()

    try:
        job = ImportJob.objects.get(pk=job_id)
        errors: List[dict] = []
        imported = 0
        updated = 0
        credentials: List[dict] = []

        def set_job_state(processed: int, total: int, message: str = ""):
            p = int((processed / total) * 100) if total else 0
            job.result = {
                **(job.result or {}),
                "progress": p,
                "processed": processed,
                "total": total,
                "message": message,
            }
            job.save(update_fields=["result"])

        def add_error(row, field, message):
            errors.append({"row": row, "field": field, "message": message})

        filename = (safe_name or "").lower()

        try:
            # ═══════════════════════════════════════════════════
            # PLANS
            # ═══════════════════════════════════════════════════
            if type == "plans":
                if filename.endswith(".xls"):
                    job.status = "FAILED"
                    job.result = {
                        **(job.result or {}),
                        "progress": 100,
                        "errors": [{"row": None, "field": "file", "message": "Convierte el plan a .xlsx"}],
                    }
                    job.save(update_fields=["status", "result"])
                    return

                if not filename.endswith(".xlsx"):
                    job.status = "FAILED"
                    job.result = {
                        **(job.result or {}),
                        "progress": 100,
                        "errors": [{"row": None, "field": "file", "message": "Sube un archivo .xlsx"}],
                    }
                    job.save(update_fields=["status", "result"])
                    return

                bio_plan = io.BytesIO(raw)
                bio_plan.name = safe_name
                plan_rows = _read_study_plan_xlsx(bio_plan)

                total = len(plan_rows)
                set_job_state(0, total, "Leyendo plan de estudios...")

                if not plan_rows:
                    add_error(None, "format", "No se detectaron filas del plan")

                created_careers = 0
                created_plans = 0
                created_courses = 0
                created_links = 0
                ranges_seen = set()

                career_idx = _build_norm_index(Career, "name")
                course_idx = _build_norm_index(Course, "name")

                plan_cache = {}
                plan_pc_seen = {}
                done = 0

                for row in plan_rows:
                    done += 1
                    if done % 50 == 0:
                        set_job_state(done, total, "Procesando plan...")

                    r = row.get("__row__", "?")
                    career_name = (row.get("career_name") or "").strip()
                    plan_start = int(row.get("plan_start_year") or date.today().year)
                    plan_end = int(row.get("plan_end_year") or plan_start)
                    semester = _to_int(row.get("semester"), None)
                    course_name = (row.get("course_name") or "").strip()
                    credits = int(row.get("credits") or 0)
                    hours = int(row.get("hours") or 0)

                    if not career_name:
                        add_error(r, "career_name", "Carrera vacía")
                        continue
                    if not semester:
                        add_error(r, "semester", "Semestre inválido")
                        continue
                    if not course_name:
                        add_error(r, "course_name", "Curso vacío")
                        continue

                    ranges_seen.add((plan_start, plan_end))

                    before = len(career_idx)
                    car = _ensure_career_cached(career_idx, career_name)
                    if len(career_idx) > before:
                        created_careers += 1

                    pkey = (car.id, plan_start, plan_end)
                    plan = plan_cache.get(pkey)
                    if not plan:
                        plan_name = f"Plan {car.name} {plan_start}-{plan_end}"
                        plan, created = Plan.objects.get_or_create(
                            career=car,
                            start_year=plan_start,
                            end_year=plan_end,
                            defaults={
                                "name": plan_name,
                                "semesters": 10,
                                "is_deleted": False,
                            },
                        )

                        if created:
                            created_plans += 1
                        else:
                            dirty_fields = []
                            if getattr(plan, "is_deleted", False):
                                plan.is_deleted = False
                                dirty_fields.append("is_deleted")
                            if plan.name != plan_name:
                                plan.name = plan_name
                                dirty_fields.append("name")
                            if int(plan.semesters or 0) <= 0:
                                plan.semesters = 10
                                dirty_fields.append("semesters")
                            if dirty_fields:
                                plan.save(update_fields=dirty_fields)
                                updated += 1

                        plan_cache[pkey] = plan

                    course_name_clean = _clean_text(_normalize_course_name(course_name))
                    is_electivos = _is_electivos_name(course_name_clean)

                    display_name_for_pc = course_name_clean
                    course_name_for_course = course_name_clean

                    if is_electivos:
                        display_name_for_pc = "ELECTIVOS"
                        course_name_for_course = f"ELECTIVOS - S{int(semester)}"

                    course = Course.objects.filter(name__iexact=course_name_for_course).order_by("id").first()

                    if not course:
                        course = _find_by_norm_cached(course_idx, course_name_for_course)

                    if not course:
                        base = _slug_code(course_name_for_course, 10)
                        code = base
                        k = 1
                        while Course.objects.filter(code=code).exists():
                            k += 1
                            code = f"{base[:8]}{k:02d}"

                        course = Course.objects.create(
                            code=code,
                            name=course_name_for_course,
                            credits=credits if credits > 0 else 3,
                        )
                        course_idx[_norm_key(course_name_for_course)] = course
                        created_courses += 1
                    else:
                        if (not is_electivos) and _clean_spaces(course.name) != course_name_for_course:
                            course.name = course_name_for_course
                            course.save(update_fields=["name"])
                            updated += 1

                        if credits > 0 and int(getattr(course, "credits", 0) or 0) == 0:
                            course.credits = credits
                            course.save(update_fields=["credits"])
                            updated += 1

                    # PlanCourse
                    if is_electivos:
                        pc, pc_created = PlanCourse.objects.update_or_create(
                            plan=plan,
                            course=course,
                            defaults={
                                "semester": max(1, int(semester)),
                                "weekly_hours": max(0, int(hours)),
                                "type": "ELECTIVE",
                                "credits": max(0, int(credits or 0)),
                                "display_code": course.code,
                                "display_name": display_name_for_pc,
                            },
                        )
                        if pc_created:
                            created_links += 1
                            imported += 1
                        else:
                            updated += 1

                        kvis = _pc_key(f"{display_name_for_pc}-S{int(semester)}")
                        plan_pc_seen[(plan.id, kvis)] = pc

                    else:
                        kvis = _pc_key(course_name_clean)
                        cache_key = (plan.id, kvis)

                        existing_pc = plan_pc_seen.get(cache_key)

                        if not existing_pc:
                            existing_pc = (
                                PlanCourse.objects
                                .select_related("course")
                                .filter(plan_id=plan.id)
                                .filter(
                                    models.Q(display_name__iexact=course_name_clean) |
                                    models.Q(course__name__iexact=course_name_for_course)
                                )
                                .order_by("id")
                                .first()
                            )
                            if existing_pc:
                                plan_pc_seen[cache_key] = existing_pc

                        if existing_pc:
                            PlanCourse.objects.filter(plan_id=plan.id).filter(
                                models.Q(display_name__iexact=course_name_clean) |
                                models.Q(course__name__iexact=course_name_for_course)
                            ).exclude(id=existing_pc.id).delete()

                            existing_pc.course = course
                            existing_pc.semester = max(1, int(semester))
                            existing_pc.weekly_hours = max(0, int(hours))
                            existing_pc.type = "MANDATORY"
                            existing_pc.credits = max(0, int(credits or 0))
                            existing_pc.display_code = course.code
                            existing_pc.display_name = course_name_clean
                            existing_pc.save()
                            updated += 1
                        else:
                            pc, pc_created = PlanCourse.objects.update_or_create(
                                plan=plan,
                                course=course,
                                defaults={
                                    "semester": max(1, int(semester)),
                                    "weekly_hours": max(0, int(hours)),
                                    "type": "MANDATORY",
                                    "credits": max(0, int(credits or 0)),
                                    "display_code": course.code,
                                    "display_name": course_name_clean,
                                },
                            )
                            plan_pc_seen[cache_key] = pc

                            if pc_created:
                                created_links += 1
                                imported += 1
                            else:
                                updated += 1

                # Ajustar semesters
                for plan in plan_cache.values():
                    mx = (PlanCourse.objects.filter(plan_id=plan.id)
                          .aggregate(mx=models.Max("semester")).get("mx")) or 0
                    if mx > 0 and int(plan.semesters or 0) != int(mx):
                        plan.semesters = int(mx)
                        plan.save(update_fields=["semesters"])

                # ── RE-LINK automático de AcademicGradeRecord sin plan_course ──
                relinked_grades = 0
                if AcademicGradeRecord is not None:
                    for plan_obj in plan_cache.values():
                        recs_to_fix = list(
                            AcademicGradeRecord.objects
                            .filter(student__plan_id=plan_obj.id, plan_course__isnull=True)
                            .select_related("student", "course")[:2000]
                        )
                        for rec in recs_to_fix:
                            pc_found = (
                                PlanCourse.objects
                                .select_related("course")
                                .filter(plan_id=plan_obj.id)
                                .filter(
                                    models.Q(course_id=rec.course_id) |
                                    models.Q(display_name__iexact=rec.course.name) |
                                    models.Q(course__name__iexact=rec.course.name)
                                )
                                .first()
                            )
                            if pc_found:
                                rec.plan_course = pc_found
                                rec.save(update_fields=["plan_course"])
                                relinked_grades += 1

                set_job_state(total, total, "Finalizando plan...")

                job.result = {
                    **(job.result or {}),
                    "imported": imported,
                    "updated": updated,
                    "errors": errors,
                    "stats": {
                        "created_careers": created_careers,
                        "created_plans": created_plans,
                        "created_courses": created_courses,
                        "created_links": created_links,
                        "relinked_grades": relinked_grades,
                        "ranges": sorted(list(ranges_seen)),
                    },
                }

            # ═══════════════════════════════════════════════════
            # STUDENTS / COURSES / GRADES
            # ═══════════════════════════════════════════════════
            elif type in ("students", "courses", "grades"):
                rows = []
                total = 0

                if type in ("students", "courses"):
                    bio = io.BytesIO(raw)
                    bio.name = safe_name
                    rows = _read_rows(bio, mapping)
                    total = len(rows)
                    set_job_state(0, total, "Leyendo archivo...")

                # ───────────────────────────────────────────────
                # STUDENTS
                # ───────────────────────────────────────────────
                if type == "students":
                    student_role, _ = Role.objects.get_or_create(name="STUDENT")
                    user_fields = {f.name for f in User._meta.fields}

                    def _email_field_info():
                        try:
                            f = User._meta.get_field("email")
                            return {
                                "exists": True,
                                "unique": bool(getattr(f, "unique", False)),
                                "null": bool(getattr(f, "null", False))
                            }
                        except Exception:
                            return {"exists": False, "unique": False, "null": False}

                    EMAIL_INFO = _email_field_info()

                    def add_error_local(row, field, message):
                        add_error(row, field, message)

                    def _safe_unique_email(username: str, email: str):
                        email = (email or "").strip().lower()
                        if email:
                            conflict = User.objects.filter(email__iexact=email).exists()
                            if not conflict:
                                return email
                            add_error_local(None, "email", f"Email duplicado '{email}'")
                        if EMAIL_INFO["null"]:
                            return None
                        if EMAIL_INFO["unique"]:
                            base = f"{username}@no-email.local"
                            e = base
                            k = 1
                            while User.objects.filter(email__iexact=e).exists():
                                k += 1
                                e = f"{username}-{k}@no-email.local"
                            return e
                        return ""

                    def _set_name_fields(u, full_name: str):
                        if "full_name" in user_fields:
                            u.full_name = full_name
                        elif "name" in user_fields:
                            u.name = full_name

                    def _ensure_user_for_student(st: Student, username: str, email: str, full_name: str, r):
                        if getattr(st, "user_id", None):
                            UserRole.objects.get_or_create(user_id=st.user_id, role_id=student_role.id)
                            return st.user, None

                        email_clean = (email or "").strip().lower()
                        user = User.objects.filter(username=username).first()
                        if not user and email_clean:
                            user = User.objects.filter(email__iexact=email_clean).first()

                        if user and Student.objects.filter(user_id=user.id).exists():
                            add_error_local(r, "user", f"user '{getattr(user,'username','')}' ya enlazado")
                            return None, None

                        temp_password = None

                        if not user:
                            uname = username
                            k = 1
                            while User.objects.filter(username=uname).exists():
                                k += 1
                                uname = f"{username}-{k}"

                            user = User(username=uname, is_active=True, is_staff=False)
                            if "email" in user_fields:
                                user.email = _safe_unique_email(uname, email_clean)
                            _set_name_fields(user, full_name)

                            temp_password = get_random_string(10) + "!"
                            user.set_password(temp_password)

                            try:
                                user.save()
                            except IntegrityError as e:
                                if "email" in str(e).lower() and "email" in user_fields:
                                    user.email = _safe_unique_email(user.username, "")
                                    user.save()
                                    add_error_local(r, "email", "Choque email (reintento ok)")
                                else:
                                    raise
                        else:
                            changed = False
                            _set_name_fields(user, full_name)
                            changed = True

                            if "email" in user_fields and email_clean:
                                conflict = User.objects.filter(email__iexact=email_clean).exclude(id=user.id).exists()
                                if not conflict and (getattr(user, "email", "") or "").lower() != email_clean:
                                    user.email = email_clean
                                    changed = True

                            if not getattr(user, "is_active", True):
                                user.is_active = True
                                changed = True

                            if changed:
                                try:
                                    user.save()
                                except IntegrityError:
                                    add_error_local(r, "email", "Update email falló")

                        UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id)
                        st.user = user
                        st.save(update_fields=["user"])
                        return user, temp_password

                    done = 0
                    for row in rows:
                        done += 1
                        if done % 50 == 0:
                            set_job_state(done, total, "Importando alumnos...")

                        r = row.get("__row__", "?")
                        num_documento = "" if row.get("num_documento") is None else str(row.get("num_documento")).strip()
                        num_documento = num_documento.replace("\x00", "")
                        num_documento = re.sub(r"\.0$", "", num_documento)
                        num_documento = re.sub(r"\D", "", num_documento)
                        if num_documento.isdigit() and len(num_documento) < 8:
                            num_documento = num_documento.zfill(8)

                        nombres = str(row.get("nombres", "")).strip()
                        ap_pat = str(row.get("apellido_paterno", "")).strip()
                        ap_mat = str(row.get("apellido_materno", "")).strip()
                        sexo = str(row.get("sexo", "")).strip()

                        if not num_documento:
                            if not nombres and not ap_pat and not ap_mat:
                                continue
                            add_error_local(r, "num_documento", "Documento requerido")
                            continue
                        if not nombres:
                            add_error_local(r, "nombres", "Nombres requerido")
                            continue

                        periodo = str(row.get("periodo", "")).strip()
                        if periodo:
                            _ensure_period(periodo)

                        programa_carrera = str(row.get("programa_carrera", "")).strip()

                        plan_obj = None
                        if programa_carrera:
                            car = match_career_robust(programa_carrera)
                            if car:
                                plan_obj = pick_plan_for_student(car, periodo)

                        st, created = Student.objects.get_or_create(
                            num_documento=num_documento,
                            defaults={"nombres": nombres, "apellido_paterno": ap_pat, "apellido_materno": ap_mat},
                        )

                        st.nombres = nombres
                        st.apellido_paterno = ap_pat
                        st.apellido_materno = ap_mat
                        st.sexo = sexo
                        st.region = str(row.get("region", "")).strip()
                        st.provincia = str(row.get("provincia", "")).strip()
                        st.distrito = str(row.get("distrito", "")).strip()
                        st.codigo_modular = str(row.get("codigo_modular", "")).strip()
                        st.nombre_institucion = str(row.get("nombre_institucion", "")).strip()
                        st.gestion = str(row.get("gestion", "")).strip()
                        st.tipo = str(row.get("tipo", "")).strip()
                        st.programa_carrera = programa_carrera
                        st.ciclo = _to_int(row.get("ciclo", None))
                        st.turno = str(row.get("turno", "")).strip()
                        st.seccion = str(row.get("seccion", "")).strip()
                        st.periodo = periodo
                        st.lengua = str(row.get("lengua", "")).strip()
                        st.discapacidad = str(row.get("discapacidad", "")).strip()
                        st.tipo_discapacidad = str(row.get("tipo_discapacidad", "")).strip()
                        st.email = str(row.get("email", "") or "").strip().lower()
                        st.celular = str(row.get("celular", "") or "").strip()
                        st.plan = plan_obj

                        username = num_documento
                        full_name = f"{nombres} {ap_pat} {ap_mat}".strip()
                        user, temp_password = _ensure_user_for_student(st, username, st.email, full_name, r)

                        st.save()

                        if temp_password and user:
                            credentials.append({
                                "row": r,
                                "num_documento": num_documento,
                                "username": getattr(user, "username", username),
                                "password": temp_password
                            })

                        if created:
                            imported += 1
                        else:
                            updated += 1

                    set_job_state(total, total, "Finalizando alumnos...")

                # ───────────────────────────────────────────────
                # COURSES
                # ───────────────────────────────────────────────
                elif type == "courses":
                    done = 0
                    for row in rows:
                        done += 1
                        if done % 50 == 0:
                            set_job_state(done, total, "Importando cursos...")

                        r = row.get("__row__", "?")
                        code = str(row.get("code", "")).strip()
                        name = str(row.get("name", "")).strip()

                        if not code:
                            add_error(r, "code", "code requerido")
                            continue
                        if not name:
                            add_error(r, "name", "name requerido")
                            continue

                        credits = _to_int(row.get("credits"), None)
                        hours = _to_int(row.get("hours", None))

                        course, _ = Course.objects.get_or_create(
                            code=code,
                            defaults={"name": name, "credits": max(0, credits or 0)}
                        )
                        course.name = name
                        if credits is not None:
                            course.credits = max(0, int(credits))
                        course.save()

                        plan_id = _to_int(row.get("plan_id", None))
                        semester = _to_int(row.get("semester", None))
                        ctype = str(row.get("type", "")).strip().upper()

                        if plan_id and semester:
                            plan = Plan.objects.filter(id=plan_id).first()
                            if not plan:
                                add_error(r, "plan_id", f"plan_id {plan_id} no existe")
                            else:
                                if ctype in ("OBLIGATORIO", "MANDATORY", "M"):
                                    type_db = "MANDATORY"
                                elif ctype in ("ELECTIVO", "ELECTIVE", "E"):
                                    type_db = "ELECTIVE"
                                else:
                                    type_db = "MANDATORY"

                                pc, pc_created = PlanCourse.objects.update_or_create(
                                    plan=plan,
                                    course=course,
                                    defaults={
                                        "semester": max(1, int(semester)),
                                        "weekly_hours": max(1, int(hours or 3)),
                                        "type": type_db,
                                        "credits": max(0, int(credits or 0)) if credits is not None else 3,
                                        "display_code": course.code,
                                        "display_name": _normalize_course_name(name),
                                    },
                                )
                                if pc_created:
                                    imported += 1
                                else:
                                    updated += 1
                        else:
                            imported += 1

                    set_job_state(total, total, "Finalizando cursos...")

                # ───────────────────────────────────────────────
                # GRADES
                # ───────────────────────────────────────────────
                elif type == "grades":
                    if AcademicGradeRecord is None:
                        job.status = "FAILED"
                        job.result = {
                            **(job.result or {}),
                            "progress": 100,
                            "errors": [{"row": None, "field": "model", "message": "AcademicGradeRecord no existe"}],
                            "imported": imported,
                            "updated": updated,
                        }
                        job.save(update_fields=["status", "result"])
                        return

                    cal_rows = _read_calificaciones_xlsx(io.BytesIO(raw))
                    total2 = len(cal_rows)
                    set_job_state(0, total2, "Leyendo calificaciones...")

                    course_idx = _build_norm_index(Course, "name")

                    _all_electivo_pcs = list(
                        PlanCourse.objects
                        .select_related("course")
                        .filter(
                            models.Q(display_name__iexact="ELECTIVOS") |
                            models.Q(course__name__istartswith="ELECTIVOS")
                        )
                    )

                    done = 0
                    for row in cal_rows:
                        done += 1
                        if done % 50 == 0:
                            set_job_state(done, total2, "Importando notas...")

                        r = row.get("__row__", "?")
                        doc = row.get("doc")
                        term = row.get("periodo")
                        ciclo = row.get("ciclo")
                        course_name = row.get("curso")
                        final = row.get("nota")

                        if not doc:
                            add_error(r, "student_document", "Documento vacío")
                            continue
                        if not term:
                            add_error(r, "term", "Periodo vacío")
                            continue
                        if not course_name:
                            add_error(r, "course", "Curso vacío")
                            continue
                        if final is None:
                            add_error(r, "final_grade", "Nota vacía")
                            continue

                        st = Student.objects.filter(num_documento=doc).first()
                        if not st:
                            add_error(r, "student_document", f"No existe alumno {doc}")
                            continue

                        # ── Asignar plan si falta ──
                        if not st.plan_id and st.programa_carrera:
                            car = match_career_robust(st.programa_carrera)
                            if car:
                                plan = pick_plan_for_student(car, st.periodo or term)
                                if plan:
                                    st.plan_id = plan.id
                                    st.save(update_fields=["plan_id"])

                        pobj = _ensure_period(term)
                        term_code = pobj.code if pobj else term

                        course = None
                        pc_match = None

                        is_electivo_target = _is_electivos_name(course_name)

                        # A) Alumno con plan → buscar en PlanCourse
                        if st.plan_id:
                            pcs = list(
                                PlanCourse.objects
                                .select_related("course")
                                .filter(plan_id=st.plan_id)
                            )

                            target = _norm_key(course_name)

                            # A.1 ELECTIVOS con ciclo
                            if is_electivo_target and ciclo:
                                ciclo_int = int(ciclo)
                                for pcx in pcs:
                                    dn = _norm_key((getattr(pcx, "display_name", "") or "").strip())
                                    if dn == "electivos" and pcx.semester == ciclo_int:
                                        pc_match = pcx
                                        course = pcx.course
                                        break

                            # A.2 ELECTIVOS sin ciclo
                            if not course and is_electivo_target:
                                for pcx in pcs:
                                    dn = _norm_key((getattr(pcx, "display_name", "") or "").strip())
                                    if dn == "electivos":
                                        pc_match = pcx
                                        course = pcx.course
                                        break

                            # A.3 Match general
                            if not course:
                                for pcx in pcs:
                                    dn = _norm_key((getattr(pcx, "display_name", "") or "").strip())
                                    dc = _norm_key((getattr(pcx, "display_code", "") or "").strip())
                                    cn = _norm_key((getattr(pcx.course, "name", "") or "").strip())
                                    cc = _norm_key((getattr(pcx.course, "code", "") or "").strip())

                                    if target in (dn, cn) or target in (dc, cc):
                                        pc_match = pcx
                                        course = pcx.course
                                        break

                            # A.4 Match con texto limpio
                            if not course:
                                t2 = _norm_key(_clean_text(_normalize_course_name(course_name)))
                                for pcx in pcs:
                                    dn2 = _norm_key(_clean_text(getattr(pcx, "display_name", "") or ""))
                                    cn2 = _norm_key(_clean_text(getattr(pcx.course, "name", "") or ""))
                                    if t2 in (dn2, cn2):
                                        pc_match = pcx
                                        course = pcx.course
                                        break

                        # B) Fallback global
                        if not course:
                            course = _find_by_norm_cached(course_idx, course_name)

                        # B.1 Fallback ELECTIVOS
                        if not course and is_electivo_target:
                            if ciclo:
                                electivo_suffixed = f"ELECTIVOS - S{int(ciclo)}"
                                course = _find_by_norm_cached(course_idx, electivo_suffixed)

                            if not course:
                                for nk, c in course_idx.items():
                                    if nk.startswith("electivos"):
                                        course = c
                                        break

                            if not course and _all_electivo_pcs:
                                if ciclo:
                                    ciclo_int = int(ciclo)
                                    for epc in _all_electivo_pcs:
                                        if epc.semester == ciclo_int:
                                            course = epc.course
                                            if not pc_match and st.plan_id and epc.plan_id == st.plan_id:
                                                pc_match = epc
                                            break

                                if not course:
                                    epc = _all_electivo_pcs[0]
                                    course = epc.course
                                    if not pc_match and st.plan_id and epc.plan_id == st.plan_id:
                                        pc_match = epc

                        if not course:
                            add_error(r, "course", f"Curso '{course_name}' no existe")
                            continue

                        if not pc_match and st.plan_id and course:
                            pc_match = match_plan_course_for_grade(st, course)

                        rec, created = AcademicGradeRecord.objects.get_or_create(
                            student=st,
                            course=course,
                            term=str(term_code),
                            defaults={
                                "final_grade": float(final),
                                "components": {},
                                "plan_course": pc_match,
                            },
                        )

                        if not created:
                            rec.final_grade = float(final)
                            rec.plan_course = pc_match
                            rec.save(update_fields=["final_grade", "plan_course"])
                            updated += 1
                        else:
                            imported += 1

                    set_job_state(total2, total2, "Finalizando notas...")

            # ═══════════════════════════════════════════════════
            # EGRESADOS  ★ Con registro_pedagogico + director/secretario global ★
            # ═══════════════════════════════════════════════════
            elif type == "egresados":
                if not filename.endswith(".xlsx"):
                    job.status = "FAILED"
                    job.result = {
                        **(job.result or {}),
                        "progress": 100,
                        "errors": [{"row": None, "field": "file", "message": "Solo se aceptan archivos .xlsx para egresados"}],
                    }
                    job.save(update_fields=["status", "result"])
                    return

                from .egresados_reader import _read_egresados_file
                from graduates.models import Graduate

                bio_eg = io.BytesIO(raw)
                bio_eg.name = safe_name
                eg_rows = _read_egresados_file(bio_eg)

                total = len(eg_rows)
                set_job_state(0, total, "Leyendo egresados...")

                if not eg_rows:
                    add_error(None, "format", "No se detectaron registros de egresados en el archivo")

                # ── Obtener tipo de grado/título por defecto del catálogo ──
                default_gt_type = _get_default_grado_titulo_type()

                linked_to_student = 0
                with_sustentacion = 0
                with_resolucion = 0
                with_diploma = 0
                with_registro_ped = 0
                format_detected = set()
                done = 0

                for row in eg_rows:
                    done += 1
                    if done % 25 == 0:
                        set_job_state(done, total, f"Procesando egresados ({done}/{total})...")

                    r = row.get("__row__", "?")
                    fmt = row.get("__format__", "egreso")
                    format_detected.add(fmt)

                    dni = (row.get("dni") or "").strip()
                    nombre = (row.get("apellidos_nombres") or "").strip()
                    especialidad = (row.get("especialidad") or "").strip()
                    nivel = (row.get("nivel") or "").strip()
                    anio_ingreso = (row.get("anio_ingreso") or "").strip()
                    anio_egreso = (row.get("anio_egreso") or "").strip()
                    fecha_sust = row.get("fecha_sustentacion")

                    # Campos adicionales (verificador)
                    grado_titulo_raw = (row.get("grado_titulo") or "").strip()
                    resolucion_acta = (row.get("resolucion_acta") or "").strip()
                    codigo_diploma = (row.get("codigo_diploma") or "").strip()
                    registro_pedagogico = (row.get("registro_pedagogico") or "").strip()
                    director_general = (row.get("director_general") or "").strip()
                    secretario_academico = (row.get("secretario_academico") or "").strip()

                    if not nombre:
                        add_error(r, "apellidos_nombres", "Nombre vacío")
                        continue

                    # Lookup: (dni + anio_egreso) o (apellidos_nombres + anio_egreso)
                    grad = None
                    if dni and anio_egreso:
                        grad = Graduate.objects.filter(
                            dni=dni, anio_egreso=anio_egreso
                        ).first()
                    if not grad and dni:
                        grad = Graduate.objects.filter(dni=dni).first()
                    if not grad and nombre and anio_egreso:
                        grad = Graduate.objects.filter(
                            apellidos_nombres__iexact=nombre,
                            anio_egreso=anio_egreso,
                        ).first()

                    created = False
                    if not grad:
                        grad = Graduate(is_active=True)
                        created = True

                    # ── Asignar campos base ──
                    if dni:
                        grad.dni = dni
                    grad.apellidos_nombres = nombre
                    if especialidad:
                        grad.especialidad = especialidad
                    if nivel:
                        grad.nivel = nivel
                    if anio_ingreso:
                        grad.anio_ingreso = anio_ingreso
                    if anio_egreso:
                        grad.anio_egreso = anio_egreso
                    if fecha_sust:
                        grad.fecha_sustentacion = fecha_sust
                        with_sustentacion += 1

                    # ── Campos de PLANTILLA_VERIFICADOR / TITULADOS ──
                    if resolucion_acta:
                        grad.resolucion_acta = resolucion_acta
                        with_resolucion += 1
                    if codigo_diploma:
                        grad.codigo_diploma = codigo_diploma
                        with_diploma += 1
                    if registro_pedagogico:
                        grad.registro_pedagogico = registro_pedagogico
                        with_registro_ped += 1
                    if director_general:
                        grad.director_general = director_general
                    if secretario_academico:
                        grad.secretario_academico = secretario_academico

                    # ── Asignar grado/título ──
                    if grado_titulo_raw:
                        grad.grado_titulo = grado_titulo_raw
                        if default_gt_type and not grad.grado_titulo_type_id:
                            grad.grado_titulo_type = default_gt_type
                    else:
                        needs_gt_update = (
                            not grad.grado_titulo
                            or not grad.grado_titulo_type_id
                            or "BACHILLER" in (grad.grado_titulo or "").upper()
                        )
                        if needs_gt_update:
                            if default_gt_type:
                                grad.grado_titulo_type = default_gt_type
                                grad.grado_titulo = default_gt_type.render(
                                    especialidad or grad.especialidad
                                )
                            else:
                                grad.grado_titulo_type = None
                                grad.grado_titulo = _derive_grado_titulo_fallback(
                                    especialidad or grad.especialidad
                                )

                    grad.save()

                    # Intentar vincular con alumno existente
                    if dni:
                        student = Student.objects.filter(num_documento=dni).first()
                        if student:
                            linked_to_student += 1

                    if created:
                        imported += 1
                    else:
                        updated += 1

                set_job_state(total, total, "Finalizando egresados...")

                job.result = {
                    **(job.result or {}),
                    "imported": imported,
                    "updated": updated,
                    "errors": errors,
                    "stats": {
                        "total_graduates": total,
                        "linked_to_student": linked_to_student,
                        "with_sustentacion": with_sustentacion,
                        "with_resolucion": with_resolucion,
                        "with_diploma": with_diploma,
                        "with_registro_ped": with_registro_ped,
                        "format_detected": sorted(list(format_detected)),
                        "grado_titulo_type": default_gt_type.name if default_gt_type else "(derivado de especialidad)",
                    },
                    "summary": {
                        "ok": imported + updated,
                        "failed": len(errors),
                        "note": (
                            f"{imported} nuevos, {updated} actualizados. "
                            f"{linked_to_student} vinculados con alumnos. "
                            f"{with_resolucion} con resolución, {with_diploma} con diploma, "
                            f"{with_registro_ped} con registro pedagógico."
                        ),
                    },
                }

            # ───────────────────────────────────────────────
            # TRASLADOS (alumno + notas históricas en 1 Excel)
            # ───────────────────────────────────────────────
            elif type == "traslados":
                if AcademicGradeRecord is None:
                    job.status = "FAILED"
                    job.result = {
                        **(job.result or {}),
                        "progress": 100,
                        "errors": [{"row": None, "field": "model", "message": "AcademicGradeRecord no existe"}],
                    }
                    job.save(update_fields=["status", "result"])
                    return

                traslado_rows = _read_traslados_xlsx(io.BytesIO(raw))
                total_t = len(traslado_rows)
                set_job_state(0, total_t, "Leyendo traslados...")

                if total_t == 0:
                    add_error(None, "file", "No se encontraron filas válidas en el Excel")

                # Agrupar filas por num_documento
                from collections import OrderedDict
                students_map = OrderedDict()
                for row in traslado_rows:
                    doc = row["doc"]
                    if doc not in students_map:
                        students_map[doc] = {
                            "info": row,
                            "grades": [],
                        }
                    students_map[doc]["grades"].append(row)

                # Setup para crear usuarios
                student_role, _ = Role.objects.get_or_create(name="STUDENT")
                user_fields = {f.name for f in User._meta.fields}

                def _email_exists():
                    try:
                        f = User._meta.get_field("email")
                        return {
                            "exists": True,
                            "unique": bool(getattr(f, "unique", False)),
                            "null": bool(getattr(f, "null", False)),
                        }
                    except Exception:
                        return {"exists": False, "unique": False, "null": False}

                email_info = _email_exists()

                # Índice de cursos
                course_idx = _build_norm_index(Course, "name")

                done_t = 0
                students_created = 0
                students_updated = 0
                grades_created = 0
                grades_updated = 0

                for doc, data in students_map.items():
                    done_t += 1
                    if done_t % 10 == 0:
                        set_job_state(done_t, len(students_map), "Procesando traslados...")

                    info = data["info"]
                    r_ref = info.get("__row__", "?")
                    nombres = info.get("nombres", "")
                    ap_pat = info.get("ap_pat", "")
                    ap_mat = info.get("ap_mat", "")
                    programa = info.get("programa", "")
                    ciclo = info.get("ciclo")
                    email = info.get("email", "")
                    celular = info.get("celular", "")

                    if not nombres:
                        add_error(r_ref, "nombres", f"Nombres vacío para {doc}")
                        continue

                    # 1) Crear o actualizar alumno
                    st, st_created = Student.objects.get_or_create(
                        num_documento=doc,
                        defaults={
                            "nombres": nombres,
                            "apellido_paterno": ap_pat,
                            "apellido_materno": ap_mat,
                        },
                    )

                    st.nombres = nombres
                    st.apellido_paterno = ap_pat
                    st.apellido_materno = ap_mat
                    if info.get("sexo"):
                        st.sexo = info["sexo"]
                    if programa:
                        st.programa_carrera = programa
                    if ciclo:
                        st.ciclo = ciclo
                    if email:
                        st.email = email
                    if celular:
                        st.celular = celular

                    # Asignar plan
                    plan_obj = None
                    if programa:
                        car = match_career_robust(programa)
                        if car:
                            periodo_ref = data["grades"][0].get("periodo", "") if data["grades"] else ""
                            plan_obj = pick_plan_for_student(car, periodo_ref)
                            if plan_obj:
                                st.plan = plan_obj

                    # Crear usuario
                    username = doc
                    full_name = f"{nombres} {ap_pat} {ap_mat}".strip()

                    if not getattr(st, "user_id", None):
                        user = User.objects.filter(username=username).first()
                        if not user and email:
                            user = User.objects.filter(email__iexact=email).first()

                        if user and Student.objects.filter(user_id=user.id).exists():
                            add_error(r_ref, "user", f"user '{getattr(user, 'username', '')}' ya enlazado a otro alumno")
                        else:
                            temp_password = None
                            if not user:
                                uname = username
                                k = 1
                                while User.objects.filter(username=uname).exists():
                                    k += 1
                                    uname = f"{username}-{k}"

                                user = User(username=uname, is_active=True, is_staff=False)
                                if "email" in user_fields:
                                    em = (email or "").strip().lower()
                                    if em and not User.objects.filter(email__iexact=em).exists():
                                        user.email = em
                                    elif email_info["null"]:
                                        user.email = None
                                    elif email_info["unique"]:
                                        user.email = f"{uname}@no-email.local"
                                    else:
                                        user.email = ""
                                if "full_name" in user_fields:
                                    user.full_name = full_name
                                elif "name" in user_fields:
                                    user.name = full_name

                                temp_password = get_random_string(10) + "!"
                                user.set_password(temp_password)
                                try:
                                    user.save()
                                except IntegrityError:
                                    user.email = f"{uname}-x@no-email.local"
                                    user.save()

                                UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id)

                                if temp_password:
                                    credentials.append({
                                        "row": r_ref,
                                        "num_documento": doc,
                                        "username": getattr(user, "username", username),
                                        "password": temp_password,
                                    })
                            else:
                                UserRole.objects.get_or_create(user_id=user.id, role_id=student_role.id)

                            if user:
                                st.user = user

                    st.save()

                    if st_created:
                        students_created += 1
                    else:
                        students_updated += 1

                    # 2) Crear notas para este alumno
                    for grade_row in data["grades"]:
                        gr = grade_row.get("__row__", "?")
                        periodo = grade_row.get("periodo", "")
                        curso_name = grade_row.get("curso", "")
                        promedio = grade_row.get("promedio")
                        components = grade_row.get("components", {})

                        if not periodo or not curso_name or promedio is None:
                            add_error(gr, "grade", "Periodo/curso/promedio vacío")
                            continue

                        pobj = _ensure_period(periodo)
                        term_code = pobj.code if pobj else periodo

                        # Resolver curso
                        course = None
                        pc_match = None
                        is_electivo_target = _is_electivos_name(curso_name)

                        if st.plan_id:
                            pcs = list(
                                PlanCourse.objects
                                .select_related("course")
                                .filter(plan_id=st.plan_id)
                            )
                            target = _norm_key(curso_name)

                            for pcx in pcs:
                                dn = _norm_key((getattr(pcx, "display_name", "") or "").strip())
                                cn = _norm_key((getattr(pcx.course, "name", "") or "").strip())
                                if target in (dn, cn):
                                    pc_match = pcx
                                    course = pcx.course
                                    break

                            if not course:
                                t2 = _norm_key(_clean_text(_normalize_course_name(curso_name)))
                                for pcx in pcs:
                                    dn2 = _norm_key(_clean_text(getattr(pcx, "display_name", "") or ""))
                                    cn2 = _norm_key(_clean_text(getattr(pcx.course, "name", "") or ""))
                                    if t2 in (dn2, cn2):
                                        pc_match = pcx
                                        course = pcx.course
                                        break

                        if not course:
                            course = _find_by_norm_cached(course_idx, curso_name)

                        if not course:
                            add_error(gr, "course", f"Curso '{curso_name}' no encontrado")
                            continue

                        if not pc_match and st.plan_id and course:
                            pc_match = match_plan_course_for_grade(st, course)

                        rec, g_created = AcademicGradeRecord.objects.get_or_create(
                            student=st,
                            course=course,
                            term=str(term_code),
                            defaults={
                                "final_grade": float(promedio),
                                "components": components,
                                "plan_course": pc_match,
                            },
                        )

                        if not g_created:
                            rec.final_grade = float(promedio)
                            rec.components = components
                            rec.plan_course = pc_match
                            rec.save(update_fields=["final_grade", "components", "plan_course"])
                            grades_updated += 1
                        else:
                            grades_created += 1

                    imported += grades_created
                    updated += grades_updated

                set_job_state(
                    len(students_map), len(students_map),
                    f"Traslados finalizados: {students_created} alumnos creados, "
                    f"{students_updated} actualizados, {grades_created} notas creadas, "
                    f"{grades_updated} notas actualizadas.",
                )

            else:
                job.status = "FAILED"
                job.result = {
                    **(job.result or {}),
                    "progress": 100,
                    "errors": [{"row": None, "field": "type", "message": "Tipo inválido"}]
                }
                job.save(update_fields=["status", "result"])
                return

            # Estado final
            if len(errors) == 0:
                final_status = "COMPLETED"
            elif (imported + updated) > 0:
                final_status = "COMPLETED_WITH_ERRORS"
            else:
                final_status = "FAILED"

            job.status = final_status
            job.result = {
                **(job.result or {}),
                "progress": 100,
                "imported": imported,
                "updated": updated,
                "errors": errors,
                "credentials": credentials[:300],
                "summary": {
                    "ok": imported + updated,
                    "failed": len(errors),
                    "note": "Corrige los errores y vuelve a importar.",
                }
            }
            job.save(update_fields=["status", "result"])

        except Exception as e:
            job.status = "FAILED"
            job.result = {
                **(job.result or {}),
                "progress": 100,
                "errors": errors + [{"row": None, "field": "exception", "message": str(e)}],
                "imported": imported,
                "updated": updated,
            }
            job.save(update_fields=["status", "result"])

    finally:
        close_old_connections()


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def imports_start(request, type: str):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    type = (type or "").lower().strip()
    up = request.FILES.get("file")

    if not up:
        return Response({"detail": "file requerido"}, status=400)

    mapping_raw = request.POST.get("mapping")
    mapping = {}
    if mapping_raw:
        try:
            import json
            mapping = json.loads(mapping_raw)
        except Exception:
            mapping = {}

    raw = up.read()
    safe_name = (getattr(up, "name", "") or "import.xlsx").split("/")[-1].split("\\")[-1]

    job = ImportJob.objects.create(
        type=type,
        file=ContentFile(raw, name=safe_name),
        mapping=mapping,
        user=request.user,
        status="RUNNING",
        result={
            "progress": 0,
            "total": 0,
            "processed": 0,
            "errors": [],
            "imported": 0,
            "updated": 0,
            "message": "Encolado...",
        }
    )

    close_old_connections()

    t = threading.Thread(
        target=_run_import_job,
        args=(job.id, raw, safe_name, type, mapping),
        daemon=True
    )
    t.start()

    return Response({"job_id": job.id}, status=202)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_status(request, jobId: int):
    try:
        job = ImportJob.objects.get(pk=jobId)
    except ImportJob.DoesNotExist:
        return Response({"detail": "job not found"}, status=404)

    result = job.result if isinstance(job.result, dict) else {}
    processed = int(result.get("processed") or 0)
    total = int(result.get("total") or 0)
    progress = result.get("progress")

    if progress is None:
        progress = int((processed / total) * 100) if total else 0

    if job.status in ("COMPLETED", "COMPLETED_WITH_ERRORS", "FAILED", "ERROR"):
        progress = 100

    return Response({
        "id": job.id,
        "state": job.status,
        "progress": int(progress),
        "processed": processed,
        "total": total,
        "message": result.get("message") or "",
        "errors": result.get("errors") or [],
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "credentials": result.get("credentials") or [],
        "summary": result.get("summary") or {},
        "stats": result.get("stats") or {},
    })