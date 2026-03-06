"""
Endpoints administrativos para gestión de egresados.
─────────────────────────────────────────────────────
catalogs/views/egresados.py

GET  /api/catalogs/egresados/           → Listado paginado con filtros
GET  /api/catalogs/egresados/stats/     → Estadísticas por año y especialidad
PATCH /api/catalogs/egresados/<id>/     → Editar campos del egresado
GET  /api/catalogs/egresados/export/    → Exportar formato PLANTILLA_VERIFICADOR
"""

import io
from datetime import date

from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from graduates.models import Graduate


# ═══════════════════════════════════════════════════════════════════════════════
# LISTADO
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def egresados_list(request):
    """
    GET /api/catalogs/egresados/?especialidad=X&year=2023&search=GARCIA&sustentado=1&page=1
    """
    qs = Graduate.objects.filter(is_active=True)

    # Filtros
    esp = request.query_params.get("especialidad", "").strip()
    if esp:
        qs = qs.filter(especialidad__iexact=esp)

    year = request.query_params.get("year", "").strip()
    if year:
        qs = qs.filter(anio_egreso__icontains=year)

    search = request.query_params.get("search", "").strip()
    if search:
        for term in search.upper().split():
            if len(term) >= 2:
                qs = qs.filter(
                    Q(apellidos_nombres__icontains=term) | Q(dni__icontains=term)
                )

    sustentado = request.query_params.get("sustentado", "").strip()
    if sustentado == "1":
        qs = qs.filter(fecha_sustentacion__isnull=False)
    elif sustentado == "0":
        qs = qs.filter(fecha_sustentacion__isnull=True)

    total = qs.count()

    # Paginación
    page = max(1, int(request.query_params.get("page", 1)))
    per_page = min(100, max(10, int(request.query_params.get("per_page", 25))))
    start = (page - 1) * per_page

    items = qs.order_by("apellidos_nombres")[start : start + per_page]
    data = [
        {
            "id": g.id,
            "dni": g.dni,
            "apellidos_nombres": g.apellidos_nombres,
            "grado_titulo": g.grado_titulo,
            "especialidad": g.especialidad,
            "nivel": g.nivel,
            "anio_ingreso": g.anio_ingreso,
            "anio_egreso": g.anio_egreso,
            "fecha_sustentacion": (
                g.fecha_sustentacion.strftime("%d/%m/%Y") if g.fecha_sustentacion else None
            ),
            "resolucion_acta": g.resolucion_acta,
            "codigo_diploma": g.codigo_diploma,
            "tiene_constancia": g.tiene_constancia,
        }
        for g in items
    ]

    return Response({
        "results": data,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if total else 1,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# ESTADÍSTICAS
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def egresados_stats(request):
    """GET /api/catalogs/egresados/stats/"""
    qs = Graduate.objects.filter(is_active=True)

    return Response({
        "total": qs.count(),
        "with_sustentacion": qs.filter(fecha_sustentacion__isnull=False).count(),
        "with_dni": qs.exclude(dni="").count(),
        "with_constancia": qs.filter(
            Q(resolucion_acta__gt="") | Q(codigo_diploma__gt="")
        ).count(),
        "by_year": list(
            qs.values("anio_egreso")
            .annotate(count=Count("id"))
            .order_by("anio_egreso")
        ),
        "by_especialidad": list(
            qs.values("especialidad")
            .annotate(count=Count("id"))
            .order_by("especialidad")
        ),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# EDICIÓN
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def egresados_update(request, pk):
    """
    PATCH /api/catalogs/egresados/<id>/
    Body: { grado_titulo, resolucion_acta, codigo_diploma, fecha_sustentacion,
            director_general, secretario_academico, dni, especialidad, ... }
    """
    try:
        grad = Graduate.objects.get(pk=pk)
    except Graduate.DoesNotExist:
        return Response({"detail": "Egresado no encontrado."}, status=404)

    EDITABLE = [
        "dni", "apellidos_nombres", "grado_titulo", "especialidad", "nivel",
        "anio_ingreso", "anio_egreso", "resolucion_acta", "codigo_diploma",
        "director_general", "secretario_academico",
    ]
    changed = False
    for field in EDITABLE:
        if field in request.data:
            setattr(grad, field, request.data[field])
            changed = True

    # fecha_sustentacion (special: date or null)
    if "fecha_sustentacion" in request.data:
        v = request.data["fecha_sustentacion"]
        if v:
            if isinstance(v, str):
                from datetime import datetime as dt
                for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        v = dt.strptime(v, fmt).date()
                        break
                    except ValueError:
                        continue
            grad.fecha_sustentacion = v
        else:
            grad.fecha_sustentacion = None
        changed = True

    if changed:
        grad.save()

    return Response({"ok": True, "id": grad.id})


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN (formato PLANTILLA_VERIFICADOR)
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def egresados_export(request):
    """
    GET /api/catalogs/egresados/export/?especialidad=X&year=2023
    Descarga Excel en formato PLANTILLA_VERIFICADOR.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    qs = Graduate.objects.filter(is_active=True)

    esp = request.query_params.get("especialidad", "").strip()
    if esp:
        qs = qs.filter(especialidad__iexact=esp)

    year = request.query_params.get("year", "").strip()
    if year:
        qs = qs.filter(anio_egreso__icontains=year)

    qs = qs.order_by("especialidad", "apellidos_nombres")

    wb = Workbook()
    ws = wb.active
    ws.title = "Verificador"

    # ── Headers ───────────────────────────────────────────────────────────
    headers = [
        "N°",
        "Director General",
        "Secretario Académico",
        "DNI",
        "Apellidos y Nombres",
        "Grado / Título / Capacitación",
        "Especialidad",
        "Año de Ingreso",
        "Año de Egreso",
        "Fecha de Sustentación",
        "Resolución / Acta",
        "Código de Diploma",
    ]

    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0C2340")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    # ── Anchos ────────────────────────────────────────────────────────────
    widths = [5, 25, 25, 10, 40, 30, 25, 12, 12, 15, 20, 20]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.auto_filter.ref = "A1:L1"

    # ── Data ──────────────────────────────────────────────────────────────
    data_font = Font(name="Calibri", size=9)
    data_align = Alignment(vertical="center", wrap_text=True)

    for idx, g in enumerate(qs, start=1):
        values = [
            idx,
            g.director_general,
            g.secretario_academico,
            g.dni,
            g.apellidos_nombres,
            g.grado_titulo,
            g.especialidad,
            g.anio_ingreso,
            g.anio_egreso,
            g.fecha_sustentacion.strftime("%d/%m/%Y") if g.fecha_sustentacion else "",
            g.resolucion_acta,
            g.codigo_diploma,
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=idx + 1, column=ci, value=v)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin

    ws.freeze_panes = "A2"

    # ── Response ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Verificador_Egresados_{date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp