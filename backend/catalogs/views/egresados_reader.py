"""
Reader para archivos de egresados (.xlsx)
──────────────────────────────────────────
catalogs/views/egresados_reader.py

Auto-detecta DOS formatos:

 A) EGRESO_XXXX.xlsx  (formato institucional)
    - Múltiples secciones por especialidad
    - Marcadores "ESPECIALIDAD: X"
    - Columnas: N° | CÓDIGO MAT. | APELLIDOS Y NOMBRES | ESP | NIVEL | INGRESO | EGRESO | SUSTENT.

 B) PLANTILLA_VERIFICADOR / TITULADOS.xlsx  (formato SUNEDU-style, 12-13 columnas)
    - Una sola tabla con headers:
      [N°] | Director General | Secretario Académico | DNI | NOMBRES Y APELLIDOS |
      GRADO/TÍTULO/CAPACITACIÓN | ESPECIALIDAD | AÑO INGRESO | AÑO EGRESO |
      Fecha Sustentación | Resolución/Acta | Código de Diploma | [Registro Pedagógico]
"""

import re
import io
from datetime import datetime, date
from typing import List, Optional


# ═══════════════════════════════════════════════════════════════
# HELPERS INTERNOS
# ═══════════════════════════════════════════════════════════════

def _norm_egreso(s) -> str:
    if s is None:
        return ""
    import unicodedata
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def _normalize_especialidad(esp: str) -> str:
    esp = esp.replace("\n", " ").replace("\\n", " ")
    esp = re.sub(r"\s+", " ", esp).strip().upper()
    esp = re.sub(r"INICIALL\b", "INICIAL", esp, flags=re.IGNORECASE)
    esp = re.sub(r"COMPUTACI[ÓO]N\s+E\s+INF[\w.]*\s*$", "COMPUTACIÓN E INFORMÁTICA", esp, flags=re.IGNORECASE)
    esp = re.sub(r"EDUCACI[ÓO]N\s+F[ÍI]SICA\b", "EDUCACIÓN FÍSICA", esp, flags=re.IGNORECASE)
    esp = re.sub(r"EDUCACI[ÓO]N\s+INICIAL\b", "EDUCACIÓN INICIAL", esp, flags=re.IGNORECASE)
    esp = re.sub(r"EDUCACI[ÓO]N\s+PRIMARIA\b", "EDUCACIÓN PRIMARIA", esp, flags=re.IGNORECASE)
    esp = re.sub(r"^E\.P\.?\s*$", "EDUCACIÓN PRIMARIA", esp, flags=re.IGNORECASE)
    esp = re.sub(r"^EDUC\.\s*FISICA\s*$", "EDUCACIÓN FÍSICA", esp, flags=re.IGNORECASE)
    return esp


def _extract_especialidad_from_marker(row_cells) -> Optional[str]:
    for cell in row_cells:
        if cell is None:
            continue
        s = str(cell).strip()
        m = re.match(r"ESPECIALIDAD\s*:\s*(.+)", s, re.IGNORECASE)
        if m:
            return _normalize_especialidad(m.group(1).strip())
    return None


def _parse_sustentacion_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_period_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        y = int(v)
        if 2000 <= y <= 2100:
            return str(y)
        return ""
    s = str(v).strip()
    s = re.sub(r"\s*[-–]\s*", "-", s)
    return s.upper()


def _clean_name(s) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _clean_dni(v) -> str:
    """Limpia DNI: quita acentos sueltos (´07289521), espacios, caracteres raros."""
    if v is None:
        return ""
    s = str(v).strip()
    # Quitar acentos sueltos y comillas al inicio/fin
    s = s.strip("´`'\"  \t\n\r\xa0")
    s = re.sub(r"\.0$", "", s)
    # Quitar todo excepto dígitos
    s = re.sub(r"\D", "", s)
    if s and len(s) < 8:
        s = s.zfill(8)
    if len(s) == 8 and s.isdigit():
        return s
    return ""


def _extract_doc_from_codigo(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D", "", s)
    if s.isdigit() and len(s) < 8:
        s = s.zfill(8)
    return s


def _clean_str_field(v) -> str:
    """Limpia un campo string genérico."""
    if v is None:
        return ""
    s = str(v).strip()
    s = s.strip("´`'\"  \t\n\r\xa0")
    return s


# ═══════════════════════════════════════════════════════════════
# DETECCIÓN DE FORMATO
# ═══════════════════════════════════════════════════════════════

def _detect_format(ws) -> str:
    """
    Analiza las primeras filas para determinar si es:
      'verificador' → PLANTILLA_VERIFICADOR / TITULADOS (tiene Director General, Resolución, Diploma, etc.)
      'egreso'      → EGRESO_XXXX (formato institucional)
    """
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        norms = [_norm_egreso(c).replace("\n", " ") for c in (row or [])]

        has_director = any("director" in n for n in norms)
        has_resolucion = any("resolucion" in n or "acta" in n for n in norms)
        has_diploma = any("diploma" in n for n in norms)
        has_grado = any("grado" in n or "titulo" in n or "capacitacion" in n for n in norms)
        has_registro = any("registro" in n and "pedagogico" in n for n in norms)

        if (has_director or has_resolucion or has_diploma or has_registro) and has_grado:
            return "verificador"

        if has_resolucion or has_diploma:
            return "verificador"

    return "egreso"


# ═══════════════════════════════════════════════════════════════
# READER: PLANTILLA_VERIFICADOR / TITULADOS
# ═══════════════════════════════════════════════════════════════

def _read_verificador_sheet(ws, sheet_name: str) -> List[dict]:
    """Lee formato PLANTILLA_VERIFICADOR / TITULADOS (12-13 columnas)."""
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return []

    col_map = {}
    header_row_idx = None

    for ridx, row in enumerate(rows_iter):
        norms = [_norm_egreso(c).replace("\n", " ").replace("\\n", " ") for c in (row or [])]

        has_dni = any(n == "dni" for n in norms)
        has_nombre = any("apellidos" in n or "nombres" in n for n in norms)
        has_grado = any("grado" in n or "titulo" in n for n in norms)

        if (has_dni or has_nombre) and has_grado:
            for i, n in enumerate(norms):
                n_clean = re.sub(r"\s+", " ", n).strip()

                if n_clean in ("n", "n°", "no"):
                    col_map["n"] = i
                elif "director" in n_clean:
                    col_map["director_general"] = i
                elif "secretario" in n_clean:
                    col_map["secretario_academico"] = i
                elif n_clean == "dni":
                    col_map["dni"] = i
                elif "apellidos" in n_clean or "nombres" in n_clean:
                    col_map["nombre"] = i
                elif "grado" in n_clean or "titulo" in n_clean or "capacitacion" in n_clean:
                    col_map["grado_titulo"] = i
                elif "especialidad" in n_clean:
                    col_map["especialidad"] = i
                elif "ingreso" in n_clean:
                    col_map["ingreso"] = i
                elif "egreso" in n_clean:
                    col_map["egreso"] = i
                elif "sustentac" in n_clean or "sustent" in n_clean:
                    col_map["sustentacion"] = i
                elif "resolucion" in n_clean or "acta" in n_clean:
                    col_map["resolucion_acta"] = i
                elif "diploma" in n_clean and "registro" not in n_clean:
                    col_map["codigo_diploma"] = i
                elif "registro" in n_clean and "pedagogico" in n_clean:
                    col_map["registro_pedagogico"] = i
                elif "codigo" in n_clean and "mat" not in n_clean and "diploma" not in n_clean:
                    # Fallback genérico para "Código de Diploma"
                    if "codigo_diploma" not in col_map:
                        col_map["codigo_diploma"] = i

            header_row_idx = ridx
            break

    if header_row_idx is None:
        return []

    # ── Director / Secretario global (primer registro que tenga dato) ──
    global_director = ""
    global_secretario = ""
    dir_col = col_map.get("director_general")
    sec_col = col_map.get("secretario_academico")

    if dir_col is not None or sec_col is not None:
        for row in rows_iter[header_row_idx + 1:]:
            if not row:
                continue
            if dir_col is not None and dir_col < len(row) and not global_director:
                v = _clean_str_field(row[dir_col])
                if v:
                    global_director = v
            if sec_col is not None and sec_col < len(row) and not global_secretario:
                v = _clean_str_field(row[sec_col])
                if v:
                    global_secretario = v
            if global_director and global_secretario:
                break

    results = []
    for ridx, row in enumerate(rows_iter[header_row_idx + 1:], start=header_row_idx + 2):
        if not row:
            continue

        def get(key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        nombre = _clean_name(get("nombre"))
        dni = _clean_dni(get("dni"))

        if not nombre and not dni:
            n_val = get("n")
            if n_val is None:
                continue
            try:
                if int(float(str(n_val).strip())) < 1:
                    continue
            except (ValueError, TypeError):
                continue
            if not nombre:
                continue

        if not nombre:
            continue

        esp_raw = get("especialidad")
        especialidad = _normalize_especialidad(str(esp_raw).strip()) if esp_raw else ""

        gt_raw = get("grado_titulo")
        grado_titulo = str(gt_raw).strip() if gt_raw else ""

        dir_raw = get("director_general")
        director = _clean_str_field(dir_raw) if dir_raw else ""

        sec_raw = get("secretario_academico")
        secretario = _clean_str_field(sec_raw) if sec_raw else ""

        res_raw = get("resolucion_acta")
        resolucion = _clean_str_field(res_raw) if res_raw else ""

        dip_raw = get("codigo_diploma")
        diploma = str(dip_raw).strip() if dip_raw else ""

        # Registro Pedagógico (columna 13 si existe)
        rp_raw = get("registro_pedagogico")
        registro_ped = _clean_str_field(rp_raw) if rp_raw else ""

        anio_ingreso = _parse_period_value(get("ingreso"))
        anio_egreso = _parse_period_value(get("egreso"))
        fecha_sustentacion = _parse_sustentacion_date(get("sustentacion"))

        # Aplicar director/secretario global si la fila no tiene
        if not director and global_director:
            director = global_director
        if not secretario and global_secretario:
            secretario = global_secretario

        results.append({
            "__row__": ridx,
            "__sheet__": sheet_name,
            "__format__": "verificador",
            "dni": dni,
            "apellidos_nombres": nombre,
            "grado_titulo": grado_titulo,
            "especialidad": especialidad,
            "nivel": "",
            "anio_ingreso": anio_ingreso,
            "anio_egreso": anio_egreso,
            "fecha_sustentacion": fecha_sustentacion,
            "resolucion_acta": resolucion,
            "codigo_diploma": diploma,
            "registro_pedagogico": registro_ped,
            "director_general": director,
            "secretario_academico": secretario,
        })

    return results


# ═══════════════════════════════════════════════════════════════
# READER: EGRESO_XXXX (original)
# ═══════════════════════════════════════════════════════════════

def _is_header_row_egreso(row_cells) -> bool:
    norms = [_norm_egreso(c).replace("\n", " ").replace("\\n", " ") for c in row_cells]
    has_n = any(n.strip() in ("n", "no", "n°", "n ") for n in norms)
    has_name = any("apellidos" in n or "nombres" in n for n in norms)
    has_codigo_as_name = any("codigo" in n and "mat" in n for n in norms)
    has_egreso = any("egreso" in n for n in norms)
    return has_n and (has_name or (has_codigo_as_name and has_egreso))


def _detect_columns_egreso(header_cells) -> dict:
    cols = {}
    for i, cell in enumerate(header_cells):
        n = _norm_egreso(cell).replace("\n", " ").replace("\\n", " ")
        n = re.sub(r"\s+", " ", n).strip()

        if n in ("n", "no", "n°", "n "):
            cols["n"] = i
        elif ("codigo" in n and "mat" in n) or n == "codigo":
            cols["codigo"] = i
        elif "apellidos" in n or "nombres" in n:
            cols["nombre"] = i
        elif n == "especialidad":
            cols["especialidad"] = i
        elif n == "nivel":
            cols["nivel"] = i
        elif "ingreso" in n:
            cols["ingreso"] = i
        elif "egreso" in n:
            cols["egreso"] = i
        elif "sustentac" in n or "sustent" in n:
            cols["sustentacion"] = i

    if "nombre" not in cols and "codigo" in cols:
        cols["nombre"] = cols["codigo"]
        del cols["codigo"]

    return cols


def _read_egreso_sheet(ws, sheet_name: str) -> List[dict]:
    """Lee formato EGRESO_XXXX (institucional, con secciones por especialidad)."""
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter or len(rows_iter) < 2:
        return []

    results = []
    current_especialidad = ""
    last_data_esp = ""
    current_cols = {}
    header_found = False

    for ridx, row in enumerate(rows_iter, start=1):
        esp = _extract_especialidad_from_marker(row)
        if esp:
            current_especialidad = esp
            header_found = False
            continue

        if _is_header_row_egreso(row):
            current_cols = _detect_columns_egreso(row)
            header_found = True
            continue

        if not header_found or not current_cols:
            continue

        n_col = current_cols.get("n")
        if n_col is not None and n_col < len(row):
            n_val = row[n_col]
            if n_val is None or (isinstance(n_val, str) and not n_val.strip()):
                continue
            try:
                if int(float(str(n_val).strip())) < 1:
                    continue
            except (ValueError, TypeError):
                continue
        else:
            continue

        def get_col(key):
            idx = current_cols.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        nombre_raw = get_col("nombre")
        codigo_raw = get_col("codigo")

        nombre = _clean_name(nombre_raw)
        dni = _extract_doc_from_codigo(codigo_raw) if codigo_raw else ""

        if not nombre and codigo_raw:
            s = str(codigo_raw).strip()
            if not s.replace(".", "").isdigit():
                nombre = _clean_name(s)
                dni = ""

        if not nombre:
            continue

        esp_fila = get_col("especialidad")
        especialidad = ""
        if esp_fila:
            especialidad = _normalize_especialidad(str(esp_fila).strip())
        if not especialidad:
            especialidad = current_especialidad
        if not especialidad:
            especialidad = last_data_esp
        if especialidad:
            last_data_esp = especialidad

        nivel_raw = get_col("nivel")
        nivel = str(nivel_raw).strip().upper() if nivel_raw else ""

        anio_ingreso = _parse_period_value(get_col("ingreso"))
        anio_egreso = _parse_period_value(get_col("egreso"))
        fecha_sustentacion = _parse_sustentacion_date(get_col("sustentacion"))

        if dni and (len(dni) != 8 or not dni.isdigit()):
            dni = ""

        results.append({
            "__row__": ridx,
            "__sheet__": sheet_name,
            "__format__": "egreso",
            "dni": dni,
            "apellidos_nombres": nombre,
            "grado_titulo": "",
            "especialidad": especialidad,
            "nivel": nivel,
            "anio_ingreso": anio_ingreso,
            "anio_egreso": anio_egreso,
            "fecha_sustentacion": fecha_sustentacion,
            "resolucion_acta": "",
            "codigo_diploma": "",
            "registro_pedagogico": "",
            "director_general": "",
            "secretario_academico": "",
        })

    return results


# ═══════════════════════════════════════════════════════════════
# READER PRINCIPAL (auto-detect)
# ═══════════════════════════════════════════════════════════════

def _read_egresados_file(file) -> List[dict]:
    """
    Lee un archivo .xlsx de egresados (auto-detecta formato).

    Retorna lista de dicts con:
    {
        __row__: int,
        __sheet__: str,
        __format__: 'egreso' | 'verificador',
        dni: str,
        apellidos_nombres: str,
        grado_titulo: str,
        especialidad: str,
        nivel: str,
        anio_ingreso: str,
        anio_egreso: str,
        fecha_sustentacion: date|None,
        resolucion_acta: str,
        codigo_diploma: str,
        registro_pedagogico: str,       ← NUEVO
        director_general: str,
        secretario_academico: str,
    }
    """
    from openpyxl import load_workbook

    if isinstance(file, bytes):
        file = io.BytesIO(file)
    elif hasattr(file, "seek"):
        try:
            file.seek(0)
        except Exception:
            pass

    wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
    results = []

    for sheet_name in wb.sheetnames:
        # Saltar hoja de instrucciones
        if "instruc" in sheet_name.lower():
            continue

        ws = wb[sheet_name]
        fmt = _detect_format(ws)

        if fmt == "verificador":
            results.extend(_read_verificador_sheet(ws, sheet_name))
        else:
            results.extend(_read_egreso_sheet(ws, sheet_name))

    return results