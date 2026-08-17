"""
Hoja de Vida del docente (Currículum Vitae, modelo institucional).

  GET    /catalogs/teachers/me/cv          → todos sus ítems, por sección
  POST   /catalogs/teachers/me/cv          → crear ítem (multipart, con archivo)
  PUT    /catalogs/teachers/me/cv/<id>     → editar (multipart)
  DELETE /catalogs/teachers/me/cv/<id>     → eliminar
  GET    /catalogs/teachers/me/cv/pdf      → CV COMPLETO en PDF:
           1) parte DESCRIPTIVA (datos personales + secciones II–VII, con el
              membrete institucional de Catálogos), firmas al final que nunca
              quedan solas en una hoja
           2) parte DOCUMENTADA: los archivos subidos, anexados en orden
              (?documentado=0 emite solo la parte descriptiva)

Cada docente solo ve y toca SU hoja de vida (se resuelve por request.user).
"""
import re
import unicodedata

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from rest_framework_simplejwt.authentication import JWTAuthentication

from catalogs.models import Teacher, TeacherCVItem

CAMPOS = ["seccion", "subseccion", "institucion", "titulo", "detalle",
          "lugar", "duracion", "fecha_inicio", "fecha_fin", "orden"]

SECCIONES_VALIDAS = {c for c, _ in TeacherCVItem.SECCIONES}
SUBSECCIONES_VALIDAS = {c for c, _ in TeacherCVItem.SUBSECCIONES}


def _teacher_de(request):
    return Teacher.ficha_de(request.user)


# ── Orden de presentación del CV (pedido de Secretaría) ──
# Formación: grados de mayor a menor (doctor → maestría → 2da especialidad →
# licenciado/título → bachiller); y dentro de cada grupo lo MÁS RECIENTE
# primero. Aplica igual en la página, el PDF y la descarga del admin.

_GRADOS_RANGO = [
    ("DOCTOR", 0),                       # doctorado / doctor(a)
    ("MAESTR", 1), ("MAGISTER", 1), ("MAGÍSTER", 1), ("MESTR", 1),
    ("SEGUNDA ESPECIAL", 2),
    ("LICENCIA", 3), ("TÍTULO", 3), ("TITULO", 3), ("PROFESOR", 3),
    ("BACHILLER", 4),
]


def _rango_grado(titulo):
    t = (titulo or "").upper()
    for marca, rango in _GRADOS_RANGO:
        if marca in t:
            return rango
    return 5


def _items_ordenados(ct):
    from datetime import date
    sec_pos = {c: i for i, (c, _) in enumerate(TeacherCVItem.SECCIONES)}
    sub_pos = {c: i for i, (c, _) in enumerate(TeacherCVItem.SUBSECCIONES)}
    vieja = date(1900, 1, 1)

    def clave(it):
        f = it.fecha_inicio or it.fecha_fin or vieja
        if it.seccion == "FORMACION":
            # El grado manda (postgrado arriba); la subsección lo acompaña
            grupo = _rango_grado(it.titulo)
        else:
            grupo = sub_pos.get(it.subseccion, 99)
        return (sec_pos.get(it.seccion, 99), grupo, -f.toordinal(), it.id)

    return sorted(ct.cv_items.all(), key=clave)


def _cv_filename(nombre, ct):
    """HOJA-DE-VIDA-APELLIDOS-NOMBRES.pdf (ASCII, sin espacios)."""
    base = unicodedata.normalize("NFKD", nombre or "").encode(
        "ascii", "ignore").decode()
    base = re.sub(r"[^A-Za-z0-9]+", "-", base).strip("-").upper()
    return f"HOJA-DE-VIDA-{base or (ct.document or ct.id)}.pdf"


def _item_dict(it, request=None):
    url = ""
    try:
        if it.archivo:
            url = it.archivo.url
            if request:
                url = request.build_absolute_uri(url)
    except Exception:
        url = ""
    return {
        "id": it.id, "seccion": it.seccion, "subseccion": it.subseccion,
        "institucion": it.institucion, "titulo": it.titulo,
        "detalle": it.detalle, "lugar": it.lugar, "duracion": it.duracion,
        "fecha_inicio": str(it.fecha_inicio) if it.fecha_inicio else "",
        "fecha_fin": str(it.fecha_fin) if it.fecha_fin else "",
        "archivo_url": url,
        "archivo_nombre": (it.archivo.name.rsplit("/", 1)[-1]
                           if it.archivo else ""),
    }


# Topes de los campos de texto (deben calzar con el modelo); detalle es
# TextField pero igual se acota para que el PDF no se deforme.
LARGOS = {"institucion": 500, "titulo": 500, "detalle": 2000,
          "lugar": 200, "duracion": 120}
ROTULOS = {"institucion": "Institución / Centro de estudios",
           "titulo": "Título / Cargo / Curso", "detalle": "Descripción",
           "lugar": "Lugar", "duracion": "Duración"}


def _aplicar_campos(it, data, files):
    for campo in CAMPOS:
        if campo not in data:
            continue
        v = data.get(campo)
        if campo in ("fecha_inicio", "fecha_fin"):
            v = (str(v).strip() or None) if v is not None else None
            setattr(it, campo, v)
        elif campo == "orden":
            try:
                it.orden = max(0, int(v or 0))
            except (TypeError, ValueError):
                pass
        else:
            v = str(v or "").strip()
            tope = LARGOS.get(campo)
            if tope and len(v) > tope:
                return (f"«{ROTULOS.get(campo, campo)}» supera el máximo de "
                        f"{tope} caracteres (tiene {len(v)}). Resúmelo para "
                        "poder guardar.")
            setattr(it, campo, v)
    if "archivo" in files:
        f = files["archivo"]
        if f.size > 10 * 1024 * 1024:
            return "El archivo no puede superar 10 MB."
        it.archivo = f
    return ""


class TeacherCVListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        ct = _teacher_de(request)
        items = [_item_dict(i, request) for i in _items_ordenados(ct)]
        return Response({"teacher_id": ct.id, "items": items})

    def post(self, request):
        ct = _teacher_de(request)
        seccion = str(request.data.get("seccion") or "").strip().upper()
        if seccion not in SECCIONES_VALIDAS:
            return Response({"detail": f"seccion inválida: {seccion}"},
                            status=400)
        sub = str(request.data.get("subseccion") or "").strip().upper()
        if sub and sub not in SUBSECCIONES_VALIDAS:
            return Response({"detail": f"subseccion inválida: {sub}"},
                            status=400)
        it = TeacherCVItem(teacher=ct, seccion=seccion, subseccion=sub)
        err = _aplicar_campos(it, request.data, request.FILES)
        if err:
            return Response({"detail": err}, status=400)
        it.save()
        return Response(_item_dict(it, request), status=201)


class TeacherCVDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def _get(self, request, item_id):
        ct = _teacher_de(request)
        return get_object_or_404(TeacherCVItem, id=item_id, teacher=ct)

    def put(self, request, item_id: int):
        it = self._get(request, item_id)
        err = _aplicar_campos(it, request.data, request.FILES)
        if err:
            return Response({"detail": err}, status=400)
        it.save()
        return Response(_item_dict(it, request))

    def delete(self, request, item_id: int):
        it = self._get(request, item_id)
        it.delete()
        return Response(status=204)


# ══════════════════════════════════════════════════════════════
# PDF — descriptivo + documentado
# ══════════════════════════════════════════════════════════════

SECCION_TITULOS = [
    ("FORMACION", "II. FORMACIÓN ACADÉMICA Y PROFESIONAL",
     ["N°", "Nivel académico / Programa", "Centro de estudios",
      "Especialidad", "Inicio", "Término", "Lugar"]),
    ("ESPECIALIZACION", "III. ESPECIALIZACIÓN Y ACTUALIZACIÓN",
     ["N°", "Curso / Programa", "Centro de estudios", "Tema",
      "Inicio", "Término", "Horas"]),
    ("EXPERIENCIA", "IV. EXPERIENCIA LABORAL",
     ["N°", "Institución", "Cargo", "Descripción", "Inicio", "Término",
      "Tiempo"]),
    ("EVENTO", "V. PARTICIPACIÓN EN EVENTOS ACADÉMICOS",
     ["N°", "Institución", "Descripción", "Fecha", "Duración (horas)"]),
    ("PUBLICACION", "PUBLICACIONES",
     ["N°", "Lugar y medio de publicación", "Título", "Participación",
      "Fecha"]),
    ("MERITO", "VI. MÉRITOS",
     ["N°", "Institución / Empresa / Entidad", "Descripción", "Fecha"]),
    ("INVESTIGACION", "VII. INVESTIGACIÓN",
     ["N°", "Lugar", "Tema", "Calidad", "Año"]),
]

SUB_LABEL = dict(TeacherCVItem.SUBSECCIONES)


def _fila_de(it, seccion, n):
    f_ini = it.fecha_inicio.strftime("%d/%m/%Y") if it.fecha_inicio else ""
    f_fin = it.fecha_fin.strftime("%d/%m/%Y") if it.fecha_fin else ""
    if seccion == "FORMACION":
        return [n, it.titulo, it.institucion, it.detalle, f_ini, f_fin, it.lugar]
    if seccion == "ESPECIALIZACION":
        return [n, it.titulo, it.institucion, it.detalle, f_ini, f_fin, it.duracion]
    if seccion == "EXPERIENCIA":
        tiempo = it.duracion
        if not tiempo and it.fecha_inicio and it.fecha_fin:
            dias = (it.fecha_fin - it.fecha_inicio).days
            anios, meses = divmod(max(dias, 0) // 30, 12)
            tiempo = (f"{anios} año(s) {meses} mes(es)" if anios
                      else f"{meses} mes(es)")
        return [n, it.institucion, it.titulo, it.detalle, f_ini, f_fin, tiempo]
    if seccion == "EVENTO":
        return [n, it.institucion, it.detalle or it.titulo, f_ini, it.duracion]
    if seccion == "PUBLICACION":
        return [n, it.institucion, it.titulo, it.detalle, f_ini]
    if seccion == "MERITO":
        return [n, it.institucion, it.detalle or it.titulo, f_ini]
    return [n, it.institucion or it.lugar, it.titulo, it.duracion,
            it.fecha_inicio.year if it.fecha_inicio else ""]


class TeacherCVPdfView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, teacher_id=None):
        from html import escape as esc
        from academic.pdf_render import html_to_pdf_bytes
        from academic.views.acta_excel import _acta_area_inst
        from academic.views.evaluation_pdf import _logo_datauris

        if teacher_id:
            # Descarga del ADMIN (panel Hojas de Vida): mismo PDF, otro docente
            from academic.views.teachers import _is_grades_admin
            if not _is_grades_admin(request.user):
                return Response({"detail": "No autorizado."}, status=403)
            ct = get_object_or_404(Teacher, id=teacher_id)
        else:
            ct = _teacher_de(request)
        items = _items_ordenados(ct)
        inst = _acta_area_inst()
        logo, logo2 = _logo_datauris()

        # Foto de perfil del docente (Mi Perfil) como data URI
        foto_uri = ""
        try:
            if ct.photo:
                import base64
                import mimetypes
                with ct.photo.open("rb") as fh:
                    foto_uri = "data:%s;base64,%s" % (
                        mimetypes.guess_type(ct.photo.name)[0] or "image/jpeg",
                        base64.b64encode(fh.read()).decode())
        except Exception:
            foto_uri = ""
        u = ct.user
        nombre = ((getattr(u, "full_name", "") or "").strip()
                  or (ct.full_name or "").strip()
                  or (u.username if u else ""))
        hoy = timezone.localtime(timezone.now())

        # ── I. Datos personales (perfil de postulante docente) ──
        # Con apellidos/nombres separados si el docente los registró;
        # si no, cae al nombre completo de siempre.
        if (ct.apellido_paterno or ct.apellido_materno or ct.nombres):
            filas_nombre = [
                ("Apellido paterno", ct.apellido_paterno or ""),
                ("Apellido materno", ct.apellido_materno or ""),
                ("Nombres", ct.nombres or ""),
            ]
        else:
            filas_nombre = [("Apellidos y nombres", nombre.upper())]
        ubigeo = " / ".join(v for v in (ct.region, ct.provincia, ct.distrito) if v)
        datos = filas_nombre + [
            ("DNI", ct.document or ""),
            ("Fecha de nacimiento",
             ct.fecha_nac.strftime("%d/%m/%Y") if ct.fecha_nac else ""),
            ("Sexo", dict(Teacher.SEXOS).get(ct.sexo, "")),
            ("Grado académico", dict(Teacher.GRADOS_ACADEMICOS).get(
                ct.grado_academico, ct.grado_academico or "")),
            ("Teléfono fijo", ct.telefono_fijo or ""),
            ("Celular", ct.phone or ""),
            ("Correo electrónico", ct.email or ""),
            ("Dirección", ct.direccion or ""),
            ("Región / Provincia / Distrito", ubigeo),
            ("Condición laboral", dict(Teacher.CONDICIONES).get(
                ct.condicion_laboral, ct.condicion_laboral or "")),
            ("N° R.D. de Nombramiento / Contrato", ct.rd_nombramiento or ""),
        ]
        filas_datos = "".join(
            f"<tr><td class='l'>{esc(k)}</td><td>{esc(str(v))}</td></tr>"
            for k, v in datos)

        # ── Secciones II–VII ──
        bloques, anexos = [], []
        for clave, titulo, headers in SECCION_TITULOS:
            del_grupo = [i for i in items if i.seccion == clave]
            if not del_grupo:
                continue
            filas, n = [], 0
            sub_actual = object()
            for it in del_grupo:
                if it.subseccion != sub_actual:
                    sub_actual = it.subseccion
                    if it.subseccion:
                        filas.append(
                            f"<tr><td colspan='{len(headers)}' class='sub'>"
                            f"{esc(SUB_LABEL.get(it.subseccion, it.subseccion))}"
                            "</td></tr>")
                n += 1
                celdas = _fila_de(it, clave, n)
                marca = " ✓" if it.archivo else ""
                filas.append("<tr>" + "".join(
                    f"<td{' class=' + chr(39) + 'c' + chr(39) if j != 1 and j != 2 else ''}>"
                    f"{esc(str(v))}{marca if j == len(celdas) - 1 else ''}</td>"
                    for j, v in enumerate(celdas)) + "</tr>")
                if it.archivo:
                    anexos.append((titulo, it))
            heads = "".join(f"<th>{esc(h)}</th>" for h in headers)
            bloques.append(f"""
<h3>{esc(titulo)}</h3>
<table><thead><tr>{heads}</tr></thead><tbody>{''.join(filas)}</tbody></table>""")

        # ── Declaración + firmas (nunca solas en una hoja: viajan con la
        #    declaración en un bloque indivisible) ──
        cierre = f"""
<div style="page-break-inside:avoid; break-inside:avoid; margin-top:26px">
<p style="font-size:10px; text-align:justify; line-height:1.7">
Yo, <b>{esc(nombre.upper())}</b>, identificado(a) con DNI N°
<b>{esc(ct.document or '')}</b>, declaro bajo juramento que la información
registrada en la presente Hoja de Vida ha sido consignada en su totalidad,
es correcta y se encuentra sustentada con los documentos que la acreditan.
</p>
<p style="font-size:9px; text-align:right; margin-top:14px">
Tarma, {hoy.strftime('%d/%m/%Y')}</p>
<table style="width:100%; margin-top:70px; border:none">
  <tr>
    <td style="border:none; text-align:center; width:33%">
      <div style="border-top:1px solid #111; width:170px; margin:0 auto; padding-top:4px">
        <b style="font-size:9px">DOCENTE</b><br>
        <span style="font-size:8px">Firma y Huella</span>
      </div>
    </td>
    <td style="border:none; text-align:center; width:33%">
      <div style="border-top:1px solid #111; width:170px; margin:0 auto; padding-top:4px">
        <b style="font-size:9px">DIRECTORA GENERAL</b><br>
        <span style="font-size:8px">Firma, Post Firma y Sello</span>
      </div>
    </td>
    <td style="border:none; text-align:center; width:33%">
      <div style="border-top:1px solid #111; width:170px; margin:0 auto; padding-top:4px">
        <b style="font-size:9px">SECRETARIO(A) ACADÉMICO(A)</b><br>
        <span style="font-size:8px">Firma, Post Firma y Sello</span>
      </div>
    </td>
  </tr>
</table>
</div>"""

        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  @page {{ size: A4; margin: 14mm 12mm 14mm 22mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 9.5px;
          color: #111; margin: 0; }}
  .head {{ display: flex; align-items: center; gap: 10px;
           border-bottom: 2px solid #1F4E79; padding-bottom: 6px; }}
  .head img {{ height: 46px; }}
  .head .tit {{ flex: 1; text-align: center; }}
  .head .tit h1 {{ font-size: 12px; margin: 0; color: #1F4E79; }}
  .head .tit p {{ margin: 2px 0 0; font-size: 9px; color: #444; }}
  h2 {{ text-align: center; font-size: 13px; letter-spacing: 1px;
        margin: 12px 0 8px; }}
  h3 {{ font-size: 10.5px; background: #E7E6E6; padding: 4px 8px;
        margin: 14px 0 4px; page-break-after: avoid; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #666; padding: 3px 5px; font-size: 9px; }}
  th {{ background: #E7E6E6; }}
  thead {{ display: table-header-group; }}
  tr {{ page-break-inside: avoid; }}
  td.l {{ background: #F3F3F3; font-weight: bold; width: 34%; }}
  td.c {{ text-align: center; }}
  td.sub {{ background: #F3F3F3; font-weight: bold; font-style: italic; }}
</style></head><body>
<div class="head">
  {f'<img src="{logo}">' if logo else ''}
  <div class="tit">
    <h1>INSTITUTO DE EDUCACIÓN SUPERIOR PEDAGÓGICO PÚBLICO<br>
        {esc(inst['nombre'])} — {esc((inst.get('provincia') or 'Tarma').upper())}</h1>
    <p>HOJA DE VIDA DEL DOCENTE</p>
  </div>
  {f'<img src="{logo2}">' if logo2 else ''}
</div>
<h2>CURRÍCULUM VITAE</h2>
<h3>I. DATOS PERSONALES</h3>
<div style="display:flex; gap:10px; align-items:stretch">
  <table style="flex:1">{filas_datos}</table>
  <div style="width:30mm; min-height:38mm; border:1px solid #666; flex-shrink:0;
              display:flex; align-items:center; justify-content:center;
              overflow:hidden; background:#fff">
    {f'<img src="{foto_uri}" style="width:100%; height:100%; object-fit:cover">'
     if foto_uri else '<span style="color:#999; font-size:9px">FOTO</span>'}
  </div>
</div>
{''.join(bloques) if bloques else
 "<p style='margin-top:14px; color:#777'>Aún no hay ítems registrados en la Hoja de Vida.</p>"}
{cierre}
</body></html>"""

        try:
            pdf = html_to_pdf_bytes(html)
        except Exception as exc:
            return Response({"detail": f"No se pudo generar el PDF: {exc}"},
                            status=500)

        # ── Parte DOCUMENTADA: anexar los archivos que acreditan ──
        documentado = str(request.query_params.get("documentado", "1")).lower() \
            not in ("0", "false", "no")
        if documentado and anexos:
            pdf = self._anexar_documentos(pdf, anexos)

        return HttpResponse(
            pdf, content_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="{_cv_filename(nombre, ct)}"'})

    @staticmethod
    def _anexar_documentos(pdf_base: bytes, anexos):
        """Anexa los archivos subidos (PDF tal cual; imágenes convertidas).

        Si `pypdf` no está instalado en el servidor, devuelve el descriptivo
        sin anexos (instalar con: pip install pypdf).
        """
        from io import BytesIO
        try:
            from pypdf import PdfReader, PdfWriter
        except ImportError:
            try:
                from PyPDF2 import PdfReader, PdfWriter   # compat
            except ImportError:
                return pdf_base

        writer = PdfWriter()
        for page in PdfReader(BytesIO(pdf_base)).pages:
            writer.add_page(page)

        for _titulo, it in anexos:
            try:
                nombre = (it.archivo.name or "").lower()
                with it.archivo.open("rb") as f:
                    data = f.read()
                if nombre.endswith(".pdf"):
                    for page in PdfReader(BytesIO(data)).pages:
                        writer.add_page(page)
                elif nombre.endswith((".jpg", ".jpeg", ".png")):
                    from PIL import Image
                    img = Image.open(BytesIO(data)).convert("RGB")
                    buf = BytesIO()
                    img.save(buf, format="PDF")
                    for page in PdfReader(BytesIO(buf.getvalue())).pages:
                        writer.add_page(page)
                # otros formatos: se omiten en silencio del anexo
            except Exception:
                continue

        out = BytesIO()
        writer.write(out)
        return out.getvalue()


# ══════════════════════════════════════════════════════════════
# ADMIN — estado de las Hojas de Vida (qué cargó cada docente)
# ══════════════════════════════════════════════════════════════

# Rótulo corto por sección para el reporte (columnas II…VII)
SECCION_CORTA = [
    ("FORMACION", "II"), ("ESPECIALIZACION", "III"), ("EXPERIENCIA", "IV"),
    ("EVENTO", "V.a"), ("PUBLICACION", "V.b"), ("MERITO", "VI"),
    ("INVESTIGACION", "VII"),
]


class TeacherCVAdminEstadoView(APIView):
    """
    GET /catalogs/teachers/cv/estado            → JSON: avance del CV por docente
    GET /catalogs/teachers/cv/estado?fmt=xlsx   → reporte en Excel
    GET /catalogs/teachers/cv/estado?fmt=pdf    → reporte en PDF
    Solo administradores (mismos roles que el módulo de Evaluación).
    Por cada docente: ítems por sección (y cuántos con documento), totales,
    % documentado y fecha del último cambio en su hoja de vida.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from academic.views.teachers import _is_grades_admin
        if not _is_grades_admin(request.user):
            return Response({"detail": "No autorizado."}, status=403)

        q = (request.query_params.get("q") or "").strip().lower()
        fmt = (request.query_params.get("fmt") or "").strip().lower()

        filas = []
        qs = (Teacher.objects.select_related("user")
              .prefetch_related("cv_items")
              .order_by("user__full_name", "full_name", "id"))
        for t in qs:
            u = t.user
            nombre = ((getattr(u, "full_name", "") or "").strip() if u else "") \
                or (t.full_name or "").strip() \
                or (u.username if u else "")
            if not nombre and not t.document:
                continue   # ficha vacía sin identidad: no aporta al reporte
            items = list(t.cv_items.all())
            if q and q not in nombre.lower() and q not in (t.document or ""):
                continue
            por_sec, ultima = {}, None
            for clave, _corta in SECCION_CORTA:
                por_sec[clave] = [0, 0]   # [ítems, con documento]
            for it in items:
                d = por_sec.get(it.seccion)
                if d is None:
                    continue
                d[0] += 1
                if it.archivo:
                    d[1] += 1
                if it.updated_at and (ultima is None or it.updated_at > ultima):
                    ultima = it.updated_at
            total = sum(v[0] for v in por_sec.values())
            con_doc = sum(v[1] for v in por_sec.values())
            filas.append({
                "teacher_id": t.id,
                "nombre": nombre.upper(),
                "dni": t.document or "",
                "secciones": {c: por_sec[c] for c, _ in SECCION_CORTA},
                "total": total,
                "con_doc": con_doc,
                "sin_doc": total - con_doc,
                "avance": round(100 * con_doc / total) if total else 0,
                "actualizado": (timezone.localtime(ultima)
                                .strftime("%d/%m/%Y %H:%M") if ultima else ""),
            })

        if fmt == "xlsx":
            return self._xlsx(filas)
        if fmt == "pdf":
            return self._pdf(filas)
        return Response({"total_docentes": len(filas), "rows": filas})

    # ── Excel ──
    @staticmethod
    def _xlsx(filas):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Hojas de Vida"
        heads = (["N°", "Docente", "DNI"]
                 + [f"{c} ítems" for _, c in SECCION_CORTA]
                 + ["Total ítems", "Con doc.", "Sin doc.", "Avance %",
                    "Última actualización"])
        fill = PatternFill("solid", start_color="1F4E79")
        for j, h in enumerate(heads, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        for i, f in enumerate(filas, 1):
            r = i + 1
            vals = ([i, f["nombre"], f["dni"]]
                    + [f"{n}" + (f" ({d}✓)" if n else "")
                       for n, d in (f["secciones"][c]
                                    for c, _ in SECCION_CORTA)]
                    + [f["total"], f["con_doc"], f["sin_doc"],
                       f["avance"], f["actualizado"]])
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(size=9)
                if j != 2:
                    c.alignment = Alignment(horizontal="center")
        widths = [4, 42, 11] + [8] * len(SECCION_CORTA) + [9, 9, 9, 9, 17]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        return HttpResponse(
            buf.getvalue(),
            content_type=("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     'attachment; filename="hojas-de-vida-docentes.xlsx"'})

    # ── PDF ──
    @staticmethod
    def _pdf(filas):
        from html import escape as esc
        from academic.pdf_render import html_to_pdf_bytes
        from academic.views.evaluation_pdf import _pdf_shell

        hoy = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
        heads = ("<th>N°</th><th>Docente</th><th>DNI</th>"
                 + "".join(f"<th>{c}</th>" for _, c in SECCION_CORTA)
                 + "<th>Total</th><th>Con<br>doc.</th><th>Sin<br>doc.</th>"
                   "<th>Avance</th><th>Última<br>actualización</th>")
        cuerpo_filas = []
        for i, f in enumerate(filas, 1):
            celdas_sec = "".join(
                "<td class='c'>{}</td>".format(
                    f"{n} ({d}✓)" if n else "—")
                for n, d in (f["secciones"][c] for c, _ in SECCION_CORTA))
            cuerpo_filas.append(
                f"<tr><td class='c'>{i}</td><td>{esc(f['nombre'])}</td>"
                f"<td class='c'>{esc(f['dni'])}</td>{celdas_sec}"
                f"<td class='c'><b>{f['total']}</b></td>"
                f"<td class='c'>{f['con_doc']}</td>"
                f"<td class='c'>{f['sin_doc']}</td>"
                f"<td class='c'>{f['avance']}%</td>"
                f"<td class='c'>{esc(f['actualizado'] or '—')}</td></tr>")
        cuerpo = f"""
<style>
  th {{ background: #E7E6E6; color: #111; }}
</style>
<h2 style="text-align:center">ESTADO DE LAS HOJAS DE VIDA DE LOS DOCENTES</h2>
<p style="font-size:8.5px; margin:2px 0 0">
  Por sección se muestra el n° de ítems registrados y, entre paréntesis,
  cuántos tienen documento que acredita (✓). Secciones: II Formación ·
  III Especialización · IV Experiencia · V.a Eventos · V.b Publicaciones ·
  VI Méritos · VII Investigación. Emitido el {hoy}.
</p>
<table>
  <thead><tr>{heads}</tr></thead>
  <tbody>{''.join(cuerpo_filas) if cuerpo_filas else
          '<tr><td colspan="15" style="color:#777">Sin docentes con hoja de vida.</td></tr>'}</tbody>
</table>
"""
        html = _pdf_shell("ESTADO DE HOJAS DE VIDA — DOCENTES", cuerpo,
                          landscape=True)
        try:
            pdf = html_to_pdf_bytes(html)
        except Exception as exc:
            return Response({"detail": f"No se pudo generar el PDF: {exc}"},
                            status=500)
        return HttpResponse(
            pdf, content_type="application/pdf",
            headers={"Content-Disposition":
                     'attachment; filename="hojas-de-vida-docentes.pdf"'})
