"""
Funciones auxiliares y helpers para el módulo de catálogos
Contiene normalización, conversión, readers y cache
─────────────────────────────────────────────────────────
✅ _ensure_period() ahora sincroniza con academic.AcademicPeriod
"""
import os
import io
import csv
import json
import re
import time
import logging
import unicodedata
from datetime import date
from typing import Optional, List, Dict
from django.conf import settings
from django.db.utils import OperationalError
from django.http import HttpResponse
from rest_framework.response import Response
from openpyxl import load_workbook
from openpyxl import Workbook

from catalogs.models import Period, Career
from academic.models import Course

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# NORMALIZACIÓN DE TEXTO
# ═══════════════════════════════════════════════════════════════

def _norm(s):
    """Normalización básica: strip, sin tildes, lowercase, espacios únicos"""
    s = "" if s is None else str(s)
    s = s.strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_key(s: str) -> str:
    """Normalización agresiva para cache keys"""
    return _norm(s).replace(".", "").replace("-", "").replace("/", "").replace("_", "").strip()


def _pc_key(name: str) -> str:
    """Clave para deduplicar cursos en plan"""
    s = "" if name is None else str(name)
    s = s.replace("\xa0", " ").replace("\u200b", "").replace("\u00ad", "")
    s = re.sub(r"\s+", " ", s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()


def _clean_spaces(s: str) -> str:
    """Limpia espacios, NBSP y múltiples espacios"""
    s = "" if s is None else str(s)
    s = s.replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _strip_nul(v):
    """Elimina NUL bytes (Postgres no los acepta)"""
    if v is None:
        return v
    if isinstance(v, str):
        return v.replace("\x00", "").replace("\u0000", "")
    return v


def _clean_text(v):
    """Limpieza completa: NUL, NBSP, espacios raros"""
    v = _strip_nul(v)
    if v is None:
        return v
    if isinstance(v, str):
        v = v.replace("\xa0", " ")
        v = re.sub(r"\s+", " ", v).strip()
        return v
    return v


def _normalize_course_name(name: str) -> str:
    """Normaliza nombre de curso con Title Case inteligente"""
    s = _clean_spaces(name)
    if not s:
        return s
    
    # Title case base
    s2 = " ".join(w.capitalize() for w in s.lower().split())
    
    # Romanos en mayúscula
    s2 = re.sub(r"\b(i|ii|iii|iv|v|vi|vii|viii|ix|x)\b", lambda m: m.group(1).upper(), s2)
    
    # Siglas comunes
    for ac in ["TIC", "TICS", "TI", "IA", "ERP", "SAP"]:
        s2 = re.sub(rf"\b{ac.title()}\b", ac, s2)
    
    # Conectores en minúscula
    s2 = re.sub(r"\b(De|Del|La|Las|Los|Y|E)\b", lambda m: m.group(1).lower(), s2)
    
    return s2


# ═══════════════════════════════════════════════════════════════
# CONVERSIÓN DE TIPOS
# ═══════════════════════════════════════════════════════════════

def _to_int(v, default=None):
    """Convierte a int de forma segura"""
    try:
        if v is None:
            return default
        s = str(v).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def _to_float(v, default=None):
    """Convierte a float seguro, acepta coma como decimal"""
    try:
        if v is None:
            return default
        s = str(v).strip().replace(",", ".")
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def _parse_date_yyyy_mm_dd(s):
    """Parse fecha YYYY-MM-DD"""
    s = (s or "").strip()
    if not s:
        return None
    try:
        y, m, d = [int(x) for x in s.split("-")]
        return date(y, m, d)
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════
# PERIODOS
# ═══════════════════════════════════════════════════════════════

_TERM_MAP = {
    "I": "I", "1": "I", "01": "I",
    "II": "II", "2": "II", "02": "II",
    "III": "III", "3": "III", "03": "III",
}


def _parse_period_code(code: str):
    """
    Acepta: 2026-I, 2026-1, 2026/1, 2026 I, 2026II, 2026-III
    Retorna (year:int, term:str) o (None, None)
    """
    s = (code or "").strip().upper()
    if not s:
        return None, None
    
    s = re.sub(r"[\/\s]+", "-", s)
    s = re.sub(r"--+", "-", s)
    
    m = re.match(r"^(20\d{2})-?([IVX]{1,3}|\d{1,2})$", s)
    if not m:
        return None, None
    
    try:
        y = int(m.group(1))
    except Exception:
        return None, None
    
    t_raw = m.group(2)
    t = _TERM_MAP.get(t_raw)
    if not t:
        return None, None
    
    return y, t


def _sync_academic_period_from_catalog(period_obj):
    """
    Sincroniza un catalogs.Period con academic.AcademicPeriod.
    Se llama después de crear/actualizar un Period en catálogos.
    """
    try:
        from academic.models import AcademicPeriod

        code = getattr(period_obj, "code", "") or ""
        code = code.strip().upper()
        if not code:
            return

        start_date = getattr(period_obj, "start_date", None)
        end_date   = getattr(period_obj, "end_date", None)

        # Fallback desde year + term
        if not start_date or not end_date:
            year = getattr(period_obj, "year", None)
            term = getattr(period_obj, "term", "") or ""
            if year:
                year = int(year)
                if term.upper() in ("I", "1"):
                    start_date = start_date or date(year, 3, 1)
                    end_date   = end_date   or date(year, 7, 31)
                elif term.upper() in ("II", "2"):
                    start_date = start_date or date(year, 8, 1)
                    end_date   = end_date   or date(year, 12, 31)
                else:
                    start_date = start_date or date(year, 1, 1)
                    end_date   = end_date   or date(year, 2, 28)
            else:
                return

        AcademicPeriod.objects.update_or_create(
            code=code,
            defaults=dict(start=start_date, end=end_date),
        )
        logger.info(f"AcademicPeriod '{code}' sincronizado desde _ensure_period")

    except Exception as e:
        logger.warning(f"Error sincronizando AcademicPeriod '{getattr(period_obj, 'code', '?')}': {e}")


def _ensure_period(code: str):
    """
    Crea Period en catálogos si no existe.
    ✅ MEJORADO: también sincroniza con academic.AcademicPeriod.
    """
    code = (code or "").strip().upper()
    y, t = _parse_period_code(code)
    if not (y and t):
        return None
    
    # Reparar Period existente con code vacío
    existing = Period.objects.filter(year=y, term=t).first()
    if existing and not (existing.code or "").strip():
        existing.code = f"{y}-{t}"
        if not (existing.label or "").strip():
            existing.label = f"{y}-{t}"
        existing.save(update_fields=["code", "label"])
        _sync_academic_period_from_catalog(existing)
        return existing
    
    defaults = {
        "year": y,
        "term": t,
        "label": f"{y}-{t}",
        "is_active": False,
    }
    
    if hasattr(Period, "start_date") and hasattr(Period, "end_date"):
        if t == "I":
            defaults["start_date"] = date(y, 3, 1)
            defaults["end_date"] = date(y, 7, 31)
        elif t == "II":
            defaults["start_date"] = date(y, 8, 1)
            defaults["end_date"] = date(y, 12, 15)
        else:  # III
            defaults["start_date"] = date(y, 1, 1)
            defaults["end_date"] = date(y, 2, 28)
    
    obj, created = Period.objects.get_or_create(code=f"{y}-{t}", defaults=defaults)

    # ✅ Sincronizar con AcademicPeriod (tanto si se creó como si ya existía)
    _sync_academic_period_from_catalog(obj)

    return obj


# ═══════════════════════════════════════════════════════════════
# CAREERS
# ═══════════════════════════════════════════════════════════════

def _slug_code(s: str, maxlen: int = 10) -> str:
    """Genera código slug alfanumérico"""
    s = _norm(s).upper()
    s = re.sub(r"[^A-Z0-9]", "", s)
    return (s[:maxlen] or "CRS")


def _make_unique_career_code(name: str, maxlen: int = 12) -> str:
    """Genera código único para Career"""
    base = _slug_code(name, maxlen=maxlen)
    code = base
    k = 1
    while Career.objects.filter(code=code).exists():
        k += 1
        code = f"{base[:max(1, maxlen-2)]}{k:02d}"
    return code


def _strip_rvm(name: str) -> str:
    s = (name or "").strip()
    s = re.sub(r"\s*\(RVM[^)]*\)\s*$", "", s, flags=re.IGNORECASE).strip()
    return s


def _norm_career(name: str) -> str:
    s = _strip_rvm(name)
    s = _norm_key(s)
    s = s.replace("educacion", "").replace("secundaria", "").replace("especialidad", "")
    s = s.replace("einformatica", "informatica")
    return s.strip()


def _ensure_career_cached(career_idx: dict, career_name: str):
    name = (career_name or "").strip().upper()
    if not name:
        raise ValueError("Career name vacío")

    car = _find_by_norm_cached(career_idx, name)
    if car:
        return car

    base_name = _strip_rvm(name).upper()
    if base_name != name:
        car = _find_by_norm_cached(career_idx, base_name)
        if car:
            career_idx[_norm_key(name)] = car
            return car

    car = Career.objects.filter(name__iexact=name).first()
    if car:
        career_idx[_norm_key(name)] = car
        return car

    if base_name != name:
        car = Career.objects.filter(name__iexact=base_name).first()
        if car:
            career_idx[_norm_key(name)] = car
            career_idx[_norm_key(base_name)] = car
            return car

    target = _norm_career(name)
    if target:
        for c in Career.objects.all().only("id", "name"):
            if _norm_career(c.name) == target:
                career_idx[_norm_key(name)] = c
                return c

    final_name = base_name if base_name else name
    code = _make_unique_career_code(final_name)
    car = Career.objects.create(name=final_name, code=code)
    career_idx[_norm_key(name)] = car
    career_idx[_norm_key(final_name)] = car
    return car


# ═══════════════════════════════════════════════════════════════
# CACHE E ÍNDICES
# ═══════════════════════════════════════════════════════════════

def _build_norm_index(model, field: str):
    idx = {}
    for o in model.objects.all().only("id", field):
        val = getattr(o, field, "") or ""
        nk = _norm_key(val)
        if nk and nk not in idx:
            idx[nk] = o
    return idx


def _find_by_norm_cached(index: dict, value: str):
    return index.get(_norm_key(value or ""))


# ═══════════════════════════════════════════════════════════════
# DATABASE HELPERS
# ═══════════════════════════════════════════════════════════════

def _retry_db(fn, tries=10, delay=0.5):
    last = None
    for _ in range(tries):
        try:
            return fn()
        except OperationalError as e:
            last = e
            if "database is locked" in str(e).lower():
                time.sleep(delay)
                continue
            raise
    if last:
        raise last


# ═══════════════════════════════════════════════════════════════
# RESPONSE HELPERS
# ═══════════════════════════════════════════════════════════════

def list_items(serializer_cls, queryset):
    return Response({"items": serializer_cls(queryset, many=True).data})


def _xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _csv_bytes(rows: List[dict], headers: List[str]) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        rr = {}
        for k in headers:
            v = r.get(k)
            if isinstance(v, (dict, list)):
                rr[k] = json.dumps(v, ensure_ascii=False)
            else:
                rr[k] = "" if v is None else v
        w.writerow(rr)
    return out.getvalue().encode("utf-8-sig")


# ═══════════════════════════════════════════════════════════════
# AUTH HELPERS
# ═══════════════════════════════════════════════════════════════

def _require_staff(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return Response({"detail": "No autorizado."}, status=403)
    return None


def _require_staff_or_teacher(request):
    if not request.user.is_authenticated:
        return Response({"detail": "No autorizado."}, status=403)
    
    if getattr(request.user, "is_staff", False):
        return None
    
    try:
        if request.user.roles.filter(name__iexact="TEACHER").exists():
            return None
    except Exception:
        pass
    
    return Response({"detail": "No autorizado."}, status=403)


# ═══════════════════════════════════════════════════════════════
# FILE READERS
# ═══════════════════════════════════════════════════════════════

STUDENT_HEADER_ALIASES = {
    "region": "region",
    "provincia": "provincia",
    "distrito": "distrito",
    "codigo modular": "codigo_modular",
    "codigo_modular": "codigo_modular",
    "nombre de la institucion": "nombre_institucion",
    "nombre institucion": "nombre_institucion",
    "gestion": "gestion",
    "tipo": "tipo",
    "programa / carrera": "programa_carrera",
    "programa carrera": "programa_carrera",
    "ciclo": "ciclo",
    "turno": "turno",
    "seccion": "seccion",
    "periodo": "periodo",
    "apellido paterno": "apellido_paterno",
    "apellido materno": "apellido_materno",
    "nombres": "nombres",
    "fecha nac": "fecha_nac",
    "fecha nac.": "fecha_nac",
    "sexo": "sexo",
    "num documento": "num_documento",
    "num. documento": "num_documento",
    "lengua": "lengua",
    "discapacidad": "discapacidad",
    "tipo de discapacidad": "tipo_discapacidad",
}


def _read_rows(file, mapping: dict):
    name = (getattr(file, "name", "") or "").lower()
    
    def apply_mapping(row: dict):
        if not mapping:
            return row
        out = dict(row)
        for field, col in (mapping or {}).items():
            if col and col in row:
                out[field] = row.get(col)
        return out
    
    # XLSX
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        try:
            file.seek(0)
        except Exception:
            pass
        
        wb = load_workbook(file, data_only=True, read_only=True, keep_links=False)
        ws = wb.active
        
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return []
        
        headers = [(_clean_text(str(h)) if h is not None else "") for h in header_row]
        headers = [h.strip() for h in headers]
        
        header_map = {}
        for j, h in enumerate(headers):
            nh = _norm(h)
            if nh in STUDENT_HEADER_ALIASES:
                header_map[j] = STUDENT_HEADER_ALIASES[nh]
        
        out = []
        for i, r in enumerate(it, start=2):
            row_raw = {}
            for j in range(len(headers)):
                if not headers[j]:
                    continue
                val = r[j] if j < len(r) else None
                val = _clean_text(val)
                row_raw[headers[j]] = "" if val is None else val
            
            row_auto = {}
            for j in range(len(headers)):
                key = header_map.get(j)
                if key:
                    val = r[j] if j < len(r) else None
                    val = _clean_text(val)
                    row_auto[key] = "" if val is None else val
            
            row_final = row_auto if row_auto else row_raw
            row_final = apply_mapping(row_final)
            out.append({**row_final, "__row__": i})
        
        return out
    
    # CSV
    raw = file.read()
    raw = raw.replace(b"\x00", b"")
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1", errors="ignore")
    
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for idx, row in enumerate(reader, start=2):
        clean = {}
        for k, v in row.items():
            if not k:
                continue
            kk = _clean_text(k)
            vv = _clean_text(v)
            clean[str(kk).strip()] = "" if vv is None else str(vv).strip()
        
        row_auto = {}
        for k, v in clean.items():
            nk = _norm(k)
            if nk in STUDENT_HEADER_ALIASES:
                row_auto[STUDENT_HEADER_ALIASES[nk]] = v
        
        row_final = row_auto if row_auto else clean
        row_final = apply_mapping(row_final)
        out.append({**row_final, "__row__": idx})
    
    return out