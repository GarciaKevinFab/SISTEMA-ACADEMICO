from django.http import HttpResponse, FileResponse, Http404
from django.db.models import Q
from django.core.files.base import ContentFile
import os
import zipfile
from datetime import datetime
from django.conf import settings
from django.core.management import call_command

# ✅ nuevos imports para importadores
import io
import csv
import json
from datetime import date
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.crypto import get_random_string
from openpyxl import Workbook, load_workbook

from rest_framework import viewsets
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from acl.models import Role
from students.models import Student
from academic.models import Career, Plan, Course, PlanCourse

# ✅ si implementaste AcademicGradeRecord (recomendado)
try:
    from academic.models import AcademicGradeRecord
except Exception:
    AcademicGradeRecord = None

from .models import *
from .serializers import *

User = get_user_model()


# ------------------ helpers ------------------
def list_items(serializer_cls, queryset):
    return Response({"items": serializer_cls(queryset, many=True).data})


def _require_staff(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return Response({"detail": "No autorizado."}, status=403)
    return None


def _xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _read_rows(file, mapping: dict):
    """
    Lee XLSX/CSV. Devuelve lista de dicts con __row__ (fila real).
    mapping: {campo: "NombreColumnaExcel"} opcional
    """
    name = (getattr(file, "name", "") or "").lower()

    def apply_mapping(row: dict):
        if not mapping:
            return row
        out = dict(row)
        for field, col in mapping.items():
            if col and col in row:
                out[field] = row.get(col)
        return out

    # XLSX
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xltx"):
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for i, r in enumerate(rows[1:], start=2):
            row = {headers[j]: r[j] for j in range(len(headers)) if headers[j]}
            row = {k: ("" if v is None else v) for k, v in row.items()}
            out.append(apply_mapping(row) | {"__row__": i})
        return out

    # CSV
    raw = file.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for idx, row in enumerate(reader, start=2):
        row = {k.strip(): ("" if v is None else str(v).strip()) for k, v in row.items() if k}
        out.append(apply_mapping(row) | {"__row__": idx})
    return out


def _to_int(v, default=None):
    try:
        if v == "" or v is None:
            return default
        return int(float(v))
    except Exception:
        return default


def _to_float(v, default=None):
    try:
        if v == "" or v is None:
            return default
        return float(v)
    except Exception:
        return default


def _parse_date_yyyy_mm_dd(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        y, m, d = [int(x) for x in s.split("-")]
        return date(y, m, d)
    except Exception:
        return None


def _ensure_student_role():
    role, _ = Role.objects.get_or_create(name="STUDENT")
    return role


def _safe_username(document: str, email: str, preferred: str = "") -> str:
    u = (preferred or "").strip()
    if u:
        return u
    if document:
        return document
    return (email or "").strip().lower()


def _get_or_create_student_user(document: str, first_name: str, last_name: str, email: str, username: str = "", password: str = ""):
    """
    Crea/recupera user:
    - username: Excel > DNI > email
    - password: Excel o generado seguro
    - asigna rol STUDENT (sin borrar otros)
    Retorna: (user, created, temp_password_or_none)
    """
    student_role = _ensure_student_role()

    full_name = f"{first_name} {last_name}".strip()
    email = (email or "").strip().lower()
    uname = _safe_username(document, email, username)

    user = User.objects.filter(username=uname).first()
    if not user and email:
        user = User.objects.filter(email=email).first()

    created = False
    temp_password = None

    if not user:
        created = True
        user = User(
            username=uname,
            email=email,
            full_name=full_name,
            is_active=True,
            is_staff=False,
        )

        if password and str(password).strip():
            temp_password = str(password).strip()
        else:
            temp_password = get_random_string(10) + "!"
        user.set_password(temp_password)
        user.save()
    else:
        changed = False
        if full_name and getattr(user, "full_name", "") != full_name:
            user.full_name = full_name
            changed = True
        if email and getattr(user, "email", "") != email:
            user.email = email
            changed = True
        if not user.is_active:
            user.is_active = True
            changed = True
        if changed:
            user.save()

        # si viene password en excel => reset
        if password and str(password).strip():
            temp_password = str(password).strip()
            user.set_password(temp_password)
            user.save(update_fields=["password"])

    if hasattr(user, "roles"):
        user.roles.add(student_role)

    return user, created, temp_password


# ------------------ Catálogos ------------------
class PeriodsViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        if hasattr(Period, "start_date"):
            return qs.order_by("-start_date")
        if hasattr(Period, "start"):
            return qs.order_by("-start")
        return qs.order_by("-id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())

    @action(detail=True, methods=["post"], url_path="active")
    def set_active(self, request, pk=None):
        is_active = bool(request.data.get("is_active", False))
        p = self.get_object()
        if is_active:
            Period.objects.update(is_active=False)
        p.is_active = is_active
        p.save(update_fields=["is_active"])
        return Response({"ok": True, "id": p.id, "is_active": p.is_active})


class CampusesViewSet(viewsets.ModelViewSet):
    queryset = Campus.objects.all().order_by("name")
    serializer_class = CampusSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())


class ClassroomsViewSet(viewsets.ModelViewSet):
    queryset = Classroom.objects.select_related("campus").all()
    serializer_class = ClassroomSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        campus_id = self.request.query_params.get("campus_id")
        if campus_id:
            qs = qs.filter(campus_id=campus_id)

        if hasattr(Classroom, "code"):
            return qs.order_by("campus__name", "code")
        return qs.order_by("campus__name", "id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())


class TeachersViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.select_related("user").all()
    serializer_class = TeacherSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "patch", "delete"]

    def get_queryset(self):
        qs = super().get_queryset()
        q = (self.request.query_params.get("q") or "").strip()

        if q:
            cond = Q()

            if hasattr(Teacher, "document"):
                cond |= Q(document__icontains=q)
            if hasattr(Teacher, "email"):
                cond |= Q(email__icontains=q)
            if hasattr(Teacher, "phone"):
                cond |= Q(phone__icontains=q)
            if hasattr(Teacher, "specialization"):
                cond |= Q(specialization__icontains=q)

            cond |= (
                Q(user__full_name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__email__icontains=q)
            )

            qs = qs.filter(cond)

        return qs.order_by("user__full_name", "user__username", "id")

    def list(self, request, *args, **kwargs):
        return list_items(self.serializer_class, self.get_queryset())


# ------------------ Ubigeo (MVP estático) ------------------
UBIGEO_DATA = {
    "LIMA": {
        "LIMA": ["LIMA", "LA MOLINA", "SURCO", "MIRAFLORES"],
        "HUAURA": ["HUACHO", "HUALMAY"],
    },
    "PIURA": {
        "PIURA": ["PIURA", "CASTILLA"],
    },
}


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_search(request):
    q = (request.query_params.get("q") or "").strip().upper()
    res = []
    if q:
        for dep, provs in UBIGEO_DATA.items():
            if q in dep:
                res.append({"department": dep})
            for prov, dists in provs.items():
                if q in prov:
                    res.append({"department": dep, "province": prov})
                for dist in dists:
                    if q in dist:
                        res.append({"department": dep, "province": prov, "district": dist})
    return Response(res[:50])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_departments(request):
    return Response(sorted(list(UBIGEO_DATA.keys())))


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_provinces(request):
    dep = (request.query_params.get("department") or "").upper()
    provs = sorted(list(UBIGEO_DATA.get(dep, {}).keys()))
    return Response(provs)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ubigeo_districts(request):
    dep = (request.query_params.get("department") or "").upper()
    prov = (request.query_params.get("province") or "").upper()
    dists = UBIGEO_DATA.get(dep, {}).get(prov, [])
    return Response(sorted(dists))


# ------------------ Institución ------------------
@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def institution_settings(request):
    obj, _ = InstitutionSetting.objects.get_or_create(pk=1)

    if request.method == "GET":
        return Response(obj.data or {})

    obj.data = {**(obj.data or {}), **(request.data or {})}
    obj.save(update_fields=["data"])
    return Response(obj.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def institution_media(request):
    file = request.FILES.get("file")
    kind = request.POST.get("kind")

    if not file or not kind:
        return Response({"detail": "file y kind requeridos"}, status=400)

    asset = MediaAsset.objects.create(kind=kind, file=file)
    return Response(MediaAssetSerializer(asset).data, status=201)


# ------------------ Importadores ------------------
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_template(request, type: str):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    type = (type or "").lower().strip()
    wb = Workbook()
    ws = wb.active

    if type == "students":
        ws.title = "students"
        headers = [
            "document","first_name","last_name","email","phone",
            "username","password",
            "codigoEstudiante","career_id","plan_id","cicloActual","turno","seccion",
            "periodoIngreso","estado",
            "sexo","fechaNacimiento",
            "departamento","provincia","distrito","direccion",
            "apoderadoNombre","apoderadoDni","apoderadoTelefono"
        ]
        ws.append(headers)
        ws.append([
            "12345678","Jesus Agustin","Pinillos Torres","awa0@gmail.com","999999999",
            "", "",
            "A0001","1","1","1","Mañana","A",
            "2026-I","activo",
            "M","2005-03-15",
            "Junín","Huancayo","Huancayo","Av. Prueba 123",
            "Juan Pinillos","87654321","988888888"
        ])
        return _xlsx_response(wb, "students_template.xlsx")

    if type == "courses":
        ws.title = "courses"
        headers = ["code","name","credits","hours","plan_id","semester","type"]
        ws.append(headers)
        ws.append(["SIS101","Introducción a Sistemas","3","3","1","1","MANDATORY"])
        return _xlsx_response(wb, "courses_template.xlsx")

    if type == "grades":
        ws.title = "grades"
        headers = ["student_document","course_code","term","final_grade","PC1","PC2","EP","EF"]
        ws.append(headers)
        ws.append(["12345678","SIS101","2026-I","15","14","16","15","15"])
        return _xlsx_response(wb, "grades_template.xlsx")

    return Response({"detail": "Tipo inválido"}, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def imports_start(request, type: str):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    file = request.FILES.get("file")
    mapping_raw = request.POST.get("mapping")
    mapping = {}

    if mapping_raw:
        try:
            mapping = json.loads(mapping_raw)
        except Exception:
            mapping = {}

    if not file:
        return Response({"detail": "file requerido"}, status=400)

    job = ImportJob.objects.create(type=type, file=file, mapping=mapping, status="RUNNING", result={})

    errors = []
    imported = 0
    updated = 0
    credentials = []

    try:
        rows = _read_rows(file, mapping)

        with transaction.atomic():
            if type == "students":
                for row in rows:
                    r = row.get("__row__", "?")

                    document = str(row.get("document", "")).strip()
                    first_name = str(row.get("first_name", "")).strip()
                    last_name = str(row.get("last_name", "")).strip()
                    email = str(row.get("email", "")).strip()
                    phone = str(row.get("phone", "")).strip()

                    username = str(row.get("username", "")).strip()
                    password = str(row.get("password", "")).strip()

                    if not document or not first_name or not last_name:
                        errors.append(f"Fila {r}: document/first_name/last_name requeridos")
                        continue

                    codigo = str(row.get("codigoEstudiante", "")).strip() or f"A-{document}"

                    if Student.objects.filter(codigo_estudiante=codigo).exclude(dni=document).exists():
                        errors.append(f"Fila {r}: codigoEstudiante '{codigo}' ya existe en otro alumno")
                        continue

                    career_id = str(row.get("career_id", "")).strip()
                    plan_id = str(row.get("plan_id", "")).strip()

                    if plan_id and not Plan.objects.filter(id=_to_int(plan_id, -1)).exists():
                        errors.append(f"Fila {r}: plan_id {plan_id} no existe")
                        continue
                    if career_id and not Career.objects.filter(id=_to_int(career_id, -1)).exists():
                        errors.append(f"Fila {r}: career_id {career_id} no existe")
                        continue

                    ciclo = _to_int(row.get("cicloActual"), None)
                    turno = str(row.get("turno", "")).strip()
                    seccion = str(row.get("seccion", "")).strip()
                    periodo_ingreso = str(row.get("periodoIngreso", "")).strip()
                    estado = str(row.get("estado", "")).strip() or "activo"

                    sexo = str(row.get("sexo", "")).strip()
                    fecha_nac = _parse_date_yyyy_mm_dd(str(row.get("fechaNacimiento", "")).strip())

                    departamento = str(row.get("departamento", "")).strip()
                    provincia = str(row.get("provincia", "")).strip()
                    distrito = str(row.get("distrito", "")).strip()
                    direccion = str(row.get("direccion", "")).strip()

                    ap_nombre = str(row.get("apoderadoNombre", "")).strip()
                    ap_dni = str(row.get("apoderadoDni", "")).strip()
                    ap_tel = str(row.get("apoderadoTelefono", "")).strip()

                    # ✅ crear/actualizar user + rol student
                    user, user_created, temp_password = _get_or_create_student_user(
                        document=document,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        username=username,
                        password=password,
                    )

                    # ✅ upsert student por DNI
                    st, created = Student.objects.get_or_create(
                        dni=document,
                        defaults={
                            "codigo_estudiante": codigo,
                            "nombres": first_name,
                            "apellidos": last_name,
                        }
                    )

                    if st.user_id and st.user_id != user.id:
                        errors.append(f"Fila {r}: DNI {document} ya está enlazado a otro usuario (user_id={st.user_id})")
                        continue

                    st.user = user
                    st.codigo_estudiante = codigo
                    st.nombres = first_name
                    st.apellidos = last_name
                    st.email = (email or "").strip().lower()
                    st.celular = phone

                    st.programa_id = str(plan_id or career_id or st.programa_id or "")
                    st.ciclo_actual = ciclo
                    st.turno = turno
                    st.seccion = seccion
                    st.periodo_ingreso = periodo_ingreso
                    st.estado = estado

                    st.sexo = sexo
                    if fecha_nac:
                        st.fecha_nacimiento = fecha_nac

                    st.departamento = departamento
                    st.provincia = provincia
                    st.distrito = distrito
                    st.direccion = direccion

                    st.apoderado_nombre = ap_nombre
                    st.apoderado_dni = ap_dni
                    st.apoderado_telefono = ap_tel

                    st.save()

                    if created:
                        imported += 1
                    else:
                        updated += 1

                    if user_created or temp_password:
                        credentials.append({
                            "row": r,
                            "dni": document,
                            "username": user.username,
                            "password": temp_password or "(sin cambio)",
                        })

            elif type == "courses":
                for row in rows:
                    r = row.get("__row__", "?")

                    code = str(row.get("code", "")).strip()
                    name = str(row.get("name", "")).strip()
                    credits = _to_int(row.get("credits"), None)
                    hours = _to_int(row.get("hours"), None)

                    if not code or not name:
                        errors.append(f"Fila {r}: code y name requeridos")
                        continue

                    course, created = Course.objects.get_or_create(
                        code=code,
                        defaults={"name": name, "credits": max(0, credits or 0)},
                    )
                    course.name = name
                    if credits is not None:
                        course.credits = max(0, credits)
                    course.save()

                    plan_id = _to_int(row.get("plan_id"), None)
                    semester = _to_int(row.get("semester"), None)
                    ctype = str(row.get("type", "")).strip().upper()

                    if plan_id and semester:
                        plan = Plan.objects.filter(id=plan_id).first()
                        if not plan:
                            errors.append(f"Fila {r}: plan_id {plan_id} no existe")
                            continue

                        if ctype in ("OBLIGATORIO", "MANDATORY", "M"):
                            type_db = "MANDATORY"
                        elif ctype in ("ELECTIVO", "ELECTIVE", "E"):
                            type_db = "ELECTIVE"
                        else:
                            type_db = "MANDATORY"

                        pc, pc_created = PlanCourse.objects.get_or_create(
                            plan=plan,
                            course=course,
                            defaults={
                                "semester": max(1, semester),
                                "weekly_hours": max(1, hours or 3),
                                "type": type_db,
                            },
                        )
                        if not pc_created:
                            pc.semester = max(1, semester)
                            pc.weekly_hours = max(1, hours or pc.weekly_hours)
                            pc.type = type_db
                            pc.save()

                    imported += 1

            elif type == "grades":
                if AcademicGradeRecord is None:
                    return Response({"detail": "AcademicGradeRecord no está implementado. Agrega el modelo y migra."}, status=400)

                for row in rows:
                    r = row.get("__row__", "?")

                    doc = str(row.get("student_document", "")).strip()
                    course_code = str(row.get("course_code", "")).strip()
                    term = str(row.get("term", "")).strip()
                    final = _to_float(row.get("final_grade", None), None)

                    if not doc or not course_code or not term or final is None:
                        errors.append(f"Fila {r}: student_document, course_code, term, final_grade requeridos")
                        continue

                    st = Student.objects.filter(dni=doc).first()
                    if not st:
                        errors.append(f"Fila {r}: no existe alumno con DNI {doc}")
                        continue

                    course = Course.objects.filter(code=course_code).first()
                    if not course:
                        errors.append(f"Fila {r}: no existe curso con code {course_code}")
                        continue

                    components = {}
                    for k in ["PC1", "PC2", "EP", "EF"]:
                        v = row.get(k)
                        if v not in (None, ""):
                            components[k] = _to_float(v, v)

                    rec, created = AcademicGradeRecord.objects.get_or_create(
                        student=st,
                        course=course,
                        term=term,
                        defaults={"final_grade": final, "components": components},
                    )
                    if not created:
                        rec.final_grade = final
                        rec.components = components
                        rec.save()

                    imported += 1

            else:
                return Response({"detail": "Tipo inválido"}, status=400)

        job.status = "COMPLETED"
        job.result = {
            "imported": imported,
            "updated": updated,
            "errors": errors,
            "credentials": credentials[:300],
        }
        job.save(update_fields=["status", "result"])
        return Response({"job_id": job.id})

    except Exception as e:
        job.status = "FAILED"
        job.result = {"imported": imported, "updated": updated, "errors": errors + [str(e)]}
        job.save(update_fields=["status", "result"])
        return Response({"job_id": job.id, "detail": "Import failed"}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def imports_status(request, jobId: int):
    try:
        job = ImportJob.objects.get(pk=jobId)
    except ImportJob.DoesNotExist:
        return Response({"detail": "job not found"}, status=404)

    result = job.result if isinstance(job.result, dict) else {}
    return Response({
        "id": job.id,
        "state": job.status,
        "progress": 100 if job.status in ("COMPLETED", "FAILED") else 0,
        "errors": result.get("errors") or [],
        "imported": result.get("imported", 0),
        "updated": result.get("updated", 0),
        "credentials": result.get("credentials") or [],
    })


# ------------------ Respaldo / Export (REAL) ------------------

def _zip_add_folder(zf: zipfile.ZipFile, folder_path: str, arc_prefix: str):
    folder_path = os.path.abspath(folder_path)
    if not os.path.exists(folder_path):
        return
    for root, _, files in os.walk(folder_path):
        for f in files:
            fp = os.path.join(root, f)
            rel = os.path.relpath(fp, folder_path).replace("\\", "/")
            zf.write(fp, f"{arc_prefix}/{rel}")


def _dumpdata_json_bytes():
    buf = io.StringIO()
    call_command("dumpdata", "--natural-foreign", "--natural-primary", "--indent", "2", stdout=buf)
    return buf.getvalue().encode("utf-8")


def _try_add_sqlite_db(zf: zipfile.ZipFile):
    try:
        db = settings.DATABASES.get("default", {})
        if db.get("ENGINE") == "django.db.backends.sqlite3":
            name = db.get("NAME")
            if name and os.path.exists(name):
                zf.write(str(name), "db/db.sqlite3")
    except Exception:
        pass


def _zip_add_folder(zf: zipfile.ZipFile, folder_path: str, arc_prefix: str):
    folder_path = os.path.abspath(folder_path)
    if not os.path.exists(folder_path):
        return
    for root, _, files in os.walk(folder_path):
        for f in files:
            fp = os.path.join(root, f)
            rel = os.path.relpath(fp, folder_path).replace("\\", "/")
            zf.write(fp, f"{arc_prefix}/{rel}")


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def backups_collection(request):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    if request.method == "GET":
        qs = BackupExport.objects.all().order_by("-created_at")
        return Response({"items": BackupExportSerializer(qs, many=True).data})

    scope = (request.data.get("scope") or "FULL").upper().strip()
    if scope not in ("FULL", "DATA_ONLY", "FILES_ONLY"):
        return Response({"detail": "scope inválido (FULL|DATA_ONLY|FILES_ONLY)"}, status=400)

    obj = BackupExport.objects.create(scope=scope)

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"backup_{scope.lower()}_{now}.zip"

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # DATA
        if scope in ("FULL", "DATA_ONLY"):
            zf.writestr("data/dumpdata.json", _dumpdata_json_bytes())
            _try_add_sqlite_db(zf)

        # FILES (MEDIA)
        if scope in ("FULL", "FILES_ONLY"):
            media_root = getattr(settings, "MEDIA_ROOT", None)
            if media_root:
                _zip_add_folder(zf, str(media_root), "media")

        zf.writestr("meta/info.txt", f"scope={scope}\ncreated_at={now}\n")

    zbuf.seek(0)
    obj.file.save(zip_name, ContentFile(zbuf.getvalue()))
    obj.save(update_fields=["file"])

    return Response({
        "id": obj.id,
        "scope": obj.scope,
        "file_url": obj.file.url if obj.file else None,
    }, status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def backup_download(request, id: int):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    try:
        b = BackupExport.objects.get(pk=id)
    except BackupExport.DoesNotExist:
        raise Http404

    if not b.file:
        raise Http404

    return FileResponse(b.file.open("rb"), as_attachment=True, filename=os.path.basename(b.file.name))

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def export_dataset(request):
    not_ok = _require_staff(request)
    if not_ok:
        return not_ok

    dataset = (request.data.get("dataset") or "DATA").upper().strip()
    now = datetime.now().strftime("%Y%m%d_%H%M%S")

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:

        if dataset == "STUDENTS":
            rows = Student.objects.all().values(
                "dni", "codigo_estudiante", "nombres", "apellidos",
                "email", "celular",
                "departamento", "provincia", "distrito", "direccion",
                "programa_id", "ciclo_actual", "turno", "seccion",
                "periodo_ingreso", "estado"
            )
            headers = [
                "dni","codigo_estudiante","nombres","apellidos",
                "email","celular",
                "departamento","provincia","distrito","direccion",
                "programa_id","ciclo_actual","turno","seccion",
                "periodo_ingreso","estado"
            ]
            zf.writestr(f"students_{now}.csv", _csv_bytes(list(rows), headers))

        elif dataset == "COURSES":
            rows = Course.objects.all().values("code", "name", "credits")
            headers = ["code","name","credits"]
            zf.writestr(f"courses_{now}.csv", _csv_bytes(list(rows), headers))

            pc_rows = PlanCourse.objects.all().values("plan_id", "course_id", "semester", "weekly_hours", "type")
            pc_headers = ["plan_id","course_id","semester","weekly_hours","type"]
            zf.writestr(f"plan_courses_{now}.csv", _csv_bytes(list(pc_rows), pc_headers))

        elif dataset == "GRADES":
            if AcademicGradeRecord is None:
                return Response({"detail": "AcademicGradeRecord no existe. Agrégalo y migra."}, status=400)

            rows = AcademicGradeRecord.objects.select_related("student", "course").all().values(
                "student__dni", "course__code", "term", "final_grade", "components"
            )
            headers = ["student__dni","course__code","term","final_grade","components"]
            zf.writestr(f"grades_{now}.csv", _csv_bytes(list(rows), headers))

        elif dataset == "CATALOGS":
            p_rows = Period.objects.all().values("code", "year", "term", "start_date", "end_date", "is_active")
            p_headers = ["code","year","term","start_date","end_date","is_active"]
            zf.writestr(f"catalog_periods_{now}.csv", _csv_bytes(list(p_rows), p_headers))

            c_rows = Campus.objects.all().values("code", "name", "address")
            c_headers = ["code","name","address"]
            zf.writestr(f"catalog_campuses_{now}.csv", _csv_bytes(list(c_rows), c_headers))

            a_rows = Classroom.objects.all().values("campus_id", "code", "name", "capacity")
            a_headers = ["campus_id","code","name","capacity"]
            zf.writestr(f"catalog_classrooms_{now}.csv", _csv_bytes(list(a_rows), a_headers))

            t_rows = Teacher.objects.all().values("document", "full_name", "email", "phone", "specialization")
            t_headers = ["document","full_name","email","phone","specialization"]
            zf.writestr(f"catalog_teachers_{now}.csv", _csv_bytes(list(t_rows), t_headers))

            ac_rows = Career.objects.all().values("id", "name")
            ac_headers = ["id","name"]
            zf.writestr(f"academic_careers_{now}.csv", _csv_bytes(list(ac_rows), ac_headers))

            pl_rows = Plan.objects.all().values("id", "career_id", "name", "start_year", "semesters")
            pl_headers = ["id","career_id","name","start_year","semesters"]
            zf.writestr(f"academic_plans_{now}.csv", _csv_bytes(list(pl_rows), pl_headers))

        else:
            return Response({"detail": "dataset inválido"}, status=400)

        zf.writestr("meta/info.txt", f"dataset={dataset}\ncreated_at={now}\n")

    zbuf.seek(0)

    # ✅ se guarda como BackupExport para que aparezca en el historial
    obj = BackupExport.objects.create(scope=f"DATASET_{dataset}")
    filename = f"export_{dataset.lower()}_{now}.zip"
    obj.file.save(filename, ContentFile(zbuf.getvalue()))
    obj.save(update_fields=["file"])

    return Response({
        "ok": True,
        "dataset": dataset,
        "backup_id": obj.id,
        "download_url": f"/catalogs/exports/backups/{obj.id}/download",
    })