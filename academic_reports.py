"""
Academic Reports System
Comprehensive reporting for students, courses, and academic consistency
"""
import logging
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional, Tuple
from io import BytesIO
import csv
from motor.motor_asyncio import AsyncIOMotorDatabase
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class AcademicReports:
    """Academic reporting system with PDF/CSV export capabilities"""
    
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
        
    async def generate_student_history_report(
        self,
        student_id: str,
        include_periods: Optional[List[str]] = None,
        format_type: str = "PDF"
    ) -> BytesIO:
        """Generate comprehensive student academic history report"""
        
        try:
            # Obtener datos del estudiante
            student = await self.db.students.find_one({"id": student_id})
            if not student:
                raise ValueError(f"Estudiante {student_id} no encontrado")
            
            # Obtener matrículas
            enrollments_query = {"student_id": student_id}
            if include_periods:
                enrollments_query["period_id"] = {"$in": include_periods}
            
            enrollments_cursor = self.db.enrollments.find(enrollments_query).sort("period_id", 1)
            enrollments = await enrollments_cursor.to_list(length=None)
            
            # Obtener calificaciones
            grades_cursor = self.db.grades.find({"student_id": student_id}).sort("period_id", 1)
            grades = await grades_cursor.to_list(length=None)
            
            # Obtener información de cursos
            course_ids = list(set(e["course_id"] for e in enrollments))
            courses_cursor = self.db.courses.find({"id": {"$in": course_ids}})
            courses = {c["id"]: c for c in await courses_cursor.to_list(length=None)}
            
            # Calcular promedios y estadísticas
            statistics = await self._calculate_student_statistics(student_id, enrollments, grades)
            
            # Generar reporte
            if format_type.upper() == "PDF":
                return await self._generate_student_pdf_report(
                    student, enrollments, grades, courses, statistics
                )
            elif format_type.upper() == "CSV":
                return await self._generate_student_csv_report(
                    student, enrollments, grades, courses, statistics
                )
            else:
                raise ValueError(f"Formato {format_type} no soportado")
                
        except Exception as e:
            logger.error(f"Error generating student history report: {e}")
            raise
    
    async def _calculate_student_statistics(
        self,
        student_id: str,
        enrollments: List[Dict],
        grades: List[Dict]
    ) -> Dict[str, Any]:
        """Calculate student academic statistics"""
        
        # Agrupar por período
        periods = {}
        total_credits = 0
        total_weighted_grade = 0
        
        # Crear diccionario de calificaciones por curso y período
        grades_dict = {}
        for grade in grades:
            key = f"{grade['course_id']}:{grade['period_id']}"
            grades_dict[key] = grade
        
        for enrollment in enrollments:
            period_id = enrollment["period_id"]
            course_id = enrollment["course_id"]
            credits = enrollment.get("credits", 0)
            
            if period_id not in periods:
                periods[period_id] = {
                    "courses": [],
                    "total_credits": 0,
                    "weighted_sum": 0,
                    "approved_credits": 0,
                    "failed_credits": 0
                }
            
            # Buscar calificación correspondiente
            grade_key = f"{course_id}:{period_id}"
            grade_info = grades_dict.get(grade_key)
            
            course_data = {
                "course_id": course_id,
                "credits": credits,
                "numerical_grade": grade_info.get("numerical_grade") if grade_info else None,
                "literal_grade": grade_info.get("literal_grade") if grade_info else None,
                "status": grade_info.get("status") if grade_info else "PENDING"
            }
            
            periods[period_id]["courses"].append(course_data)
            periods[period_id]["total_credits"] += credits
            
            if grade_info and grade_info.get("numerical_grade"):
                numerical_grade = float(grade_info["numerical_grade"])
                periods[period_id]["weighted_sum"] += numerical_grade * credits
                
                if grade_info.get("status") == "APPROVED":
                    periods[period_id]["approved_credits"] += credits
                else:
                    periods[period_id]["failed_credits"] += credits
                
                total_credits += credits
                total_weighted_grade += numerical_grade * credits
        
        # Calcular promedios por período
        for period_data in periods.values():
            if period_data["total_credits"] > 0:
                period_data["period_average"] = period_data["weighted_sum"] / period_data["total_credits"]
            else:
                period_data["period_average"] = 0.0
        
        # Promedio acumulado
        cumulative_average = total_weighted_grade / total_credits if total_credits > 0 else 0.0
        
        return {
            "periods": periods,
            "cumulative_average": round(cumulative_average, 2),
            "total_credits_attempted": total_credits,
            "total_credits_approved": sum(p["approved_credits"] for p in periods.values()),
            "total_credits_failed": sum(p["failed_credits"] for p in periods.values()),
            "completion_rate": (sum(p["approved_credits"] for p in periods.values()) / total_credits * 100) if total_credits > 0 else 0.0
        }
    
    async def _generate_student_pdf_report(
        self,
        student: Dict,
        enrollments: List[Dict],
        grades: List[Dict],
        courses: Dict[str, Dict],
        statistics: Dict[str, Any]
    ) -> BytesIO:
        """Generate PDF report for student history"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkgreen
        )
        
        story = []
        
        # Título
        story.append(Paragraph("HISTORIAL ACADÉMICO COMPLETO", title_style))
        story.append(Spacer(1, 12))
        
        # Información del estudiante
        story.append(Paragraph("DATOS DEL ESTUDIANTE", header_style))
        
        student_data = [
            ["Código:", student.get("id", "")],
            ["Nombre completo:", student.get("full_name", "")],
            ["DNI:", student.get("dni", "")],
            ["Email:", student.get("email", "")],
            ["Programa:", student.get("program", "")],
            ["Estado:", student.get("status", "")]
        ]
        
        student_table = Table(student_data, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(student_table)
        story.append(Spacer(1, 20))
        
        # Resumen académico
        story.append(Paragraph("RESUMEN ACADÉMICO", header_style))
        
        summary_data = [
            ["Promedio ponderado acumulado:", f"{statistics['cumulative_average']:.2f}"],
            ["Créditos intentados:", str(statistics['total_credits_attempted'])],
            ["Créditos aprobados:", str(statistics['total_credits_approved'])],
            ["Créditos desaprobados:", str(statistics['total_credits_failed'])],
            ["Tasa de aprobación:", f"{statistics['completion_rate']:.1f}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Historial por períodos
        story.append(Paragraph("HISTORIAL POR PERÍODOS", header_style))
        
        for period_id in sorted(statistics['periods'].keys()):
            period_data = statistics['periods'][period_id]
            
            story.append(Paragraph(f"<b>Período: {period_id}</b>", styles['Heading3']))
            story.append(Paragraph(f"Promedio del período: {period_data['period_average']:.2f}", styles['Normal']))
            story.append(Spacer(1, 6))
            
            # Tabla de cursos del período
            course_data = [["Código", "Curso", "Créditos", "Nota", "Literal", "Estado"]]
            
            for course_info in period_data['courses']:
                course_id = course_info['course_id']
                course_name = courses.get(course_id, {}).get('name', 'Curso no encontrado')
                
                course_data.append([
                    course_id,
                    course_name,
                    str(course_info['credits']),
                    str(course_info['numerical_grade']) if course_info['numerical_grade'] else "-",
                    course_info['literal_grade'] or "-",
                    course_info['status']
                ])
            
            course_table = Table(course_data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 0.7*inch, 0.7*inch, 1*inch])
            course_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(course_table)
            story.append(Spacer(1, 15))
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    async def _generate_student_csv_report(
        self,
        student: Dict,
        enrollments: List[Dict],
        grades: List[Dict],
        courses: Dict[str, Dict],
        statistics: Dict[str, Any]
    ) -> BytesIO:
        """Generate CSV report for student history"""
        
        buffer = BytesIO()
        
        # Crear workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Académico"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Información del estudiante
        ws['A1'] = "HISTORIAL ACADÉMICO - " + student.get("full_name", "")
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Datos del estudiante
        row = 3
        student_info = [
            ("Código:", student.get("id", "")),
            ("DNI:", student.get("dni", "")),
            ("Email:", student.get("email", "")),
            ("Programa:", student.get("program", "")),
            ("Promedio Acumulado:", f"{statistics['cumulative_average']:.2f}"),
            ("Créditos Aprobados:", f"{statistics['total_credits_approved']}/{statistics['total_credits_attempted']}")
        ]
        
        for label, value in student_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Separador
        row += 2
        
        # Encabezados de tabla
        headers = ["Período", "Código Curso", "Nombre Curso", "Créditos", "Nota Numérica", "Nota Literal", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        
        # Datos de cursos
        grades_dict = {f"{g['course_id']}:{g['period_id']}": g for g in grades}
        
        for enrollment in sorted(enrollments, key=lambda x: x['period_id']):
            period_id = enrollment['period_id']
            course_id = enrollment['course_id']
            course_name = courses.get(course_id, {}).get('name', 'Curso no encontrado')
            
            grade_key = f"{course_id}:{period_id}"
            grade_info = grades_dict.get(grade_key, {})
            
            ws[f'A{row}'] = period_id
            ws[f'B{row}'] = course_id
            ws[f'C{row}'] = course_name
            ws[f'D{row}'] = enrollment.get('credits', 0)
            ws[f'E{row}'] = grade_info.get('numerical_grade', '-')
            ws[f'F{row}'] = grade_info.get('literal_grade', '-')
            ws[f'G{row}'] = grade_info.get('status', 'PENDING')
            
            row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def generate_course_outcomes_report(
        self,
        course_id: str,
        period_id: str,
        teacher_id: Optional[str] = None,
        format_type: str = "PDF"
    ) -> BytesIO:
        """Generate course outcomes report with grade distribution and attendance"""
        
        try:
            # Obtener información del curso
            course = await self.db.courses.find_one({"id": course_id})
            if not course:
                raise ValueError(f"Curso {course_id} no encontrado")
            
            # Obtener matrículas del curso en el período
            enrollments_query = {"course_id": course_id, "period_id": period_id}
            enrollments_cursor = self.db.enrollments.find(enrollments_query)
            enrollments = await enrollments_cursor.to_list(length=None)
            
            # Obtener calificaciones
            grades_cursor = self.db.grades.find({"course_id": course_id, "period_id": period_id})
            grades = await grades_cursor.to_list(length=None)
            
            # Obtener asistencias
            attendance_cursor = self.db.attendance.find({"course_id": course_id, "period_id": period_id})
            attendance_records = await attendance_cursor.to_list(length=None)
            
            # Obtener información de estudiantes
            student_ids = [e["student_id"] for e in enrollments]
            students_cursor = self.db.students.find({"id": {"$in": student_ids}})
            students = {s["id"]: s for s in await students_cursor.to_list(length=None)}
            
            # Analizar resultados
            analysis = await self._analyze_course_outcomes(
                course, enrollments, grades, attendance_records, students
            )
            
            if format_type.upper() == "PDF":
                return await self._generate_course_pdf_report(course, period_id, analysis)
            elif format_type.upper() == "CSV":
                return await self._generate_course_csv_report(course, period_id, analysis)
            else:
                raise ValueError(f"Formato {format_type} no soportado")
                
        except Exception as e:
            logger.error(f"Error generating course outcomes report: {e}")
            raise
    
    async def _analyze_course_outcomes(
        self,
        course: Dict,
        enrollments: List[Dict],
        grades: List[Dict],
        attendance_records: List[Dict],
        students: Dict[str, Dict]
    ) -> Dict[str, Any]:
        """Analyze course outcomes and generate statistics"""
        
        # Crear diccionarios de lookup
        grades_dict = {g["student_id"]: g for g in grades}
        
        # Calcular asistencia por estudiante
        attendance_dict = {}
        for record in attendance_records:
            student_id = record["student_id"]
            if student_id not in attendance_dict:
                attendance_dict[student_id] = {"total": 0, "present": 0}
            
            attendance_dict[student_id]["total"] += 1
            if record["status"] in ["PRESENT", "LATE"]:
                attendance_dict[student_id]["present"] += 1
        
        # Analizar cada estudiante
        student_results = []
        grade_distribution = {"AD": 0, "A": 0, "B": 0, "C": 0}
        status_distribution = {"APPROVED": 0, "FAILED": 0, "PENDING": 0}
        total_numerical_grades = []
        
        for enrollment in enrollments:
            student_id = enrollment["student_id"]
            student_info = students.get(student_id, {})
            grade_info = grades_dict.get(student_id, {})
            attendance_info = attendance_dict.get(student_id, {"total": 0, "present": 0})
            
            # Calcular porcentaje de asistencia
            attendance_percentage = 0
            if attendance_info["total"] > 0:
                attendance_percentage = (attendance_info["present"] / attendance_info["total"]) * 100
            
            result = {
                "student_id": student_id,
                "student_name": student_info.get("full_name", ""),
                "numerical_grade": grade_info.get("numerical_grade"),
                "literal_grade": grade_info.get("literal_grade"),
                "status": grade_info.get("status", "PENDING"),
                "attendance_percentage": round(attendance_percentage, 1),
                "total_classes": attendance_info["total"],
                "attended_classes": attendance_info["present"]
            }
            
            student_results.append(result)
            
            # Actualizar distribuciones
            if grade_info.get("literal_grade"):
                grade_distribution[grade_info["literal_grade"]] += 1
            
            status_distribution[result["status"]] += 1
            
            if grade_info.get("numerical_grade"):
                total_numerical_grades.append(float(grade_info["numerical_grade"]))
        
        # Calcular estadísticas generales
        total_students = len(enrollments)
        approved_count = status_distribution["APPROVED"]
        failed_count = status_distribution["FAILED"]
        
        approval_rate = (approved_count / total_students * 100) if total_students > 0 else 0
        failure_rate = (failed_count / total_students * 100) if total_students > 0 else 0
        
        # Estadísticas de notas
        grade_stats = {}
        if total_numerical_grades:
            grade_stats = {
                "average": round(sum(total_numerical_grades) / len(total_numerical_grades), 2),
                "min": min(total_numerical_grades),
                "max": max(total_numerical_grades),
                "count": len(total_numerical_grades)
            }
        
        # Estadísticas de asistencia
        attendance_percentages = [r["attendance_percentage"] for r in student_results if r["attendance_percentage"] > 0]
        attendance_stats = {}
        if attendance_percentages:
            attendance_stats = {
                "average": round(sum(attendance_percentages) / len(attendance_percentages), 1),
                "min": min(attendance_percentages),
                "max": max(attendance_percentages)
            }
        
        return {
            "student_results": student_results,
            "total_students": total_students,
            "approval_rate": round(approval_rate, 1),
            "failure_rate": round(failure_rate, 1),
            "grade_distribution": grade_distribution,
            "status_distribution": status_distribution,
            "grade_statistics": grade_stats,
            "attendance_statistics": attendance_stats
        }
    
    async def _generate_course_pdf_report(
        self,
        course: Dict,
        period_id: str,
        analysis: Dict[str, Any]
    ) -> BytesIO:
        """Generate PDF report for course outcomes"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        story = []
        
        # Título
        title = f"REPORTE DE RESULTADOS - {course.get('name', 'Curso')}"
        story.append(Paragraph(title, styles['Title']))
        story.append(Paragraph(f"Período: {period_id}", styles['Heading2']))
        story.append(Spacer(1, 20))
        
        # Resumen estadístico
        story.append(Paragraph("RESUMEN ESTADÍSTICO", styles['Heading2']))
        
        summary_data = [
            ["Total de estudiantes:", str(analysis['total_students'])],
            ["Tasa de aprobación:", f"{analysis['approval_rate']}%"],
            ["Tasa de desaprobación:", f"{analysis['failure_rate']}%"]
        ]
        
        if analysis['grade_statistics']:
            summary_data.extend([
                ["Promedio de notas:", str(analysis['grade_statistics']['average'])],
                ["Nota más alta:", str(analysis['grade_statistics']['max'])],
                ["Nota más baja:", str(analysis['grade_statistics']['min'])]
            ])
        
        if analysis['attendance_statistics']:
            summary_data.extend([
                ["Asistencia promedio:", f"{analysis['attendance_statistics']['average']}%"],
                ["Asistencia máxima:", f"{analysis['attendance_statistics']['max']}%"],
                ["Asistencia mínima:", f"{analysis['attendance_statistics']['min']}%"]
            ])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Distribución de calificaciones
        story.append(Paragraph("DISTRIBUCIÓN DE CALIFICACIONES", styles['Heading2']))
        
        dist_data = [["Calificación", "Cantidad", "Porcentaje"]]
        total = analysis['total_students']
        
        for grade, count in analysis['grade_distribution'].items():
            percentage = (count / total * 100) if total > 0 else 0
            dist_data.append([grade, str(count), f"{percentage:.1f}%"])
        
        dist_table = Table(dist_data)
        dist_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))
        
        story.append(dist_table)
        story.append(Spacer(1, 20))
        
        # Listado detallado de estudiantes
        story.append(Paragraph("DETALLE POR ESTUDIANTE", styles['Heading2']))
        
        student_data = [["Código", "Nombre", "Nota", "Literal", "Estado", "Asistencia"]]
        
        for result in sorted(analysis['student_results'], key=lambda x: x['student_name']):
            student_data.append([
                result['student_id'],
                result['student_name'],
                str(result['numerical_grade']) if result['numerical_grade'] else "-",
                result['literal_grade'] or "-",
                result['status'],
                f"{result['attendance_percentage']}%"
            ])
        
        student_table = Table(student_data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER')
        ]))
        
        story.append(student_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    async def _generate_course_csv_report(
        self,
        course: Dict,
        period_id: str,
        analysis: Dict[str, Any]
    ) -> BytesIO:
        """Generate CSV report for course outcomes"""
        
        buffer = BytesIO()
        
        # Crear workbook Excel
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        ws1['A1'] = f"REPORTE DE RESULTADOS - {course.get('name', '')}"
        ws1['A1'].font = Font(bold=True, size=14)
        
        ws1['A3'] = "Período:"
        ws1['B3'] = period_id
        
        # Estadísticas
        stats = [
            ("Total estudiantes:", analysis['total_students']),
            ("Tasa aprobación:", f"{analysis['approval_rate']}%"),
            ("Tasa desaprobación:", f"{analysis['failure_rate']}%")
        ]
        
        if analysis['grade_statistics']:
            stats.extend([
                ("Promedio notas:", analysis['grade_statistics']['average']),
                ("Nota máxima:", analysis['grade_statistics']['max']),
                ("Nota mínima:", analysis['grade_statistics']['min'])
            ])
        
        row = 5
        for label, value in stats:
            ws1[f'A{row}'] = label
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Hoja 2: Detalle estudiantes
        ws2 = wb.create_sheet("Detalle Estudiantes")
        
        headers = ["Código", "Nombre", "Nota Numérica", "Nota Literal", "Estado", "Asistencia %", "Clases Asistidas", "Total Clases"]
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row, result in enumerate(analysis['student_results'], 2):
            ws2[f'A{row}'] = result['student_id']
            ws2[f'B{row}'] = result['student_name']
            ws2[f'C{row}'] = result['numerical_grade'] if result['numerical_grade'] else '-'
            ws2[f'D{row}'] = result['literal_grade'] if result['literal_grade'] else '-'
            ws2[f'E{row}'] = result['status']
            ws2[f'F{row}'] = result['attendance_percentage']
            ws2[f'G{row}'] = result['attended_classes']
            ws2[f'H{row}'] = result['total_classes']
        
        # Ajustar columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def check_academic_consistency(
        self,
        period_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Check academic data consistency and generate anomaly report"""
        
        try:
            anomalies = []
            
            # 1. Verificar matrículas sin calificaciones
            enrollment_query = {}
            if period_id:
                enrollment_query["period_id"] = period_id
            
            enrollments_cursor = self.db.enrollments.find(enrollment_query)
            enrollments = await enrollments_cursor.to_list(length=None)
            
            grades_cursor = self.db.grades.find(enrollment_query if period_id else {})
            grades = await grades_cursor.to_list(length=None)
            
            # Crear sets para comparación rápida
            enrollment_keys = {f"{e['student_id']}:{e['course_id']}:{e['period_id']}" for e in enrollments}
            grade_keys = {f"{g['student_id']}:{g['course_id']}:{g['period_id']}" for g in grades}
            
            # Matrículas sin calificaciones
            missing_grades = enrollment_keys - grade_keys
            for key in missing_grades:
                student_id, course_id, period = key.split(":")
                anomalies.append({
                    "type": "MISSING_GRADE",
                    "description": f"Matrícula sin calificación: {student_id} en {course_id} ({period})",
                    "student_id": student_id,
                    "course_id": course_id,
                    "period_id": period,
                    "severity": "HIGH"
                })
            
            # Calificaciones sin matrícula
            orphan_grades = grade_keys - enrollment_keys
            for key in orphan_grades:
                student_id, course_id, period = key.split(":")
                anomalies.append({
                    "type": "ORPHAN_GRADE",
                    "description": f"Calificación sin matrícula: {student_id} en {course_id} ({period})",
                    "student_id": student_id,
                    "course_id": course_id,
                    "period_id": period,
                    "severity": "CRITICAL"
                })
            
            # 2. Verificar secciones con capacidad excedida
            sections_cursor = self.db.sections.find({})
            async for section in sections_cursor:
                section_enrollments = await self.db.enrollments.count_documents({
                    "section_id": section["id"],
                    "status": "ACTIVE"
                })
                
                if section_enrollments > section.get("capacity", 0):
                    anomalies.append({
                        "type": "SECTION_OVERCAPACITY", 
                        "description": f"Sección {section['id']} excede capacidad: {section_enrollments}/{section.get('capacity', 0)}",
                        "section_id": section["id"],
                        "current_enrollment": section_enrollments,
                        "capacity": section.get("capacity", 0),
                        "severity": "MEDIUM"
                    })
            
            # 3. Verificar conflictos de horario
            schedule_conflicts = await self._check_schedule_conflicts(period_id)
            anomalies.extend(schedule_conflicts)
            
            # 4. Verificar calificaciones inválidas
            invalid_grades = await self._check_invalid_grades(period_id)
            anomalies.extend(invalid_grades)
            
            # Clasificar anomalías por severidad
            severity_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
            for anomaly in anomalies:
                severity_counts[anomaly.get("severity", "LOW")] += 1
            
            result = {
                "check_date": datetime.now(timezone.utc).isoformat(),
                "period_id": period_id,
                "total_anomalies": len(anomalies),
                "severity_distribution": severity_counts,
                "anomalies": anomalies,
                "consistency_score": max(0, 100 - len(anomalies) * 2)  # Penalizar 2 puntos por anomalía
            }
            
            # Guardar resultado
            await self.db.consistency_checks.insert_one(result)
            
            return result
            
        except Exception as e:
            logger.error(f"Error in consistency check: {e}")
            raise
    
    async def _check_schedule_conflicts(self, period_id: Optional[str]) -> List[Dict[str, Any]]:
        """Check for schedule conflicts"""
        
        conflicts = []
        
        # Obtener horarios activos
        schedule_query = {"status": "ACTIVE"}
        if period_id:
            schedule_query["period_id"] = period_id
        
        schedules_cursor = self.db.schedules.find(schedule_query)
        schedules = await schedules_cursor.to_list(length=None)
        
        # Agrupar por día y verificar solapamientos
        by_day = {}
        for schedule in schedules:
            day = schedule.get("day_of_week")
            if day not in by_day:
                by_day[day] = []
            by_day[day].append(schedule)
        
        for day, day_schedules in by_day.items():
            # Ordenar por hora de inicio
            day_schedules.sort(key=lambda x: x.get("start_time", ""))
            
            for i in range(len(day_schedules) - 1):
                current = day_schedules[i]
                next_schedule = day_schedules[i + 1]
                
                # Verificar solapamiento (simplificado)
                if (current.get("end_time", "") > next_schedule.get("start_time", "") and
                    current.get("classroom") == next_schedule.get("classroom")):
                    
                    conflicts.append({
                        "type": "SCHEDULE_CONFLICT",
                        "description": f"Conflicto de horario en aula {current.get('classroom')} el {day}",
                        "schedule_1": current["id"],
                        "schedule_2": next_schedule["id"],
                        "classroom": current.get("classroom"),
                        "day": day,
                        "severity": "HIGH"
                    })
        
        return conflicts
    
    async def _check_invalid_grades(self, period_id: Optional[str]) -> List[Dict[str, Any]]:
        """Check for invalid grade values"""
        
        invalid_grades = []
        
        grade_query = {}
        if period_id:
            grade_query["period_id"] = period_id
        
        grades_cursor = self.db.grades.find(grade_query)
        
        async for grade in grades_cursor:
            numerical = grade.get("numerical_grade")
            literal = grade.get("literal_grade")
            
            # Verificar rango numérico
            if numerical is not None:
                if numerical < 0 or numerical > 20:
                    invalid_grades.append({
                        "type": "INVALID_NUMERICAL_GRADE",
                        "description": f"Nota numérica fuera de rango: {numerical} para {grade['student_id']}",
                        "student_id": grade["student_id"],
                        "course_id": grade["course_id"],
                        "grade_value": numerical,
                        "severity": "HIGH"
                    })
                
                # Verificar consistencia numérica-literal
                expected_literal = self._get_expected_literal_grade(numerical)
                if literal and literal != expected_literal:
                    invalid_grades.append({
                        "type": "GRADE_CONVERSION_MISMATCH",
                        "description": f"Inconsistencia numérica-literal: {numerical}->{literal} (esperado: {expected_literal})",
                        "student_id": grade["student_id"],
                        "course_id": grade["course_id"],
                        "numerical_grade": numerical,
                        "actual_literal": literal,
                        "expected_literal": expected_literal,
                        "severity": "MEDIUM"
                    })
        
        return invalid_grades
    
    def _get_expected_literal_grade(self, numerical_grade: float) -> str:
        """Get expected literal grade for numerical grade"""
        if numerical_grade >= 18.0:
            return "AD"
        elif numerical_grade >= 14.0:
            return "A"  
        elif numerical_grade >= 11.0:
            return "B"
        else:
            return "C"